"""
Пример использования парсера Excel файлов предложений страховых компаний
"""
import os
import tempfile
from openpyxl import Workbook
from .offer_excel_parser import parse_offer_file, validate_offer_data, OfferParsingError


def create_example_offer_file():
    """Создает пример файла предложения для демонстрации"""
    wb = Workbook()
    ws = wb.active
    
    # Устанавливаем название компании в объединенной ячейке A3:A5
    ws.merge_cells('A3:A5')
    ws['A3'] = 'Альфа Страхование'
    
    # Данные по годам страхования
    # Строка 3 - первый год
    ws['B3'] = '1 год'
    ws['C3'] = 1500000  # Страховая сумма
    ws['E3'] = 45000    # Премия с франшизой
    ws['F3'] = 50000    # Франшиза (вариант 1)
    ws['H3'] = 60000    # Премия без франшизы
    ws['I3'] = 0        # Франшиза (вариант 2)
    
    # Строка 4 - второй год
    ws['B4'] = '2 год'
    ws['C4'] = 1200000
    ws['E4'] = 36000
    ws['F4'] = 50000
    ws['H4'] = 48000
    ws['I4'] = 0
    
    # Строка 5 - третий год
    ws['B5'] = '3 год'
    ws['C5'] = 1000000
    ws['E5'] = 30000
    ws['F5'] = 50000
    ws['H5'] = 40000
    ws['I5'] = 0
    
    # Сохраняем файл
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    return temp_file.name


def main():
    """Демонстрация использования парсера"""
    print("=== Пример использования парсера Excel файлов предложений ===\n")
    
    # Создаем пример файла
    print("1. Создание примера файла предложения...")
    file_path = create_example_offer_file()
    print(f"   Файл создан: {file_path}\n")
    
    try:
        # Парсим файл
        print("2. Парсинг файла предложения...")
        offer_data = parse_offer_file(file_path)
        print(f"   ✓ Файл успешно обработан\n")
        
        # Выводим результаты
        print("3. Результаты парсинга:")
        print(f"   Компания: {offer_data['company_name']}")
        print(f"   Количество лет: {len(offer_data['years_data'])}\n")
        
        for i, year_data in enumerate(offer_data['years_data'], 1):
            print(f"   Год {i} ({year_data['year']}):")
            print(f"     Страховая сумма: {year_data['insurance_sum']:,} ₽" if year_data['insurance_sum'] else "     Страховая сумма: не указана")
            print(f"     Премия с франшизой: {year_data['premium_with_franchise']:,} ₽" if year_data['premium_with_franchise'] else "     Премия с франшизой: не указана")
            print(f"     Франшиза (вар. 1): {year_data['franchise_variant1']:,} ₽" if year_data['franchise_variant1'] else "     Франшиза (вар. 1): не указана")
            print(f"     Премия без франшизы: {year_data['premium_without_franchise']:,} ₽" if year_data['premium_without_franchise'] else "     Премия без франшизы: не указана")
            print(f"     Франшиза (вар. 2): {year_data['franchise_variant2']:,} ₽" if year_data['franchise_variant2'] else "     Франшиза (вар. 2): не указана")
            print()
        
        # Валидируем данные
        print("4. Валидация данных...")
        validated_data = validate_offer_data(offer_data)
        print("   ✓ Данные прошли валидацию\n")
        
        # Демонстрируем структуру данных для сохранения в базу
        print("5. Структура данных для сохранения в базу данных:")
        print("   Для каждого года будет создана отдельная запись InsuranceOffer:")
        
        for year_data in offer_data['years_data']:
            print(f"   - InsuranceOffer(")
            print(f"       company_name='{offer_data['company_name']}',")
            print(f"       insurance_year='{year_data['year']}',")
            print(f"       insurance_sum={year_data['insurance_sum']},")
            print(f"       yearly_premium_with_franchise={year_data['premium_with_franchise']},")
            print(f"       yearly_premium_without_franchise={year_data['premium_without_franchise']},")
            print(f"       franchise_amount_variant1={year_data['franchise_variant1']},")
            print(f"       franchise_amount_variant2={year_data['franchise_variant2']}")
            print(f"     )")
        
    except OfferParsingError as e:
        print(f"   ✗ Ошибка парсинга: {e}")
    except Exception as e:
        print(f"   ✗ Неожиданная ошибка: {e}")
    
    finally:
        # Удаляем временный файл
        try:
            os.unlink(file_path)
            print(f"\n6. Временный файл удален: {file_path}")
        except:
            pass


if __name__ == '__main__':
    main()