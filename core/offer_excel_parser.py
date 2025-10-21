"""
Парсер Excel файлов предложений страховых компаний
"""
from typing import Dict, Any, List, Optional, Union
import logging
from openpyxl import load_workbook
import pandas as pd
from decimal import Decimal, InvalidOperation

logger = logging.getLogger(__name__)


class OfferParsingError(Exception):
    """Базовый класс для ошибок парсинга предложений"""
    pass


class FileParsingError(OfferParsingError):
    """Ошибка парсинга файла"""
    pass


class DataValidationError(OfferParsingError):
    """Ошибка валидации данных"""
    pass


class OfferExcelParser:
    """
    Парсер Excel файлов предложений страховых компаний.
    
    Извлекает данные из специфических ячеек согласно требованиям:
    - A3:A5 (объединенная ячейка) - Название страховой компании
    - B3, B4, B5 - Года страхования ("1 год", "2 год", "3 год")
    - C3, C4, C5 - Страховые стоимости по годам
    - E3, E4, E5 - Страховые премии с франшизой
    - F3, F4, F5 - Размеры франшиз (вариант 1)
    - H3, H4, H5 - Страховые премии без франшизы
    - I3, I4, I5 - Размеры франшиз (вариант 2)
    """
    
    def __init__(self, file_path: str):
        """
        Инициализация парсера.
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self.df = None
        self.use_openpyxl = True
        
    def parse_offer(self) -> Dict[str, Any]:
        """
        Извлекает данные предложения из Excel файла.
        
        Returns:
            Словарь с данными предложения в формате:
            {
                'company_name': str,
                'years_data': [
                    {
                        'year': str,
                        'insurance_sum': Decimal or None,
                        'premium_with_franchise': Decimal or None,
                        'franchise_variant1': Decimal or None,
                        'premium_without_franchise': Decimal or None,
                        'franchise_variant2': Decimal or None
                    },
                    ...
                ]
            }
            
        Raises:
            FileParsingError: Если не удается прочитать файл
            DataValidationError: Если данные не прошли валидацию
        """
        logger.info(f"Starting to parse offer file: {self.file_path}")
        
        try:
            # Пробуем загрузить файл с openpyxl (для .xlsx)
            self._load_file_openpyxl()
            logger.info(f"Successfully loaded file with openpyxl: {self.file_path}")
        except Exception as openpyxl_error:
            logger.warning(f"openpyxl failed for {self.file_path}, trying pandas: {str(openpyxl_error)}")
            try:
                # Если не получилось, пробуем pandas (для .xls)
                self._load_file_pandas()
                self.use_openpyxl = False
                logger.info(f"Successfully loaded file with pandas: {self.file_path}")
            except Exception as pandas_error:
                error_msg = f"Failed to load file {self.file_path}. openpyxl: {str(openpyxl_error)}, pandas: {str(pandas_error)}"
                logger.error(error_msg)
                raise FileParsingError(error_msg)
        
        try:
            # Извлекаем данные
            if self.use_openpyxl:
                data = self._extract_data_openpyxl()
            else:
                data = self._extract_data_pandas()
            
            # Валидируем извлеченные данные
            self._validate_extracted_data(data)
            
            logger.info(f"Successfully parsed offer from {self.file_path}: company='{data['company_name']}', years={len(data['years_data'])}")
            return data
            
        except Exception as e:
            error_msg = f"Error extracting data from {self.file_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise DataValidationError(error_msg)
    
    def _load_file_openpyxl(self):
        """Загружает файл с помощью openpyxl"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            self.sheet = self.workbook.active
        except Exception as e:
            raise FileParsingError(f"Cannot load file with openpyxl: {str(e)}")
    
    def _load_file_pandas(self):
        """Загружает файл с помощью pandas"""
        try:
            self.df = pd.read_excel(self.file_path, sheet_name=0, header=None)
        except Exception as e:
            raise FileParsingError(f"Cannot load file with pandas: {str(e)}")
    
    def _extract_data_openpyxl(self) -> Dict[str, Any]:
        """Извлекает данные используя openpyxl"""
        logger.debug(f"Extracting data with openpyxl from {self.file_path}")
        
        # Извлекаем название компании
        company_name = self._extract_company_name_openpyxl()
        
        # Извлекаем данные по годам
        years_data = self._extract_yearly_data_openpyxl()
        
        return {
            'company_name': company_name,
            'years_data': years_data
        }
    
    def _extract_data_pandas(self) -> Dict[str, Any]:
        """Извлекает данные используя pandas"""
        logger.debug(f"Extracting data with pandas from {self.file_path}")
        
        # Извлекаем название компании
        company_name = self._extract_company_name_pandas()
        
        # Извлекаем данные по годам
        years_data = self._extract_yearly_data_pandas()
        
        return {
            'company_name': company_name,
            'years_data': years_data
        }
    
    def _extract_company_name_openpyxl(self) -> str:
        """
        Извлекает название компании из объединенной ячейки A3:A5 (openpyxl).
        
        Returns:
            Название компании или 'Компания не указана' если не найдено
        """
        try:
            # Проверяем объединенные ячейки
            for merged_range in self.sheet.merged_cells.ranges:
                if merged_range.min_row <= 3 <= merged_range.max_row and merged_range.min_col == 1:
                    # Получаем значение из первой ячейки объединенного диапазона
                    cell_value = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    if cell_value:
                        company_name = str(cell_value).strip()
                        logger.debug(f"Found company name in merged cell A{merged_range.min_row}:A{merged_range.max_row}: '{company_name}'")
                        return company_name
            
            # Если объединенных ячеек нет, проверяем A3, A4, A5 по отдельности
            for row in [3, 4, 5]:
                cell_value = self.sheet.cell(row=row, column=1).value  # Столбец A = 1
                if cell_value:
                    company_name = str(cell_value).strip()
                    logger.debug(f"Found company name in cell A{row}: '{company_name}'")
                    return company_name
            
            logger.warning(f"Company name not found in A3:A5 range in {self.file_path}")
            return 'Компания не указана'
            
        except Exception as e:
            logger.error(f"Error extracting company name with openpyxl from {self.file_path}: {str(e)}")
            return 'Компания не указана'
    
    def _extract_company_name_pandas(self) -> str:
        """
        Извлекает название компании из ячеек A3:A5 (pandas).
        
        Returns:
            Название компании или 'Компания не указана' если не найдено
        """
        try:
            # Проверяем ячейки A3, A4, A5 (индексы 2, 3, 4 для pandas)
            for row_idx in [2, 3, 4]:  # A3, A4, A5
                if row_idx < len(self.df):
                    cell_value = self.df.iloc[row_idx, 0]  # Столбец A = 0
                    if pd.notna(cell_value) and str(cell_value).strip():
                        company_name = str(cell_value).strip()
                        logger.debug(f"Found company name in cell A{row_idx + 1}: '{company_name}'")
                        return company_name
            
            logger.warning(f"Company name not found in A3:A5 range in {self.file_path}")
            return 'Компания не указана'
            
        except Exception as e:
            logger.error(f"Error extracting company name with pandas from {self.file_path}: {str(e)}")
            return 'Компания не указана'
    
    def _extract_yearly_data_openpyxl(self) -> List[Dict[str, Any]]:
        """
        Извлекает данные по годам страхования (openpyxl).
        
        Returns:
            Список словарей с данными по каждому году
        """
        years_data = []
        
        # Обрабатываем строки 3, 4, 5 (соответствуют 1, 2, 3 годам)
        for row in [3, 4, 5]:
            try:
                year_data = self._extract_year_row_openpyxl(row)
                if year_data:  # Добавляем только если есть хотя бы какие-то данные
                    years_data.append(year_data)
            except Exception as e:
                logger.warning(f"Error extracting data for row {row} in {self.file_path}: {str(e)}")
                continue
        
        logger.debug(f"Extracted {len(years_data)} years of data with openpyxl from {self.file_path}")
        return years_data
    
    def _extract_yearly_data_pandas(self) -> List[Dict[str, Any]]:
        """
        Извлекает данные по годам страхования (pandas).
        
        Returns:
            Список словарей с данными по каждому году
        """
        years_data = []
        
        # Обрабатываем строки 3, 4, 5 (индексы 2, 3, 4 для pandas)
        for row_idx in [2, 3, 4]:  # Строки 3, 4, 5
            try:
                year_data = self._extract_year_row_pandas(row_idx)
                if year_data:  # Добавляем только если есть хотя бы какие-то данные
                    years_data.append(year_data)
            except Exception as e:
                logger.warning(f"Error extracting data for row {row_idx + 1} in {self.file_path}: {str(e)}")
                continue
        
        logger.debug(f"Extracted {len(years_data)} years of data with pandas from {self.file_path}")
        return years_data
    
    def _extract_year_row_openpyxl(self, row: int) -> Optional[Dict[str, Any]]:
        """
        Извлекает данные для одного года из указанной строки (openpyxl).
        
        Args:
            row: Номер строки (3, 4, или 5)
            
        Returns:
            Словарь с данными года или None если нет данных
        """
        # Извлекаем значения из соответствующих столбцов
        year = self._get_cell_value_openpyxl(row, 2)  # B = 2
        insurance_sum = self._get_cell_value_openpyxl(row, 3)  # C = 3
        premium_with_franchise = self._get_cell_value_openpyxl(row, 5)  # E = 5
        franchise_variant1 = self._get_cell_value_openpyxl(row, 6)  # F = 6
        premium_without_franchise = self._get_cell_value_openpyxl(row, 8)  # H = 8
        franchise_variant2 = self._get_cell_value_openpyxl(row, 9)  # I = 9
        
        # Проверяем, есть ли хотя бы какие-то данные в строке
        if not any([year, insurance_sum, premium_with_franchise, premium_without_franchise]):
            return None
        
        # Определяем год страхования
        year_str = self._normalize_year_string(year, row)
        
        year_data = {
            'year': year_str,
            'insurance_sum': self._parse_decimal_value(insurance_sum),
            'premium_with_franchise': self._parse_decimal_value(premium_with_franchise),
            'franchise_variant1': self._parse_decimal_value(franchise_variant1),
            'premium_without_franchise': self._parse_decimal_value(premium_without_franchise),
            'franchise_variant2': self._parse_decimal_value(franchise_variant2)
        }
        
        logger.debug(f"Extracted year data for row {row}: {year_data}")
        return year_data
    
    def _extract_year_row_pandas(self, row_idx: int) -> Optional[Dict[str, Any]]:
        """
        Извлекает данные для одного года из указанной строки (pandas).
        
        Args:
            row_idx: Индекс строки (2, 3, или 4 для строк 3, 4, 5)
            
        Returns:
            Словарь с данными года или None если нет данных
        """
        if row_idx >= len(self.df):
            return None
        
        # Извлекаем значения из соответствующих столбцов
        year = self._get_cell_value_pandas(row_idx, 1)  # B = 1
        insurance_sum = self._get_cell_value_pandas(row_idx, 2)  # C = 2
        premium_with_franchise = self._get_cell_value_pandas(row_idx, 4)  # E = 4
        franchise_variant1 = self._get_cell_value_pandas(row_idx, 5)  # F = 5
        premium_without_franchise = self._get_cell_value_pandas(row_idx, 7)  # H = 7
        franchise_variant2 = self._get_cell_value_pandas(row_idx, 8)  # I = 8
        
        # Проверяем, есть ли хотя бы какие-то данные в строке
        if not any([year, insurance_sum, premium_with_franchise, premium_without_franchise]):
            return None
        
        # Определяем год страхования
        year_str = self._normalize_year_string(year, row_idx + 1)
        
        year_data = {
            'year': year_str,
            'insurance_sum': self._parse_decimal_value(insurance_sum),
            'premium_with_franchise': self._parse_decimal_value(premium_with_franchise),
            'franchise_variant1': self._parse_decimal_value(franchise_variant1),
            'premium_without_franchise': self._parse_decimal_value(premium_without_franchise),
            'franchise_variant2': self._parse_decimal_value(franchise_variant2)
        }
        
        logger.debug(f"Extracted year data for row {row_idx + 1}: {year_data}")
        return year_data
    
    def _get_cell_value_openpyxl(self, row: int, col: int) -> Any:
        """Получает значение ячейки с помощью openpyxl"""
        try:
            return self.sheet.cell(row=row, column=col).value
        except Exception as e:
            logger.debug(f"Error getting cell value at row {row}, col {col}: {str(e)}")
            return None
    
    def _get_cell_value_pandas(self, row_idx: int, col_idx: int) -> Any:
        """Получает значение ячейки с помощью pandas"""
        try:
            if row_idx < len(self.df) and col_idx < len(self.df.columns):
                value = self.df.iloc[row_idx, col_idx]
                return value if pd.notna(value) else None
            return None
        except Exception as e:
            logger.debug(f"Error getting cell value at row {row_idx}, col {col_idx}: {str(e)}")
            return None
    
    def _normalize_year_string(self, year_value: Any, row_number: int) -> str:
        """
        Нормализует строку года страхования.
        
        Args:
            year_value: Значение из ячейки
            row_number: Номер строки для определения года по умолчанию
            
        Returns:
            Нормализованная строка года ("1 год", "2 год", "3 год")
        """
        if year_value:
            year_str = str(year_value).strip().lower()
            # Пытаемся извлечь номер года из строки
            if '1' in year_str or 'первый' in year_str or 'один' in year_str:
                return '1 год'
            elif '2' in year_str or 'второй' in year_str or 'два' in year_str:
                return '2 год'
            elif '3' in year_str or 'третий' in year_str or 'три' in year_str:
                return '3 год'
        
        # Если не удалось определить из значения, используем номер строки
        year_mapping = {3: '1 год', 4: '2 год', 5: '3 год'}
        return year_mapping.get(row_number, f'{row_number - 2} год')
    
    def _parse_decimal_value(self, value: Any) -> Optional[Decimal]:
        """
        Парсит значение в Decimal с обработкой ошибок.
        
        Args:
            value: Значение для парсинга
            
        Returns:
            Decimal значение или None если парсинг не удался
        """
        if value is None:
            return None
        
        try:
            # Если это уже число
            if isinstance(value, (int, float)):
                return Decimal(str(value))  # Возвращаем даже 0
            
            # Если это строка
            if isinstance(value, str):
                # Убираем пробелы и заменяем запятые на точки
                cleaned_value = value.strip().replace(',', '.').replace(' ', '')
                
                # Убираем символы валют и другие нечисловые символы
                import re
                numeric_value = re.sub(r'[^\d.-]', '', cleaned_value)
                
                if not numeric_value or numeric_value in ['-', '.']:
                    return None
                
                decimal_value = Decimal(numeric_value)
                return decimal_value
            
            # Для других типов пытаемся преобразовать в строку и парсить
            return self._parse_decimal_value(str(value))
            
        except (InvalidOperation, ValueError, TypeError) as e:
            logger.debug(f"Cannot parse decimal value '{value}': {str(e)}")
            return None
    
    def _validate_extracted_data(self, data: Dict[str, Any]):
        """
        Валидирует извлеченные данные.
        
        Args:
            data: Словарь с извлеченными данными
            
        Raises:
            DataValidationError: Если данные не прошли валидацию
        """
        if not isinstance(data, dict):
            raise DataValidationError("Extracted data must be a dictionary")
        
        if 'company_name' not in data:
            raise DataValidationError("Company name is required")
        
        if not data['company_name'] or data['company_name'] == 'Компания не указана':
            logger.warning(f"Company name not found in {self.file_path}")
        
        if 'years_data' not in data:
            raise DataValidationError("Years data is required")
        
        if not isinstance(data['years_data'], list):
            raise DataValidationError("Years data must be a list")
        
        if len(data['years_data']) == 0:
            logger.warning(f"No year data found in {self.file_path}")
        
        # Валидируем каждый год
        for i, year_data in enumerate(data['years_data']):
            if not isinstance(year_data, dict):
                raise DataValidationError(f"Year data {i} must be a dictionary")
            
            required_fields = ['year', 'insurance_sum', 'premium_with_franchise', 
                             'franchise_variant1', 'premium_without_franchise', 'franchise_variant2']
            
            for field in required_fields:
                if field not in year_data:
                    raise DataValidationError(f"Year data {i} missing field: {field}")
            
            # Проверяем, что есть хотя бы какие-то числовые данные
            numeric_fields = ['insurance_sum', 'premium_with_franchise', 'premium_without_franchise']
            has_numeric_data = any(year_data.get(field) is not None for field in numeric_fields)
            
            if not has_numeric_data:
                logger.warning(f"Year {year_data.get('year', i)} has no numeric data in {self.file_path}")
        
        logger.debug(f"Data validation passed for {self.file_path}")


def parse_offer_file(file_path: str) -> Dict[str, Any]:
    """
    Удобная функция для парсинга файла предложения.
    
    Args:
        file_path: Путь к Excel файлу
        
    Returns:
        Словарь с данными предложения
        
    Raises:
        OfferParsingError: При ошибках парсинга
    """
    parser = OfferExcelParser(file_path)
    return parser.parse_offer()


def validate_offer_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Валидирует данные предложения.
    
    Args:
        data: Словарь с данными предложения
        
    Returns:
        Валидированные данные
        
    Raises:
        DataValidationError: При ошибках валидации
    """
    # Дополнительная валидация бизнес-логики
    if not data.get('company_name') or data['company_name'] == 'Компания не указана':
        raise DataValidationError("Название компании обязательно для заполнения")
    
    years_data = data.get('years_data', [])
    if not years_data:
        raise DataValidationError("Должны быть данные хотя бы по одному году страхования")
    
    # Проверяем, что есть хотя бы одно предложение с числовыми данными
    has_valid_offer = False
    for year_data in years_data:
        if (year_data.get('insurance_sum') or 
            year_data.get('premium_with_franchise') or 
            year_data.get('premium_without_franchise')):
            has_valid_offer = True
            break
    
    if not has_valid_offer:
        raise DataValidationError("Должно быть хотя бы одно предложение с числовыми данными")
    
    # Проверяем логическую связность данных
    for year_data in years_data:
        insurance_sum = year_data.get('insurance_sum')
        premium_with = year_data.get('premium_with_franchise')
        premium_without = year_data.get('premium_without_franchise')
        
        # Если есть и премия с франшизой, и без франшизы, то премия без франшизы должна быть больше
        if (premium_with and premium_without and 
            premium_with > premium_without):
            logger.warning(f"Премия с франшизой ({premium_with}) больше премии без франшизы ({premium_without}) для года {year_data.get('year')}")
        
        # Премия не должна превышать страховую сумму более чем в 2 раза (разумная проверка)
        if insurance_sum and premium_with and premium_with > insurance_sum * 2:
            logger.warning(f"Премия с франшизой ({premium_with}) подозрительно высока относительно страховой суммы ({insurance_sum}) для года {year_data.get('year')}")
        
        if insurance_sum and premium_without and premium_without > insurance_sum * 2:
            logger.warning(f"Премия без франшизы ({premium_without}) подозрительно высока относительно страховой суммы ({insurance_sum}) для года {year_data.get('year')}")
    
    logger.info(f"Offer data validation passed for company: {data['company_name']}")
    return data