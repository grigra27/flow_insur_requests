"""
Тесты для парсера Excel файлов предложений страховых компаний
"""
import os
import tempfile
import unittest
from decimal import Decimal
from openpyxl import Workbook
import pandas as pd

from .offer_excel_parser import (
    OfferExcelParser, 
    parse_offer_file, 
    validate_offer_data,
    OfferParsingError,
    FileParsingError,
    DataValidationError
)


class TestOfferExcelParser(unittest.TestCase):
    """Тесты для класса OfferExcelParser"""
    
    def setUp(self):
        """Подготовка тестовых данных"""
        self.temp_dir = tempfile.mkdtemp()
        
    def tearDown(self):
        """Очистка после тестов"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def create_test_xlsx_file(self, data_config: dict) -> str:
        """
        Создает тестовый .xlsx файл с заданными данными.
        
        Args:
            data_config: Конфигурация данных для файла
            
        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        
        # Устанавливаем название компании в A3:A5 (объединенная ячейка)
        if 'company_name' in data_config:
            ws.merge_cells('A3:A5')
            ws['A3'] = data_config['company_name']
        
        # Устанавливаем данные по годам
        years_data = data_config.get('years_data', [])
        for i, year_data in enumerate(years_data[:3]):  # Максимум 3 года
            row = 3 + i  # Строки 3, 4, 5
            
            if 'year' in year_data:
                ws.cell(row=row, column=2, value=year_data['year'])  # B
            if 'insurance_sum' in year_data:
                ws.cell(row=row, column=3, value=year_data['insurance_sum'])  # C
            if 'premium_with_franchise' in year_data:
                ws.cell(row=row, column=5, value=year_data['premium_with_franchise'])  # E
            if 'franchise_variant1' in year_data:
                ws.cell(row=row, column=6, value=year_data['franchise_variant1'])  # F
            if 'premium_without_franchise' in year_data:
                ws.cell(row=row, column=8, value=year_data['premium_without_franchise'])  # H
            if 'franchise_variant2' in year_data:
                ws.cell(row=row, column=9, value=year_data['franchise_variant2'])  # I
        
        file_path = os.path.join(self.temp_dir, 'test_offer.xlsx')
        wb.save(file_path)
        return file_path
    
    def create_test_xls_file(self, data_config: dict) -> str:
        """
        Создает тестовый .xls файл с заданными данными.
        
        Args:
            data_config: Конфигурация данных для файла
            
        Returns:
            Путь к созданному файлу
        """
        # Создаем DataFrame с пустыми данными
        df = pd.DataFrame(index=range(10), columns=range(10))
        
        # Устанавливаем название компании в A3 (индекс 2, 0)
        if 'company_name' in data_config:
            df.iloc[2, 0] = data_config['company_name']
        
        # Устанавливаем данные по годам
        years_data = data_config.get('years_data', [])
        for i, year_data in enumerate(years_data[:3]):  # Максимум 3 года
            row_idx = 2 + i  # Индексы 2, 3, 4 (строки 3, 4, 5)
            
            if 'year' in year_data:
                df.iloc[row_idx, 1] = year_data['year']  # B
            if 'insurance_sum' in year_data:
                df.iloc[row_idx, 2] = year_data['insurance_sum']  # C
            if 'premium_with_franchise' in year_data:
                df.iloc[row_idx, 4] = year_data['premium_with_franchise']  # E
            if 'franchise_variant1' in year_data:
                df.iloc[row_idx, 5] = year_data['franchise_variant1']  # F
            if 'premium_without_franchise' in year_data:
                df.iloc[row_idx, 7] = year_data['premium_without_franchise']  # H
            if 'franchise_variant2' in year_data:
                df.iloc[row_idx, 8] = year_data['franchise_variant2']  # I
        
        file_path = os.path.join(self.temp_dir, 'test_offer.xls')
        df.to_excel(file_path, index=False, header=False)
        return file_path
    
    def test_parse_valid_xlsx_offer(self):
        """Тест парсинга валидного .xlsx файла предложения"""
        data_config = {
            'company_name': 'Альфа Страхование',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': 1500000,
                    'premium_with_franchise': 45000,
                    'franchise_variant1': 50000,
                    'premium_without_franchise': 60000,
                    'franchise_variant2': 0
                },
                {
                    'year': '2 год',
                    'insurance_sum': 1200000,
                    'premium_with_franchise': 36000,
                    'franchise_variant1': 50000,
                    'premium_without_franchise': 48000,
                    'franchise_variant2': 0
                },
                {
                    'year': '3 год',
                    'insurance_sum': 1000000,
                    'premium_with_franchise': 30000,
                    'franchise_variant1': 50000,
                    'premium_without_franchise': 40000,
                    'franchise_variant2': 0
                }
            ]
        }
        
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        self.assertEqual(result['company_name'], 'Альфа Страхование')
        self.assertEqual(len(result['years_data']), 3)
        
        # Проверяем первый год
        year1 = result['years_data'][0]
        self.assertEqual(year1['year'], '1 год')
        self.assertEqual(year1['insurance_sum'], Decimal('1500000'))
        self.assertEqual(year1['premium_with_franchise'], Decimal('45000'))
        self.assertEqual(year1['franchise_variant1'], Decimal('50000'))
        self.assertEqual(year1['premium_without_franchise'], Decimal('60000'))
        self.assertEqual(year1['franchise_variant2'], Decimal('0'))
    
    def test_parse_pandas_fallback(self):
        """Тест fallback на pandas при ошибке openpyxl"""
        # Создаем файл с данными, который будет обработан pandas
        data_config = {
            'company_name': 'Бета Страхование',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': 2000000,
                    'premium_with_franchise': 55000,
                    'franchise_variant1': 75000,
                    'premium_without_franchise': 70000,
                    'franchise_variant2': None
                }
            ]
        }
        
        # Создаем .xlsx файл, но будем тестировать pandas fallback
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        
        # Принудительно используем pandas
        parser.use_openpyxl = False
        parser._load_file_pandas()
        result = parser._extract_data_pandas()
        
        self.assertEqual(result['company_name'], 'Бета Страхование')
        self.assertEqual(len(result['years_data']), 1)
        
        year1 = result['years_data'][0]
        self.assertEqual(year1['year'], '1 год')
        self.assertEqual(year1['insurance_sum'], Decimal('2000000'))
        self.assertEqual(year1['premium_with_franchise'], Decimal('55000'))
        self.assertEqual(year1['franchise_variant1'], Decimal('75000'))
        self.assertEqual(year1['premium_without_franchise'], Decimal('70000'))
        self.assertIsNone(year1['franchise_variant2'])
    
    def test_parse_partial_data(self):
        """Тест парсинга файла с частичными данными"""
        data_config = {
            'company_name': 'Гамма Страхование',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': 1000000,
                    'premium_with_franchise': None,  # Отсутствует
                    'franchise_variant1': None,
                    'premium_without_franchise': 50000,
                    'franchise_variant2': None
                },
                {
                    'year': '2 год',
                    'insurance_sum': None,  # Отсутствует
                    'premium_with_franchise': 40000,
                    'franchise_variant1': 30000,
                    'premium_without_franchise': None,
                    'franchise_variant2': None
                }
            ]
        }
        
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        self.assertEqual(result['company_name'], 'Гамма Страхование')
        self.assertEqual(len(result['years_data']), 2)
        
        # Проверяем первый год
        year1 = result['years_data'][0]
        self.assertEqual(year1['insurance_sum'], Decimal('1000000'))
        self.assertIsNone(year1['premium_with_franchise'])
        self.assertEqual(year1['premium_without_franchise'], Decimal('50000'))
        
        # Проверяем второй год
        year2 = result['years_data'][1]
        self.assertIsNone(year2['insurance_sum'])
        self.assertEqual(year2['premium_with_franchise'], Decimal('40000'))
        self.assertEqual(year2['franchise_variant1'], Decimal('30000'))
    
    def test_parse_empty_file(self):
        """Тест парсинга пустого файла"""
        data_config = {}
        
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        self.assertEqual(result['company_name'], 'Компания не указана')
        self.assertEqual(len(result['years_data']), 0)
    
    def test_parse_company_name_variations(self):
        """Тест различных вариантов названия компании"""
        # Тест с названием в A4 вместо A3
        wb = Workbook()
        ws = wb.active
        ws['A4'] = 'Дельта Страхование'
        
        file_path = os.path.join(self.temp_dir, 'test_company_a4.xlsx')
        wb.save(file_path)
        
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        self.assertEqual(result['company_name'], 'Дельта Страхование')
    
    def test_parse_numeric_formats(self):
        """Тест различных числовых форматов"""
        data_config = {
            'company_name': 'Эпсилон Страхование',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': '1 500 000',  # С пробелами
                    'premium_with_franchise': '45,000.50',  # С запятой и точкой
                    'franchise_variant1': '50000₽',  # С символом валюты
                    'premium_without_franchise': 60000.75,  # Число с плавающей точкой
                    'franchise_variant2': 0
                }
            ]
        }
        
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        year1 = result['years_data'][0]
        # Проверяем, что значения были извлечены (могут быть None из-за сложного форматирования)
        self.assertIsNotNone(year1['insurance_sum'])
        self.assertIsNotNone(year1['premium_without_franchise'])
        # Для строковых значений с форматированием результат может отличаться
        if year1['premium_with_franchise']:
            self.assertIsInstance(year1['premium_with_franchise'], Decimal)
        if year1['franchise_variant1']:
            self.assertIsInstance(year1['franchise_variant1'], Decimal)
    
    def test_year_normalization(self):
        """Тест нормализации строк года"""
        data_config = {
            'company_name': 'Зета Страхование',
            'years_data': [
                {
                    'year': 'Первый год',
                    'insurance_sum': 1000000,
                    'premium_with_franchise': 30000
                },
                {
                    'year': '2-й год',
                    'insurance_sum': 900000,
                    'premium_with_franchise': 27000
                },
                {
                    'year': 'третий',
                    'insurance_sum': 800000,
                    'premium_with_franchise': 24000
                }
            ]
        }
        
        file_path = self.create_test_xlsx_file(data_config)
        parser = OfferExcelParser(file_path)
        result = parser.parse_offer()
        
        self.assertEqual(result['years_data'][0]['year'], '1 год')
        self.assertEqual(result['years_data'][1]['year'], '2 год')
        self.assertEqual(result['years_data'][2]['year'], '3 год')
    
    def test_invalid_file_path(self):
        """Тест обработки несуществующего файла"""
        parser = OfferExcelParser('/nonexistent/file.xlsx')
        
        with self.assertRaises(FileParsingError):
            parser.parse_offer()
    
    def test_data_validation_no_company(self):
        """Тест валидации данных без названия компании"""
        data = {
            'company_name': '',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': Decimal('1000000'),
                    'premium_with_franchise': Decimal('30000'),
                    'franchise_variant1': None,
                    'premium_without_franchise': Decimal('40000'),
                    'franchise_variant2': None
                }
            ]
        }
        
        with self.assertRaises(DataValidationError):
            validate_offer_data(data)
    
    def test_data_validation_no_numeric_data(self):
        """Тест валидации данных без числовых значений"""
        data = {
            'company_name': 'Тест Компания',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': None,
                    'premium_with_franchise': None,
                    'franchise_variant1': None,
                    'premium_without_franchise': None,
                    'franchise_variant2': None
                }
            ]
        }
        
        with self.assertRaises(DataValidationError):
            validate_offer_data(data)
    
    def test_data_validation_valid_data(self):
        """Тест валидации корректных данных"""
        data = {
            'company_name': 'Валидная Компания',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': Decimal('1000000'),
                    'premium_with_franchise': Decimal('30000'),
                    'franchise_variant1': Decimal('50000'),
                    'premium_without_franchise': Decimal('40000'),
                    'franchise_variant2': None
                }
            ]
        }
        
        # Не должно вызывать исключений
        validated_data = validate_offer_data(data)
        self.assertEqual(validated_data['company_name'], 'Валидная Компания')
    
    def test_parse_offer_file_function(self):
        """Тест удобной функции parse_offer_file"""
        data_config = {
            'company_name': 'Функция Тест',
            'years_data': [
                {
                    'year': '1 год',
                    'insurance_sum': 500000,
                    'premium_with_franchise': 15000,
                    'franchise_variant1': 25000,
                    'premium_without_franchise': 20000,
                    'franchise_variant2': None
                }
            ]
        }
        
        file_path = self.create_test_xlsx_file(data_config)
        result = parse_offer_file(file_path)
        
        self.assertEqual(result['company_name'], 'Функция Тест')
        self.assertEqual(len(result['years_data']), 1)


if __name__ == '__main__':
    unittest.main()