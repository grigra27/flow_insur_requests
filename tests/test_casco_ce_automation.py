"""
Тесты для автоматического определения КАСКО кат. C/E
"""
import unittest
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
import tempfile
import os
from openpyxl import Workbook
import django
from django.conf import settings

# Настройка Django для тестов
if not settings.configured:
    settings.configure(
        USE_TZ=True,
        TIME_ZONE='UTC',
        SECRET_KEY='test-secret-key',
        INSTALLED_APPS=[
            'django.contrib.auth',
            'django.contrib.contenttypes',
        ],
        DATABASES={
            'default': {
                'ENGINE': 'django.db.backends.sqlite3',
                'NAME': ':memory:',
            }
        }
    )
    django.setup()

from core.excel_utils import ExcelReader


class TestCascoCEAutomation(unittest.TestCase):
    """Тесты для автоматического определения КАСКО кат. C/E"""
    
    def setUp(self):
        """Настройка тестов"""
        self.temp_files = []
    
    def tearDown(self):
        """Очистка после тестов"""
        for temp_file in self.temp_files:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
    
    def create_temp_xlsx_file(self, row_45_data=None):
        """Создает временный .xlsx файл с тестовыми данными"""
        wb = Workbook()
        ws = wb.active
        
        # Добавляем базовые данные для корректной работы ExcelReader
        ws['D7'] = 'Тестовый клиент'
        ws['D9'] = '1234567890'
        ws['D21'] = 'КАСКО'  # Тип страхования
        ws['N17'] = '1 год'  # Период страхования
        
        # Добавляем данные в строку 45 если указаны
        if row_45_data:
            columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
            for i, value in enumerate(row_45_data):
                if i < len(columns) and value is not None:
                    ws[f'{columns[i]}45'] = value
        
        # Сохраняем во временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        self.temp_files.append(temp_file.name)
        return temp_file.name
    
    def create_temp_xls_dataframe(self, row_45_data=None):
        """Создает DataFrame для имитации .xls файла"""
        # Создаем DataFrame с 50 строками и 15 столбцами
        df = pd.DataFrame(index=range(50), columns=range(15))
        
        # Заполняем базовые данные
        df.iloc[6, 3] = 'Тестовый клиент'  # D7
        df.iloc[8, 3] = '1234567890'       # D9
        df.iloc[20, 3] = 'КАСКО'           # D21
        df.iloc[16, 13] = '1 год'          # N17
        
        # Заполняем строку 45 (индекс 44) если указаны данные
        if row_45_data:
            for i, value in enumerate(row_45_data):
                if i < 7 and value is not None:  # Столбцы C-I (индексы 2-8)
                    df.iloc[44, i + 2] = value
        
        return df
    
    def test_determine_casco_ce_openpyxl_with_filled_cells(self):
        """Тест определения КАСКО C/E с заполненными ячейками (openpyxl)"""
        # Тестируем различные комбинации заполненных ячеек
        test_cases = [
            ['значение', None, None, None, None, None, None],  # Только C45
            [None, 'значение', None, None, None, None, None],  # Только D45
            [None, None, None, None, None, None, 'значение'],  # Только I45
            ['знач1', 'знач2', None, None, None, None, None],  # C45 и D45
            ['знач1', 'знач2', 'знач3', 'знач4', 'знач5', 'знач6', 'знач7'],  # Все ячейки
        ]
        
        for row_data in test_cases:
            with self.subTest(row_data=row_data):
                temp_file = self.create_temp_xlsx_file(row_data)
                reader = ExcelReader(temp_file)
                
                # Загружаем workbook и получаем sheet
                from openpyxl import load_workbook
                workbook = load_workbook(temp_file, data_only=True)
                sheet = workbook.active
                
                result = reader._determine_casco_ce_openpyxl(sheet)
                self.assertTrue(result, f"Should return True for row data: {row_data}")
    
    def test_determine_casco_ce_openpyxl_with_empty_cells(self):
        """Тест определения КАСКО C/E с пустыми ячейками (openpyxl)"""
        # Тестируем различные варианты пустых значений
        test_cases = [
            [None, None, None, None, None, None, None],  # Все None
            ['', '', '', '', '', '', ''],                # Все пустые строки
            [' ', '  ', '', None, '', ' ', ''],         # Пробелы и пустые значения
        ]
        
        for row_data in test_cases:
            with self.subTest(row_data=row_data):
                temp_file = self.create_temp_xlsx_file(row_data)
                reader = ExcelReader(temp_file)
                
                # Загружаем workbook и получаем sheet
                from openpyxl import load_workbook
                workbook = load_workbook(temp_file, data_only=True)
                sheet = workbook.active
                
                result = reader._determine_casco_ce_openpyxl(sheet)
                self.assertFalse(result, f"Should return False for row data: {row_data}")
    
    def test_determine_casco_ce_pandas_with_filled_cells(self):
        """Тест определения КАСКО C/E с заполненными ячейками (pandas)"""
        # Тестируем различные комбинации заполненных ячеек
        test_cases = [
            ['значение', None, None, None, None, None, None],  # Только C45
            [None, 'значение', None, None, None, None, None],  # Только D45
            [None, None, None, None, None, None, 'значение'],  # Только I45
            ['знач1', 'знач2', None, None, None, None, None],  # C45 и D45
            ['знач1', 'знач2', 'знач3', 'знач4', 'знач5', 'знач6', 'знач7'],  # Все ячейки
        ]
        
        for row_data in test_cases:
            with self.subTest(row_data=row_data):
                df = self.create_temp_xls_dataframe(row_data)
                reader = ExcelReader('dummy_path')  # Путь не используется для pandas тестов
                
                result = reader._determine_casco_ce_pandas(df)
                self.assertTrue(result, f"Should return True for row data: {row_data}")
    
    def test_determine_casco_ce_pandas_with_empty_cells(self):
        """Тест определения КАСКО C/E с пустыми ячейками (pandas)"""
        # Тестируем различные варианты пустых значений
        test_cases = [
            [None, None, None, None, None, None, None],  # Все None
            ['', '', '', '', '', '', ''],                # Все пустые строки
            [' ', '  ', '', None, '', ' ', ''],         # Пробелы и пустые значения
        ]
        
        for row_data in test_cases:
            with self.subTest(row_data=row_data):
                df = self.create_temp_xls_dataframe(row_data)
                reader = ExcelReader('dummy_path')  # Путь не используется для pandas тестов
                
                result = reader._determine_casco_ce_pandas(df)
                self.assertFalse(result, f"Should return False for row data: {row_data}")
    
    def test_determine_casco_ce_pandas_with_partially_filled_cells(self):
        """Тест определения КАСКО C/E с частично заполненными ячейками (pandas)"""
        # Тестируем случаи, когда только некоторые ячейки заполнены
        test_cases = [
            ['значение', '', None, '', None, '', ''],     # Смешанные пустые и заполненные
            [None, None, 'значение', None, None, None, None],  # Только E45
            ['', '', '', 'значение', '', '', ''],         # Только F45
            [None, '', 'знач1', '', 'знач2', '', None],   # Несколько заполненных
        ]
        
        for row_data in test_cases:
            with self.subTest(row_data=row_data):
                df = self.create_temp_xls_dataframe(row_data)
                reader = ExcelReader('dummy_path')
                
                result = reader._determine_casco_ce_pandas(df)
                self.assertTrue(result, f"Should return True for partially filled row data: {row_data}")
    
    def test_determine_casco_ce_openpyxl_error_handling(self):
        """Тест обработки ошибок при чтении ячеек (openpyxl)"""
        reader = ExcelReader('dummy_path')
        
        # Создаем mock sheet, который вызывает исключение при обращении к ячейкам
        mock_sheet = Mock()
        
        # Настраиваем mock для вызова исключения при обращении к ячейкам
        def mock_getitem(key):
            raise Exception("Cell access error")
        
        mock_sheet.__getitem__ = mock_getitem
        
        # Метод должен вернуть False при ошибке
        result = reader._determine_casco_ce_openpyxl(mock_sheet)
        self.assertFalse(result)
    
    def test_determine_casco_ce_pandas_error_handling(self):
        """Тест обработки ошибок при чтении ячеек (pandas)"""
        reader = ExcelReader('dummy_path')
        
        # Создаем DataFrame, который вызывает исключение при обращении к ячейкам
        mock_df = Mock()
        
        # Настраиваем mock для вызова исключения при обращении к ячейкам
        def mock_getitem(key):
            raise Exception("DataFrame access error")
        
        mock_iloc = Mock()
        mock_iloc.__getitem__ = mock_getitem
        mock_df.iloc = mock_iloc
        
        # Метод должен вернуть False при ошибке
        result = reader._determine_casco_ce_pandas(mock_df)
        self.assertFalse(result)
    
    def test_integration_with_read_insurance_request_openpyxl(self):
        """Интеграционный тест с методом read_insurance_request (openpyxl)"""
        # Создаем файл с заполненной строкой 45
        row_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file = self.create_temp_xlsx_file(row_data)
        
        reader = ExcelReader(temp_file)
        result = reader.read_insurance_request()
        
        # Проверяем, что has_casco_ce установлено в True
        self.assertTrue(result.get('has_casco_ce', False))
        
        # Проверяем, что другие поля тоже корректно извлечены
        self.assertEqual(result.get('client_name'), 'Тестовый клиент')
        self.assertEqual(result.get('insurance_type'), 'КАСКО')
    
    def test_integration_with_read_insurance_request_pandas(self):
        """Интеграционный тест с методом read_insurance_request (pandas)"""
        # Создаем временный .xlsx файл для тестирования pandas пути
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        self.temp_files.append(temp_file.name)
        
        # Создаем DataFrame с заполненной строкой 45
        row_data = [None, 'КАСКО C/E', None, None, None, None, None]
        df = self.create_temp_xls_dataframe(row_data)
        
        # Сохраняем DataFrame в файл (используем xlsx формат)
        df.to_excel(temp_file.name, index=False, header=False, engine='openpyxl')
        
        reader = ExcelReader(temp_file.name)
        
        # Мокаем openpyxl чтобы форсировать использование pandas
        with patch('openpyxl.load_workbook', side_effect=Exception("Force pandas")):
            result = reader.read_insurance_request()
        
        # Проверяем, что has_casco_ce установлено в True
        self.assertTrue(result.get('has_casco_ce', False))
        
        # Проверяем, что другие поля тоже корректно извлечены
        self.assertEqual(result.get('client_name'), 'Тестовый клиент')
        self.assertEqual(result.get('insurance_type'), 'КАСКО')
    
    def test_integration_with_empty_row_45(self):
        """Интеграционный тест с пустой строкой 45"""
        # Создаем файл без данных в строке 45
        temp_file = self.create_temp_xlsx_file()  # Без row_45_data
        
        reader = ExcelReader(temp_file)
        result = reader.read_insurance_request()
        
        # Проверяем, что has_casco_ce установлено в False
        self.assertFalse(result.get('has_casco_ce', True))  # Default True для проверки
        
        # Проверяем, что другие поля корректно извлечены
        self.assertEqual(result.get('client_name'), 'Тестовый клиент')
        self.assertEqual(result.get('insurance_type'), 'КАСКО')


if __name__ == '__main__':
    unittest.main()