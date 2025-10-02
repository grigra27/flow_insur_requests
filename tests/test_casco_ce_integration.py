"""
Integration tests for CASCO C/E automatic detection workflow
"""
import unittest
import tempfile
import os
from datetime import date, datetime, timedelta
from unittest.mock import patch, Mock
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
import pandas as pd

from insurance_requests.models import InsuranceRequest, RequestAttachment
from core.excel_utils import ExcelReader


class TestCascoCEIntegrationWorkflow(TestCase):
    """Integration tests for CASCO C/E automatic detection in complete workflow"""
    
    def setUp(self):
        """Set up test fixtures"""
        # Create test user with proper permissions
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        # Create user group for permissions
        user_group, created = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        
        # Create admin user
        self.admin_user = User.objects.create_user(
            username='admin',
            email='admin@example.com',
            password='adminpass123',
            is_staff=True,
            is_superuser=True
        )
        
        admin_group, created = Group.objects.get_or_create(name='Администраторы')
        self.admin_user.groups.add(admin_group)
        
        self.client = Client()
        self.temp_files = []
    
    def tearDown(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
    
    def create_test_xlsx_file(self, row_45_data=None, client_name='Тестовый клиент', 
                             dfa_number='ДФА-TEST-001', insurance_type='КАСКО'):
        """Create a test .xlsx file with specified CASCO C/E data in row 45"""
        wb = Workbook()
        ws = wb.active
        
        # Add basic required data for ExcelReader
        ws['D7'] = client_name  # Client name
        ws['D9'] = '1234567890'  # INN
        ws['D21'] = insurance_type  # Insurance type
        ws['N17'] = '1 год'  # Insurance period
        
        # DFA number goes in cells H2, I2, J2 (split the number)
        dfa_parts = dfa_number.split('-') if dfa_number else ['ДФА', 'TEST', '001']
        if len(dfa_parts) >= 1:
            ws['H2'] = dfa_parts[0]
        if len(dfa_parts) >= 2:
            ws['I2'] = dfa_parts[1]
        if len(dfa_parts) >= 3:
            ws['J2'] = dfa_parts[2]
        
        # Branch goes in cells C4, D4, E4, F4
        ws['C4'] = 'Казанский'
        ws['D4'] = 'филиал'
        
        ws['D23'] = 'Toyota Camry 2023'  # Vehicle info
        
        # Add CASCO C/E data in row 45 if specified
        if row_45_data:
            columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
            for i, value in enumerate(row_45_data):
                if i < len(columns) and value is not None:
                    ws[f'{columns[i]}45'] = value
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        self.temp_files.append(temp_file.name)
        return temp_file.name
    
    def test_excel_upload_with_casco_ce_detection_true(self):
        """Test complete Excel upload workflow with CASCO C/E detection returning True"""
        # Create Excel file with filled row 45 cells
        row_45_data = ['КАСКО C/E', None, 'значение', None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест КАСКО C/E',
            dfa_number='ДФА-CASCO-CE-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Create uploaded file object
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_casco_ce.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        # Post to upload_excel view
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should redirect to request detail on success
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created
        request = InsuranceRequest.objects.filter(client_name='ООО Тест КАСКО C/E').first()
        self.assertIsNotNone(request)
        
        # Verify CASCO C/E was automatically detected and set to True
        self.assertTrue(request.has_casco_ce)
        
        # Verify other fields were correctly extracted
        self.assertEqual(request.client_name, 'ООО Тест КАСКО C/E')
        self.assertEqual(request.dfa_number, 'ДФА CASCO CE')  # DFA parts are joined with spaces
        self.assertEqual(request.insurance_type, 'КАСКО')
        self.assertEqual(request.inn, '1234567890')
        self.assertEqual(request.branch, 'Казань')  # Should be mapped from 'Казанский филиал'
        self.assertEqual(request.created_by, self.user)
        
        # Verify attachment was created
        attachment = RequestAttachment.objects.filter(request=request).first()
        self.assertIsNotNone(attachment)
        self.assertEqual(attachment.original_filename, 'test_casco_ce.xlsx')
    
    def test_excel_upload_with_casco_ce_detection_false(self):
        """Test complete Excel upload workflow with CASCO C/E detection returning False"""
        # Create Excel file with empty row 45 cells
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=None,  # No data in row 45
            client_name='ООО Тест без КАСКО C/E',
            dfa_number='ДФА-NO-CASCO-CE-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Create uploaded file object
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_no_casco_ce.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        # Post to upload_excel view
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should redirect to request detail on success
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created
        request = InsuranceRequest.objects.filter(client_name='ООО Тест без КАСКО C/E').first()
        self.assertIsNotNone(request)
        
        # Verify CASCO C/E was automatically detected and set to False
        self.assertFalse(request.has_casco_ce)
        
        # Verify other fields were correctly extracted
        self.assertEqual(request.client_name, 'ООО Тест без КАСКО C/E')
        self.assertEqual(request.dfa_number, 'ДФА NO CASCO')  # DFA parts are joined with spaces
        self.assertEqual(request.insurance_type, 'КАСКО')
    
    def test_excel_upload_with_partially_filled_row_45(self):
        """Test Excel upload with partially filled row 45 cells (should return True)"""
        # Create Excel file with some filled and some empty cells in row 45
        row_45_data = [None, '', 'значение', '', None, ' ', '']  # Only one meaningful value
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест частичное заполнение',
            dfa_number='ДФА-PARTIAL-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Create uploaded file object
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_partial_casco_ce.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        # Post to upload_excel view
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should redirect to request detail on success
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created
        request = InsuranceRequest.objects.filter(client_name='ООО Тест частичное заполнение').first()
        self.assertIsNotNone(request)
        
        # Verify CASCO C/E was automatically detected and set to True (any non-empty cell)
        self.assertTrue(request.has_casco_ce)
    
    def test_excel_upload_with_all_empty_row_45(self):
        """Test Excel upload with all empty/whitespace cells in row 45 (should return False)"""
        # Create Excel file with only empty/whitespace cells in row 45
        row_45_data = [None, '', ' ', '  ', '', None, '']  # All empty or whitespace
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест пустые ячейки',
            dfa_number='ДФА-EMPTY-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Create uploaded file object
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_empty_casco_ce.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        # Post to upload_excel view
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should redirect to request detail on success
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created
        request = InsuranceRequest.objects.filter(client_name='ООО Тест пустые ячейки').first()
        self.assertIsNotNone(request)
        
        # Verify CASCO C/E was automatically detected and set to False (all empty)
        self.assertFalse(request.has_casco_ce)
    
    def test_database_save_verification(self):
        """Test that CASCO C/E value is correctly saved to database"""
        # Test with True value
        row_45_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест сохранение True',
            dfa_number='ДФА-SAVE-TRUE-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload file
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_save_true.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Get request from database
        request = InsuranceRequest.objects.filter(client_name='ООО Тест сохранение True').first()
        self.assertIsNotNone(request)
        
        # Verify database field
        self.assertTrue(request.has_casco_ce)
        
        # Refresh from database to ensure it's persisted
        request.refresh_from_db()
        self.assertTrue(request.has_casco_ce)
        
        # Test with False value
        temp_file_path_false = self.create_test_xlsx_file(
            row_45_data=None,  # Empty row 45
            client_name='ООО Тест сохранение False',
            dfa_number='ДФА-SAVE-FALSE-001'
        )
        
        with open(temp_file_path_false, 'rb') as f:
            uploaded_file_false = SimpleUploadedFile(
                name='test_save_false.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file_false
        })
        
        # Get request from database
        request_false = InsuranceRequest.objects.filter(client_name='ООО Тест сохранение False').first()
        self.assertIsNotNone(request_false)
        
        # Verify database field
        self.assertFalse(request_false.has_casco_ce)
        
        # Refresh from database to ensure it's persisted
        request_false.refresh_from_db()
        self.assertFalse(request_false.has_casco_ce)
    
    def test_excel_upload_error_handling(self):
        """Test error handling during Excel upload with CASCO C/E detection"""
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Test with invalid file
        invalid_file = SimpleUploadedFile(
            name='invalid.txt',
            content=b'This is not an Excel file',
            content_type='text/plain'
        )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': invalid_file
        })
        
        # Should stay on upload page with error
        self.assertEqual(response.status_code, 200)
        
        # No request should be created
        self.assertEqual(InsuranceRequest.objects.count(), 0)
    
    def test_excel_upload_with_different_file_formats(self):
        """Test Excel upload with different file formats (.xlsx and .xls)"""
        # Test .xlsx format (already tested above, but let's be explicit)
        row_45_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест XLSX формат',
            dfa_number='ДФА-XLSX-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload .xlsx file
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_xlsx_format.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should redirect to request detail on success
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created with correct CASCO C/E detection
        request = InsuranceRequest.objects.filter(client_name='ООО Тест XLSX формат').first()
        self.assertIsNotNone(request)
        self.assertTrue(request.has_casco_ce)
    
    def test_multiple_excel_uploads_independence(self):
        """Test that multiple Excel uploads work independently with different CASCO C/E values"""
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload first file with CASCO C/E = True
        row_45_data_true = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_1 = self.create_test_xlsx_file(
            row_45_data=row_45_data_true,
            client_name='ООО Первый клиент',
            dfa_number='ДФА-FIRST-001'
        )
        
        with open(temp_file_1, 'rb') as f:
            uploaded_file_1 = SimpleUploadedFile(
                name='first_client.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response_1 = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file_1
        })
        self.assertEqual(response_1.status_code, 302)
        
        # Upload second file with CASCO C/E = False
        temp_file_2 = self.create_test_xlsx_file(
            row_45_data=None,  # Empty row 45
            client_name='ООО Второй клиент',
            dfa_number='ДФА-SECOND-001'
        )
        
        with open(temp_file_2, 'rb') as f:
            uploaded_file_2 = SimpleUploadedFile(
                name='second_client.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response_2 = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file_2
        })
        self.assertEqual(response_2.status_code, 302)
        
        # Verify both requests were created with correct CASCO C/E values
        request_1 = InsuranceRequest.objects.filter(client_name='ООО Первый клиент').first()
        request_2 = InsuranceRequest.objects.filter(client_name='ООО Второй клиент').first()
        
        self.assertIsNotNone(request_1)
        self.assertIsNotNone(request_2)
        
        # Verify independent CASCO C/E detection
        self.assertTrue(request_1.has_casco_ce)
        self.assertFalse(request_2.has_casco_ce)
        
        # Verify they are different requests
        self.assertNotEqual(request_1.id, request_2.id)
    
    def test_excel_upload_logging(self):
        """Test that CASCO C/E detection is properly logged"""
        # Create Excel file with CASCO C/E data
        row_45_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест логирование',
            dfa_number='ДФА-LOG-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Mock logger to capture log messages
        with patch('insurance_requests.views.logger') as mock_logger:
            with open(temp_file_path, 'rb') as f:
                uploaded_file = SimpleUploadedFile(
                    name='test_logging.xlsx',
                    content=f.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            response = self.client.post(reverse('insurance_requests:upload_excel'), {
                'excel_file': uploaded_file
            })
            
            # Should redirect on success
            self.assertEqual(response.status_code, 302)
            
            # Verify logging was called for CASCO C/E detection
            mock_logger.info.assert_called_with(
                'CASCO C/E automatically detected for file: test_logging.xlsx'
            )
    
    def test_backward_compatibility(self):
        """Test that existing functionality still works with CASCO C/E detection"""
        # Create Excel file without CASCO C/E data (backward compatibility)
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=None,
            client_name='ООО Обратная совместимость',
            dfa_number='ДФА-COMPAT-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload file
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_compatibility.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        
        # Should work as before
        self.assertEqual(response.status_code, 302)
        
        # Verify request was created
        request = InsuranceRequest.objects.filter(client_name='ООО Обратная совместимость').first()
        self.assertIsNotNone(request)
        
        # Verify all existing fields still work
        self.assertEqual(request.client_name, 'ООО Обратная совместимость')
        self.assertEqual(request.dfa_number, 'ДФА COMPAT 001')  # DFA parts are joined with spaces
        self.assertEqual(request.insurance_type, 'КАСКО')
        self.assertEqual(request.inn, '1234567890')
        self.assertEqual(request.branch, 'Казань')
        
        # Verify CASCO C/E defaults to False
        self.assertFalse(request.has_casco_ce)


if __name__ == '__main__':
    unittest.main()


class TestCascoCEManualEditingIntegration(TestCase):
    """Integration tests for manual editing of CASCO C/E field after automatic detection"""
    
    def setUp(self):
        """Set up test fixtures"""
        # Create test user with proper permissions
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        # Create user group for permissions
        user_group, created = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        
        self.client = Client()
        self.temp_files = []
    
    def tearDown(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
    
    def create_test_xlsx_file(self, row_45_data=None, client_name='Тестовый клиент', 
                             dfa_number='ДФА-TEST-001', insurance_type='КАСКО'):
        """Create a test .xlsx file with specified CASCO C/E data in row 45"""
        wb = Workbook()
        ws = wb.active
        
        # Add basic required data for ExcelReader
        ws['D7'] = client_name  # Client name
        ws['D9'] = '1234567890'  # INN
        ws['D21'] = insurance_type  # Insurance type
        ws['N17'] = '1 год'  # Insurance period
        
        # DFA number goes in cells H2, I2, J2 (split the number)
        dfa_parts = dfa_number.split('-') if dfa_number else ['ДФА', 'TEST', '001']
        if len(dfa_parts) >= 1:
            ws['H2'] = dfa_parts[0]
        if len(dfa_parts) >= 2:
            ws['I2'] = dfa_parts[1]
        if len(dfa_parts) >= 3:
            ws['J2'] = dfa_parts[2]
        
        # Branch goes in cells C4, D4, E4, F4
        ws['C4'] = 'Казанский'
        ws['D4'] = 'филиал'
        
        ws['D23'] = 'Toyota Camry 2023'  # Vehicle info
        
        # Add CASCO C/E data in row 45 if specified
        if row_45_data:
            columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
            for i, value in enumerate(row_45_data):
                if i < len(columns) and value is not None:
                    ws[f'{columns[i]}45'] = value
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        self.temp_files.append(temp_file.name)
        return temp_file.name
    
    def test_manual_edit_casco_ce_from_true_to_false(self):
        """Test manually changing CASCO C/E from True (auto-detected) to False"""
        # First, create a request with automatic CASCO C/E detection = True
        row_45_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест ручное изменение True->False',
            dfa_number='ДФА-MANUAL-TRUE-FALSE-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload Excel file (should auto-detect CASCO C/E = True)
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_manual_edit.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        self.assertEqual(response.status_code, 302)
        
        # Get the created request
        request = InsuranceRequest.objects.filter(client_name='ООО Тест ручное изменение True->False').first()
        self.assertIsNotNone(request)
        self.assertTrue(request.has_casco_ce)  # Should be auto-detected as True
        
        # Now manually edit the request to change CASCO C/E to False
        edit_response = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': 'Казань',  # Use valid branch choice
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': False,  # Manually change to False
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes or ''
        })
        
        # Should redirect to detail page on success
        self.assertEqual(edit_response.status_code, 302)
        
        # Verify the change was saved
        request.refresh_from_db()
        self.assertFalse(request.has_casco_ce)  # Should now be False
        
        # Verify other fields remain unchanged
        self.assertEqual(request.client_name, 'ООО Тест ручное изменение True->False')
        self.assertEqual(request.dfa_number, 'ДФА MANUAL TRUE')  # DFA parts are joined with spaces
    
    def test_manual_edit_casco_ce_from_false_to_true(self):
        """Test manually changing CASCO C/E from False (auto-detected) to True"""
        # First, create a request with automatic CASCO C/E detection = False
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=None,  # Empty row 45 -> should auto-detect as False
            client_name='ООО Тест ручное изменение False->True',
            dfa_number='ДФА-MANUAL-FALSE-TRUE-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload Excel file (should auto-detect CASCO C/E = False)
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_manual_edit_false_true.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        self.assertEqual(response.status_code, 302)
        
        # Get the created request
        request = InsuranceRequest.objects.filter(client_name='ООО Тест ручное изменение False->True').first()
        self.assertIsNotNone(request)
        self.assertFalse(request.has_casco_ce)  # Should be auto-detected as False
        
        # Now manually edit the request to change CASCO C/E to True
        edit_response = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': 'Казань',  # Use valid branch choice
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': True,  # Manually change to True
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes or ''
        })
        
        # Should redirect to detail page on success
        self.assertEqual(edit_response.status_code, 302)
        
        # Verify the change was saved
        request.refresh_from_db()
        self.assertTrue(request.has_casco_ce)  # Should now be True
        
        # Verify other fields remain unchanged
        self.assertEqual(request.client_name, 'ООО Тест ручное изменение False->True')
        self.assertEqual(request.dfa_number, 'ДФА MANUAL FALSE')  # DFA parts are joined with spaces
    
    def test_edit_form_displays_current_casco_ce_value(self):
        """Test that the edit form correctly displays the current CASCO C/E value"""
        # Create a request with CASCO C/E = True
        request = InsuranceRequest.objects.create(
            client_name='ООО Тест отображение формы',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА-FORM-DISPLAY-001',
            branch='Казань',
            has_casco_ce=True,  # Set to True
            created_by=self.user
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Get the edit form
        response = self.client.get(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}))
        self.assertEqual(response.status_code, 200)
        
        # Check that the form contains the correct checkbox state
        self.assertContains(response, 'name="has_casco_ce"')
        self.assertContains(response, 'checked')  # Should be checked since has_casco_ce=True
        
        # Now test with CASCO C/E = False
        request.has_casco_ce = False
        request.save()
        
        response_false = self.client.get(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}))
        self.assertEqual(response_false.status_code, 200)
        
        # Check that the form contains the checkbox but it's not checked
        self.assertContains(response_false, 'name="has_casco_ce"')
        # The checkbox should not be checked (no 'checked' attribute in the has_casco_ce input)
        form_content = response_false.content.decode()
        # Look for the specific checkbox input for has_casco_ce
        import re
        casco_ce_input_pattern = r'<input[^>]*name="has_casco_ce"[^>]*>'
        casco_ce_input_match = re.search(casco_ce_input_pattern, form_content)
        self.assertIsNotNone(casco_ce_input_match)
        casco_ce_input = casco_ce_input_match.group()
        self.assertNotIn('checked', casco_ce_input)  # Should not be checked
    
    def test_multiple_manual_edits_persistence(self):
        """Test that multiple manual edits of CASCO C/E field persist correctly"""
        # Create a request with automatic detection
        row_45_data = ['КАСКО C/E', None, None, None, None, None, None]
        temp_file_path = self.create_test_xlsx_file(
            row_45_data=row_45_data,
            client_name='ООО Тест множественные изменения',
            dfa_number='ДФА-MULTIPLE-EDITS-001'
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Upload Excel file (should auto-detect CASCO C/E = True)
        with open(temp_file_path, 'rb') as f:
            uploaded_file = SimpleUploadedFile(
                name='test_multiple_edits.xlsx',
                content=f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        response = self.client.post(reverse('insurance_requests:upload_excel'), {
            'excel_file': uploaded_file
        })
        self.assertEqual(response.status_code, 302)
        
        # Get the created request
        request = InsuranceRequest.objects.filter(client_name='ООО Тест множественные изменения').first()
        self.assertIsNotNone(request)
        self.assertTrue(request.has_casco_ce)  # Initially True
        
        # First edit: Change to False
        edit_response_1 = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': 'Казань',  # Use valid branch choice
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': False,  # Change to False
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': 'Первое изменение - убрали КАСКО C/E'
        })
        self.assertEqual(edit_response_1.status_code, 302)
        
        # Verify first change
        request.refresh_from_db()
        self.assertFalse(request.has_casco_ce)
        self.assertEqual(request.notes, 'Первое изменение - убрали КАСКО C/E')
        
        # Second edit: Change back to True
        edit_response_2 = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': 'Казань',  # Use valid branch choice
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': True,  # Change back to True
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': 'Второе изменение - вернули КАСКО C/E'
        })
        self.assertEqual(edit_response_2.status_code, 302)
        
        # Verify second change
        request.refresh_from_db()
        self.assertTrue(request.has_casco_ce)
        self.assertEqual(request.notes, 'Второе изменение - вернули КАСКО C/E')
        
        # Third edit: Change to False again
        edit_response_3 = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': 'Казань',  # Use valid branch choice
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': False,  # Change to False again
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': 'Третье изменение - окончательно убрали КАСКО C/E'
        })
        self.assertEqual(edit_response_3.status_code, 302)
        
        # Verify final change
        request.refresh_from_db()
        self.assertFalse(request.has_casco_ce)
        self.assertEqual(request.notes, 'Третье изменение - окончательно убрали КАСКО C/E')
    
    def test_manual_edit_preserves_other_fields(self):
        """Test that manual editing of CASCO C/E doesn't affect other fields"""
        # Create a request with all fields populated
        request = InsuranceRequest.objects.create(
            client_name='ООО Тест сохранение полей',
            inn='9876543210',
            insurance_type='страхование спецтехники',
            insurance_period='на весь срок лизинга',
            vehicle_info='Экскаватор Caterpillar 320D',
            dfa_number='ДФА-PRESERVE-FIELDS-001',
            branch='Нижний Новгород',
            has_franchise=True,
            has_installment=True,
            has_autostart=False,
            has_casco_ce=False,  # Initially False
            notes='Исходные примечания',
            created_by=self.user
        )
        
        # Store original values for comparison
        original_values = {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': request.insurance_period,
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'notes': request.notes
        }
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Edit only the CASCO C/E field
        edit_response = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': request.insurance_period,
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': True,  # Only change this field
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes
        })
        self.assertEqual(edit_response.status_code, 302)
        
        # Verify CASCO C/E changed
        request.refresh_from_db()
        self.assertTrue(request.has_casco_ce)  # Should be changed to True
        
        # Verify all other fields remain unchanged
        self.assertEqual(request.client_name, original_values['client_name'])
        self.assertEqual(request.inn, original_values['inn'])
        self.assertEqual(request.insurance_type, original_values['insurance_type'])
        self.assertEqual(request.insurance_period, original_values['insurance_period'])
        self.assertEqual(request.vehicle_info, original_values['vehicle_info'])
        self.assertEqual(request.dfa_number, original_values['dfa_number'])
        self.assertEqual(request.branch, original_values['branch'])
        self.assertEqual(request.has_franchise, original_values['has_franchise'])
        self.assertEqual(request.has_installment, original_values['has_installment'])
        self.assertEqual(request.has_autostart, original_values['has_autostart'])
        self.assertEqual(request.notes, original_values['notes'])
    
    def test_manual_edit_form_validation(self):
        """Test form validation when manually editing CASCO C/E field"""
        # Create a request
        request = InsuranceRequest.objects.create(
            client_name='ООО Тест валидация формы',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА-VALIDATION-001',
            branch='Казань',
            has_casco_ce=True,
            created_by=self.user
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Test valid form submission
        valid_response = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': 'ООО Обновленное название',  # Change client name
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': False,  # Change CASCO C/E
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': 'Обновленные примечания'
        })
        
        # Should redirect on success
        self.assertEqual(valid_response.status_code, 302)
        
        # Verify changes were saved
        request.refresh_from_db()
        self.assertEqual(request.client_name, 'ООО Обновленное название')
        self.assertFalse(request.has_casco_ce)
        self.assertEqual(request.notes, 'Обновленные примечания')
        
        # Test invalid form submission (missing required field)
        invalid_response = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': '',  # Empty required field
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': True,  # Try to change CASCO C/E back
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes
        })
        
        # Should stay on edit page with errors
        self.assertEqual(invalid_response.status_code, 200)
        self.assertContains(invalid_response, 'form-control')  # Form should be displayed
        
        # Verify changes were NOT saved due to validation error
        request.refresh_from_db()
        self.assertEqual(request.client_name, 'ООО Обновленное название')  # Should remain unchanged
        self.assertFalse(request.has_casco_ce)  # Should remain unchanged
    
    def test_edit_form_checkbox_behavior(self):
        """Test checkbox behavior in edit form for CASCO C/E field"""
        # Create a request with CASCO C/E = False
        request = InsuranceRequest.objects.create(
            client_name='ООО Тест поведение чекбокса',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА-CHECKBOX-001',
            branch='Казань',
            has_casco_ce=False,
            created_by=self.user
        )
        
        # Login user
        self.client.login(username='testuser', password='testpass123')
        
        # Test submitting form without has_casco_ce field (unchecked checkbox)
        # When checkbox is unchecked, it's not included in POST data
        response_unchecked = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            # has_casco_ce is intentionally omitted (unchecked checkbox)
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes or ''
        })
        
        # Should redirect on success
        self.assertEqual(response_unchecked.status_code, 302)
        
        # Verify CASCO C/E remains False (unchecked)
        request.refresh_from_db()
        self.assertFalse(request.has_casco_ce)
        
        # Test submitting form with has_casco_ce field (checked checkbox)
        response_checked = self.client.post(reverse('insurance_requests:edit_request', kwargs={'pk': request.pk}), {
            'client_name': request.client_name,
            'inn': request.inn,
            'insurance_type': request.insurance_type,
            'insurance_period': '1 год',
            'vehicle_info': request.vehicle_info,
            'dfa_number': request.dfa_number,
            'branch': request.branch,
            'has_franchise': request.has_franchise,
            'has_installment': request.has_installment,
            'has_autostart': request.has_autostart,
            'has_casco_ce': True,  # Explicitly set to True (checked checkbox)
            'response_deadline': request.response_deadline.strftime('%Y-%m-%dT%H:%M') if request.response_deadline else '',
            'notes': request.notes or ''
        })
        
        # Should redirect on success
        self.assertEqual(response_checked.status_code, 302)
        
        # Verify CASCO C/E is now True (checked)
        request.refresh_from_db()
        self.assertTrue(request.has_casco_ce)


if __name__ == '__main__':
    unittest.main()