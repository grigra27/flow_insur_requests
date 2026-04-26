"""Сервис аналитики по сотрудникам (Django-юзерам, загружающим заявки).

Phase 1: filters, KPI, табличные метрики, time-to-*, charts, funnel, базовые алерты.
Phase 2: качество данных, портфель, heatmap-связи, аномалии цикла, просадка объёма,
композитный quality-score.

Реализация компромиссная: для прошлых данных нет точного таймстемпа смены статуса,
поэтому стадии funnel считаются по снимку (current status + наличие summary).
Будущие данные пишутся в StatusEvent — на них метрики time-to-* станут точными.
"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from statistics import median
from typing import Any, Iterable

from django.contrib.auth.models import User
from django.db.models import Count, Q, QuerySet
from django.utils import timezone

from insurance_requests.models import InsuranceRequest
from ..models import InsuranceSummary

# --- Константы --------------------------------------------------------------

DEFAULT_PERIOD_DAYS = 365
PERIOD_CHOICES = ('all', '30', '90', '365', 'custom')

QUALITY_SCORE_WEIGHTS: dict[str, float] = {
    'completeness': 0.40,
    'win_rate': 0.30,
    'speed': 0.20,
    'volume': 0.10,
}

# Активные статусы (не завершённые) для подсчёта backlog’а.
ACTIVE_REQUEST_STATUSES = {'uploaded', 'email_generated'}
ACTIVE_SUMMARY_STATUSES = {'collecting', 'ready', 'sent'}
TERMINAL_SUMMARY_STATUSES = {'completed_accepted', 'completed_rejected'}

# Пороги алертов
NO_ACTIVITY_DAYS_DEFAULT = 14
STUCK_SUMMARY_DAYS_DEFAULT = 30
VOLUME_DROP_THRESHOLD_PCT = -30.0  # % изменения объёма к прошлому окну: алерт ≤ −30%
CYCLE_OUTLIER_FACTOR = 1.5  # threshold = factor * team_p75 (устойчиво к малым выборкам)
CYCLE_OUTLIER_FLOOR_HOURS = 7 * 24  # пол threshold’а — 7 дней, чтобы не ловить мелочёвку
EDIT_THRESHOLD_SECONDS = 60  # updated_at - created_at > порог → заявка считается отредактированной

# Группы полей для completeness-score
FIELDS_BASE = (
    'dfa_number', 'branch', 'manager_name', 'vehicle_info',
    'notes', 'insurance_period', 'response_deadline',
)
FIELDS_CASCO = (
    'key_completeness', 'pts_psm', 'creditor_bank',
    'usage_purposes', 'telematics_complex',
    'manufacturing_year', 'asset_status',
)
FIELDS_PROPERTY = ('insurance_territory',)
CASCO_INSURANCE_TYPES = {'КАСКО', 'страхование спецтехники'}
PROPERTY_INSURANCE_TYPES = {'страхование имущества'}

HEATMAP_TOP_LIMIT = 8  # сколько значений измерения показывать в heatmap’е
HEATMAP_OUTLIER_LIMIT = 50


# --- Парсер фильтров --------------------------------------------------------


def parse_filters(get_params) -> dict[str, Any]:
    """Разбирает request.GET. По умолчанию — период 365 дней.

    Возвращает dict, чтобы было легко передавать в шаблон без распаковки.
    """
    errors: list[str] = []
    period = (get_params.get('period') or str(DEFAULT_PERIOD_DAYS)).strip().lower()
    if period not in PERIOD_CHOICES:
        period = str(DEFAULT_PERIOD_DAYS)

    start_date_str = (get_params.get('start_date') or '').strip()
    end_date_str = (get_params.get('end_date') or '').strip()

    def _parse_date(value: str, field: str) -> date | None:
        if not value:
            return None
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f'Некорректная дата в поле «{field}»')
            return None

    start_date: date | None = None
    end_date: date | None = None

    if start_date_str or end_date_str:
        start_date = _parse_date(start_date_str, 'Дата с')
        end_date = _parse_date(end_date_str, 'Дата по')
        period = 'custom'
    elif period in {'30', '90', '365'}:
        days = int(period)
        end_date = timezone.localdate()
        start_date = end_date - timedelta(days=days - 1)
    elif period == 'all':
        start_date = None
        end_date = None

    if start_date and end_date and start_date > end_date:
        errors.append('Дата начала позже даты окончания')
        start_date = end_date = None
        period = 'all'

    user_ids: list[int] = []
    for raw in get_params.getlist('user_ids') if hasattr(get_params, 'getlist') else []:
        for token in str(raw).split(','):
            token = token.strip()
            if not token:
                continue
            try:
                user_ids.append(int(token))
            except ValueError:
                errors.append(f'Некорректный id сотрудника: {token!r}')

    return {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_str': start_date.isoformat() if start_date else '',
        'end_date_str': end_date.isoformat() if end_date else '',
        'user_ids': user_ids,
        'branch': (get_params.get('branch') or '').strip(),
        'insurance_type': (get_params.get('insurance_type') or '').strip(),
        'deal_status': (get_params.get('deal_status') or '').strip(),
        'include_unassigned': (get_params.get('include_unassigned') or '1').strip() != '0',
        'errors': errors,
    }


# --- Helpers ----------------------------------------------------------------


def _apply_window(qs: QuerySet, field: str, filters: dict) -> QuerySet:
    if filters['start_date']:
        qs = qs.filter(**{f'{field}__date__gte': filters['start_date']})
    if filters['end_date']:
        qs = qs.filter(**{f'{field}__date__lte': filters['end_date']})
    return qs


def _build_request_qs(filters: dict) -> QuerySet:
    qs = InsuranceRequest.objects.select_related('created_by')
    qs = _apply_window(qs, 'created_at', filters)
    if filters['branch']:
        qs = qs.filter(branch=filters['branch'])
    if filters['insurance_type']:
        qs = qs.filter(insurance_type=filters['insurance_type'])
    if filters['deal_status']:
        qs = qs.filter(deal_status=filters['deal_status'])
    if filters['user_ids']:
        if filters['include_unassigned']:
            qs = qs.filter(Q(created_by_id__in=filters['user_ids']) | Q(created_by__isnull=True))
        else:
            qs = qs.filter(created_by_id__in=filters['user_ids'])
    elif not filters['include_unassigned']:
        qs = qs.filter(created_by__isnull=False)
    return qs


def _build_summary_qs(filters: dict) -> QuerySet:
    """Своды, чьи заявки попадают в окно и фильтры."""
    qs = InsuranceSummary.objects.select_related('request', 'request__created_by').prefetch_related('offers')
    request_ids = list(_build_request_qs(filters).values_list('id', flat=True))
    return qs.filter(request_id__in=request_ids)


def _user_display(user: User | None) -> str:
    if user is None:
        return 'Без автора'
    full = (user.get_full_name() or '').strip()
    return full or user.username


def _hours_between(start, end) -> float | None:
    if not start or not end:
        return None
    delta = end - start
    return delta.total_seconds() / 3600.0


def _avg(values: Iterable[float | None]) -> float | None:
    cleaned = [v for v in values if v is not None]
    if not cleaned:
        return None
    return sum(cleaned) / len(cleaned)


def _median(values: Iterable[float | None]) -> float | None:
    cleaned = [v for v in values if v is not None]
    if not cleaned:
        return None
    return float(median(cleaned))


def _percentile(values: list[float], q: float) -> float | None:
    """Линейная интерполяция p-квантили. q ∈ [0, 1]."""
    if not values:
        return None
    if len(values) == 1:
        return float(values[0])
    sorted_v = sorted(values)
    pos = q * (len(sorted_v) - 1)
    lo = int(pos)
    hi = min(lo + 1, len(sorted_v) - 1)
    frac = pos - lo
    return float(sorted_v[lo] + (sorted_v[hi] - sorted_v[lo]) * frac)


def _is_field_filled(value) -> bool:
    """Поле считается заполненным, если есть непустое значение (без учёта пробелов)."""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, bool):
        # Булевы поля «заполнены» всегда — это часть схемы.
        return True
    return True


def _is_inn_valid(inn: str) -> bool:
    """Проверка длины ИНН: 10 (юрлицо) или 12 (ИП/физлицо)."""
    if not inn:
        return False
    cleaned = inn.strip()
    return len(cleaned) in (10, 12) and cleaned.isdigit()


def _completeness_for_request(req: InsuranceRequest) -> dict[str, Any]:
    """Возвращает заполненность полей для одной заявки.

    Возвращает:
        {
            'base_filled': int, 'base_total': int,
            'casco_filled': int|None, 'casco_total': int|None,  # None если не КАСКО
            'property_filled': int|None, 'property_total': int|None,
            'inn_valid': bool,
            'is_other_type': bool,
            'edited_after_create': bool,
        }
    """
    base_filled = sum(1 for f in FIELDS_BASE if _is_field_filled(getattr(req, f, None)))
    casco_filled = casco_total = None
    property_filled = property_total = None
    if req.insurance_type in CASCO_INSURANCE_TYPES:
        casco_filled = sum(1 for f in FIELDS_CASCO if _is_field_filled(getattr(req, f, None)))
        casco_total = len(FIELDS_CASCO)
    if req.insurance_type in PROPERTY_INSURANCE_TYPES:
        property_filled = sum(1 for f in FIELDS_PROPERTY if _is_field_filled(getattr(req, f, None)))
        property_total = len(FIELDS_PROPERTY)

    edited = False
    if req.created_at and req.updated_at:
        delta = (req.updated_at - req.created_at).total_seconds()
        edited = delta > EDIT_THRESHOLD_SECONDS

    return {
        'base_filled': base_filled,
        'base_total': len(FIELDS_BASE),
        'casco_filled': casco_filled,
        'casco_total': casco_total,
        'property_filled': property_filled,
        'property_total': property_total,
        'inn_valid': _is_inn_valid(req.inn),
        'is_other_type': req.insurance_type == 'другое',
        'edited_after_create': edited,
    }


def _aggregate_completeness(requests: list[InsuranceRequest]) -> dict[str, Any]:
    """Агрегирует completeness-метрики по списку заявок одного сотрудника."""
    if not requests:
        return {
            'base_pct': None, 'casco_pct': None, 'property_pct': None,
            'overall_pct': None,
            'inn_invalid_count': 0, 'inn_invalid_pct': None,
            'other_type_count': 0, 'other_type_pct': None,
            'edited_count': 0, 'edited_pct': None,
        }

    base_filled = base_total = 0
    casco_filled = casco_total = 0
    property_filled = property_total = 0
    inn_invalid = 0
    other_type = 0
    edited = 0

    for req in requests:
        c = _completeness_for_request(req)
        base_filled += c['base_filled']
        base_total += c['base_total']
        if c['casco_total']:
            casco_filled += c['casco_filled']
            casco_total += c['casco_total']
        if c['property_total']:
            property_filled += c['property_filled']
            property_total += c['property_total']
        if not c['inn_valid']:
            inn_invalid += 1
        if c['is_other_type']:
            other_type += 1
        if c['edited_after_create']:
            edited += 1

    base_pct = base_filled / base_total * 100 if base_total else None
    casco_pct = casco_filled / casco_total * 100 if casco_total else None
    property_pct = property_filled / property_total * 100 if property_total else None

    # Общий процент — среднее по доступным группам
    available = [pct for pct in (base_pct, casco_pct, property_pct) if pct is not None]
    overall_pct = sum(available) / len(available) if available else None

    n = len(requests)
    return {
        'base_pct': base_pct,
        'casco_pct': casco_pct,
        'property_pct': property_pct,
        'overall_pct': overall_pct,
        'inn_invalid_count': inn_invalid,
        'inn_invalid_pct': inn_invalid / n * 100,
        'other_type_count': other_type,
        'other_type_pct': other_type / n * 100,
        'edited_count': edited,
        'edited_pct': edited / n * 100,
    }


def _portfolio_for_user(requests: list[InsuranceRequest]) -> dict[str, Any]:
    """Распределение по типу страхования, deal_status и булевым флагам."""
    if not requests:
        return {
            'by_insurance_type': {},
            'prolongation_pct': None, 'new_deal_pct': None,
            'flag_pct': {},
        }

    by_type: dict[str, int] = defaultdict(int)
    new_count = prolong_count = 0
    flag_counters = {
        'has_franchise': 0, 'has_installment': 0,
        'has_autostart': 0, 'has_casco_ce': 0,
        'has_transportation': 0, 'has_construction_work': 0,
    }
    n = len(requests)
    for req in requests:
        by_type[req.insurance_type or 'не указан'] += 1
        if req.deal_status == 'prolongation':
            prolong_count += 1
        else:
            new_count += 1
        for flag in flag_counters:
            if getattr(req, flag, False):
                flag_counters[flag] += 1

    return {
        'by_insurance_type': dict(by_type),
        'new_deal_pct': new_count / n * 100,
        'prolongation_pct': prolong_count / n * 100,
        'flag_pct': {k: v / n * 100 for k, v in flag_counters.items()},
    }


LATE_HOUR_THRESHOLD = 22  # «поздняя» загрузка с 22:00
WEEKEND_DAYS = (5, 6)  # суббота, воскресенье

# Корзины возраста backlog’а в днях
BACKLOG_AGE_BUCKETS = [
    (0, 3, '0-3 дн.'),
    (4, 7, '4-7 дн.'),
    (8, 14, '8-14 дн.'),
    (15, 30, '15-30 дн.'),
    (31, None, '>30 дн.'),
]


def _pattern_metrics_for_user(
    requests: list[InsuranceRequest],
    *,
    now=None,
) -> dict[str, Any]:
    """Поведенческие метрики: cadence, late/weekend, days_since_last."""
    if not requests:
        return {
            'cadence_hours': None,
            'late_pct': None,
            'weekend_pct': None,
            'days_since_last': None,
        }
    now = now or timezone.now()

    sorted_dts = sorted((r.created_at for r in requests if r.created_at))
    if len(sorted_dts) >= 2:
        deltas = [
            (sorted_dts[i + 1] - sorted_dts[i]).total_seconds() / 3600.0
            for i in range(len(sorted_dts) - 1)
        ]
        cadence_hours = sum(deltas) / len(deltas)
    else:
        cadence_hours = None

    n = len(requests)
    late = sum(1 for r in requests if r.created_at and _is_late(r.created_at))
    weekend = sum(1 for r in requests if r.created_at and r.created_at.weekday() in WEEKEND_DAYS)

    last_dt = sorted_dts[-1] if sorted_dts else None
    days_since_last = (now - last_dt).total_seconds() / 86400.0 if last_dt else None

    return {
        'cadence_hours': cadence_hours,
        'late_pct': late / n * 100,
        'weekend_pct': weekend / n * 100,
        'days_since_last': days_since_last,
    }


def _is_late(dt) -> bool:
    """Загрузка после 22:00 в локальной таймзоне."""
    try:
        local = timezone.localtime(dt)
    except (ValueError, TypeError):
        local = dt
    return local.hour >= LATE_HOUR_THRESHOLD


def _backlog_age_buckets(filters: dict, *, now=None) -> dict[str, Any]:
    """Распределение возрастов активных заявок и сводов по корзинам."""
    now = now or timezone.now()

    request_qs = _build_request_qs(filters).filter(status__in=list(ACTIVE_REQUEST_STATUSES))
    summary_qs = _build_summary_qs(filters).filter(status__in=list(ACTIVE_SUMMARY_STATUSES))

    items = []
    for r in request_qs:
        if r.created_at:
            age_days = (now - r.created_at).total_seconds() / 86400.0
            items.append({'kind': 'request', 'age_days': age_days})
    for s in summary_qs:
        anchor = s.updated_at or s.created_at
        if anchor:
            age_days = (now - anchor).total_seconds() / 86400.0
            items.append({'kind': 'summary', 'age_days': age_days})

    buckets = []
    for low, high, label in BACKLOG_AGE_BUCKETS:
        if high is None:
            count = sum(1 for it in items if it['age_days'] >= low)
        else:
            count = sum(1 for it in items if low <= it['age_days'] <= high)
        buckets.append({'label': label, 'count': count})

    return {
        'buckets': buckets,
        'total': len(items),
        'request_count': sum(1 for it in items if it['kind'] == 'request'),
        'summary_count': sum(1 for it in items if it['kind'] == 'summary'),
    }


def _team_trend(filters: dict) -> dict[str, Any]:
    """WoW/MoM: сравнение последних 7/30 дней с предыдущими 7/30.

    Считается всегда от «сегодня», независимо от выбранного периода фильтра —
    даёт быстрый sense check состояния «прямо сейчас».
    """
    now = timezone.now()
    today = timezone.localdate()

    def _stats(start: date, end: date) -> dict[str, Any]:
        qs = InsuranceRequest.objects.filter(
            created_at__date__gte=start, created_at__date__lte=end
        )
        # фильтры user_ids/branch/insurance_type/deal_status тоже применяем
        if filters['user_ids']:
            qs = qs.filter(created_by_id__in=filters['user_ids'])
        if filters['branch']:
            qs = qs.filter(branch=filters['branch'])
        if filters['insurance_type']:
            qs = qs.filter(insurance_type=filters['insurance_type'])
        if filters['deal_status']:
            qs = qs.filter(deal_status=filters['deal_status'])

        requests_count = qs.count()
        summary_qs = InsuranceSummary.objects.filter(request_id__in=qs.values('id'))
        accepted_qs = summary_qs.filter(status='completed_accepted')
        accepted_count = accepted_qs.count()

        premium_total = Decimal('0')
        for s in accepted_qs.prefetch_related('offers'):
            premium_total += _accepted_premium(s)

        return {
            'requests': requests_count,
            'accepted': accepted_count,
            'premium': premium_total,
        }

    def _delta_pct(cur: float | int | Decimal, prev: float | int | Decimal) -> float | None:
        if not prev:
            return None
        return float((cur - prev) / prev * 100)

    week_cur = _stats(today - timedelta(days=6), today)
    week_prev = _stats(today - timedelta(days=13), today - timedelta(days=7))
    month_cur = _stats(today - timedelta(days=29), today)
    month_prev = _stats(today - timedelta(days=59), today - timedelta(days=30))

    return {
        'week': {
            'current': week_cur,
            'previous': week_prev,
            'delta_pct': {
                'requests': _delta_pct(week_cur['requests'], week_prev['requests']),
                'accepted': _delta_pct(week_cur['accepted'], week_prev['accepted']),
                'premium': _delta_pct(week_cur['premium'], week_prev['premium']),
            },
        },
        'month': {
            'current': month_cur,
            'previous': month_prev,
            'delta_pct': {
                'requests': _delta_pct(month_cur['requests'], month_prev['requests']),
                'accepted': _delta_pct(month_cur['accepted'], month_prev['accepted']),
                'premium': _delta_pct(month_cur['premium'], month_prev['premium']),
            },
        },
    }


def _moving_average(values: list[int | float], window: int) -> list[float | None]:
    """Скользящее среднее. Левая граница (меньше окна) — None."""
    if window <= 1:
        return [float(v) for v in values]
    out: list[float | None] = []
    acc = 0.0
    queue: list[float] = []
    for v in values:
        queue.append(float(v))
        acc += float(v)
        if len(queue) > window:
            acc -= queue.pop(0)
        out.append(acc / len(queue) if len(queue) >= window else None)
    return out


def _day_hour_heatmap(filters: dict) -> dict[str, Any]:
    """Heatmap 7×24: количество загрузок по дням недели и часам (локальное время)."""
    request_qs = _build_request_qs(filters).values_list('created_at', flat=True)
    grid = [[0] * 24 for _ in range(7)]
    max_value = 0
    total = 0
    for created_at in request_qs:
        if not created_at:
            continue
        try:
            local = timezone.localtime(created_at)
        except (ValueError, TypeError):
            local = created_at
        d = local.weekday()
        h = local.hour
        grid[d][h] += 1
        total += 1
        if grid[d][h] > max_value:
            max_value = grid[d][h]
    day_labels = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    rows = [
        {'label': day_labels[i], 'cells': grid[i]}
        for i in range(7)
    ]
    return {
        'rows': rows,
        'max_value': max_value,
        'total': total,
        'hours': list(range(24)),
        'header_marks': [0, 3, 6, 9, 12, 15, 18, 21],
        # для обратной совместимости (старый ключ тестов и т.п.):
        'grid': grid,
        'day_labels': day_labels,
    }


def _accepted_premium(summary: InsuranceSummary) -> Decimal:
    """Σ премии accepted-свода по выбранной СК и варианту франшизы."""
    if summary.status != 'completed_accepted' or not summary.selected_company:
        return Decimal('0')
    variant = summary.selected_franchise_variant or 1
    total = Decimal('0')
    for offer in summary.offers.all():
        if not offer.is_valid or offer.company_name != summary.selected_company:
            continue
        premium = (
            offer.premium_with_franchise_2
            if variant == 2 and offer.premium_with_franchise_2 is not None
            else offer.premium_with_franchise_1
        )
        if premium is not None:
            total += premium
    return total


def _accepted_sum(summary: InsuranceSummary) -> Decimal:
    """Σ страховой суммы accepted-свода (по offers выбранной СК)."""
    if summary.status != 'completed_accepted' or not summary.selected_company:
        return Decimal('0')
    total = Decimal('0')
    for offer in summary.offers.all():
        if not offer.is_valid or offer.company_name != summary.selected_company:
            continue
        if offer.insurance_sum is not None:
            total += offer.insurance_sum
    return total


# --- Funnel / Time-to-* per summary -----------------------------------------


def _summary_time_to_metrics(summary: InsuranceSummary) -> dict[str, float | None]:
    """Считает доступные time-to-* интервалы (в часах)."""
    upload_at = summary.request.created_at if summary.request else None
    summary_at = summary.created_at
    sent_at = summary.sent_to_client_at
    completed_at = summary.completed_at if summary.status == 'completed_accepted' else None

    first_offer_at = None
    offers = list(summary.offers.all())
    if offers:
        valid_dates = [o.received_at for o in offers if o.received_at]
        if valid_dates:
            first_offer_at = min(valid_dates)

    return {
        'upload_to_summary_h': _hours_between(upload_at, summary_at),
        'summary_to_first_offer_h': _hours_between(summary_at, first_offer_at),
        'summary_to_sent_h': _hours_between(summary_at, sent_at),
        'sent_to_completed_h': _hours_between(sent_at, completed_at),
        'total_cycle_h': _hours_between(upload_at, completed_at),
    }


# --- Charts -----------------------------------------------------------------


def _daily_counts(request_qs: QuerySet, filters: dict) -> dict[str, list]:
    """Возвращает labels (даты) и series (по сотрудникам) для линии загрузок по дням."""
    if filters['start_date'] and filters['end_date']:
        start, end = filters['start_date'], filters['end_date']
    else:
        # period='all' и нет дат: показываем последние 90 дней.
        end = timezone.localdate()
        start = end - timedelta(days=89)

    days = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur += timedelta(days=1)

    rows = list(
        request_qs.values('created_at__date', 'created_by_id', 'created_by__username',
                          'created_by__first_name', 'created_by__last_name')
        .annotate(c=Count('id'))
    )
    # series_key → label
    series_keys: dict[str, str] = {}
    series_data: dict[str, dict[date, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        if row['created_by_id'] is None:
            key = '__unassigned__'
            label = 'Без автора'
        else:
            key = f"u{row['created_by_id']}"
            full = (f"{row['created_by__first_name']} {row['created_by__last_name']}").strip()
            label = full or row['created_by__username'] or f'user#{row["created_by_id"]}'
        series_keys[key] = label
        series_data[key][row['created_at__date']] += row['c']

    series = []
    for key, label in sorted(series_keys.items(), key=lambda kv: kv[1].lower()):
        series.append({
            'key': key,
            'label': label,
            'data': [series_data[key].get(d, 0) for d in days],
        })

    cumulative_total = []
    running = 0
    daily_total = []
    for d in days:
        day_sum = sum(series_data[k].get(d, 0) for k in series_data)
        running += day_sum
        cumulative_total.append(running)
        daily_total.append(day_sum)

    return {
        'labels': [d.isoformat() for d in days],
        'series': series,
        'cumulative_total': cumulative_total,
        'moving_average_28d': _moving_average(daily_total, window=28),
    }


def _weekly_stacked(request_qs: QuerySet) -> dict[str, list]:
    rows = list(
        request_qs.values('created_at__date', 'created_by_id', 'created_by__username',
                          'created_by__first_name', 'created_by__last_name')
    )
    if not rows:
        return {'labels': [], 'series': []}

    weeks: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    series_labels: dict[str, str] = {}
    week_set: set[str] = set()

    for row in rows:
        d = row['created_at__date']
        # Неделя как ISO «YYYY-Www»
        iso_year, iso_week, _ = d.isocalendar()
        week_key = f'{iso_year}-W{iso_week:02d}'
        week_set.add(week_key)
        if row['created_by_id'] is None:
            key = '__unassigned__'
            label = 'Без автора'
        else:
            key = f"u{row['created_by_id']}"
            full = (f"{row['created_by__first_name']} {row['created_by__last_name']}").strip()
            label = full or row['created_by__username'] or f'user#{row["created_by_id"]}'
        series_labels[key] = label
        weeks[week_key][key] += 1

    labels = sorted(week_set)
    series = []
    for key, label in sorted(series_labels.items(), key=lambda kv: kv[1].lower()):
        series.append({
            'key': key,
            'label': label,
            'data': [weeks[w].get(key, 0) for w in labels],
        })
    return {'labels': labels, 'series': series}


# --- Per-manager rows -------------------------------------------------------


def _build_manager_rows(filters: dict) -> tuple[list[dict], dict]:
    """Возвращает (rows, team_aggregate). Каждая строка — метрики на сотрудника."""
    request_qs = _build_request_qs(filters)
    summary_qs = _build_summary_qs(filters)

    # Группируем заявки по created_by
    requests_by_user: dict[int | None, list[InsuranceRequest]] = defaultdict(list)
    for req in request_qs:
        requests_by_user[req.created_by_id].append(req)

    summaries_by_user: dict[int | None, list[InsuranceSummary]] = defaultdict(list)
    for s in summary_qs:
        summaries_by_user[s.request.created_by_id if s.request else None].append(s)

    rows: list[dict] = []
    now = timezone.now()

    for user_id, requests in requests_by_user.items():
        user = requests[0].created_by if requests and requests[0].created_by_id else None
        summaries = summaries_by_user.get(user_id, [])

        accepted = [s for s in summaries if s.status == 'completed_accepted']
        rejected = [s for s in summaries if s.status == 'completed_rejected']
        completed = accepted + rejected
        active_summaries = [s for s in summaries if s.status in ACTIVE_SUMMARY_STATUSES]
        active_requests = [r for r in requests if r.status in ACTIVE_REQUEST_STATUSES]

        win_rate = (len(accepted) / len(completed) * 100) if completed else None

        premium_total = sum((_accepted_premium(s) for s in accepted), Decimal('0'))
        sum_total = sum((_accepted_sum(s) for s in accepted), Decimal('0'))
        avg_ticket = (premium_total / len(accepted)) if accepted else Decimal('0')

        # Time-to-*
        per_summary_metrics = [_summary_time_to_metrics(s) for s in summaries]
        cycle_values = [m['total_cycle_h'] for m in per_summary_metrics if m['total_cycle_h'] is not None]

        time_to = {
            'upload_to_summary_h': _avg(m['upload_to_summary_h'] for m in per_summary_metrics),
            'summary_to_first_offer_h': _avg(m['summary_to_first_offer_h'] for m in per_summary_metrics),
            'summary_to_sent_h': _avg(m['summary_to_sent_h'] for m in per_summary_metrics),
            'sent_to_completed_h': _avg(m['sent_to_completed_h'] for m in per_summary_metrics),
            'avg_cycle_h': _avg(cycle_values),
            'p50_cycle_h': _median(cycle_values),
            'p90_cycle_h': _percentile(cycle_values, 0.9),
        }

        # Просрочка по response_deadline (по активным заявкам)
        overdue_count = sum(
            1 for r in requests
            if r.response_deadline and r.response_deadline < now
            and r.status != 'emails_sent'
        )

        last_activity = max((r.created_at for r in requests), default=None)

        completeness = _aggregate_completeness(requests)
        portfolio = _portfolio_for_user(requests)
        patterns = _pattern_metrics_for_user(requests, now=now)

        rows.append({
            'user_id': user_id,
            'display': _user_display(user),
            'username': user.username if user else None,
            'is_unassigned': user_id is None,
            'requests_total': len(requests),
            'summaries_total': len(summaries),
            'accepted': len(accepted),
            'rejected': len(rejected),
            'win_rate': win_rate,
            'premium_total': premium_total,
            'sum_total': sum_total,
            'avg_ticket': avg_ticket,
            'time_to': time_to,
            'active_requests': len(active_requests),
            'active_summaries': len(active_summaries),
            'active_total': len(active_requests) + len(active_summaries),
            'overdue_count': overdue_count,
            'last_activity': last_activity,
            'completeness': completeness,
            'portfolio': portfolio,
            'patterns': patterns,
        })

    # Сортируем: сначала реальные сотрудники по объёму, в конце «Без автора»
    rows.sort(key=lambda r: (r['is_unassigned'], -r['requests_total'], r['display']))

    # Команда (агрегация)
    real_rows = [r for r in rows if not r['is_unassigned']]
    all_requests = [req for reqs in requests_by_user.values() for req in reqs]

    team = {
        'requests_total': sum(r['requests_total'] for r in rows),
        'summaries_total': sum(r['summaries_total'] for r in rows),
        'accepted': sum(r['accepted'] for r in rows),
        'rejected': sum(r['rejected'] for r in rows),
        'premium_total': sum((r['premium_total'] for r in rows), Decimal('0')),
        'sum_total': sum((r['sum_total'] for r in rows), Decimal('0')),
        'active_total': sum(r['active_total'] for r in rows),
        'overdue_count': sum(r['overdue_count'] for r in rows),
        'managers_count': sum(1 for r in rows if not r['is_unassigned']),
        'completeness': _aggregate_completeness(all_requests),
        'portfolio': _portfolio_for_user(all_requests),
    }
    completed = team['accepted'] + team['rejected']
    team['win_rate'] = (team['accepted'] / completed * 100) if completed else None
    team['avg_ticket'] = (team['premium_total'] / team['accepted']) if team['accepted'] else Decimal('0')

    # Composite quality-score (нужны командные ориентиры)
    team_max_volume = max((r['requests_total'] for r in real_rows), default=0)
    cycle_values = [r['time_to']['avg_cycle_h'] for r in real_rows if r['time_to']['avg_cycle_h']]
    team_best_cycle = min(cycle_values) if cycle_values else None
    for row in rows:
        row['quality_score'] = _compute_quality_score(row, team_max_volume, team_best_cycle)
    if team_max_volume > 0:
        # Команда — невзвешенное среднее по реальным сотрудникам
        scores = [r['quality_score'] for r in real_rows if r['quality_score'] is not None]
        team['quality_score'] = sum(scores) / len(scores) if scores else None
    else:
        team['quality_score'] = None

    # Радар-координаты на каждую строку
    for row in rows:
        row['radar'] = _radar_axes_for_row(row, team)

    return rows, team


def _compute_quality_score(row: dict, team_max_volume: int, team_best_cycle: float | None) -> float | None:
    """Composite-score 0..100 по 4 компонентам (см. план §«Решения», п.4)."""
    completeness = row['completeness']['overall_pct']
    if completeness is None:
        completeness = 0.0
    win_rate = row['win_rate'] or 0.0
    avg_cycle = row['time_to']['avg_cycle_h']
    if avg_cycle and team_best_cycle and avg_cycle > 0:
        speed = min(100.0, 100.0 * team_best_cycle / avg_cycle)
    else:
        speed = 0.0
    if team_max_volume > 0:
        volume = min(100.0, 100.0 * row['requests_total'] / team_max_volume)
    else:
        volume = 0.0

    score = (
        QUALITY_SCORE_WEIGHTS['completeness'] * completeness
        + QUALITY_SCORE_WEIGHTS['win_rate'] * win_rate
        + QUALITY_SCORE_WEIGHTS['speed'] * speed
        + QUALITY_SCORE_WEIGHTS['volume'] * volume
    )
    return round(max(0.0, min(100.0, score)), 1)


# --- Funnel -----------------------------------------------------------------


def _build_funnel(filters: dict) -> dict[str, Any]:
    request_qs = _build_request_qs(filters)
    summary_qs = _build_summary_qs(filters)

    total_requests = request_qs.count()
    request_with_summary = request_qs.filter(summary__isnull=False).count()
    summaries_ready_plus = summary_qs.filter(
        status__in=['ready', 'sent', 'completed_accepted', 'completed_rejected']
    ).count()
    summaries_sent_plus = summary_qs.filter(
        status__in=['sent', 'completed_accepted', 'completed_rejected']
    ).count()
    summaries_completed = summary_qs.filter(status__in=list(TERMINAL_SUMMARY_STATUSES)).count()
    summaries_accepted = summary_qs.filter(status='completed_accepted').count()

    stages = [
        {'key': 'uploaded', 'label': 'Загружено', 'count': total_requests},
        {'key': 'sent_emails', 'label': 'Письма отправлены', 'count': request_with_summary},
        {'key': 'ready', 'label': 'Свод готов', 'count': summaries_ready_plus},
        {'key': 'sent_to_client', 'label': 'Отправлен клиенту', 'count': summaries_sent_plus},
        {'key': 'completed', 'label': 'Сделка закрыта', 'count': summaries_completed},
        {'key': 'accepted', 'label': 'Акцепт', 'count': summaries_accepted},
    ]
    # Доля от первой стадии и drop-off от предыдущей
    base = max(stages[0]['count'], 1)
    prev = stages[0]['count']
    for stage in stages:
        stage['pct_of_top'] = round(stage['count'] / base * 100, 1)
        stage['drop_off'] = max(prev - stage['count'], 0)
        prev = stage['count']

    return {
        'stages': stages,
        'win_rate': round(summaries_accepted / summaries_completed * 100, 1) if summaries_completed else None,
    }


# --- Heatmaps ---------------------------------------------------------------


def _build_heatmap(
    pairs: Iterable[tuple[str, str]],
    *,
    top_limit: int = HEATMAP_TOP_LIMIT,
) -> dict[str, Any]:
    """Принимает итератор пар (manager_label, dim_value) и собирает heatmap.

    Возвращает {rows, columns, cells, max_value}.
    rows: [{'label': str, 'total': int, 'cells': [int, ...]}],
    columns: [str, ...] (топ-N по сумме, остальное в «Прочее»).
    """
    grid: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    column_totals: dict[str, int] = defaultdict(int)
    row_labels: set[str] = set()

    for manager_label, dim_value in pairs:
        if not dim_value:
            dim_value = 'не указано'
        grid[manager_label][dim_value] += 1
        column_totals[dim_value] += 1
        row_labels.add(manager_label)

    if not grid:
        return {'rows': [], 'columns': [], 'cells': [], 'max_value': 0}

    sorted_cols = sorted(column_totals.items(), key=lambda kv: (-kv[1], kv[0]))
    head = [c for c, _ in sorted_cols[:top_limit]]
    tail = [c for c, _ in sorted_cols[top_limit:]]

    columns = list(head)
    has_other = bool(tail)
    if has_other:
        columns.append('Прочее')

    rows = []
    max_value = 0
    for label in sorted(row_labels, key=str.lower):
        cells = []
        for col in head:
            value = grid[label].get(col, 0)
            cells.append(value)
            if value > max_value:
                max_value = value
        if has_other:
            other_sum = sum(grid[label].get(c, 0) for c in tail)
            cells.append(other_sum)
            if other_sum > max_value:
                max_value = other_sum
        rows.append({'label': label, 'total': sum(cells), 'cells': cells})

    return {'rows': rows, 'columns': columns, 'cells': len(rows) * len(columns), 'max_value': max_value}


def _compute_heatmaps(filters: dict) -> dict[str, dict]:
    """4 heatmap’а: manager × branch / insurance_type / alliance / selected_company."""
    request_qs = _build_request_qs(filters).select_related('created_by')
    summary_qs = _build_summary_qs(filters).select_related('request', 'request__created_by')

    pairs_branch = (
        (_user_display(r.created_by), r.branch)
        for r in request_qs
    )
    pairs_type = (
        (_user_display(r.created_by), r.insurance_type)
        for r in request_qs
    )
    pairs_alliance = (
        (_user_display(r.created_by), r.manager_name)
        for r in request_qs if r.manager_name
    )
    pairs_company = (
        (
            _user_display(s.request.created_by) if s.request else 'Без автора',
            s.selected_company,
        )
        for s in summary_qs
        if s.status == 'completed_accepted' and s.selected_company
    )

    return {
        'branch': _build_heatmap(pairs_branch),
        'insurance_type': _build_heatmap(pairs_type),
        'alliance_manager': _build_heatmap(pairs_alliance),
        'selected_company': _build_heatmap(pairs_company),
    }


# --- Алерты -----------------------------------------------------------------


def build_alerts(filters: dict, *,
                 no_activity_days: int = NO_ACTIVITY_DAYS_DEFAULT,
                 stuck_days: int = STUCK_SUMMARY_DAYS_DEFAULT) -> dict[str, list]:
    now = timezone.now()
    no_activity_threshold = now - timedelta(days=no_activity_days)
    stuck_threshold = now - timedelta(days=stuck_days)

    # Просроченные active-заявки (response_deadline < now и статус ещё не emails_sent)
    overdue_qs = (
        _build_request_qs(filters)
        .filter(response_deadline__lt=now)
        .exclude(status='emails_sent')
        .order_by('response_deadline')
    )
    overdue = [
        {
            'request_id': r.id,
            'dfa_number': r.dfa_number or f'#{r.id}',
            'client_name': r.client_name,
            'manager': _user_display(r.created_by),
            'manager_id': r.created_by_id,
            'deadline': r.response_deadline,
            'overdue_hours': round((now - r.response_deadline).total_seconds() / 3600, 1),
            'status': r.get_status_display(),
        }
        for r in overdue_qs[:50]
    ]

    # Зависшие сделки: активные своды, у которых updated_at старше порога.
    stuck_qs = (
        _build_summary_qs(filters)
        .filter(status__in=list(ACTIVE_SUMMARY_STATUSES))
        .filter(updated_at__lt=stuck_threshold)
        .order_by('updated_at')
    )
    stuck = [
        {
            'summary_id': s.id,
            'request_id': s.request_id,
            'dfa_number': (s.request.dfa_number if s.request else None) or f'#{s.id}',
            'client_name': s.request.client_name if s.request else '',
            'manager': _user_display(s.request.created_by) if s.request else 'Без автора',
            'manager_id': s.request.created_by_id if s.request else None,
            'days_inactive': (now - s.updated_at).days,
            'status': s.get_status_display(),
        }
        for s in stuck_qs[:50]
    ]

    # Сотрудники без активности: были заявки в окне, но ни одной свежее threshold.
    request_qs = _build_request_qs(filters)
    user_ids = [
        uid for uid in request_qs.values_list('created_by_id', flat=True).distinct()
        if uid is not None
    ]
    inactive_users = []
    for user_id in user_ids:
        recent = request_qs.filter(
            created_by_id=user_id, created_at__gt=no_activity_threshold
        ).exists()
        if recent:
            continue
        user = User.objects.filter(pk=user_id).first()
        if not user:
            continue
        last_req = (
            InsuranceRequest.objects.filter(created_by_id=user_id)
            .order_by('-created_at').first()
        )
        inactive_users.append({
            'user_id': user_id,
            'display': _user_display(user),
            'last_activity': last_req.created_at if last_req else None,
        })

    volume_drop = _volume_drop_alert(filters)
    cycle_outliers = _cycle_outlier_alert(filters)

    return {
        'overdue': overdue,
        'overdue_total': overdue_qs.count(),
        'stuck': stuck,
        'stuck_total': stuck_qs.count(),
        'inactive_users': inactive_users,
        'no_activity_days': no_activity_days,
        'stuck_days': stuck_days,
        'volume_drop': volume_drop,
        'cycle_outliers': cycle_outliers,
    }


def _volume_drop_alert(filters: dict) -> dict[str, Any]:
    """Сравнивает объём текущего окна с предыдущим равной длины.

    Алерт срабатывает на сотрудника, если падение ≤ VOLUME_DROP_THRESHOLD_PCT.
    Не работает для period='all' (нет смысла сравнивать с «прошлым всем временем»).
    """
    if not (filters['start_date'] and filters['end_date']):
        return {'enabled': False, 'reason': 'Период «всё время» — сравнение недоступно', 'flagged': []}

    span_days = (filters['end_date'] - filters['start_date']).days + 1
    prev_end = filters['start_date'] - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span_days - 1)

    prev_filters = dict(filters)
    prev_filters['start_date'] = prev_start
    prev_filters['end_date'] = prev_end

    cur_qs = _build_request_qs(filters)
    prev_qs = _build_request_qs(prev_filters)

    cur_counts: dict[int | None, int] = defaultdict(int)
    for uid, c in cur_qs.values_list('created_by_id').annotate(c=Count('id')):
        cur_counts[uid] = c
    prev_counts: dict[int | None, int] = defaultdict(int)
    for uid, c in prev_qs.values_list('created_by_id').annotate(c=Count('id')):
        prev_counts[uid] = c

    flagged = []
    for uid in set(list(cur_counts.keys()) + list(prev_counts.keys())):
        if uid is None:
            continue
        prev = prev_counts.get(uid, 0)
        cur = cur_counts.get(uid, 0)
        if prev < 3:  # слишком мало данных для сравнения
            continue
        change_pct = (cur - prev) / prev * 100
        if change_pct <= VOLUME_DROP_THRESHOLD_PCT:
            user = User.objects.filter(pk=uid).first()
            flagged.append({
                'user_id': uid,
                'display': _user_display(user),
                'previous_count': prev,
                'current_count': cur,
                'change_pct': round(change_pct, 1),
            })

    flagged.sort(key=lambda x: x['change_pct'])
    return {
        'enabled': True,
        'threshold_pct': VOLUME_DROP_THRESHOLD_PCT,
        'previous_window': {'start': prev_start, 'end': prev_end},
        'flagged': flagged,
    }


def _cycle_outlier_alert(filters: dict) -> dict[str, Any]:
    """Аномально длинные сделки. Threshold = max(p75 × factor, 7 дней).

    p75 устойчив к одиночным выбросам, в отличие от p90. Floor — чтобы при
    очень коротких циклах не ловить тривиальные отклонения.
    """
    summary_qs = (
        _build_summary_qs(filters)
        .filter(status='completed_accepted')
        .select_related('request', 'request__created_by')
        .prefetch_related('offers')
    )

    cycles = []
    per_summary = []
    for s in summary_qs:
        cycle = _hours_between(
            s.request.created_at if s.request else None,
            s.completed_at,
        )
        if cycle is None:
            continue
        cycles.append(cycle)
        per_summary.append((s, cycle))

    if not cycles:
        return {'threshold_h': None, 'team_baseline_h': None, 'outliers': []}

    team_p75 = _percentile(cycles, 0.75)
    if team_p75 is None or team_p75 <= 0:
        return {'threshold_h': None, 'team_baseline_h': None, 'outliers': []}

    threshold = max(team_p75 * CYCLE_OUTLIER_FACTOR, CYCLE_OUTLIER_FLOOR_HOURS)
    outliers = []
    for s, cycle in per_summary:
        if cycle <= threshold:
            continue
        outliers.append({
            'summary_id': s.id,
            'request_id': s.request_id,
            'dfa_number': (s.request.dfa_number if s.request else None) or f'#{s.id}',
            'client_name': s.request.client_name if s.request else '',
            'manager': _user_display(s.request.created_by) if s.request else 'Без автора',
            'manager_id': s.request.created_by_id if s.request else None,
            'cycle_hours': round(cycle, 1),
            'cycle_days': round(cycle / 24, 1),
        })
    outliers.sort(key=lambda x: -x['cycle_hours'])
    return {
        'threshold_h': round(threshold, 1),
        'team_baseline_h': round(team_p75, 1),
        'outliers': outliers[:HEATMAP_OUTLIER_LIMIT],
    }


# --- Available filter values ------------------------------------------------


def _available_filters() -> dict[str, list]:
    branches = list(
        InsuranceRequest.objects
        .exclude(branch__isnull=True).exclude(branch__exact='')
        .values_list('branch', flat=True).distinct().order_by('branch')
    )
    insurance_types = [v for v, _ in InsuranceRequest.INSURANCE_TYPE_CHOICES]
    deal_statuses = [{'value': v, 'label': l} for v, l in InsuranceRequest.DEAL_STATUS_CHOICES]
    users = [
        {'id': u.pk, 'display': _user_display(u)}
        for u in User.objects.filter(insurancerequest__isnull=False).distinct().order_by('username')
    ]
    return {
        'branches': branches,
        'insurance_types': insurance_types,
        'deal_statuses': deal_statuses,
        'users': users,
    }


# --- Public API -------------------------------------------------------------


def build_overview_payload(filters: dict) -> dict[str, Any]:
    rows, team = _build_manager_rows(filters)
    funnel = _build_funnel(filters)
    request_qs = _build_request_qs(filters)
    daily = _daily_counts(request_qs, filters)
    weekly = _weekly_stacked(request_qs)

    if filters['start_date'] and filters['end_date']:
        days_span = max((filters['end_date'] - filters['start_date']).days + 1, 1)
    else:
        days_span = max(team['requests_total'] and 365 or 1, 1)
    avg_per_day = round(team['requests_total'] / days_span, 2) if days_span else 0
    avg_per_week = round(avg_per_day * 7, 2)
    avg_per_month = round(avg_per_day * 30, 2)

    heatmaps = _compute_heatmaps(filters)
    backlog = _backlog_age_buckets(filters)
    day_hour = _day_hour_heatmap(filters)
    trend = _team_trend(filters)
    team_radar = _team_radar(rows)

    payload = {
        'filters': filters,
        'kpi': {
            'total_requests': team['requests_total'],
            'total_summaries': team['summaries_total'],
            'accepted': team['accepted'],
            'win_rate': team['win_rate'],
            'premium_total': team['premium_total'],
            'avg_ticket': team['avg_ticket'],
            'active_total': team['active_total'],
            'overdue_count': team['overdue_count'],
            'managers_count': team['managers_count'],
            'avg_per_day': avg_per_day,
            'avg_per_week': avg_per_week,
            'avg_per_month': avg_per_month,
            'team_quality_score': team.get('quality_score'),
        },
        'rows': rows,
        'team': team,
        'funnel': funnel,
        'charts': {
            'daily': daily,
            'weekly': weekly,
        },
        'heatmaps': heatmaps,
        'backlog': backlog,
        'day_hour': day_hour,
        'trend': trend,
        'team_radar': team_radar,
        'available_filters': _available_filters(),
    }
    radar_payload = {
        'axes': [label for _, label in RADAR_AXES],
        'team': team_radar,
        'managers': [
            {
                'user_id': r['user_id'],
                'display': r['display'],
                'values': [r['radar'][k] for k, _ in RADAR_AXES],
            }
            for r in rows if not r['is_unassigned']
        ],
    }
    payload['radar'] = radar_payload
    payload['charts_json'] = json.dumps(
        {
            'daily': daily,
            'weekly': weekly,
            'day_hour': day_hour,
            'radar': radar_payload,
        },
        ensure_ascii=False,
        default=str,
    )
    return payload


RADAR_AXES = (
    ('volume', 'Volume'),
    ('speed', 'Speed'),
    ('win_rate', 'Win-rate'),
    ('quality', 'Quality'),
    ('money', 'Money'),
    ('activity', 'Activity'),
)


def _radar_axes_for_row(row: dict, team: dict) -> dict[str, float]:
    """Нормализованные оси радара 0..100."""
    real_max_volume = max((team['requests_total'], 1), key=lambda x: x)
    if isinstance(real_max_volume, tuple):
        real_max_volume = team['requests_total'] or 1
    volume = (row['requests_total'] / real_max_volume * 100) if real_max_volume else 0.0
    win_rate = row.get('win_rate') or 0.0

    avg_cycle = row['time_to'].get('avg_cycle_h')
    # speed: чем меньше cycle, тем больше score. Базис — 24 часа = 100 баллов.
    if avg_cycle and avg_cycle > 0:
        speed = min(100.0, 24.0 / avg_cycle * 100.0)
    else:
        speed = 0.0

    quality = row.get('quality_score') or 0.0

    team_premium = float(team.get('premium_total') or 0)
    row_premium = float(row.get('premium_total') or 0)
    money = (row_premium / team_premium * 100) if team_premium else 0.0

    days_since_last = (row.get('patterns') or {}).get('days_since_last')
    if days_since_last is None:
        activity = 0.0
    else:
        # Чем меньше дней без активности, тем выше. 0 дней → 100, 14 дней → 0.
        activity = max(0.0, min(100.0, 100.0 - (days_since_last / 14.0 * 100.0)))

    return {
        'volume': round(min(100.0, volume), 1),
        'speed': round(speed, 1),
        'win_rate': round(min(100.0, win_rate), 1),
        'quality': round(quality, 1),
        'money': round(min(100.0, money), 1),
        'activity': round(activity, 1),
    }


def _team_radar(rows: list[dict]) -> dict[str, float]:
    """Среднее по реальным сотрудникам — для overlay-фигуры на радаре."""
    real = [r for r in rows if not r['is_unassigned'] and r.get('radar')]
    if not real:
        return {axis_key: 0.0 for axis_key, _ in RADAR_AXES}
    out = {}
    for axis_key, _ in RADAR_AXES:
        vals = [r['radar'][axis_key] for r in real]
        out[axis_key] = round(sum(vals) / len(vals), 1)
    return out


def _personal_status_events(user_id: int, *, limit: int = 50) -> list[dict[str, Any]]:
    """Возвращает события смены статуса заявок и сводов сотрудника.

    События сразу обогащаем display-данными цели (DFA, клиент), чтобы
    не дёргать БД из шаблона.
    """
    from django.contrib.contenttypes.models import ContentType
    from ..models import StatusEvent

    request_ct = ContentType.objects.get_for_model(InsuranceRequest)
    summary_ct = ContentType.objects.get_for_model(InsuranceSummary)

    request_ids = list(InsuranceRequest.objects.filter(created_by_id=user_id).values_list('id', flat=True))
    summary_ids = list(
        InsuranceSummary.objects.filter(request__created_by_id=user_id).values_list('id', flat=True)
    )

    events = StatusEvent.objects.filter(
        Q(content_type=request_ct, object_id__in=request_ids)
        | Q(content_type=summary_ct, object_id__in=summary_ids)
    ).order_by('-changed_at')[:limit]

    request_map = {
        r.id: r
        for r in InsuranceRequest.objects.filter(id__in=request_ids).only(
            'id', 'dfa_number', 'client_name', 'status'
        )
    }
    summary_map = {
        s.id: s
        for s in InsuranceSummary.objects.filter(id__in=summary_ids).select_related('request').only(
            'id', 'request_id', 'request__dfa_number', 'request__client_name', 'status'
        )
    }

    out: list[dict[str, Any]] = []
    for evt in events:
        kind: str
        target_label: str
        target_id: int | None = None
        request_id: int | None = None
        if evt.content_type_id == request_ct.id:
            kind = 'request'
            req = request_map.get(evt.object_id)
            if req:
                target_id = req.id
                request_id = req.id
                target_label = req.dfa_number or f'#{req.id}'
                client_name = req.client_name
            else:
                target_label = f'#{evt.object_id}'
                client_name = ''
        elif evt.content_type_id == summary_ct.id:
            kind = 'summary'
            sm = summary_map.get(evt.object_id)
            if sm:
                target_id = sm.id
                request_id = sm.request_id
                target_label = (sm.request.dfa_number if sm.request else None) or f'#{sm.id}'
                client_name = sm.request.client_name if sm.request else ''
            else:
                target_label = f'#{evt.object_id}'
                client_name = ''
        else:
            continue

        out.append({
            'changed_at': evt.changed_at,
            'kind': kind,
            'target_id': target_id,
            'request_id': request_id,
            'target_label': target_label,
            'client_name': client_name,
            'from_status': evt.from_status,
            'to_status': evt.to_status,
        })
    return out


def build_manager_profile_payload(user_id: int, filters: dict) -> dict[str, Any]:
    """Полное досье сотрудника.

    Возвращает: profile (User-данные), kpi (включая радар + сравнение с командой),
    активные/просроченные/зависшие, топ-5 СК и филиалов, timeline.
    """
    overview = build_overview_payload(filters)

    user = User.objects.filter(pk=user_id).first()
    row = next(
        (r for r in overview['rows'] if not r['is_unassigned'] and r['user_id'] == user_id),
        None,
    )
    if user is None and row is None:
        return {
            'filters': filters,
            'user_id': user_id,
            'user': None,
            'row': None,
            'team': overview['team'],
            'available_filters': overview['available_filters'],
            'not_found': True,
        }

    now = timezone.now()
    request_qs = InsuranceRequest.objects.filter(created_by_id=user_id)
    if filters['start_date']:
        request_qs = request_qs.filter(created_at__date__gte=filters['start_date'])
    if filters['end_date']:
        request_qs = request_qs.filter(created_at__date__lte=filters['end_date'])

    summary_qs = (
        InsuranceSummary.objects
        .filter(request__created_by_id=user_id)
        .select_related('request')
        .prefetch_related('offers')
    )
    if filters['start_date']:
        summary_qs = summary_qs.filter(request__created_at__date__gte=filters['start_date'])
    if filters['end_date']:
        summary_qs = summary_qs.filter(request__created_at__date__lte=filters['end_date'])

    # Активные сделки
    active_summaries = summary_qs.filter(status__in=list(ACTIVE_SUMMARY_STATUSES)).order_by('-updated_at')
    active_requests = request_qs.filter(status__in=list(ACTIVE_REQUEST_STATUSES)).order_by('-created_at')

    active_list = []
    for r in active_requests[:25]:
        active_list.append({
            'kind': 'request',
            'id': r.id,
            'dfa_number': r.dfa_number or f'#{r.id}',
            'client_name': r.client_name,
            'status': r.get_status_display(),
            'updated_at': r.updated_at,
            'age_days': (now - r.created_at).days if r.created_at else None,
        })
    for s in active_summaries[:25]:
        active_list.append({
            'kind': 'summary',
            'id': s.id,
            'request_id': s.request_id,
            'dfa_number': (s.request.dfa_number if s.request else None) or f'#{s.id}',
            'client_name': s.request.client_name if s.request else '',
            'status': s.get_status_display(),
            'updated_at': s.updated_at,
            'age_days': (now - (s.updated_at or s.created_at)).days,
        })

    # Просроченные (response_deadline < now & status != emails_sent)
    overdue_qs = (
        request_qs.filter(response_deadline__lt=now)
        .exclude(status='emails_sent')
        .order_by('response_deadline')
    )
    overdue = [
        {
            'request_id': r.id,
            'dfa_number': r.dfa_number or f'#{r.id}',
            'client_name': r.client_name,
            'deadline': r.response_deadline,
            'overdue_hours': round((now - r.response_deadline).total_seconds() / 3600, 1),
            'status': r.get_status_display(),
        }
        for r in overdue_qs[:25]
    ]

    # Зависшие
    stuck_threshold = now - timedelta(days=STUCK_SUMMARY_DAYS_DEFAULT)
    stuck_qs = (
        summary_qs.filter(status__in=list(ACTIVE_SUMMARY_STATUSES))
        .filter(updated_at__lt=stuck_threshold)
        .order_by('updated_at')
    )
    stuck = [
        {
            'summary_id': s.id,
            'request_id': s.request_id,
            'dfa_number': (s.request.dfa_number if s.request else None) or f'#{s.id}',
            'client_name': s.request.client_name if s.request else '',
            'days_inactive': (now - s.updated_at).days,
            'status': s.get_status_display(),
        }
        for s in stuck_qs[:25]
    ]

    # Топ-5 СК (по accepted)
    company_counts: dict[str, int] = defaultdict(int)
    for s in summary_qs.filter(status='completed_accepted'):
        if s.selected_company:
            company_counts[s.selected_company] += 1
    top_companies = [
        {'name': name, 'count': cnt}
        for name, cnt in sorted(company_counts.items(), key=lambda x: -x[1])[:5]
    ]

    # Топ-5 филиалов (по всем заявкам)
    branch_counts: dict[str, int] = defaultdict(int)
    for r in request_qs:
        branch_counts[r.branch or 'не указан'] += 1
    top_branches = [
        {'name': name, 'count': cnt}
        for name, cnt in sorted(branch_counts.items(), key=lambda x: -x[1])[:5]
    ]

    timeline = _personal_status_events(user_id, limit=50)

    # Self-benchmark vs предыдущий период
    self_benchmark = None
    if filters['start_date'] and filters['end_date']:
        span_days = (filters['end_date'] - filters['start_date']).days + 1
        prev_end = filters['start_date'] - timedelta(days=1)
        prev_start = prev_end - timedelta(days=span_days - 1)
        prev_qs = InsuranceRequest.objects.filter(
            created_by_id=user_id,
            created_at__date__gte=prev_start,
            created_at__date__lte=prev_end,
        )
        prev_count = prev_qs.count()
        cur_count = request_qs.count()
        if prev_count:
            change_pct = round((cur_count - prev_count) / prev_count * 100, 1)
        else:
            change_pct = None
        self_benchmark = {
            'previous_window': {'start': prev_start, 'end': prev_end},
            'previous_count': prev_count,
            'current_count': cur_count,
            'change_pct': change_pct,
        }

    return {
        'filters': filters,
        'user_id': user_id,
        'user': user,
        'row': row,
        'team': overview['team'],
        'team_radar': overview['team_radar'],
        'active_items': active_list,
        'overdue': overdue,
        'stuck': stuck,
        'top_companies': top_companies,
        'top_branches': top_branches,
        'timeline': timeline,
        'self_benchmark': self_benchmark,
        'available_filters': overview['available_filters'],
        'radar_json': json.dumps(
            {
                'axes': [label for _, label in RADAR_AXES],
                'team': overview['team_radar'],
                'manager': row['radar'] if row else None,
                'display': row['display'] if row else (user.get_full_name() if user else f'#{user_id}'),
            },
            ensure_ascii=False,
            default=str,
        ),
    }


def build_compare_payload(user_ids: list[int], filters: dict) -> dict[str, Any]:
    """Side-by-side сравнение нескольких сотрудников.

    Если user_ids пуст — берём всех активных сотрудников за период.
    """
    overview = build_overview_payload(filters)
    rows = overview['rows']
    real_rows = [r for r in rows if not r['is_unassigned']]

    if user_ids:
        selected = [r for r in real_rows if r['user_id'] in user_ids]
    else:
        selected = real_rows

    radar_data = {
        'axes': [label for _, label in RADAR_AXES],
        'team': overview.get('team_radar', {}),
        'managers': [
            {
                'user_id': r['user_id'],
                'display': r['display'],
                'values': [r['radar'][k] for k, _ in RADAR_AXES],
            }
            for r in selected
        ],
    }
    return {
        'filters': filters,
        'user_ids': user_ids,
        'managers': selected,
        'available_users': overview['available_filters']['users'],
        'team': overview['team'],
        'radar': radar_data,
        'radar_json': json.dumps(radar_data, ensure_ascii=False, default=str),
    }


def export_overview_xlsx(filters: dict):
    """Многолистовой XLSX обзора. Возвращает BytesIO."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    payload = build_overview_payload(filters)
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')

    def _set_header(ws, headers: list[str]):
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

    # Sheet 1: KPI
    ws_kpi = wb.active
    ws_kpi.title = 'KPI'
    ws_kpi.append(['Метрика', 'Значение'])
    for cell in ws_kpi[1]:
        cell.font = header_font
        cell.fill = header_fill
    kpi = payload['kpi']
    rows = [
        ('Заявок', kpi['total_requests']),
        ('Сводов', kpi['total_summaries']),
        ('Акцепт', kpi['accepted']),
        ('Win-rate, %', kpi['win_rate']),
        ('Σ премии', float(kpi['premium_total'] or 0)),
        ('Средний чек', float(kpi['avg_ticket'] or 0)),
        ('Активно сейчас', kpi['active_total']),
        ('Просрочек', kpi['overdue_count']),
        ('Активных сотрудников', kpi['managers_count']),
        ('Среднее заявок/день', kpi['avg_per_day']),
        ('Среднее заявок/неделя', kpi['avg_per_week']),
        ('Среднее заявок/месяц', kpi['avg_per_month']),
        ('Quality-score команды', kpi.get('team_quality_score')),
    ]
    for label, value in rows:
        ws_kpi.append([label, value])

    # Sheet 2: Сотрудники
    ws_mgr = wb.create_sheet('Сотрудники')
    _set_header(ws_mgr, [
        'Сотрудник', 'Заявок', 'Сводов', 'Акцепт', 'Не будет',
        'Win-rate, %', 'Σ премии', 'Средний чек', 'Avg цикл, ч',
        'Активно', 'Просрочек', 'Quality',
    ])
    for r in payload['rows']:
        ws_mgr.append([
            r['display'],
            r['requests_total'],
            r['summaries_total'],
            r['accepted'],
            r['rejected'],
            r['win_rate'],
            float(r['premium_total'] or 0),
            float(r.get('avg_ticket') or 0),
            r['time_to'].get('avg_cycle_h'),
            r['active_total'],
            r['overdue_count'],
            r.get('quality_score'),
        ])

    # Sheet 3: Качество данных
    ws_q = wb.create_sheet('Качество данных')
    _set_header(ws_q, [
        'Сотрудник', 'База, %', 'КАСКО-доп, %', 'Имущ.-доп, %',
        'Общая, %', 'Битый ИНН', 'Тип «другое»', 'Отредактировано',
    ])
    for r in payload['rows']:
        c = r['completeness']
        ws_q.append([
            r['display'],
            c['base_pct'], c['casco_pct'], c['property_pct'], c['overall_pct'],
            c['inn_invalid_count'], c['other_type_count'], c['edited_count'],
        ])

    # Sheet 4: Funnel
    ws_f = wb.create_sheet('Funnel')
    _set_header(ws_f, ['Стадия', 'Количество', '% от загрузок', 'Drop-off'])
    for stage in payload['funnel']['stages']:
        ws_f.append([stage['label'], stage['count'], stage['pct_of_top'], stage['drop_off']])

    # Sheet 5: Heatmap branch
    for hm_key, hm_label in [
        ('branch', 'Heatmap - Филиалы'),
        ('insurance_type', 'Heatmap - Типы'),
        ('alliance_manager', 'Heatmap - Менеджер Альянса'),
        ('selected_company', 'Heatmap - Выбранная СК'),
    ]:
        hm = payload['heatmaps'].get(hm_key) or {}
        ws_h = wb.create_sheet(hm_label[:31])  # Excel sheet name limit
        if not hm.get('rows'):
            ws_h.append(['Нет данных'])
            continue
        _set_header(ws_h, ['Сотрудник'] + list(hm['columns']) + ['Σ'])
        for hrow in hm['rows']:
            ws_h.append([hrow['label']] + list(hrow['cells']) + [hrow['total']])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_manager_dossier_xlsx(user_id: int, filters: dict):
    """XLSX-досье одного сотрудника."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    profile = build_manager_profile_payload(user_id, filters)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')

    def _set_header(ws, headers: list[str]):
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

    row = profile.get('row')
    user = profile.get('user')
    display = (row['display'] if row else
               (user.get_full_name() if user else f'user#{user_id}'))

    ws_kpi = wb.active
    ws_kpi.title = 'KPI'
    ws_kpi.append(['Сотрудник', display])
    if user:
        ws_kpi.append(['Username', user.username])
        ws_kpi.append(['Email', user.email])
    if row:
        ws_kpi.append(['Заявок', row['requests_total']])
        ws_kpi.append(['Сводов', row['summaries_total']])
        ws_kpi.append(['Акцепт', row['accepted']])
        ws_kpi.append(['Не будет', row['rejected']])
        ws_kpi.append(['Win-rate, %', row['win_rate']])
        ws_kpi.append(['Σ премии', float(row['premium_total'] or 0)])
        ws_kpi.append(['Σ страховой суммы', float(row['sum_total'] or 0)])
        ws_kpi.append(['Средний чек', float(row['avg_ticket'] or 0)])
        ws_kpi.append(['Avg цикл, ч', row['time_to'].get('avg_cycle_h')])
        ws_kpi.append(['P50 цикла, ч', row['time_to'].get('p50_cycle_h')])
        ws_kpi.append(['P90 цикла, ч', row['time_to'].get('p90_cycle_h')])
        ws_kpi.append(['Активно сейчас', row['active_total']])
        ws_kpi.append(['Просрочек', row['overdue_count']])
        ws_kpi.append(['Quality-score', row.get('quality_score')])

    ws_active = wb.create_sheet('Активные')
    _set_header(ws_active, ['Тип', 'ID', 'DFA', 'Клиент', 'Статус', 'Возраст, дн.'])
    for item in profile.get('active_items') or []:
        ws_active.append([
            'Заявка' if item['kind'] == 'request' else 'Свод',
            item['id'], item['dfa_number'], item['client_name'],
            item['status'], item.get('age_days'),
        ])

    ws_over = wb.create_sheet('Просроченные')
    _set_header(ws_over, ['DFA', 'Клиент', 'Статус', 'Дедлайн', 'Просрочка, ч'])
    for item in profile.get('overdue') or []:
        ws_over.append([
            item['dfa_number'], item['client_name'], item['status'],
            item['deadline'], item['overdue_hours'],
        ])

    ws_stuck = wb.create_sheet('Зависшие')
    _set_header(ws_stuck, ['DFA', 'Клиент', 'Статус', 'Дней без движения'])
    for item in profile.get('stuck') or []:
        ws_stuck.append([
            item['dfa_number'], item['client_name'], item['status'], item['days_inactive'],
        ])

    ws_top = wb.create_sheet('Топ СК и Филиалы')
    ws_top.append(['Топ СК (accepted)', '', 'Топ Филиалы'])
    for cell in ws_top[1]:
        cell.font = header_font
    top_companies = profile.get('top_companies') or []
    top_branches = profile.get('top_branches') or []
    for i in range(max(len(top_companies), len(top_branches), 1)):
        c = top_companies[i] if i < len(top_companies) else {'name': '', 'count': ''}
        b = top_branches[i] if i < len(top_branches) else {'name': '', 'count': ''}
        ws_top.append([f'{c["name"]} ({c["count"]})' if c['name'] else '', '',
                       f'{b["name"]} ({b["count"]})' if b['name'] else ''])

    ws_tl = wb.create_sheet('Активность')
    _set_header(ws_tl, ['Когда', 'Тип', 'DFA', 'Клиент', 'Изменение'])
    for evt in profile.get('timeline') or []:
        ws_tl.append([
            evt['changed_at'].strftime('%d.%m.%Y %H:%M') if evt.get('changed_at') else '',
            'Заявка' if evt['kind'] == 'request' else 'Свод',
            evt['target_label'], evt['client_name'],
            f'{evt["from_status"] or "—"} → {evt["to_status"]}',
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_leaderboard_payload(filters: dict) -> dict[str, Any]:
    """Рейтинг по composite quality-score. Только реальные сотрудники."""
    overview = build_overview_payload(filters)
    real = [r for r in overview['rows'] if not r['is_unassigned']]

    ranked = sorted(
        real,
        key=lambda r: (r.get('quality_score') is None, -(r.get('quality_score') or 0)),
    )
    rows = []
    for idx, r in enumerate(ranked, start=1):
        rows.append({
            'rank': idx,
            'user_id': r['user_id'],
            'display': r['display'],
            'username': r.get('username'),
            'quality_score': r.get('quality_score'),
            'breakdown': {
                'completeness': (r.get('completeness') or {}).get('overall_pct'),
                'win_rate': r.get('win_rate'),
                'avg_cycle_h': (r.get('time_to') or {}).get('avg_cycle_h'),
                'requests_total': r['requests_total'],
                'premium_total': r.get('premium_total'),
                'accepted': r.get('accepted'),
            },
        })

    return {
        'filters': filters,
        'rows': rows,
        'team_score': overview['team'].get('quality_score'),
        'weights': QUALITY_SCORE_WEIGHTS,
    }
