"""
Сервисы для работы со сводами предложений
"""

import logging
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, Any, List
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from django.conf import settings
from django.db import transaction
from django.core.exceptions import ValidationError

from ..models import InsuranceSummary, InsuranceOffer
from ..exceptions import DuplicateOfferError
from insurance_requests.models import InsuranceRequest


logger = logging.getLogger(__name__)


class ExcelExportServiceError(Exception):
    """Базовое исключение для ошибок ExcelExportService"""
    pass


class TemplateNotFoundError(ExcelExportServiceError):
    """Исключение для случаев, когда шаблон Excel не найден"""
    pass


class InvalidSummaryDataError(ExcelExportServiceError):
    """Исключение для случаев, когда данные свода некорректны"""
    pass


# Исключения для обработки Excel файлов с ответами компаний
class ExcelProcessingError(Exception):
    """Базовое исключение для ошибок обработки Excel файлов"""
    pass


class InvalidFileFormatError(ExcelProcessingError):
    """Ошибка неверного формата файла"""
    pass


class MissingDataError(ExcelProcessingError):
    """Ошибка отсутствующих данных"""
    def __init__(self, missing_cells):
        self.missing_cells = missing_cells
        super().__init__(f"Отсутствуют данные в ячейках: {', '.join(missing_cells)}")


class InvalidDataError(ExcelProcessingError):
    """Ошибка некорректных данных"""
    def __init__(self, field_name, value, expected_format):
        self.field_name = field_name
        self.value = value
        self.expected_format = expected_format
        super().__init__(f"Некорректное значение в поле '{field_name}': '{value}'. Ожидается: {expected_format}")


class RowProcessingError(ExcelProcessingError):
    """Ошибка обработки конкретной строки"""
    def __init__(self, row_number, field_name, error_message, cell_address=None):
        self.row_number = row_number
        self.field_name = field_name
        self.cell_address = cell_address
        self.error_message = error_message
        
        # Формируем детальное сообщение об ошибке
        if cell_address:
            super().__init__(f"Ошибка в строке {row_number}, ячейка {cell_address}, поле '{field_name}': {error_message}")
        else:
            super().__init__(f"Ошибка в строке {row_number}, поле '{field_name}': {error_message}")


class ExcelExportService:
    """Сервис для генерации Excel-файлов сводов предложений"""
    
    # Константы для маппинга колонок и строк данных компаний
    # Полный шаблон (существующий)
    FULL_TEMPLATE_COLUMNS = {
        'company_name': 'A',      # Название компании
        'year': 'B',              # Год страхования
        'insurance_sum': 'C',     # Страховая сумма
        'rate_1': 'E',            # Страховой тариф-1 (в процентах)
        'premium_1': 'F',         # Премия-1
        'franchise_1': 'G',       # Франшиза-1
        'installment_1': 'H',     # Рассрочка-1
        'premium_1_summary': 'I', # Сумма премий-1
        'rate_2': 'K',            # Страховой тариф-2 (в процентах)
        'premium_2': 'L',         # Премия-2
        'franchise_2': 'M',       # Франшиза-2
        'installment_2': 'N',     # Рассрочка-2
        'premium_2_summary': 'O', # Сумма премий-2
        'notes': 'Q'              # Примечания
    }
    
    # Упрощенный шаблон (новый)
    SIMPLIFIED_TEMPLATE_COLUMNS = {
        'company_name': 'A',      # Название компании
        'year': 'B',              # Год страхования
        'insurance_sum': 'C',     # Страховая сумма
        'rate_1': 'E',            # Страховой тариф-1 (в процентах)
        'premium_1': 'F',         # Премия-1
        'franchise_1': 'G',       # Франшиза-1
        'installment_1': 'H',     # Рассрочка-1
        'premium_1_summary': 'I', # Сумма премий-1
        'notes': 'K'              # Примечания (сдвинутая колонка)
    }
    
    # Обратная совместимость
    COMPANY_DATA_COLUMNS = FULL_TEMPLATE_COLUMNS
    
    SEPARATOR_ROW = 9  # Номер строки-разделителя для копирования
    FIRST_DATA_ROW = 10  # Первая строка для данных компаний
    
    # Константы для технического листа (tech_info)
    TECH_INFO_CELLS = {
        'request_number': 'B5',      # Номер заявки
        'client_name': 'B6',         # Название клиента
        'client_inn': 'B7',          # ИНН клиента
        'branch': 'B8',              # Филиал
        'insurance_type': 'B11',     # Тип страхования
        'vehicle_info': 'B12',       # Информация о предмете лизинга
        'creditor_bank': 'B13',      # Банк-кредитор
        'franchise_info': 'B14',     # Информация о франшизе
        'installment_info': 'B15',   # Информация о рассрочке
        'usage_purposes': 'B16',     # Цели использования
        'letter_text': 'B17',        # Текст письма заявки
        'autostart_info': 'B21',     # Информация об автозапуске
        'key_completeness': 'B22',   # Комплектность ключей
        'pts_psm': 'B23',           # ПТС/ПСМ
        'telematics_complex': 'B24', # Телематический комплекс
        'insurance_territory': 'B27', # Территория страхования
        'transportation_info': 'B28', # Информация о перевозке
        'construction_work_info': 'B29', # Информация о СМР
    }
    
    # Константы для обработки граничных случаев и валидации
    MAX_ROWS_LIMIT = 1000  # Максимальное количество строк для предотвращения зависания
    MAX_COMPANIES_LIMIT = 100  # Максимальное количество компаний
    MAX_YEARS_PER_COMPANY = 10  # Максимальное количество лет на компанию
    MAX_NOTES_LENGTH = 1000  # Максимальная длина примечаний
    MIN_INSURANCE_SUM = Decimal('1')  # Минимальная страховая сумма
    MAX_INSURANCE_SUM = Decimal('1000000000')  # Максимальная страховая сумма (1 млрд)
    
    def __init__(self, template_path: str):
        """
        Инициализация сервиса
        
        Args:
            template_path: Путь к файлу шаблона Excel
        """
        self.template_path = Path(template_path)
        self._validate_template()
    
    def _validate_template(self) -> None:
        """Проверяет доступность шаблона Excel"""
        if not self.template_path.exists():
            error_msg = f"Шаблон Excel не найден по пути: {self.template_path}"
            logger.error(error_msg)
            raise TemplateNotFoundError(error_msg)
        
        if not self.template_path.is_file():
            error_msg = f"Путь к шаблону не является файлом: {self.template_path}"
            logger.error(error_msg)
            raise TemplateNotFoundError(error_msg)
    
    def generate_summary_excel(self, summary: InsuranceSummary, is_client_version: bool = False) -> BytesIO:
        """
        Генерирует Excel-файл для свода предложений
        
        Args:
            summary: Объект свода предложений
            is_client_version: Если True, генерируется клиентская версия (без технического листа)
            
        Returns:
            BytesIO: Сгенерированный Excel-файл в памяти
            
        Raises:
            InvalidSummaryDataError: Если данные свода некорректны
            ExcelExportServiceError: При ошибках работы с Excel
        """
        version_type = "клиентского" if is_client_version else "полного"
        logger.info(f"Начинаем генерацию {version_type} Excel-файла для свода ID: {summary.id}")
        
        try:
            # Валидация данных свода
            self._validate_summary_data(summary)
            
            # Определение типа шаблона на основе данных свода
            template_type = self._determine_template_type_safe(summary)
            
            # Загрузка соответствующего шаблона (клиентского или обычного)
            workbook = self._load_template(template_type, is_client_version)
            
            # Заполнение данными с учетом типа шаблона
            self._fill_template_data(workbook, summary, template_type, is_client_version)
            
            # Сохранение в память
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"{version_type.capitalize()} Excel-файл успешно сгенерирован для свода ID: {summary.id}")
            return excel_buffer
            
        except (InvalidSummaryDataError, TemplateNotFoundError):
            # Переброс известных исключений
            raise
        except Exception as e:
            error_msg = f"Ошибка при генерации {version_type} Excel-файла для свода ID {summary.id}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _validate_summary_data(self, summary: InsuranceSummary) -> None:
        """
        Валидирует данные свода перед генерацией
        
        Args:
            summary: Объект свода предложений
            
        Raises:
            InvalidSummaryDataError: Если данные некорректны
        """
        missing_fields = []
        
        # Проверка наличия связанной заявки
        if not summary.request:
            missing_fields.append("связанная заявка")
        else:
            request = summary.request
            
            # Проверка обязательных полей заявки
            if not request.dfa_number or request.dfa_number.strip() == '':
                missing_fields.append("номер заявки (dfa_number)")
            
            if not request.vehicle_info or request.vehicle_info.strip() == '':
                missing_fields.append("информация о предмете лизинга (vehicle_info)")
            
            if not request.client_name or request.client_name.strip() == '':
                missing_fields.append("название клиента (client_name)")
        
        if missing_fields:
            error_msg = f"Отсутствуют обязательные данные: {', '.join(missing_fields)}"
            logger.error(f"Валидация данных свода ID {summary.id} не пройдена: {error_msg}")
            raise InvalidSummaryDataError(error_msg)
        
        logger.debug(f"Валидация данных свода ID {summary.id} успешно пройдена")
    
    def _load_template(self, template_type: str = 'full', is_client_version: bool = False) -> Workbook:
        """
        Загружает соответствующий шаблон Excel из файла
        
        Args:
            template_type: Тип шаблона ('full' или 'simplified')
            is_client_version: Если True, загружается клиентский шаблон
        
        Returns:
            Workbook: Загруженная книга Excel
            
        Raises:
            ExcelExportServiceError: При ошибках загрузки шаблона
        """
        try:
            template_path = self._get_template_path(template_type, is_client_version)
            version_label = "клиентский" if is_client_version else "обычный"
            logger.debug(f"Загружаем {version_label} шаблон из файла: {template_path}")
            workbook = load_workbook(template_path)
            logger.debug(f"{version_label.capitalize()} шаблон типа '{template_type}' успешно загружен")
            return workbook
        except Exception as e:
            version_label = "клиентского" if is_client_version else "обычного"
            error_msg = f"Ошибка при загрузке {version_label} шаблона Excel типа '{template_type}': {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _fill_template_data(self, workbook: Workbook, summary: InsuranceSummary, template_type: str = 'full', is_client_version: bool = False) -> None:
        """
        Заполняет шаблон данными из свода
        
        Args:
            workbook: Книга Excel для заполнения
            summary: Объект свода предложений
            template_type: Тип шаблона ('full' или 'simplified')
            is_client_version: Если True, технический лист не заполняется
            
        Raises:
            ExcelExportServiceError: При ошибках заполнения данных
        """
        try:
            version_label = "клиентской версии" if is_client_version else "полной версии"
            logger.debug(f"Начинаем заполнение данных для свода ID: {summary.id} с шаблоном типа '{template_type}' ({version_label})")
            
            # Получаем рабочий лист (используем первый лист или ищем по имени)
            worksheet = self._get_target_worksheet(workbook)
            
            request = summary.request
            
            # Заполнение заголовков согласно требованиям:
            # CDE1 - номер заявки
            self._set_merged_cell_value(worksheet, 'C1', request.dfa_number)
            logger.debug(f"Записан номер заявки в C1: {request.dfa_number}")
            
            # CDE2 - информация о предмете лизинга с годом выпуска
            vehicle_info_with_year = request.vehicle_info
            if request.manufacturing_year and request.manufacturing_year.strip():
                vehicle_info_with_year = f"{request.vehicle_info}, {request.manufacturing_year}"
            self._set_merged_cell_value(worksheet, 'C2', vehicle_info_with_year)
            logger.debug(f"Записана информация о предмете лизинга в C2: {vehicle_info_with_year[:50]}...")
            
            # CDE3 - название клиента
            self._set_merged_cell_value(worksheet, 'C3', request.client_name)
            logger.debug(f"Записано название клиента в C3: {request.client_name}")
            
            # CDE4 - цели использования
            usage_purposes = request.usage_purposes if request.usage_purposes else ''
            self._set_merged_cell_value(worksheet, 'C4', usage_purposes)
            logger.debug(f"Записаны цели использования в C4: {usage_purposes[:50] if usage_purposes else '(пусто)'}...")
            
            # CDE5 - общее примечание к своду
            summary_notes = summary.notes if summary.notes else ''
            self._set_merged_cell_value(worksheet, 'C5', summary_notes)
            logger.debug(f"Записано общее примечание к своду в C5: {summary_notes[:50] if summary_notes else '(пусто)'}...")
            
            # Заполнение данных компаний с учетом типа шаблона
            self._fill_company_data(workbook, summary, template_type)
            
            # Заполнение технического листа (tech_info) - только для полной версии
            if not is_client_version:
                self._fill_tech_info_sheet(workbook, summary)
                logger.debug("Технический лист заполнен (полная версия)")
            else:
                logger.debug("Технический лист пропущен (клиентская версия)")
            
            logger.info(f"Данные успешно заполнены для свода ID: {summary.id} с шаблоном типа '{template_type}' ({version_label})")
            
        except Exception as e:
            error_msg = f"Ошибка при заполнении данных в Excel: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _get_target_worksheet(self, workbook: Workbook):
        """
        Получает целевой рабочий лист для заполнения
        
        Args:
            workbook: Книга Excel
            
        Returns:
            Worksheet: Рабочий лист для заполнения
            
        Raises:
            ExcelExportServiceError: Если не удается найти подходящий лист
        """
        try:
            # Сначала пытаемся найти лист с именем "summary_template_sheet"
            if 'summary_template_sheet' in workbook.sheetnames:
                logger.debug("Найден лист 'summary_template_sheet'")
                return workbook['summary_template_sheet']
            
            # Если не найден, используем первый лист
            if workbook.worksheets:
                worksheet = workbook.worksheets[0]
                logger.debug(f"Используем первый лист: {worksheet.title}")
                return worksheet
            
            # Если нет листов вообще
            raise ExcelExportServiceError("В шаблоне Excel не найдено ни одного рабочего листа")
            
        except Exception as e:
            error_msg = f"Ошибка при получении рабочего листа: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _set_merged_cell_value(self, worksheet, cell_address: str, value: str) -> None:
        """
        Устанавливает значение в ячейку (может быть объединенной)
        
        Args:
            worksheet: Рабочий лист
            cell_address: Адрес ячейки (например, 'C1')
            value: Значение для записи
        """
        try:
            cell = worksheet[cell_address]
            cell.value = value
            logger.debug(f"Установлено значение в ячейку {cell_address}: {value}")
        except Exception as e:
            error_msg = f"Ошибка при записи в ячейку {cell_address}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _determine_template_type(self, summary: InsuranceSummary) -> str:
        """
        Определяет тип шаблона на основе анализа предложений в своде
        
        Args:
            summary: Объект свода предложений
        
        Returns:
            str: 'full' если есть хотя бы одно предложение с премией-2,
                 'simplified' если нет ни одного предложения с премией-2
        """
        try:
            logger.debug(f"Анализируем предложения свода ID: {summary.id} для определения типа шаблона")
            
            # Получаем все предложения свода
            offers = summary.offers.all()
            total_offers = offers.count()
            
            # Проверяем наличие хотя бы одного предложения с премией-2
            offers_with_premium_2 = offers.filter(
                premium_with_franchise_2__isnull=False,
                premium_with_franchise_2__gt=0
            ).count()
            
            # Логируем статистику
            logger.info(f"Свод ID: {summary.id} - всего предложений: {total_offers}, с премией-2: {offers_with_premium_2}")
            
            if offers_with_premium_2 > 0:
                logger.info(f"Найдено {offers_with_premium_2} предложений с премией-2, используем полный шаблон")
                return 'full'
            else:
                logger.info("Предложений с премией-2 не найдено, используем упрощенный шаблон")
                return 'simplified'
                
        except Exception as e:
            logger.error(f"Ошибка при анализе предложений для определения типа шаблона: {e}")
            raise
    
    def _determine_template_type_safe(self, summary: InsuranceSummary) -> str:
        """
        Безопасное определение типа шаблона с fallback к полному шаблону
        
        Args:
            summary: Объект свода предложений
        
        Returns:
            str: Тип шаблона ('full' или 'simplified')
        """
        try:
            template_type = self._determine_template_type(summary)
            self._log_template_selection(template_type, summary)
            return template_type
        except Exception as e:
            logger.error(f"Ошибка при определении типа шаблона для свода ID {summary.id}: {e}")
            logger.info("Используем полный шаблон по умолчанию")
            return 'full'  # Fallback к полному шаблону
    
    def _get_template_path(self, template_type: str, is_client_version: bool = False) -> str:
        """
        Возвращает путь к соответствующему шаблону
        
        Args:
            template_type: Тип шаблона ('full' или 'simplified')
            is_client_version: Если True, используется клиентский шаблон
        
        Returns:
            str: Путь к файлу шаблона
        """
        if is_client_version:
            # Клиентские шаблоны (без технического листа)
            if template_type == 'simplified':
                return str(settings.BASE_DIR / 'templates' / 'client_summary_template_simplified.xlsx')
            else:
                return str(settings.BASE_DIR / 'templates' / 'client_summary_template.xlsx')
        else:
            # Обычные шаблоны (с техническим листом)
            if template_type == 'simplified':
                return str(settings.BASE_DIR / 'templates' / 'summary_template_simplified.xlsx')
            else:
                return str(settings.BASE_DIR / 'templates' / 'summary_template.xlsx')
    
    def _get_columns_mapping(self, template_type: str) -> dict:
        """
        Возвращает маппинг колонок для указанного типа шаблона
        
        Args:
            template_type: Тип шаблона ('full' или 'simplified')
        
        Returns:
            dict: Маппинг колонок
        """
        if template_type == 'simplified':
            return self.SIMPLIFIED_TEMPLATE_COLUMNS
        else:
            return self.FULL_TEMPLATE_COLUMNS
    
    def _log_template_selection(self, template_type: str, summary: InsuranceSummary) -> None:
        """
        Логирует информацию о выборе шаблона
        
        Args:
            template_type: Выбранный тип шаблона
            summary: Объект свода предложений
        """
        try:
            total_offers = summary.offers.count()
            offers_with_premium_2 = summary.offers.filter(
                premium_with_franchise_2__isnull=False,
                premium_with_franchise_2__gt=0
            ).count()
            
            logger.info(f"=== ВЫБОР ШАБЛОНА ДЛЯ СВОДА ID: {summary.id} ===")
            logger.info(f"Выбран тип шаблона: {template_type}")
            logger.info(f"Всего предложений в своде: {total_offers}")
            logger.info(f"Предложений с премией-2: {offers_with_premium_2}")
            
            if template_type == 'simplified':
                logger.info("Будет использован упрощенный шаблон без колонок для второго предложения")
                logger.info("Колонки J, K, L, M, N, O будут отсутствовать")
                logger.info("Примечания будут в колонке K вместо Q")
            else:
                logger.info("Будет использован полный шаблон с колонками для обоих предложений")
                logger.info("Все колонки A-Q будут присутствовать")
            
            logger.info("=" * 50)
            
        except Exception as e:
            logger.warning(f"Ошибка при логировании выбора шаблона: {e}")
    
    def _get_tech_info_worksheet(self, workbook: Workbook):
        """
        Получает лист tech_info из книги Excel
        
        Args:
            workbook: Книга Excel
            
        Returns:
            Worksheet или None если лист не найден
        """
        try:
            if 'tech_info' in workbook.sheetnames:
                logger.debug("Найден лист 'tech_info'")
                return workbook['tech_info']
            else:
                logger.debug("Лист 'tech_info' не найден в шаблоне")
                return None
        except Exception as e:
            logger.warning(f"Ошибка при поиске листа tech_info: {e}")
            return None
    
    def _fill_tech_cell(self, worksheet, cell_address: str, value, field_name: str) -> None:
        """
        Безопасно заполняет ячейку на техническом листе
        
        Args:
            worksheet: Рабочий лист
            cell_address: Адрес ячейки
            value: Значение для записи
            field_name: Название поля для логирования
        """
        try:
            if value is not None and str(value).strip():
                worksheet[cell_address].value = str(value)
                logger.debug(f"Записано {field_name} в ячейку {cell_address}: {str(value)[:50]}...")
            else:
                logger.debug(f"Значение {field_name} отсутствует, ячейка {cell_address} остается пустой")
        except Exception as e:
            logger.warning(f"Ошибка при записи {field_name} в ячейку {cell_address}: {e}")
    
    def _get_franchise_text_for_tech_info(self, franchise_type: str) -> str:
        """
        Возвращает текст о франшизе для технического листа
        
        Args:
            franchise_type: Тип франшизы ('none', 'with_franchise', 'both_variants')
        
        Returns:
            str: Соответствующий текст для технического листа
        """
        franchise_texts = {
            'none': 'Запрос котировок без франшизы',
            'with_franchise': 'Запрос котировок с франшизой',
            'both_variants': 'Запрос котировок включая оба варианта (с франшизой и без)'
        }
        return franchise_texts.get(franchise_type, 'не указано')
    
    def _get_autostart_text_for_tech_info(self, has_autostart: bool) -> str:
        """
        Возвращает текст об автозапуске для технического листа
        
        Args:
            has_autostart: Наличие автозапуска
        
        Returns:
            str: Соответствующий текст для технического листа
        """
        if has_autostart:
            return 'Запрос с условием автозапуска'
        else:
            return 'Запрос без автозапуска'
    
    def _get_installment_text_for_tech_info(self, has_installment: bool) -> str:
        """
        Возвращает текст о рассрочке для технического листа
        
        Args:
            has_installment: Наличие рассрочки
        
        Returns:
            str: Соответствующий текст для технического листа
        """
        if has_installment:
            return 'Запрос тарифа с рассрочкой'
        else:
            return 'Запрос тарифа без рассрочки'
    
    def _get_insurance_description_for_tech_info(self, request) -> str:
        """
        Формирует описание страхования для технического листа (аналог ins_type + casco_type_ce)
        
        Args:
            request: Объект заявки
        
        Returns:
            str: Описание типа страхования с дополнительными рисками
        """
        # Маппинг типов страхования (аналогично core/templates.py)
        insurance_type_descriptions = {
            'КАСКО': 'Стандартный запрос по КАСКО. ',
            'страхование спецтехники': (
                'Запрос по страхованию спецтехники (максимальный пакет с транспортировкой (погрузкой/выгрузкой) на весь срок '
                'страхования).\n Запрашивались:\n - риск кражи / угона в ночное время с неохраняемой стоянки.\n - риск просадки '
                'грунта, провала дорог или мостов, обвала тоннелей\n - риск провала под лед, затопления специальной техники, '
                'дополнительного оборудования.'
            ),
            'страхование имущества': (
                'Запрос по страхованию имущества ("полный пакет рисков").\nОтдельно запрашивались:\n'
                '1. Риски РНПК\n'
                '2. Ограничения по выплате страхового возмещения в той степени, в которой предоставление '
                'такого покрытия, возмещение такого убытка или предоставление такой компенсации подвергло бы Страховщика '
                'действиям любых санкций, запретов или ограничений установленных (резолюциями Организации Объединенных '
                'Наций; законами или правилами Европейского союза, Соединенного Королевства Великобритании или Соединенных '
                'Штатов Америки; законодательством РФ, указами Президента РФ и/или иными нормативными подзаконными актами '
                'РФ, принятыми в соответствии с резолюциями СБ ООН, указами Президента РФ и/или иными нормативными '
                'подзаконными актами РФ).\n'
                '3. Наличие рисков:\n а) Бой стекол \n б) Риск повреждения животными'
            ),
            'другое': 'Иной запрос по страхованию предмета лизинга.'
        }
        
        # Получаем базовое описание типа страхования
        insurance_type = request.insurance_type or 'КАСКО'
        base_description = insurance_type_descriptions.get(insurance_type, insurance_type)
        
        # Добавляем информацию о КАСКО C/E если необходимо
        casco_ce_text = ''
        if request.has_casco_ce:
            casco_ce_text = (
                'Также запрашивались доп. риски для категории C/E: \n'
                '- страхование вне дорог общего пользования, \n'
                '- провал грунта, \n'
                '- переворот \n'
                '- опрокидывание.'
            )
        
        # Объединяем описания
        result = base_description
        if casco_ce_text:
            result += casco_ce_text
        
        return result.strip()
    
    def _fill_tech_info_sheet(self, workbook: Workbook, summary: InsuranceSummary) -> None:
        """
        Заполняет лист tech_info техническими данными о заявке и своде
        
        Args:
            workbook: Книга Excel
            summary: Объект свода предложений
        """
        try:
            logger.debug(f"Начинаем заполнение листа tech_info для свода ID: {summary.id}")
            
            # Ищем лист tech_info
            tech_sheet = self._get_tech_info_worksheet(workbook)
            if not tech_sheet:
                logger.warning("Лист tech_info не найден в шаблоне, пропускаем заполнение")
                return
            
            request = summary.request
            if not request:
                logger.warning("Связанная заявка не найдена, пропускаем заполнение tech_info")
                return
            
            # Заполняем основные данные заявки
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['request_number'], 
                               request.dfa_number, 'номер заявки')
            
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['client_name'], 
                               request.client_name, 'название клиента')
            
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['client_inn'], 
                               request.inn, 'ИНН клиента')
            
            # Заполняем дополнительные данные
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['branch'], 
                               request.branch, 'филиал')
            
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['insurance_type'], 
                               request.get_insurance_type_display(), 'тип страхования')
            
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['vehicle_info'], 
                               request.vehicle_info, 'информация о предмете лизинга')
            
            # Заполняем описание страхования (вместо полного текста письма)
            insurance_description = self._get_insurance_description_for_tech_info(request)
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['letter_text'], 
                               insurance_description, 'описание страхования')
            
            # Заполняем информацию о франшизе (B15)
            franchise_text = self._get_franchise_text_for_tech_info(request.franchise_type)
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['franchise_info'], 
                               franchise_text, 'информация о франшизе')
            
            # Заполняем информацию об автозапуске (B21) - только для КАСКО/спецтехники
            if request.insurance_type in ['КАСКО', 'страхование спецтехники']:
                autostart_text = self._get_autostart_text_for_tech_info(request.has_autostart)
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['autostart_info'], 
                                   autostart_text, 'информация об автозапуске')
                logger.debug(f"Заполнена информация об автозапуске для КАСКО/спецтехники: {autostart_text}")
            else:
                logger.debug(f"Автозапуск не заполняется для типа страхования: {request.insurance_type}")
            
            # Заполняем информацию о рассрочке (B17)
            installment_text = self._get_installment_text_for_tech_info(request.has_installment)
            self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['installment_info'], 
                               installment_text, 'информация о рассрочке')
            
            # Заполняем дополнительные параметры в зависимости от типа заявки
            self._fill_additional_parameters(tech_sheet, request)
            
            logger.info(f"Лис�� tech_info успешно заполнен техническими данными для свода ID: {summary.id}")
            
        except Exception as e:
            logger.warning(f"Ошибка при заполнении листа tech_info для свода ID {summary.id}: {e}")
            # Не прерываем выполнение, продолжаем работу
    
    def _fill_additional_parameters(self, tech_sheet, request) -> None:
        """
        Заполняет дополнительные параметры на техническом листе в зависимости от типа заявки
        
        Args:
            tech_sheet: Рабочий лист tech_info
            request: Объект заявки
        """
        try:
            # Общие параметры для КАСКО/спецтехника и страхования имущества
            if request.insurance_type in ['КАСКО', 'страхование спецтехники', 'страхование имущества']:
                logger.debug(f"Заполняем общие дополнительные параметры для заявки {request.dfa_number}")
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['creditor_bank'], 
                                   request.creditor_bank, 'банк-кредитор')
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['usage_purposes'], 
                                   request.usage_purposes, 'цели использования')
            
            # Специфичные параметры для КАСКО/спецтехника
            if request.insurance_type in ['КАСКО', 'страхование спецтехники']:
                logger.debug(f"Заполняем специфичные параметры КАСКО/спецтехника для заявки {request.dfa_number}")
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['key_completeness'], 
                                   request.key_completeness, 'комплектность ключей')
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['pts_psm'], 
                                   request.pts_psm, 'ПТС/ПСМ')
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['telematics_complex'], 
                                   request.telematics_complex, 'телематический комплекс')
                
                logger.debug(f"Специфичные параметры КАСКО/спецтехника успешно заполнены для заявки {request.dfa_number}")
            
            # Специфичные параметры для страхования имущества
            if request.insurance_type == 'страхование имущества':
                logger.debug(f"Заполняем специфичные параметры страхования имущества для заявки {request.dfa_number}")
                
                self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['insurance_territory'], 
                                   request.insurance_territory, 'территория страхования')
                
                # Заполняем информацию о перевозке (B28)
                if request.has_transportation:
                    transportation_text = "Запрос с условием перевозки"
                    self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['transportation_info'], 
                                       transportation_text, 'информация о перевозке')
                    logger.debug(f"Заполнена информация о перевозке для заявки {request.dfa_number}: {transportation_text}")
                
                # Заполняем информацию о СМР (B29)
                if request.has_construction_work:
                    construction_work_text = "Запрос с условием СМР"
                    self._fill_tech_cell(tech_sheet, self.TECH_INFO_CELLS['construction_work_info'], 
                                       construction_work_text, 'информация о СМР')
                    logger.debug(f"Заполнена информация о СМР для заявки {request.dfa_number}: {construction_work_text}")
                
                logger.debug(f"Специфичные параметры страхования имущества успешно заполнены для заявки {request.dfa_number}")
            
            if request.insurance_type in ['КАСКО', 'страхование спецтехники', 'страхование имущества']:
                logger.debug(f"Дополнительные параметры успешно заполнены для заявки {request.dfa_number}")
            else:
                logger.debug(f"Заявка {request.dfa_number} не поддерживает дополнительные параметры")
            
        except Exception as e:
            logger.warning(f"Ошибка при заполнении дополнительных параметров для заявки {request.dfa_number}: {e}")
            # Не прерываем выполнение, продолжаем работу
    
    def _get_companies_sorted_data(self, summary: InsuranceSummary) -> Dict[str, List]:
        """
        Получает данные компаний, отсортированные по алфавиту
        
        Args:
            summary: Объект свода предложений
            
        Returns:
            Dict: Словарь с компаниями и их предложениями, отсортированными по алфавиту
            Формат: {'Компания': [offer_year1, offer_year2, ...], ...}
            
        Raises:
            ExcelExportServiceError: При ошибках получения данных
        """
        try:
            logger.debug(f"Получаем отсортированные данные компаний для свода ID: {summary.id}")
            
            # Используем существующий метод модели для получения сгруппированных данных
            companies_data = summary.get_offers_grouped_by_company()
            
            logger.info(f"Найдено {len(companies_data)} компаний для свода ID: {summary.id}")
            
            # Логируем информацию о каждой компании
            for company_name, offers in companies_data.items():
                years = [offer.insurance_year for offer in offers]
                logger.debug(f"Компания '{company_name}': {len(offers)} предложений для лет {years}")
            
            return companies_data
            
        except Exception as e:
            error_msg = f"Ошибка при получении данных компаний для свода ID {summary.id}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _fill_company_data(self, workbook: Workbook, summary: InsuranceSummary, template_type: str = 'full') -> None:
        """
        Основная логика заполнения данных компаний в Excel
        
        Args:
            workbook: Книга Excel для заполнения
            summary: Объект свода предложений
            template_type: Тип шаблона ('full' или 'simplified')
            
        Raises:
            ExcelExportServiceError: При ошибках заполнения данных компаний
        """
        try:
            logger.info(f"Начинаем заполнение данных компаний для свода ID: {summary.id} с шаблоном типа '{template_type}'")
            
            # Получаем маппинг колонок для текущего типа шаблона
            columns = self._get_columns_mapping(template_type)
            logger.debug(f"Используем маппинг колонок для шаблона '{template_type}': {columns}")
            
            # Получаем отсортированные данные компаний
            raw_companies_data = self._get_companies_sorted_data(summary)
            
            # Валидируем и обрабатываем граничные случаи
            companies_data = self._validate_companies_data(raw_companies_data)
            
            # Проверяем, есть ли данные для заполнения после валидации
            if not companies_data:
                logger.warning(f"Нет валидных предложений от компаний для свода ID: {summary.id}, пропускаем заполнение")
                return
            
            # Получаем рабочий лист
            worksheet = self._get_target_worksheet(workbook)
            
            current_row = self.FIRST_DATA_ROW
            total_companies = len(companies_data)
            
            logger.info(f"Будет обработано {total_companies} компаний, начиная со строки {current_row}")
            
            # Обрабатываем каждую компанию
            for company_index, (company_name, offers) in enumerate(companies_data.items()):
                logger.debug(f"Обрабатываем компанию {company_index + 1}/{total_companies}: '{company_name}'")
                
                # Проверяем ограничение на количество строк перед обработкой компании
                estimated_rows_needed = len(offers) + (1 if company_index < total_companies - 1 else 0)  # +1 для разделителя
                if current_row + estimated_rows_needed > self.MAX_ROWS_LIMIT:
                    logger.warning(f"Достигнут лимит строк ({self.MAX_ROWS_LIMIT}), останавливаем обработку на компании '{company_name}'")
                    break
                
                # Запоминаем начальную строку компании для объединения ячеек
                company_start_row = current_row
                
                # Заполняем данные по годам для текущей компании
                for year_index, offer in enumerate(offers):
                    # Дополнительная проверка лимита строк
                    if current_row >= self.MAX_ROWS_LIMIT:
                        logger.warning(f"Достигнут лимит строк ({self.MAX_ROWS_LIMIT}), останавливаем заполнение")
                        break
                    
                    year_display = f"{offer.insurance_year} год"
                    
                    logger.debug(f"Заполняем строку {current_row} для компании '{company_name}', {year_display}")
                    
                    # Название компании записываем только в первую строку
                    company_name_for_row = company_name if year_index == 0 else None
                    
                    # Заполняем строку с данными года
                    self._fill_company_year_row(worksheet, current_row, company_name_for_row, offer, year_display, columns)
                    current_row += 1
                
                # Объединяем ячейки с названием компании если у неё несколько лет
                if len(offers) > 1:
                    self._merge_company_name_cells(worksheet, company_start_row, current_row - 1, company_name, columns)
                    # Объединяем ячейки с суммами премий
                    self._merge_premium_summary_cells(worksheet, company_start_row, current_row - 1, company_name, offers, columns)
                    # Объединяем ячейки с примечаниями
                    self._merge_notes_cells(worksheet, company_start_row, current_row - 1, company_name, offers, columns)
                else:
                    # Для компаний с одним годом заполняем столбцы I и O значениями из F и L
                    self._fill_single_year_premium_summary(worksheet, company_start_row, company_name, offers[0], columns)
                
                # Добавляем разделитель между компаниями (кроме последней)
                if company_index < total_companies - 1 and current_row < self.MAX_ROWS_LIMIT:
                    logger.debug(f"Добавляем разделитель после компании '{company_name}' в строку {current_row}")
                    self._copy_separator_row(worksheet, self.SEPARATOR_ROW, current_row)
                    current_row += 1
            
            logger.info(f"Заполнение данных компаний завершено. Обработано {total_companies} компаний, заполнено строк до {current_row - 1}")
            
        except Exception as e:
            error_msg = f"Ошибка при заполнении данных компаний для свода ID {summary.id}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _validate_companies_data(self, companies_data: Dict[str, List]) -> Dict[str, List]:
        """
        Валидирует данные компаний и применяет ограничения для предотвращения проблем
        
        Обрабатывает граничные случаи:
        - Отсутствие предложений от компаний
        - Превышение лимитов по количеству компаний и лет
        - Компании с одним годом страхования
        
        Args:
            companies_data: Словарь с данными компаний
            
        Returns:
            Dict: Валидированные и ограниченные данные компаний
            
        Raises:
            InvalidSummaryDataError: При критических проблемах с данными
        """
        try:
            logger.debug(f"Начинаем валидацию данных {len(companies_data)} компаний")
            
            # Обработка случая отсутствия предложений от компаний (требование 7.1)
            if not companies_data:
                logger.warning("Нет предложений от компаний для обработки")
                return {}
            
            # Применяем ограничение на количество компаний
            if len(companies_data) > self.MAX_COMPANIES_LIMIT:
                logger.warning(f"Количество компаний ({len(companies_data)}) превышает лимит ({self.MAX_COMPANIES_LIMIT}), обрезаем")
                # Берем первые N компаний по алфавиту
                sorted_companies = sorted(companies_data.keys())[:self.MAX_COMPANIES_LIMIT]
                companies_data = {name: companies_data[name] for name in sorted_companies}
            
            validated_data = {}
            total_rows_estimated = 0
            
            for company_name, offers in companies_data.items():
                # Валидация названия компании
                if not company_name or not company_name.strip():
                    logger.warning(f"Пропускаем компанию с пустым названием")
                    continue
                
                # Обработка компаний с одним годом страхования (требование 7.2)
                if len(offers) == 1:
                    logger.info(f"Компания '{company_name}' имеет только один год страхования - это нормально")
                
                # Применяем ограничение на количество лет на компанию
                if len(offers) > self.MAX_YEARS_PER_COMPANY:
                    logger.warning(f"Компания '{company_name}' имеет {len(offers)} лет, ограничиваем до {self.MAX_YEARS_PER_COMPANY}")
                    offers = offers[:self.MAX_YEARS_PER_COMPANY]
                
                # Валидируем каждое предложение
                valid_offers = []
                for offer in offers:
                    if self._validate_offer_data(offer, company_name):
                        valid_offers.append(offer)
                
                # Добавляем компанию только если есть валидные предложения
                if valid_offers:
                    validated_data[company_name] = valid_offers
                    # Оцениваем количество строк (предложения + разделители)
                    total_rows_estimated += len(valid_offers) + 1  # +1 для разделителя
                else:
                    logger.warning(f"Компания '{company_name}' не имеет валидных предложений, пропускаем")
            
            # Проверяем общее ограничение на количество строк (требование 7.4)
            if total_rows_estimated > self.MAX_ROWS_LIMIT:
                logger.warning(f"Оценочное количество строк ({total_rows_estimated}) превышает лимит ({self.MAX_ROWS_LIMIT})")
                # Обрезаем данные до достижения лимита
                validated_data = self._limit_data_by_rows(validated_data)
            
            logger.info(f"Валидация завершена: {len(validated_data)} компаний прошли проверку")
            return validated_data
            
        except Exception as e:
            error_msg = f"Ошибка при валидации данных компаний: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise InvalidSummaryDataError(error_msg) from e
    
    def _validate_offer_data(self, offer, company_name: str) -> bool:
        """
        Валидирует данные отдельного предложения
        
        Проверяет:
        - Наличие обязательных данных
        - Корректность числовых значений
        - Обработка отсутствующих данных по франшизе или премии (требование 7.3)
        
        Args:
            offer: Объект предложения InsuranceOffer
            company_name: Название компании для логирования
            
        Returns:
            bool: True если предложение валидно, False если нужно пропустить
        """
        try:
            # Проверяем наличие года страхования
            if not hasattr(offer, 'insurance_year') or offer.insurance_year is None:
                logger.warning(f"Компания '{company_name}': предложение без года страхования, пропускаем")
                return False
            
            # Проверяем корректность года
            if not isinstance(offer.insurance_year, int) or offer.insurance_year < 1 or offer.insurance_year > 20:
                logger.warning(f"Компания '{company_name}': некорректный год страхования {offer.insurance_year}, пропускаем")
                return False
            
            # Проверяем страховую сумму
            insurance_sum = getattr(offer, 'insurance_sum', None)
            if insurance_sum is None or insurance_sum <= 0:
                logger.warning(f"Компания '{company_name}', год {offer.insurance_year}: отсутствует или некорректная страховая сумма, пропускаем")
                return False
            
            # Валидируем числовые данные перед записью в Excel (требование 7.3)
            premium_1 = getattr(offer, 'premium_with_franchise_1', None)
            franchise_1 = getattr(offer, 'franchise_1', None)
            
            # Для первого варианта премия обязательна
            if premium_1 is None or premium_1 <= 0:
                logger.warning(f"Компания '{company_name}', год {offer.insurance_year}: отсутствует премия-1, пропускаем предложение")
                return False
            
            # Франшиза может отсутствовать (будет 0 по умолчанию)
            if franchise_1 is None:
                logger.debug(f"Компания '{company_name}', год {offer.insurance_year}: франшиза-1 отсутствует, будет использован 0")
            
            # Проверяем второй вариант если он есть
            if offer.has_second_franchise_variant():
                premium_2 = getattr(offer, 'premium_with_franchise_2', None)
                franchise_2 = getattr(offer, 'franchise_2', None)
                
                if premium_2 is None or premium_2 <= 0:
                    logger.warning(f"Компания '{company_name}', год {offer.insurance_year}: некорректная премия-2, второй вариант будет пропущен")
                
                if franchise_2 is None:
                    logger.debug(f"Компания '{company_name}', год {offer.insurance_year}: франшиза-2 отсутствует")
            
            logger.debug(f"Компания '{company_name}', год {offer.insurance_year}: предложение прошло валидацию")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при валидации предложения компании '{company_name}': {str(e)}")
            return False
    
    def _limit_data_by_rows(self, companies_data: Dict[str, List]) -> Dict[str, List]:
        """
        Ограничивает данные компаний по количеству строк для предотвращения превышения лимитов
        
        Args:
            companies_data: Данные компаний
            
        Returns:
            Dict: Ограниченные данные компаний
        """
        try:
            limited_data = {}
            current_rows = self.FIRST_DATA_ROW
            
            logger.info(f"Применяем ограничение по строкам, лимит: {self.MAX_ROWS_LIMIT}")
            
            for company_name, offers in companies_data.items():
                # Рассчитываем сколько строк займет эта компания
                company_rows = len(offers) + 1  # +1 для разделителя
                
                # Проверяем, поместится ли компания в лимит
                if current_rows + company_rows > self.MAX_ROWS_LIMIT:
                    logger.warning(f"Компания '{company_name}' не помещается в лимит строк, останавливаем обработку")
                    break
                
                limited_data[company_name] = offers
                current_rows += company_rows
                
                logger.debug(f"Компания '{company_name}': {len(offers)} предложений, текущая строка: {current_rows}")
            
            logger.info(f"Ограничение применено: {len(limited_data)} компаний, ожидаемая строка: {current_rows}")
            return limited_data
            
        except Exception as e:
            logger.error(f"Ошибка при ограничении данных по строкам: {str(e)}")
            return companies_data  # Возвращаем исходные данные при ошибке
    
    def _validate_numeric_data_before_write(self, value, field_name: str, company_name: str, year: int) -> bool:
        """
        Валидирует числовые данные перед записью в Excel
        
        Args:
            value: Значение для проверки
            field_name: Название поля
            company_name: Название компании
            year: Год страхования
            
        Returns:
            bool: True если данные валидны для записи
        """
        try:
            if value is None:
                logger.debug(f"Компания '{company_name}', год {year}, поле '{field_name}': значение None")
                return True  # None допустимо, будет пропущено при записи
            
            # Проверяем, что это число
            if not isinstance(value, (int, float, Decimal)):
                try:
                    float(value)
                except (ValueError, TypeError):
                    logger.warning(f"Компания '{company_name}', год {year}, поле '{field_name}': не числовое значение '{value}'")
                    return False
            
            # Проверяем разумные пределы
            numeric_value = float(value)
            if numeric_value < 0:
                logger.warning(f"Компания '{company_name}', год {year}, поле '{field_name}': отрицательное значение {numeric_value}")
                return False
            
            if numeric_value > float(self.MAX_INSURANCE_SUM):
                logger.warning(f"Компания '{company_name}', год {year}, поле '{field_name}': значение {numeric_value} превышает максимум")
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка валидации числовых данных для поля '{field_name}': {str(e)}")
            return False
    
    def _fill_company_year_row(self, worksheet, row_num: int, company_name: str, 
                              offer, year_display: str, columns_mapping: dict = None) -> None:
        """
        Заполняет строку с данными года страхования с корректным форматированием
        
        Использует специализированные методы форматирования для обеспечения
        корректного отображения числовых данных в Excel.
        
        Args:
            worksheet: Рабочий лист Excel
            row_num: Номер строки для заполнения
            company_name: Название страховой компании (может быть None для объединенных ячеек)
            offer: Объект предложения InsuranceOffer
            year_display: Отображаемое значение года (например, "1 год")
            columns_mapping: Маппинг колонок (если None, используется COMPANY_DATA_COLUMNS)
            
        Raises:
            ExcelExportServiceError: При ошибках заполнения строки
        """
        try:
            # Используем переданный маппинг или значение по умолчанию
            columns = columns_mapping if columns_mapping is not None else self.COMPANY_DATA_COLUMNS
            
            logger.debug(f"Заполняем строку {row_num} для компании '{company_name}', {year_display}")
            
            # Копируем стили из строки 10 (первой строки с данными) перед заполнением данных
            if row_num != self.FIRST_DATA_ROW:
                self._copy_row_styles(worksheet, self.FIRST_DATA_ROW, row_num)
            
            # Основные данные
            # Название компании записываем только если передано (для первой строки компании)
            if company_name is not None:
                worksheet[f"{columns['company_name']}{row_num}"].value = company_name
            worksheet[f"{columns['year']}{row_num}"].value = year_display
            
            # Страховая сумма с форматированием и валидацией
            insurance_sum = self._format_insurance_sum(offer)
            if insurance_sum is not None and self._validate_numeric_data_before_write(
                insurance_sum, 'страховая сумма', company_name, getattr(offer, 'insurance_year', 0)
            ):
                worksheet[f"{columns['insurance_sum']}{row_num}"].value = insurance_sum
                logger.debug(f"Записана страховая сумма: {insurance_sum}")
            else:
                logger.warning(f"Страховая сумма не записана для компании '{company_name}' из-за ошибок валидации")
            
            # Предложение 1 (обязательное) с форматированием и валидацией
            premium_1 = self._format_premium(offer, 1)
            franchise_1 = self._format_franchise(offer, 1)
            installment_1 = self._format_installment_payments(offer, 1)
            
            # Заполняем страховой тариф-1 (колонка E)
            self._fill_insurance_rate(worksheet, row_num, columns['rate_1'], 1)
            
            # Валидация и запись премии-1
            if premium_1 is not None and self._validate_numeric_data_before_write(
                premium_1, 'премия-1', company_name, getattr(offer, 'insurance_year', 0)
            ):
                worksheet[f"{columns['premium_1']}{row_num}"].value = premium_1
                logger.debug(f"Записана премия-1: {premium_1}")
            elif premium_1 is None:
                logger.warning(f"Премия-1 отсутствует для компании '{company_name}', ячейка останется пустой")
            
            # Валидация и запись франшизы-1 (может быть 0 или отсутствовать)
            if franchise_1 is not None and self._validate_numeric_data_before_write(
                franchise_1, 'франшиза-1', company_name, getattr(offer, 'insurance_year', 0)
            ):
                worksheet[f"{columns['franchise_1']}{row_num}"].value = franchise_1
                logger.debug(f"Записана франшиза-1: {franchise_1}")
            elif franchise_1 is None:
                # Для франшизы None допустимо - ячейка останется пустой
                logger.debug(f"Франшиза-1 отсутствует для компании '{company_name}', ячейка останется пустой")
            
            # Рассрочка всегда записывается (минимум 1)
            worksheet[f"{columns['installment_1']}{row_num}"].value = installment_1
            logger.debug(f"Записана рассрочка-1: {installment_1} платежей в год")
            
            # Предложение 2 (если есть и поддерживается шаблоном) с форматированием и валидацией
            if offer.has_second_franchise_variant() and 'rate_2' in columns:
                premium_2 = self._format_premium(offer, 2)
                franchise_2 = self._format_franchise(offer, 2)
                installment_2 = self._format_installment_payments(offer, 2)
                
                # Заполняем страховой тариф-2 (только если колонка существует)
                self._fill_insurance_rate(worksheet, row_num, columns['rate_2'], 2)
                
                # Валидация и запись премии-2
                if premium_2 is not None and self._validate_numeric_data_before_write(
                    premium_2, 'премия-2', company_name, getattr(offer, 'insurance_year', 0)
                ):
                    worksheet[f"{columns['premium_2']}{row_num}"].value = premium_2
                    logger.debug(f"Записана премия-2: {premium_2}")
                elif premium_2 is None:
                    logger.warning(f"Премия-2 отсутствует для компании '{company_name}', ячейка останется пустой")
                
                # Валидация и запись франшизы-2
                if franchise_2 is not None and self._validate_numeric_data_before_write(
                    franchise_2, 'франшиза-2', company_name, getattr(offer, 'insurance_year', 0)
                ):
                    worksheet[f"{columns['franchise_2']}{row_num}"].value = franchise_2
                    logger.debug(f"Записана франшиза-2: {franchise_2}")
                elif franchise_2 is None:
                    logger.debug(f"Франшиза-2 отсутствует для компании '{company_name}', ячейка останется пустой")
                
                # Рассрочка-2 всегда записывается
                worksheet[f"{columns['installment_2']}{row_num}"].value = installment_2
                logger.debug(f"Записана рассрочка-2: {installment_2} платежей в год")
            else:
                logger.debug("Второй вариант франшизы отсутствует, пропускаем заполнение")
            
            # Примечания (если есть) с обрезкой и валидацией
            if hasattr(offer, 'notes') and offer.notes:
                notes = str(offer.notes).strip()
                if notes:
                    # Ограничиваем длину примечаний для Excel
                    if len(notes) > self.MAX_NOTES_LENGTH:
                        notes = notes[:self.MAX_NOTES_LENGTH] + "..."
                        logger.warning(f"Примечания обрезаны до {self.MAX_NOTES_LENGTH} символов для компании '{company_name}'")
                    
                    # Дополнительная валидация примечаний
                    # Удаляем потенциально проблемные символы для Excel
                    notes = notes.replace('\x00', '').replace('\x01', '').replace('\x02', '')
                    
                    if notes:  # Проверяем, что после очистки что-то осталось
                        worksheet[f"{columns['notes']}{row_num}"].value = notes
                        logger.debug(f"Записаны примечания: {notes[:50]}...")
                    else:
                        logger.debug(f"Примечания для компании '{company_name}' пусты после очистки")
                else:
                    logger.debug(f"Примечания для компании '{company_name}' пусты после обрезки пробелов")
            
            logger.info(f"Строка {row_num} успешно заполнена для компании '{company_name}', {year_display}")
            
        except Exception as e:
            error_msg = f"Ошибка при заполнении строки {row_num} для компании '{company_name}': {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _format_installment_payments(self, offer, variant: int = 1) -> int:
        """
        Форматирует количество платежей в год для указанного варианта рассрочки
        
        Логика определения:
        - 1 если рассрочки нет (единовременный платеж)
        - Фактическое количество платежей в год если есть рассрочка (2, 3, 4, 12 и т.д.)
        
        Args:
            offer: Объект предложения InsuranceOffer
            variant: Номер варианта (1 или 2)
            
        Returns:
            int: Количество платежей в год (1 если нет рассрочки, иначе фактическое количество)
            
        Raises:
            ValueError: Если variant не равен 1 или 2
        """
        try:
            # Валидация входного параметра
            if variant not in [1, 2]:
                logger.error(f"Некорректный номер варианта: {variant}. Ожидается 1 или 2")
                raise ValueError(f"Номер варианта должен быть 1 или 2, получен: {variant}")
            
            # Получаем данные для соответствующего варианта
            if variant == 1:
                payments_per_year = getattr(offer, 'payments_per_year_variant_1', 1)
                has_installment = getattr(offer, 'installment_variant_1', False)
                logger.debug(f"Вариант 1: платежей в год={payments_per_year}, рассрочка={has_installment}")
            else:  # variant == 2
                payments_per_year = getattr(offer, 'payments_per_year_variant_2', 1)
                has_installment = getattr(offer, 'installment_variant_2', False)
                logger.debug(f"Вариант 2: платежей в год={payments_per_year}, рассрочка={has_installment}")
            
            # Проверяем валидность данных
            if payments_per_year is None:
                payments_per_year = 1
                logger.warning(f"Количество платежей в год для варианта {variant} равно None, используем 1")
            
            # Если рассрочки нет или только один платеж в год - возвращаем 1
            if not has_installment or payments_per_year <= 1:
                logger.debug(f"Рассрочка недоступна для варианта {variant}: has_installment={has_installment}, payments_per_year={payments_per_year}")
                return 1
            
            # Возвращаем фактическое количество платежей в год без преобразований
            logger.debug(f"Рассрочка вариант {variant}: {payments_per_year} платежей в год")
            
            return payments_per_year
            
        except ValueError:
            # Переброс ошибок валидации
            raise
        except Exception as e:
            error_msg = f"Ошибка при форматировании рассрочки для варианта {variant}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            # В случае ошибки возвращаем 1 (единовременный платеж)
            return 1
    
    def _format_numeric_value(self, value, field_name: str = "значение") -> Optional[Decimal]:
        """
        Форматирует числовое значение для записи в Excel
        
        Обеспечивает корректное форматирование премий, франшиз и страховых сумм.
        Обрабатывает различные типы входных данных и валидирует их.
        
        Args:
            value: Числовое значение (может быть Decimal, float, int, str или None)
            field_name: Название поля для логирования (для отладки)
            
        Returns:
            Optional[Decimal]: Отформатированное значение или None если значение пустое/некорректное
        """
        try:
            # Если значение None или пустое - возвращаем None
            if value is None:
                logger.debug(f"Поле '{field_name}': значение None, возвращаем None")
                return None
            
            # Если это строка - пытаемся преобразовать
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    logger.debug(f"Поле '{field_name}': пустая строка, возвращаем None")
                    return None
                
                # Убираем пробелы и запятые для парсинга
                value = value.replace(' ', '').replace(',', '')
                
                try:
                    value = Decimal(value)
                except (ValueError, InvalidOperation) as e:
                    logger.warning(f"Поле '{field_name}': не удалось преобразовать строку '{value}' в число: {e}")
                    return None
            
            # Преобразуем в Decimal для точности
            if isinstance(value, (int, float)):
                value = Decimal(str(value))
            elif not isinstance(value, Decimal):
                logger.warning(f"Поле '{field_name}': неподдерживаемый тип данных {type(value)}, пытаемся преобразовать")
                value = Decimal(str(value))
            
            # Валидация: проверяем, что значение не отрицательное
            if value < 0:
                logger.warning(f"Поле '{field_name}': отрицательное значение {value}, возвращаем 0")
                return Decimal('0')
            
            # Валидация: проверяем разумные пределы
            if value > self.MAX_INSURANCE_SUM:
                logger.warning(f"Поле '{field_name}': значение {value} превышает максимальный лимит {self.MAX_INSURANCE_SUM}")
                return self.MAX_INSURANCE_SUM
            
            # Проверяем минимальные пределы
            if value < self.MIN_INSURANCE_SUM and value > 0:
                logger.warning(f"Поле '{field_name}': значение {value} меньше минимального лимита {self.MIN_INSURANCE_SUM}")
                return self.MIN_INSURANCE_SUM
            
            # Округляем до 2 знаков после запятой для денежных значений
            result = value.quantize(Decimal('0.01'))
            
            logger.debug(f"Поле '{field_name}': успешно отформатировано значение {result}")
            return result
            
        except Exception as e:
            error_msg = f"Ошибка при форматировании числового значения для поля '{field_name}': {str(e)}"
            logger.error(error_msg, exc_info=True)
            return None
    
    def _format_insurance_sum(self, offer) -> Optional[Decimal]:
        """
        Форматирует страховую сумму для записи в Excel
        
        Args:
            offer: Объект предложения InsuranceOffer
            
        Returns:
            Optional[Decimal]: Отформатированная страховая сумма
        """
        try:
            insurance_sum = getattr(offer, 'insurance_sum', None)
            return self._format_numeric_value(insurance_sum, 'страховая сумма')
        except Exception as e:
            logger.error(f"Ошибка при форматировании страховой суммы: {str(e)}")
            return None
    
    def _format_premium(self, offer, variant: int = 1) -> Optional[Decimal]:
        """
        Форматирует премию для указанного варианта
        
        Args:
            offer: Объект предложения InsuranceOffer
            variant: Номер варианта (1 или 2)
            
        Returns:
            Optional[Decimal]: Отформатированная премия
        """
        try:
            if variant == 1:
                premium = getattr(offer, 'premium_with_franchise_1', None)
                field_name = 'премия-1'
            else:
                premium = getattr(offer, 'premium_with_franchise_2', None)
                field_name = 'премия-2'
            
            return self._format_numeric_value(premium, field_name)
        except Exception as e:
            logger.error(f"Ошибка при форматировании премии для варианта {variant}: {str(e)}")
            return None
    
    def _format_franchise(self, offer, variant: int = 1) -> Optional[Decimal]:
        """
        Форматирует франшизу для указанного варианта
        
        Args:
            offer: Объект предложения InsuranceOffer
            variant: Номер варианта (1 или 2)
            
        Returns:
            Optional[Decimal]: Отформатированная франшиза
        """
        try:
            if variant == 1:
                franchise = getattr(offer, 'franchise_1', None)
                field_name = 'франшиза-1'
            else:
                franchise = getattr(offer, 'franchise_2', None)
                field_name = 'франшиза-2'
            
            # Для франшизы 0 - это валидное значение, не None
            formatted_value = self._format_numeric_value(franchise, field_name)
            
            # Если франшиза None, но это вариант 1 - возвращаем 0 (по умолчанию)
            if formatted_value is None and variant == 1:
                logger.debug(f"Франшиза-1 равна None, используем значение по умолчанию 0")
                return Decimal('0')
            
            return formatted_value
        except Exception as e:
            logger.error(f"Ошибка при форматировании франшизы для варианта {variant}: {str(e)}")
            return Decimal('0') if variant == 1 else None
    
    def _copy_formula_to_row(self, worksheet, source_row: int, target_row: int, column: str) -> None:
        """
        Копирует формулу из исходной ячейки в целевую ячейку с адаптацией к новой строке
        
        Поскольку openpyxl не автоматически адаптирует формулы при копировании,
        мы вручную заменяем номера строк в формуле.
        
        Args:
            worksheet: Рабочий лист Excel
            source_row: Номер исходной строки (обычно 10)
            target_row: Номер целевой строки
            column: Буква колонки ('E' для тарифа-1 или 'K' для тарифа-2)
            
        Raises:
            ExcelExportServiceError: При критических ошибках копирования
        """
        try:
            logger.debug(f"Копируем формулу из ячейки {column}{source_row} в {column}{target_row}")
            
            source_cell = worksheet[f'{column}{source_row}']
            target_cell = worksheet[f'{column}{target_row}']
            
            # Проверяем, есть ли формула в исходной ячейке
            if hasattr(source_cell, 'data_type') and source_cell.data_type == 'f' and source_cell.value:
                # Получаем формулу и адаптируем её к новой строке
                source_formula = str(source_cell.value)
                logger.debug(f"Исходная формула: {source_formula}")
                
                # Заменяем номер строки в формуле более точно (только ссылки на ячейки)
                # Заменяем конкретные ссылки на ячейки
                adapted_formula = source_formula
                adapted_formula = adapted_formula.replace(f'C{source_row}', f'C{target_row}')
                adapted_formula = adapted_formula.replace(f'F{source_row}', f'F{target_row}')
                adapted_formula = adapted_formula.replace(f'L{source_row}', f'L{target_row}')
                target_cell.value = adapted_formula
                
                logger.debug(f"Адаптированная формула в ячейке {column}{target_row}: {adapted_formula}")
            else:
                # Если формулы нет, создаем новую
                logger.debug(f"Формула не найдена в {column}{source_row}, создаем новую для {column}{target_row}")
                self._create_rate_formula(worksheet, target_row, column)
                
        except Exception as e:
            logger.warning(f"Не удалось скопировать формулу в ячейку {column}{target_row}: {str(e)}")
            # Fallback: создаем формулу вручную
            try:
                self._create_rate_formula(worksheet, target_row, column)
            except Exception as fallback_error:
                logger.warning(f"Не удалось создать формулу для ячейки {column}{target_row}: {str(fallback_error)}")
    
    def _create_rate_formula(self, worksheet, row_num: int, column: str) -> None:
        """
        Создает формулу для вычисления страхового тарифа в процентах
        
        Формула вычисляет тариф как отношение премии к страховой сумме, умноженное на 100.
        Включает проверки на деление на ноль и пустые ячейки.
        
        Args:
            worksheet: Рабочий лист Excel
            row_num: Номер строки
            column: Колонка тарифа ('E' для варианта 1, 'K' для варианта 2)
            
        Raises:
            ValueError: При неизвестной колонке тарифа
        """
        try:
            # Определяем колонку премии в зависимости от варианта тарифа
            if column == 'E':  # Тариф для варианта 1
                premium_column = 'F'
                variant_name = "тариф-1"
            elif column == 'K':  # Тариф для варианта 2
                premium_column = 'L'
                variant_name = "тариф-2"
            else:
                error_msg = f"Неизвестная колонка тарифа: {column}. Ожидается 'E' или 'K'"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Создаем формулу с проверками на пустые ячейки и деление на ноль
            # Формула: ЕСЛИ(И(страховая_сумма<>0, премия<>0), премия/страховая_сумма*100, "")
            formula = f'=IF(AND(C{row_num}<>0,{premium_column}{row_num}<>0),{premium_column}{row_num}/C{row_num}*100,"")'
            
            worksheet[f'{column}{row_num}'].value = formula
            logger.debug(f"Создана формула для {variant_name} в ячейке {column}{row_num}: {formula}")
            
        except ValueError:
            # Переброс ошибок валидации
            raise
        except Exception as e:
            error_msg = f"Ошибка при создании формулы для ячейки {column}{row_num}: {str(e)}"
            logger.warning(error_msg)
            raise
    
    def _fill_insurance_rate(self, worksheet, row_num: int, column: str, variant: int) -> None:
        """
        Заполняет ячейку страхового тарифа формулой или вычисленным значением
        
        Основной метод для заполнения тарифов. Сначала пытается скопировать формулу
        из строки 10, если не получается - создает новую формулу.
        
        Args:
            worksheet: Рабочий лист Excel
            row_num: Номер строки для заполнения
            column: Колонка тарифа ('E' или 'K')
            variant: Номер варианта предложения (1 или 2)
        """
        try:
            logger.debug(f"Заполняем страховой тариф для варианта {variant} в ячейке {column}{row_num}")
            
            # Если заполняем строку 10, ничего не делаем (там уже должна быть формула из шаблона)
            if row_num == self.FIRST_DATA_ROW:
                logger.debug(f"Строка {row_num} - исходная строка данных, пропускаем заполнение тарифа")
                return
            
            # Пытаемся скопировать формулу из строки 10
            self._copy_formula_to_row(worksheet, self.FIRST_DATA_ROW, row_num, column)
            
            logger.debug(f"Страховой тариф для варианта {variant} успешно заполнен в ячейке {column}{row_num}")
            
        except Exception as e:
            logger.warning(f"Ошибка при заполнении тарифа в ячейке {column}{row_num}: {str(e)}")
            # Не выбрасываем исключение, чтобы не прерывать заполнение других данных
    
    def _copy_row_styles(self, worksheet, source_row: int, target_row: int) -> None:
        """
        Копирует стили форматирования из исходной строки в целевую строку
        
        Копирует все атрибуты форматирования включая:
        - Шрифт (размер, цвет, жирность, курсив и т.д.)
        - Границы ячеек
        - Заливку и цвет фона
        - Выравнивание текста
        - Формат чисел
        - Защиту ячеек
        - Высоту строки и другие свойства строки
        
        Args:
            worksheet: Рабочий лист Excel
            source_row: Номер исходной строки для копирования стилей
            target_row: Номер целевой строки для применения стилей
            
        Raises:
            ExcelExportServiceError: При критических ошибках копирования стилей
        """
        try:
            logger.debug(f"Копируем стили из строки {source_row} в строку {target_row}")
            
            # Копируем свойства самой строки (высота, скрытость, группировка)
            try:
                source_row_dimension = worksheet.row_dimensions[source_row]
                target_row_dimension = worksheet.row_dimensions[target_row]
                
                # Копируем высоту строки если она задана
                if source_row_dimension.height is not None:
                    target_row_dimension.height = source_row_dimension.height
                    logger.debug(f"Скопирована высота строки: {source_row_dimension.height}")
                
                # Копируем скрытость строки
                target_row_dimension.hidden = source_row_dimension.hidden
                
                # Копируем уровень группировки
                target_row_dimension.outline_level = source_row_dimension.outline_level
                
            except Exception as row_error:
                logger.warning(f"Не удалось скопировать свойства строки из {source_row} в {target_row}: {str(row_error)}")
            
            # Определяем диапазон колонок для копирования (A-Z, достаточно для большинства случаев)
            columns_to_copy = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            cells_copied = 0
            cells_with_errors = 0
            
            for col_letter in columns_to_copy:
                try:
                    source_cell = worksheet[f'{col_letter}{source_row}']
                    target_cell = worksheet[f'{col_letter}{target_row}']
                    
                    # Копируем стили только если у исходной ячейки есть стили
                    if source_cell.has_style:
                        # Копируем шрифт
                        if source_cell.font:
                            target_cell.font = source_cell.font.copy()
                        
                        # Копируем границы
                        if source_cell.border:
                            target_cell.border = source_cell.border.copy()
                        
                        # Копируем заливку
                        if source_cell.fill:
                            target_cell.fill = source_cell.fill.copy()
                        
                        # Копируем выравнивание
                        if source_cell.alignment:
                            target_cell.alignment = source_cell.alignment.copy()
                        
                        # Копируем защиту
                        if source_cell.protection:
                            target_cell.protection = source_cell.protection.copy()
                        
                        # Копируем формат чисел
                        if source_cell.number_format:
                            target_cell.number_format = source_cell.number_format
                        
                        cells_copied += 1
                        
                except Exception as cell_error:
                    cells_with_errors += 1
                    logger.warning(f"Не удалось скопировать стили ячейки {col_letter}{source_row} в {col_letter}{target_row}: {str(cell_error)}")
                    continue
            
            logger.debug(f"Копирование стилей завершено: скопировано {cells_copied} ячеек, ошибок {cells_with_errors}")
            
            if cells_with_errors > 0:
                logger.warning(f"При копировании стилей из строки {source_row} в строку {target_row} возникло {cells_with_errors} ошибок")
            
        except Exception as e:
            error_msg = f"Критическая ошибка при копировании стилей из строки {source_row} в строку {target_row}: {str(e)}"
            logger.warning(error_msg)
            # Не выбрасываем исключение, чтобы не прерывать заполнение данных
            # Заполнение может продолжиться без стилей
    
    def _copy_separator_row(self, worksheet, source_row: int, target_row: int) -> None:
        """
        Копирует строку-разделитель с сохранением всех стилей форматирования и содержимого
        
        Копирует все ячейки из исходной строки в целевую строку, включая:
        - Значения ячеек
        - Все стили форматирования (через _copy_row_styles)
        - Комментарии ячеек
        
        Args:
            worksheet: Рабочий лист Excel
            source_row: Номер исходной строки для копирования (строка 9 по умолчанию)
            target_row: Номер целевой строки для вставки разделителя
            
        Raises:
            ExcelExportServiceError: При ошибках копирования строки или стилей
        """
        try:
            logger.debug(f"Копируем строку-разделитель из строки {source_row} в строку {target_row}")
            
            # Сначала копируем все стили с помощью специализированного метода
            self._copy_row_styles(worksheet, source_row, target_row)
            
            # Затем копируем содержимое ячеек и комментарии
            source_row_obj = worksheet[source_row]
            cells_copied = 0
            cells_with_errors = 0
            
            for cell in source_row_obj:
                # Копируем содержимое всех ячеек, включая пустые с комментариями
                if cell.value is not None or cell.comment:
                    try:
                        # Получаем целевую ячейку
                        target_cell = worksheet.cell(row=target_row, column=cell.column)
                        
                        # Копируем значение
                        if cell.value is not None:
                            target_cell.value = cell.value
                        
                        # Копируем комментарии (если есть)
                        if cell.comment:
                            target_cell.comment = cell.comment
                        
                        cells_copied += 1
                        
                    except Exception as cell_error:
                        cells_with_errors += 1
                        logger.warning(f"Не удалось скопировать содержимое ячейки {cell.coordinate}: {str(cell_error)}")
                        continue
            
            logger.info(f"Строка-разделитель успешно скопирована из строки {source_row} в строку {target_row}")
            logger.debug(f"Скопировано ячеек с содержимым: {cells_copied}, ошибок при копировании: {cells_with_errors}")
            
        except Exception as e:
            error_msg = f"Ошибка при копировании строки-разделителя из строки {source_row} в строку {target_row}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _merge_company_name_cells(self, worksheet, start_row: int, end_row: int, company_name: str, columns_mapping: dict = None) -> None:
        """
        Объединяет ячейки в столбце A для названия страховой компании
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            company_name: Название страховой компании
            columns_mapping: Маппинг колонок (если None, используется COMPANY_DATA_COLUMNS)
            
        Raises:
            ExcelExportServiceError: При критических ошибках объединения
        """
        try:
            # Используем переданный маппинг или значение по умолчанию
            columns = columns_mapping if columns_mapping is not None else self.COMPANY_DATA_COLUMNS
            company_column = columns['company_name']
            
            logger.debug(f"Объединяем ячейки для компании '{company_name}' в диапазоне {company_column}{start_row}:{company_column}{end_row}")
            
            # Определяем диапазон для объединения
            merge_range = f'{company_column}{start_row}:{company_column}{end_row}'
            
            # Объединяем ячейки
            worksheet.merge_cells(merge_range)
            
            # Записываем название компании в объединенную ячейку
            merged_cell = worksheet[f'{company_column}{start_row}']
            merged_cell.value = company_name
            
            # Применяем вертикальное выравнивание по центру
            if merged_cell.alignment:
                # Сохраняем существующее выравнивание и добавляем вертикальное
                from copy import copy
                alignment = copy(merged_cell.alignment)
                alignment.vertical = 'center'
                merged_cell.alignment = alignment
            else:
                # Создаем новое выравнивание
                from openpyxl.styles import Alignment
                merged_cell.alignment = Alignment(vertical='center')
            
            # Применяем границы к объединенной ячейке
            self._apply_borders_to_merged_cell(worksheet, start_row, end_row)
            
            logger.info(f"Объединены ячейки {merge_range} для компании '{company_name}'")
            
        except Exception as e:
            logger.warning(f"Не удалось объединить ячейки для компании '{company_name}' в диапазоне A{start_row}:A{end_row}: {e}")
            # Fallback: записываем название в каждую ячейку
            self._fill_company_name_fallback(worksheet, start_row, end_row, company_name)
    
    def _apply_borders_to_merged_cell(self, worksheet, start_row: int, end_row: int) -> None:
        """
        Применяет границы к объединенной ячейке, копируя стиль из первой строки данных
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка объединенной ячейки
            end_row: Последняя строка объединенной ячейки
        """
        try:
            # Получаем стиль границ из первой строки данных (строка 10)
            source_cell = worksheet['A10']
            merged_cell = worksheet[f'A{start_row}']
            
            # Копируем границы
            if source_cell.border:
                from copy import copy
                merged_cell.border = copy(source_cell.border)
            
            logger.debug(f"Применены границы к объединенной ячейке A{start_row}:A{end_row}")
            
        except Exception as e:
            logger.warning(f"Не удалось применить границы к объединенной ячейке A{start_row}:A{end_row}: {e}")
    
    def _fill_company_name_fallback(self, worksheet, start_row: int, end_row: int, company_name: str) -> None:
        """
        Fallback метод: заполняет название компании в каждую ячейку если объединение не удалось
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            company_name: Название страховой компании
        """
        try:
            # Сначала попытаемся разъединить ячейки если они были объединены
            try:
                merge_range = f'A{start_row}:A{end_row}'
                if merge_range in [str(r) for r in worksheet.merged_cells.ranges]:
                    worksheet.unmerge_cells(merge_range)
                    logger.debug(f"Разъединены ячейки {merge_range} для fallback")
            except Exception as unmerge_error:
                logger.debug(f"Не удалось разъединить ячейки (возможно, они не были объединены): {unmerge_error}")
            
            # Теперь заполняем каждую ячейку
            for row_num in range(start_row, end_row + 1):
                try:
                    worksheet[f'A{row_num}'].value = company_name
                except Exception as cell_error:
                    logger.warning(f"Не удалось записать в ячейку A{row_num}: {cell_error}")
                    continue
            
            logger.info(f"Использован fallback для компании '{company_name}': название записано в каждую строку {start_row}-{end_row}")
            
        except Exception as e:
            logger.error(f"Критическая ошибка при fallback заполнении названия компании '{company_name}': {e}")
    
    def _merge_premium_summary_cells(self, worksheet, start_row: int, end_row: int, 
                                    company_name: str, offers: List, columns_mapping: dict = None) -> None:
        """
        Объединяет ячейки в столбцах I и O с суммированием премий по всем годам страхования
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            company_name: Название страховой компании
            offers: Список предложений компании
            columns_mapping: Маппинг колонок (если None, используется COMPANY_DATA_COLUMNS)
            
        Raises:
            ExcelExportServiceError: При критических ошибках объединения
        """
        try:
            # Используем переданный маппинг или значение по умолчанию
            columns = columns_mapping if columns_mapping is not None else self.COMPANY_DATA_COLUMNS
            
            logger.debug(f"Объединяем ячейки премий для компании '{company_name}' в диапазоне {start_row}:{end_row}")
            
            # Вычисляем суммы премий
            premium_1_total = self._calculate_premium_sum(offers, 1)
            premium_2_total = self._calculate_premium_sum(offers, 2)
            
            # Объединяем и заполняем ячейки в столбце премий-1
            if premium_1_total is not None:
                premium_1_column = columns['premium_1_summary']
                merge_range_i = f'{premium_1_column}{start_row}:{premium_1_column}{end_row}'
                worksheet.merge_cells(merge_range_i)
                merged_cell_i = worksheet[f'{premium_1_column}{start_row}']
                merged_cell_i.value = premium_1_total
                
                # Применяем форматирование
                self._apply_premium_cell_formatting(worksheet, f'{premium_1_column}{start_row}', self.FIRST_DATA_ROW)
                
                logger.debug(f"Объединены ячейки {merge_range_i} с суммой премий-1: {premium_1_total}")
            
            # Объединяем и заполняем ячейки в столбце премий-2 только если есть данные и колонка существует
            if premium_2_total is not None and premium_2_total > 0 and 'premium_2_summary' in columns:
                premium_2_column = columns['premium_2_summary']
                merge_range_o = f'{premium_2_column}{start_row}:{premium_2_column}{end_row}'
                worksheet.merge_cells(merge_range_o)
                merged_cell_o = worksheet[f'O{start_row}']
                merged_cell_o.value = premium_2_total
                
                # Применяем форматирование
                self._apply_premium_cell_formatting(worksheet, f'O{start_row}', self.FIRST_DATA_ROW)
                
                logger.debug(f"Объединены ячейки {merge_range_o} с суммой премий-2: {premium_2_total}")
            else:
                # Если нет премий-2, все равно объединяем ячейки но оставляем пустыми
                merge_range_o = f'O{start_row}:O{end_row}'
                worksheet.merge_cells(merge_range_o)
                merged_cell_o = worksheet[f'O{start_row}']
                merged_cell_o.value = None
                
                # Применяем форматирование
                self._apply_premium_cell_formatting(worksheet, f'O{start_row}', self.FIRST_DATA_ROW)
                
                logger.debug(f"Объединены ячейки {merge_range_o} без данных (премии-2 отсутствуют)")
            
            logger.info(f"Объединение ячеек премий для компании '{company_name}' завершено успешно")
            
        except Exception as e:
            logger.warning(f"Не удалось объединить ячейки премий для компании '{company_name}': {e}")
            # Fallback: заполняем премии в отдельные ячейки
            self._fill_premium_summary_fallback(worksheet, start_row, end_row, offers)
    
    def _calculate_premium_sum(self, offers: List, premium_type: int) -> Optional[Decimal]:
        """
        Вычисляет сумму премий по всем годам страхования для указанного типа
        
        Args:
            offers: Список предложений компании
            premium_type: Тип премии (1 или 2)
            
        Returns:
            Optional[Decimal]: Сумма премий или None если нет валидных данных
        """
        try:
            logger.debug(f"Вычисляем сумму премий типа {premium_type} для {len(offers)} предложений")
            
            total_sum = Decimal('0')
            valid_premiums_count = 0
            
            for offer in offers:
                try:
                    if premium_type == 1:
                        premium = self._format_premium(offer, 1)
                    else:
                        premium = self._format_premium(offer, 2)
                    
                    if premium is not None and premium > 0:
                        total_sum += premium
                        valid_premiums_count += 1
                        logger.debug(f"Добавлена премия-{premium_type}: {premium}, текущая сумма: {total_sum}")
                    else:
                        logger.debug(f"Премия-{premium_type} отсутствует или равна 0 для года {getattr(offer, 'insurance_year', 'неизвестно')}")
                        
                except Exception as offer_error:
                    logger.warning(f"Ошибка при обработке премии-{premium_type} для года {getattr(offer, 'insurance_year', 'неизвестно')}: {offer_error}")
                    continue
            
            if valid_premiums_count == 0:
                logger.debug(f"Не найдено валидных премий типа {premium_type}")
                return None
            
            logger.debug(f"Сумма премий-{premium_type}: {total_sum} (из {valid_premiums_count} валидных значений)")
            return total_sum
            
        except Exception as e:
            logger.error(f"Ошибка при вычислении суммы премий типа {premium_type}: {e}")
            return None
    
    def _apply_premium_cell_formatting(self, worksheet, cell_address: str, template_row: int) -> None:
        """
        Применяет форматирование к ячейкам с суммами премий
        
        Args:
            worksheet: Рабочий лист Excel
            cell_address: Адрес ячейки для форматирования
            template_row: Номер строки-шаблона для копирования стилей
        """
        try:
            logger.debug(f"Применяем форматирование к ячейке {cell_address}")
            
            # Получаем колонку из адреса ячейки
            column = cell_address[0]  # Первый символ - это колонка (I или O)
            
            # Копируем стили из соответствующей ячейки шаблона
            template_cell = worksheet[f'{column}{template_row}']
            target_cell = worksheet[cell_address]
            
            # Копируем все стили
            if template_cell.has_style:
                from copy import copy
                
                if template_cell.font:
                    target_cell.font = copy(template_cell.font)
                if template_cell.border:
                    target_cell.border = copy(template_cell.border)
                if template_cell.fill:
                    target_cell.fill = copy(template_cell.fill)
                if template_cell.number_format:
                    target_cell.number_format = template_cell.number_format
                if template_cell.protection:
                    target_cell.protection = copy(template_cell.protection)
            
            # Применяем вертикальное выравнивание по центру
            from openpyxl.styles import Alignment
            if target_cell.alignment:
                from copy import copy
                alignment = copy(target_cell.alignment)
                alignment.vertical = 'center'
                target_cell.alignment = alignment
            else:
                target_cell.alignment = Alignment(vertical='center')
            
            logger.debug(f"Форматирование применено к ячейке {cell_address}")
            
        except Exception as e:
            logger.warning(f"Не удалось применить форматирование к ячейке {cell_address}: {e}")
    
    def _fill_premium_summary_fallback(self, worksheet, start_row: int, end_row: int, offers: List) -> None:
        """
        Fallback метод: заполняет премии в отдельные ячейки при ошибке объединения
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            offers: Список предложений компании
        """
        try:
            logger.info(f"Используем fallback для заполнения премий в строках {start_row}-{end_row}")
            
            # Сначала пытаемся разъединить ячейки если они были объединены
            try:
                for column in ['I', 'O']:
                    merge_range = f'{column}{start_row}:{column}{end_row}'
                    if merge_range in [str(r) for r in worksheet.merged_cells.ranges]:
                        worksheet.unmerge_cells(merge_range)
                        logger.debug(f"Разъединены ячейки {merge_range} для fallback")
            except Exception as unmerge_error:
                logger.debug(f"Не удалось разъединить ячейки (возможно, они не были объединены): {unmerge_error}")
            
            # Заполняем каждую строку отдельно
            for i, offer in enumerate(offers):
                if i >= (end_row - start_row + 1):
                    break  # Защита от выхода за границы
                
                current_row = start_row + i
                
                try:
                    # Заполняем премию-1
                    premium_1 = self._format_premium(offer, 1)
                    if premium_1 is not None:
                        worksheet[f'I{current_row}'].value = premium_1
                    
                    # Заполняем премию-2 если есть
                    if offer.has_second_franchise_variant():
                        premium_2 = self._format_premium(offer, 2)
                        if premium_2 is not None:
                            worksheet[f'O{current_row}'].value = premium_2
                    
                except Exception as row_error:
                    logger.warning(f"Не удалось заполнить премии для строки {current_row}: {row_error}")
                    continue
            
            logger.info(f"Fallback заполнение премий завершено для строк {start_row}-{end_row}")
            
        except Exception as e:
            logger.error(f"Критическая ошибка при fallback заполнении премий: {e}")
    
    def _fill_single_year_premium_summary(self, worksheet, row_num: int, company_name: str, 
                                         offer, columns_mapping: dict) -> None:
        """
        Заполняет столбцы I и O для компаний с одним годом страхования
        
        Для компаний с одним годом не нужно объединение ячеек, но нужно 
        скопировать значения премий из столбцов F и L в столбцы I и O соответственно.
        
        Args:
            worksheet: Рабочий лист Excel
            row_num: Номер строки компании
            company_name: Название страховой компании
            offer: Предложение компании (единственное)
            columns_mapping: Маппинг колонок
        """
        try:
            logger.debug(f"Заполняем столбцы премий для компании с одним годом '{company_name}' в строке {row_num}")
            
            # Получаем значения премий из столбцов F и L
            premium_1_cell = f"{columns_mapping['premium_1']}{row_num}"
            premium_1_value = worksheet[premium_1_cell].value
            
            # Заполняем столбец I значением из столбца F
            if premium_1_value is not None:
                premium_1_summary_cell = f"{columns_mapping['premium_1_summary']}{row_num}"
                worksheet[premium_1_summary_cell].value = premium_1_value
                logger.debug(f"Скопирована премия-1 из {premium_1_cell} в {premium_1_summary_cell}: {premium_1_value}")
            
            # Заполняем столбец O значением из столбца L (если есть и колонка существует)
            if 'premium_2' in columns_mapping and 'premium_2_summary' in columns_mapping:
                premium_2_cell = f"{columns_mapping['premium_2']}{row_num}"
                premium_2_value = worksheet[premium_2_cell].value
                
                if premium_2_value is not None:
                    premium_2_summary_cell = f"{columns_mapping['premium_2_summary']}{row_num}"
                    worksheet[premium_2_summary_cell].value = premium_2_value
                    logger.debug(f"Скопирована премия-2 из {premium_2_cell} в {premium_2_summary_cell}: {premium_2_value}")
                else:
                    logger.debug(f"Премия-2 отсутствует для компании '{company_name}', столбец O остается пустым")
            else:
                logger.debug(f"Столбцы для премии-2 отсутствуют в шаблоне (упрощенный шаблон)")
            
            logger.info(f"Столбцы премий заполнены для компании с одним годом '{company_name}' в строке {row_num}")
            
        except Exception as e:
            logger.warning(f"Ошибка при заполнении столбцов премий для компании '{company_name}' в строке {row_num}: {e}")
            # Не прерываем выполнение, продолжаем работу
    
    def _merge_notes_cells(self, worksheet, start_row: int, end_row: int, 
                          company_name: str, offers: List, columns_mapping: dict = None) -> None:
        """
        Объединяет ячейки в столбце примечаний с консолидацией примечаний по всем годам страхования
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            company_name: Название страховой компании
            offers: Список предложений компании
            columns_mapping: Маппинг колонок (если None, используется COMPANY_DATA_COLUMNS)
            
        Raises:
            ExcelExportServiceError: При критических ошибках объединения
        """
        try:
            # Используем переданный маппинг или значение по умолчанию
            columns = columns_mapping if columns_mapping is not None else self.COMPANY_DATA_COLUMNS
            notes_column = columns['notes']
            
            logger.debug(f"Объединяем ячейки примечаний для компании '{company_name}' в диапазоне {notes_column}{start_row}:{notes_column}{end_row}")
            
            # Консолидируем примечания из всех предложений
            consolidated_notes = self._consolidate_notes(offers)
            
            # Определяем диапазон для объединения в столбце примечаний
            merge_range_q = f'{notes_column}{start_row}:{notes_column}{end_row}'
            
            # Объединяем ячейки
            worksheet.merge_cells(merge_range_q)
            merged_cell_q = worksheet[f'{notes_column}{start_row}']
            
            # Записываем консолидированные примечания
            if consolidated_notes:
                merged_cell_q.value = consolidated_notes
                logger.debug(f"Записаны консолидированные примечания: {consolidated_notes[:50]}...")
            else:
                merged_cell_q.value = None
                logger.debug("Примечания отсутствуют, ячейка оставлена пустой")
            
            # Применяем форматирование
            self._apply_notes_cell_formatting(worksheet, f'{notes_column}{start_row}', self.FIRST_DATA_ROW)
            
            logger.info(f"Объединены ячейки {merge_range_q} для примечаний компании '{company_name}'")
            
        except Exception as e:
            logger.warning(f"Не удалось объединить ячейки примечаний для компании '{company_name}': {e}")
            # Fallback: заполняем примечания в отдельные ячейки
            self._fill_notes_fallback(worksheet, start_row, end_row, offers)
    
    def _consolidate_notes(self, offers: List) -> Optional[str]:
        """
        Объединяет все примечания по годам в единый текст с разделением пробелом
        
        Пример:
        - Год 1: "тест1"
        - Год 2: "тест2" 
        - Год 3: "" (пустое)
        Результат: "тест1 тест2"
        
        Args:
            offers: Список предложений компании
            
        Returns:
            Optional[str]: Объединенный текст примечаний или None если нет примечаний
        """
        try:
            logger.debug(f"Консолидируем примечания из {len(offers)} предложений")
            
            consolidated_parts = []
            
            # Сортируем предложения по годам для правильного порядка
            sorted_offers = sorted(offers, key=lambda x: getattr(x, 'insurance_year', 0))
            
            for offer in sorted_offers:
                try:
                    # Получаем примечания из предложения
                    notes = getattr(offer, 'notes', None)
                    
                    if notes:
                        # Преобразуем в строку и очищаем от лишних пробелов
                        notes_str = str(notes).strip()
                        
                        if notes_str:  # Пропускаем пустые примечания
                            # Удаляем потенциально проблемные символы для Excel
                            notes_str = notes_str.replace('\x00', '').replace('\x01', '').replace('\x02', '')
                            notes_str = notes_str.replace('\n', ' ').replace('\r', ' ')
                            
                            # Убираем множественные пробелы
                            notes_str = ' '.join(notes_str.split())
                            
                            if notes_str:  # Проверяем, что после очистки что-то осталось
                                consolidated_parts.append(notes_str)
                                logger.debug(f"Добавлены примечания для года {getattr(offer, 'insurance_year', 'неизвестно')}: {notes_str[:30]}...")
                            else:
                                logger.debug(f"Примечания для года {getattr(offer, 'insurance_year', 'неизвестно')} пусты после очистки")
                        else:
                            logger.debug(f"Примечания для года {getattr(offer, 'insurance_year', 'неизвестно')} пусты")
                    else:
                        logger.debug(f"Примечания отсутствуют для года {getattr(offer, 'insurance_year', 'неизвестно')}")
                        
                except Exception as offer_error:
                    logger.warning(f"Ошибка при обработке примечаний для года {getattr(offer, 'insurance_year', 'неизвестно')}: {offer_error}")
                    continue
            
            # Объединяем все части с разделением пробелом
            if consolidated_parts:
                consolidated_text = ' '.join(consolidated_parts)
                
                # Ограничиваем длину для Excel (максимум 32767 символов)
                max_excel_length = 32767
                if len(consolidated_text) > max_excel_length:
                    consolidated_text = consolidated_text[:max_excel_length - 3] + "..."
                    logger.warning(f"Консолидированные примечания обрезаны до {max_excel_length} символов")
                
                logger.debug(f"Консолидированы примечания: {len(consolidated_parts)} частей, итоговая длина: {len(consolidated_text)}")
                return consolidated_text
            else:
                logger.debug("Нет примечаний для консолидации")
                return None
                
        except Exception as e:
            logger.error(f"Ошибка при консолидации примечаний: {e}")
            return None
    
    def _apply_notes_cell_formatting(self, worksheet, cell_address: str, template_row: int) -> None:
        """
        Применяет форматирование к ячейкам с объединенными примечаниями, 
        копируя стили из строки 10
        
        Args:
            worksheet: Рабочий лист Excel
            cell_address: Адрес ячейки для форматирования (например, 'Q10')
            template_row: Номер строки-шаблона для копирования стилей (обычно 10)
        """
        try:
            logger.debug(f"Применяем форматирование к ячейке примечаний {cell_address}")
            
            # Получаем колонку из адреса ячейки (должна быть Q)
            column = cell_address[0]  # Первый символ - это колонка
            
            # Копируем стили из соответствующей ячейки шаблона (Q10)
            template_cell = worksheet[f'{column}{template_row}']
            target_cell = worksheet[cell_address]
            
            # Копируем все стили из шаблона
            if template_cell.has_style:
                from copy import copy
                
                if template_cell.font:
                    target_cell.font = copy(template_cell.font)
                if template_cell.border:
                    target_cell.border = copy(template_cell.border)
                if template_cell.fill:
                    target_cell.fill = copy(template_cell.fill)
                if template_cell.number_format:
                    target_cell.number_format = template_cell.number_format
                if template_cell.protection:
                    target_cell.protection = copy(template_cell.protection)
            
            # Применяем вертикальное выравнивание по центру
            from openpyxl.styles import Alignment
            if target_cell.alignment:
                from copy import copy
                alignment = copy(target_cell.alignment)
                alignment.vertical = 'center'
                target_cell.alignment = alignment
            else:
                target_cell.alignment = Alignment(vertical='center')
            
            logger.debug(f"Форматирование применено к ячейке примечаний {cell_address}")
            
        except Exception as e:
            logger.warning(f"Не удалось применить форматирование к ячейке примечаний {cell_address}: {e}")
    
    def _fill_notes_fallback(self, worksheet, start_row: int, end_row: int, offers: List) -> None:
        """
        Fallback метод: заполняет примечания в отдельные ячейки при ошибке объединения
        
        Args:
            worksheet: Рабочий лист Excel
            start_row: Первая строка компании
            end_row: Последняя строка компании
            offers: Список предложений компании
        """
        try:
            logger.info(f"Используем fallback для заполнения примечаний в строках {start_row}-{end_row}")
            
            # Сначала пытаемся разъединить ячейки если они были объединены
            try:
                merge_range = f'Q{start_row}:Q{end_row}'
                if merge_range in [str(r) for r in worksheet.merged_cells.ranges]:
                    worksheet.unmerge_cells(merge_range)
                    logger.debug(f"Разъединены ячейки {merge_range} для fallback")
            except Exception as unmerge_error:
                logger.debug(f"Не удалось разъединить ячейки (возможно, они не были объединены): {unmerge_error}")
            
            # Заполняем каждую строку отдельно
            for i, offer in enumerate(offers):
                if i >= (end_row - start_row + 1):
                    break  # Защита от выхода за границы
                
                current_row = start_row + i
                
                try:
                    # Получаем примечания для текущего предложения
                    notes = getattr(offer, 'notes', None)
                    
                    if notes:
                        notes_str = str(notes).strip()
                        if notes_str:
                            # Ограничиваем длину примечаний для Excel
                            if len(notes_str) > self.MAX_NOTES_LENGTH:
                                notes_str = notes_str[:self.MAX_NOTES_LENGTH] + "..."
                                logger.warning(f"Примечания обрезаны до {self.MAX_NOTES_LENGTH} символов для строки {current_row}")
                            
                            # Очищаем проблемные символы
                            notes_str = notes_str.replace('\x00', '').replace('\x01', '').replace('\x02', '')
                            
                            if notes_str:  # Проверяем, что после очистки что-то осталось
                                worksheet[f'Q{current_row}'].value = notes_str
                                logger.debug(f"Записаны примечания в строку {current_row}: {notes_str[:30]}...")
                    
                except Exception as row_error:
                    logger.warning(f"Не удалось заполнить примечания для строки {current_row}: {row_error}")
                    continue
            
            logger.info(f"Fallback заполнение примечаний завершено для строк {start_row}-{end_row}")
            
        except Exception as e:
            logger.error(f"Критическая ошибка при fallback заполнении примечаний: {e}")


class ExcelResponseProcessor:
    """Сервис для обработки Excel файлов с ответами страховых компаний"""
    
    # Конфигурация для поддерживаемых строк
    MIN_YEAR_ROW = 6  # Минимальная строка для данных лет
    MAX_YEAR_ROW = 10  # Максимальная строка для данных лет
    
    # Маппинг колонок для данных лет
    YEAR_COLUMNS = {
        'year': 'A',
        'insurance_sum': 'B',
        'premium': 'D',
        'franchise': 'E',
        'installment': 'F',
        'premium_2': 'H',
        'franchise_2': 'I',
        'installment_2': 'J'
    }
    
    # Динамическая конфигурация маппинга ячеек
    CELL_MAPPING = {
        'company_name': 'B2',
        'notes_first_year': 'F2',  # Объединенная ячейка FGHIJ2 для примечаний первого года
        'year_rows': {
            'start_row': MIN_YEAR_ROW,
            'end_row': MAX_YEAR_ROW,
            'columns': YEAR_COLUMNS
        }
    }
    
    # Допустимые значения рассрочки
    VALID_INSTALLMENT_VALUES = [1, 2, 3, 4, 6, 12]
    
    def __init__(self):
        """Инициализация процессора"""
        self.logger = logging.getLogger(f'{__name__}.{self.__class__.__name__}')
        # Импортируем здесь, чтобы избежать циклических импортов
        from .company_matcher import create_company_matcher
        self.company_matcher = create_company_matcher()
    
    def _generate_year_mappings(self) -> Dict[str, Dict[str, str]]:
        """
        Генерирует маппинги ячеек для всех поддерживаемых лет (строки 6-10)
        
        Returns:
            Dict с маппингами для каждого года в формате:
            {
                'year_1': {'year': 'A6', 'insurance_sum': 'B6', ...},
                'year_2': {'year': 'A7', 'insurance_sum': 'B7', ...},
                ...
            }
        """
        mappings = {}
        year_config = self.CELL_MAPPING['year_rows']
        
        for row_num in range(year_config['start_row'], year_config['end_row'] + 1):
            year_key = f"year_{row_num - year_config['start_row'] + 1}"
            mappings[year_key] = {}
            
            for field_name, column_letter in year_config['columns'].items():
                cell_address = f"{column_letter}{row_num}"
                mappings[year_key][field_name] = cell_address
        
        self.logger.debug(f"Сгенерированы маппинги для {len(mappings)} лет: {list(mappings.keys())}")
        return mappings
    
    def _detect_available_years(self, worksheet) -> List[int]:
        """
        Обнаруживает, в каких строках присутствуют данные лет
        
        Args:
            worksheet: Рабочий лист Excel
            
        Returns:
            Список номеров строк с валидными данными года
        """
        available_rows = []
        skipped_rows = []
        year_config = self.CELL_MAPPING['year_rows']
        year_column = year_config['columns']['year']
        
        self.logger.info(f"Начинаем обнаружение данных лет в строках {year_config['start_row']}-{year_config['end_row']}")
        
        for row_num in range(year_config['start_row'], year_config['end_row'] + 1):
            year_cell = f"{year_column}{row_num}"
            year_value = self._get_cell_value(worksheet, year_cell)
            
            self.logger.debug(f"Строка {row_num}: проверяем ячейку {year_cell}, значение: '{year_value}'")
            
            # Проверяем, есть ли валидное значение года
            if year_value is not None and str(year_value).strip():
                try:
                    year_int = int(year_value)
                    if 1 <= year_int <= 10:  # Разумные ограничения для года страхования
                        available_rows.append(row_num)
                        self.logger.info(f"Строка {row_num}: найден валидный год страхования {year_int}")
                    else:
                        skipped_rows.append(row_num)
                        self.logger.warning(f"Строка {row_num}: год {year_int} вне допустимого диапазона (1-10), строка пропущена")
                except (ValueError, TypeError):
                    skipped_rows.append(row_num)
                    self.logger.warning(f"Строка {row_num}: некорректное значение года '{year_value}', строка пропущена")
                    continue
            else:
                skipped_rows.append(row_num)
                self.logger.debug(f"Строка {row_num}: пустое значение года, строка пропущена")
        
        self.logger.info(f"Обнаружение завершено: найдено {len(available_rows)} строк с данными, пропущено {len(skipped_rows)} строк")
        self.logger.info(f"Строки с данными: {available_rows}")
        if skipped_rows:
            self.logger.info(f"Пропущенные строки: {skipped_rows}")
        
        return available_rows
    
    def _extract_all_years_data(self, worksheet) -> Dict[str, Any]:
        """
        Извлекает данные для всех обнаруженных лет с информацией об обработке
        
        Args:
            worksheet: Рабочий лист Excel
            
        Returns:
            Dict с данными по годам и информацией об обработке:
            {
                'years': List[Dict[str, Any]],  # Данные по годам
                'processing_info': {
                    'total_rows_checked': int,
                    'rows_with_data': List[int],
                    'rows_skipped': List[int],
                    'years_processed': List[int]
                }
            }
        """
        self.logger.info("Начинаем извлечение данных по всем годам страхования")
        
        all_years_data = []
        available_rows = self._detect_available_years(worksheet)
        year_mappings = self._generate_year_mappings()
        year_config = self.CELL_MAPPING['year_rows']
        
        # Информация об обработке
        total_rows_checked = year_config['end_row'] - year_config['start_row'] + 1
        all_possible_rows = list(range(year_config['start_row'], year_config['end_row'] + 1))
        rows_skipped = [row for row in all_possible_rows if row not in available_rows]
        years_processed = []
        processing_errors = []
        
        self.logger.info(f"Будет обработано {len(available_rows)} строк с данными: {available_rows}")
        
        for row_num in available_rows:
            try:
                self.logger.debug(f"Обрабатываем строку {row_num}")
                
                # Создаем маппинг для конкретной строки
                row_mapping = {}
                for field_name, column_letter in year_config['columns'].items():
                    cell_address = f"{column_letter}{row_num}"
                    row_mapping[field_name] = cell_address
                
                self.logger.debug(f"Строка {row_num}: маппинг ячеек создан - {row_mapping}")
                
                # Извлекаем данные для этой строки
                year_data = self._extract_year_data(worksheet, row_mapping, row_num)
                if year_data:
                    all_years_data.append(year_data)
                    years_processed.append(year_data['year'])
                    self.logger.info(f"Строка {row_num}: успешно извлечены данные для {year_data['year']} года (сумма: {year_data['insurance_sum']}, премия: {year_data['premium']})")
                else:
                    self.logger.warning(f"Строка {row_num}: данные не извлечены (пустая строка)")
                    
            except RowProcessingError as e:
                # Логируем ошибку обработки строки и добавляем строку в пропущенные
                error_info = f"Строка {row_num}: {str(e)}"
                processing_errors.append(error_info)
                self.logger.error(f"Ошибка обработки - {error_info}")
                
                if row_num in available_rows:
                    available_rows.remove(row_num)
                if row_num not in rows_skipped:
                    rows_skipped.append(row_num)
                continue
        
        processing_info = {
            'total_rows_checked': total_rows_checked,
            'rows_with_data': [row for row in available_rows if any(year['year'] for year in all_years_data)],
            'rows_skipped': rows_skipped,
            'years_processed': years_processed,
            'processing_errors': processing_errors
        }
        
        # Итоговое логирование
        self.logger.info(f"Извлечение данных завершено:")
        self.logger.info(f"  - Всего проверено строк: {total_rows_checked}")
        self.logger.info(f"  - Успешно обработано лет: {len(all_years_data)}")
        self.logger.info(f"  - Годы страхования: {sorted(years_processed) if years_processed else 'нет'}")
        self.logger.info(f"  - Пропущено строк: {len(rows_skipped)} {rows_skipped if rows_skipped else ''}")
        
        if processing_errors:
            self.logger.warning(f"  - Ошибки обработки ({len(processing_errors)}): {processing_errors}")
        
        return {
            'years': all_years_data,
            'processing_info': processing_info
        }
    
    def process_excel_file(self, file, summary: InsuranceSummary) -> Dict[str, Any]:
        """
        Обрабатывает Excel файл и создает предложения
        
        Args:
            file: Загруженный файл Excel
            summary: Свод предложений для связи
            
        Returns:
            Dict с результатами обработки
            
        Raises:
            ExcelProcessingError: При ошибках обработки файла
        """
        self.logger.info(f"=== НАЧАЛО ОБРАБОТКИ EXCEL ФАЙЛА ===")
        self.logger.info(f"Свод ID: {summary.id}, Файл: {file.name}")
        
        try:
            # Загружаем Excel файл
            self.logger.info("Этап 1: Загрузка Excel файла")
            workbook = self._load_excel_file(file)
            worksheet = self._get_worksheet(workbook)
            
            # Извлекаем данные компании
            self.logger.info("Этап 2: Извлечение данных компании")
            company_data = self.extract_company_data(worksheet)
            
            # Логируем результаты извлечения
            processing_info = company_data.get('processing_info', {})
            self.logger.info(f"Результаты извлечения данных:")
            self.logger.info(f"  - Компания: {company_data['company_name']}")
            self.logger.info(f"  - Найдено лет: {len(company_data['years'])}")
            self.logger.info(f"  - Обработано строк: {processing_info.get('rows_with_data', [])}")
            self.logger.info(f"  - Пропущено строк: {processing_info.get('rows_skipped', [])}")
            
            # Валидируем извлеченные данные
            self.logger.info("Этап 3: Валидация извлеченных данных")
            self.validate_extracted_data(company_data)
            
            # Создаем предложения
            self.logger.info("Этап 4: Создание предложений в базе данных")
            created_offers = self.create_offers(company_data, summary)
            
            result = {
                'success': True,
                'company_name': company_data['company_name'],
                'offers_created': len(created_offers),
                'years': [offer.insurance_year for offer in created_offers],
                'skipped_rows': processing_info.get('rows_skipped', []),
                'processed_rows': processing_info.get('rows_with_data', []),
                'company_matching_info': company_data.get('company_matching_info', {}),
                'processing_errors': processing_info.get('processing_errors', [])
            }
            
            self.logger.info(f"=== ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО ===")
            self.logger.info(f"Создано предложений: {len(created_offers)} для лет: {sorted([offer.insurance_year for offer in created_offers])}")
            
            return result
            
        except (ExcelProcessingError, DuplicateOfferError) as e:
            self.logger.error(f"=== ОБРАБОТКА ЗАВЕРШЕНА С ОШИБКОЙ ===")
            self.logger.error(f"Тип ошибки: {type(e).__name__}, Сообщение: {str(e)}")
            # Переброс известных исключений
            raise
        except Exception as e:
            error_msg = f"Неожиданная ошибка при обработке Excel файла: {str(e)}"
            self.logger.error(f"=== ОБРАБОТКА ЗАВЕРШЕНА С КРИТИЧЕСКОЙ ОШИБКОЙ ===")
            self.logger.error(error_msg, exc_info=True)
            raise ExcelProcessingError(error_msg) from e
    
    def _load_excel_file(self, file):
        """
        Загружает Excel файл из загруженного файла
        
        Args:
            file: Загруженный файл
            
        Returns:
            Workbook: Загруженная книга Excel
            
        Raises:
            InvalidFileFormatError: При ошибках загрузки файла
        """
        try:
            self.logger.debug("Загружаем Excel файл")
            
            # Проверяем расширение файла
            if not file.name.lower().endswith('.xlsx'):
                raise InvalidFileFormatError("Файл должен иметь расширение .xlsx")
            
            # Загружаем файл
            workbook = load_workbook(file, data_only=True)
            self.logger.debug("Excel файл успешно загружен")
            return workbook
            
        except InvalidFileFormatError:
            raise
        except Exception as e:
            error_msg = f"Ошибка при загрузке Excel файла: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise InvalidFileFormatError(error_msg) from e
    
    def _get_worksheet(self, workbook):
        """
        Получает рабочий лист для обработки
        
        Args:
            workbook: Книга Excel
            
        Returns:
            Worksheet: Рабочий лист
            
        Raises:
            InvalidFileFormatError: Если не удается найти подходящий лист
        """
        try:
            if not workbook.worksheets:
                raise InvalidFileFormatError("В Excel файле не найдено ни одного рабочего листа")
            
            # Используем первый лист
            worksheet = workbook.worksheets[0]
            self.logger.debug(f"Используем рабочий лист: {worksheet.title}")
            return worksheet
            
        except Exception as e:
            error_msg = f"Ошибка при получении рабочего листа: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise InvalidFileFormatError(error_msg) from e
    
    def extract_company_data(self, worksheet) -> Dict[str, Any]:
        """
        Извлекает данные компании из Excel файла с сопоставлением названия компании
        
        Args:
            worksheet: Рабочий лист Excel
            
        Returns:
            Dict с извлеченными данными
            
        Raises:
            MissingDataError: При отсутствии обязательных данных
        """
        self.logger.info("Начинаем извлечение данных компании из Excel файла")
        
        try:
            data = {}
            
            # Извлекаем название компании
            self.logger.debug(f"Извлекаем название компании из ячейки {self.CELL_MAPPING['company_name']}")
            raw_company_name = self._get_cell_value(worksheet, self.CELL_MAPPING['company_name'])
            if not raw_company_name:
                self.logger.error(f"Название компании не найдено в ячейке {self.CELL_MAPPING['company_name']}")
                raise MissingDataError([self.CELL_MAPPING['company_name']])
            
            # Сопоставляем название компании с закрытым списком
            raw_name_str = str(raw_company_name).strip()
            self.logger.info(f"Исходное название компании: '{raw_name_str}'")
            
            standardized_name = self.company_matcher.match_company_name(raw_name_str)
            data['company_name'] = standardized_name
            
            # Сохраняем информацию о сопоставлении для пользователя
            matching_info = {
                'original_name': raw_name_str,
                'standardized_name': standardized_name,
                'was_matched': standardized_name != raw_name_str,
                'assigned_other': standardized_name == 'другое' and raw_name_str.lower() != 'другое'
            }
            data['company_matching_info'] = matching_info
            
            # Логируем процесс сопоставления
            if standardized_name == 'другое' and raw_name_str.lower() != 'другое':
                self.logger.warning(f"Название компании '{raw_name_str}' не найдено в закрытом списке, присвоено значение 'другое'")
            elif standardized_name != raw_name_str:
                self.logger.info(f"Название компании сопоставлено: '{raw_name_str}' -> '{standardized_name}'")
            else:
                self.logger.info(f"Название компании '{raw_name_str}' точно совпадает с элементом закрытого списка")
            
            # Извлекаем данные по годам с использованием динамического обнаружения
            self.logger.info("Начинаем извлечение данных по годам страхования")
            years_result = self._extract_all_years_data(worksheet)
            data['years'] = years_result['years']
            data['processing_info'] = years_result['processing_info']
            
            # Извлекаем примечания для первого года из объединенной ячейки FGHIJ2
            self.logger.debug(f"Извлекаем примечания первого года из ячейки {self.CELL_MAPPING['notes_first_year']}")
            notes_first_year = self._get_cell_value(worksheet, self.CELL_MAPPING['notes_first_year'])
            if notes_first_year and str(notes_first_year).strip():
                notes_text = str(notes_first_year).strip()
                self.logger.info(f"Найдены примечания для первого года: {notes_text[:50]}...")
                # Добавляем примечания к первому году (год = 1)
                for year_data in data['years']:
                    if year_data['year'] == 1:
                        year_data['notes'] = notes_text
                        self.logger.debug(f"Примечания добавлены к году {year_data['year']}")
                        break
            else:
                self.logger.debug("Примечания для первого года не найдены или пусты")
            
            # Проверяем, что есть хотя бы один год с данными
            if not data['years']:
                self.logger.error("Не найдено ни одного года с валидными данными страхования")
                raise MissingDataError(['данные по годам страхования'])
            
            self.logger.info(f"Извлечение данных компании завершено успешно:")
            self.logger.info(f"  - Компания: '{data['company_name']}'")
            self.logger.info(f"  - Количество лет: {len(data['years'])}")
            self.logger.info(f"  - Годы: {[year['year'] for year in data['years']]}")
            
            return data
            
        except (MissingDataError, InvalidDataError):
            raise
        except Exception as e:
            error_msg = f"Ошибка при извлечении данных компании: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise ExcelProcessingError(error_msg) from e
    
    def _extract_year_data(self, worksheet, year_mapping: Dict[str, str], row_number: int) -> Optional[Dict[str, Any]]:
        """
        Извлекает данные для конкретного года
        
        Args:
            worksheet: Рабочий лист Excel
            year_mapping: Маппинг ячеек для года
            row_number: Номер строки Excel (для ошибок и логирования)
            
        Returns:
            Dict с данными года или None, если данных нет
            
        Raises:
            RowProcessingError: При ошибках валидации данных в строке
        """
        try:
            # Проверяем, есть ли год в ячейке
            year_value = self._get_cell_value(worksheet, year_mapping['year'])
            if not year_value:
                self.logger.debug(f"Строка {row_number}: год не указан, пропускаем")
                return None
            
            # Извлекаем основные данные года с обработкой ошибок для конкретной строки
            year_data = {
                'year': self._parse_year_with_row(year_value, year_mapping['year'], row_number),
                'insurance_sum': self._parse_decimal_with_row(
                    self._get_cell_value(worksheet, year_mapping['insurance_sum']),
                    year_mapping['insurance_sum'],
                    'страховая сумма',
                    row_number
                ),
                'premium': self._parse_decimal_with_row(
                    self._get_cell_value(worksheet, year_mapping['premium']),
                    year_mapping['premium'],
                    'премия',
                    row_number
                ),
                'franchise': self._parse_decimal_with_row(
                    self._get_cell_value(worksheet, year_mapping['franchise']),
                    year_mapping['franchise'],
                    'франшиза',
                    row_number,
                    default_value=Decimal('0')
                ),
                'installment': self._parse_installment_with_row(
                    self._get_cell_value(worksheet, year_mapping['installment']),
                    year_mapping['installment'],
                    row_number
                )
            }
            
            # Извлекаем дополнительные поля (H, I, J), если они присутствуют в маппинге
            if 'premium_2' in year_mapping:
                year_data['premium_2'] = self._parse_decimal_with_row(
                    self._get_cell_value(worksheet, year_mapping['premium_2']),
                    year_mapping['premium_2'],
                    'премия-2',
                    row_number,
                    required=False,
                    default_value=None
                )
            
            if 'franchise_2' in year_mapping:
                year_data['franchise_2'] = self._parse_decimal_with_row(
                    self._get_cell_value(worksheet, year_mapping['franchise_2']),
                    year_mapping['franchise_2'],
                    'франшиза-2',
                    row_number,
                    required=False,
                    default_value=None
                )
            
            if 'installment_2' in year_mapping:
                year_data['installment_2'] = self._parse_installment_with_row(
                    self._get_cell_value(worksheet, year_mapping['installment_2']),
                    year_mapping['installment_2'],
                    row_number,
                    required=False
                )
            
            self.logger.debug(f"Строка {row_number}: извлечены данные для {year_data['year']} года")
            return year_data
            
        except RowProcessingError:
            # Переброс ошибок строки как есть
            raise
        except Exception as e:
            # Неожиданные ошибки оборачиваем в RowProcessingError
            error_msg = f"Неожиданная ошибка при обработке данных: {str(e)}"
            self.logger.error(f"Строка {row_number}: {error_msg}")
            raise RowProcessingError(row_number, 'общая обработка', error_msg)
    
    def _get_cell_value(self, worksheet, cell_address: str):
        """
        Получает значение ячейки
        
        Args:
            worksheet: Рабочий лист
            cell_address: Адрес ячейки (например, 'B2')
            
        Returns:
            Значение ячейки или None
        """
        try:
            cell = worksheet[cell_address]
            return cell.value
        except Exception as e:
            self.logger.warning(f"Ошибка при чтении ячейки {cell_address}: {str(e)}")
            return None
    
    def _parse_year(self, value, cell_address: str) -> int:
        """
        Парсит значение года (устаревший метод для обратной совместимости)
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            
        Returns:
            Номер года
            
        Raises:
            InvalidDataError: При некорректном значении года
        """
        try:
            if value is None:
                raise InvalidDataError('год', value, 'положительное число')
            
            year = int(value)
            if year < 1 or year > 10:  # Разумные ограничения для года страхования
                raise InvalidDataError('год', value, 'число от 1 до 10')
            
            return year
            
        except (ValueError, TypeError):
            raise InvalidDataError('год', value, 'положительное число')
    
    def _parse_year_with_row(self, value, cell_address: str, row_number: int) -> int:
        """
        Парсит значение года с указанием номера строки
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            row_number: Номер строки Excel
            
        Returns:
            Номер года
            
        Raises:
            RowProcessingError: При некорректном значении года
        """
        try:
            if value is None:
                raise RowProcessingError(row_number, 'год', 'значение не указано', cell_address)
            
            year = int(value)
            if year < 1 or year > 10:  # Разумные ограничения для года страхования
                raise RowProcessingError(row_number, 'год', f'значение {value} вне допустимого диапазона (1-10)', cell_address)
            
            return year
            
        except (ValueError, TypeError):
            raise RowProcessingError(row_number, 'год', f'некорректное значение "{value}", ожидается положительное число', cell_address)
    
    def _parse_decimal(self, value, cell_address: str, field_name: str, default_value: Optional[Decimal] = None) -> Decimal:
        """
        Парсит десятичное значение (устаревший метод для обратной совместимости)
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            field_name: Название поля для ошибок
            default_value: Значение по умолчанию
            
        Returns:
            Десятичное значение
            
        Raises:
            InvalidDataError: При некорректном значении
        """
        try:
            if value is None or value == '':
                if default_value is not None:
                    return default_value
                raise InvalidDataError(field_name, value, 'числовое значение')
            
            # Преобразуем в Decimal
            decimal_value = Decimal(str(value))
            
            # Проверяем, что значение положительное (кроме франшизы, которая может быть 0)
            if decimal_value < 0:
                raise InvalidDataError(field_name, value, 'положительное число')
            
            return decimal_value
            
        except (InvalidOperation, ValueError, TypeError):
            raise InvalidDataError(field_name, value, 'числовое значение')
    
    def _parse_decimal_with_row(self, value, cell_address: str, field_name: str, row_number: int, default_value: Optional[Decimal] = None, required: bool = True) -> Optional[Decimal]:
        """
        Парсит десятичное значение с указанием номера строки
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            field_name: Название поля для ошибок
            row_number: Номер строки Excel
            default_value: Значение по умолчанию
            required: Обязательно ли поле (если False, пустые значения возвращают None)
            
        Returns:
            Десятичное значение или None для необязательных полей
            
        Raises:
            RowProcessingError: При некорректном значении
        """
        try:
            if value is None or value == '':
                if default_value is not None:
                    return default_value
                if not required:
                    return None
                raise RowProcessingError(row_number, field_name, 'значение не указано', cell_address)
            
            # Преобразуем в Decimal
            decimal_value = Decimal(str(value))
            
            # Проверяем, что значение положительное (кроме франшизы, которая может быть 0)
            if decimal_value < 0:
                raise RowProcessingError(row_number, field_name, f'отрицательное значение {value}, ожидается положительное число', cell_address)
            
            return decimal_value
            
        except (InvalidOperation, ValueError, TypeError):
            # Специальная обработка для франшизы: текстовые значения интерпретируются как 0
            if 'франшиз' in field_name.lower():
                self.logger.info(f"Строка {row_number}: текстовое значение '{value}' в поле '{field_name}' интерпретировано как 0")
                return Decimal('0')
            
            raise RowProcessingError(row_number, field_name, f'некорректное значение "{value}", ожидается числовое значение', cell_address)
    
    def _parse_installment(self, value, cell_address: str) -> int:
        """
        Парсит значение рассрочки (устаревший метод для обратной совместимости)
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            
        Returns:
            Количество платежей в год
            
        Raises:
            InvalidDataError: При некорректном значении рассрочки
        """
        try:
            if value is None or value == '':
                return 1  # По умолчанию - единовременная оплата
            
            installment = int(value)
            
            if installment not in self.VALID_INSTALLMENT_VALUES:
                raise InvalidDataError(
                    'рассрочка', 
                    value, 
                    f'одно из значений: {", ".join(map(str, self.VALID_INSTALLMENT_VALUES))}'
                )
            
            return installment
            
        except (ValueError, TypeError):
            raise InvalidDataError(
                'рассрочка', 
                value, 
                f'одно из значений: {", ".join(map(str, self.VALID_INSTALLMENT_VALUES))}'
            )
    
    def _parse_installment_with_row(self, value, cell_address: str, row_number: int, required: bool = True) -> Optional[int]:
        """
        Парсит значение рассрочки с указанием номера строки
        
        Args:
            value: Значение из ячейки
            cell_address: Адрес ячейки для ошибок
            row_number: Номер строки Excel
            required: Обязательно ли поле (если False, пустые значения возвращают None)
            
        Returns:
            Количество платежей в год или None для необязательных полей
            
        Raises:
            RowProcessingError: При некорректном значении рассрочки
        """
        try:
            if value is None or value == '':
                if not required:
                    return None
                return 1  # По умолчанию - единовременная оплата
            
            installment = int(value)
            
            if installment not in self.VALID_INSTALLMENT_VALUES:
                valid_values = ", ".join(map(str, self.VALID_INSTALLMENT_VALUES))
                raise RowProcessingError(
                    row_number, 
                    'рассрочка', 
                    f'недопустимое значение {value}, ожидается одно из: {valid_values}',
                    cell_address
                )
            
            return installment
            
        except (ValueError, TypeError):
            valid_values = ", ".join(map(str, self.VALID_INSTALLMENT_VALUES))
            raise RowProcessingError(
                row_number, 
                'рассрочка', 
                f'некорректное значение "{value}", ожидается одно из: {valid_values}',
                cell_address
            )
    
    def validate_extracted_data(self, data: Dict[str, Any]) -> None:
        """
        Валидирует извлеченные данные
        
        Args:
            data: Извлеченные данные
            
        Raises:
            InvalidDataError: При некорректных данных
        """
        self.logger.debug("Валидируем извлеченные данные")
        
        # Проверяем название компании
        if not data.get('company_name') or not data['company_name'].strip():
            raise InvalidDataError('название компании', data.get('company_name'), 'непустая строка')
        
        # Проверяем данные по годам
        if not data.get('years') or len(data['years']) == 0:
            raise InvalidDataError('данные по годам', 'отсутствуют', 'хотя бы один год с данными')
        
        # Валидируем каждый год
        for year_data in data['years']:
            self._validate_year_data(year_data)
        
        self.logger.debug("Валидация данных успешно завершена")
    
    def _validate_year_data(self, year_data: Dict[str, Any]) -> None:
        """
        Валидирует данные конкретного года
        
        Args:
            year_data: Данные года
            
        Raises:
            InvalidDataError: При некорректных данных
        """
        # Проверяем обязательные поля
        required_fields = ['year', 'insurance_sum', 'premium']
        for field in required_fields:
            if field not in year_data or year_data[field] is None:
                raise InvalidDataError(f'{field} (год {year_data.get("year", "?")})', None, 'обязательное поле')
        
        # Проверяем, что премия не больше страховой суммы (разумная проверка)
        if year_data['premium'] > year_data['insurance_sum']:
            raise InvalidDataError(
                f'премия (год {year_data["year"]})', 
                year_data['premium'], 
                f'не больше страховой суммы ({year_data["insurance_sum"]})'
            )
    
    def create_offers(self, data: Dict[str, Any], summary: InsuranceSummary) -> List[InsuranceOffer]:
        """
        Создает записи предложений в базе данных
        
        Args:
            data: Извлеченные и валидированные данные
            summary: Свод предложений
            
        Returns:
            Список созданных предложений
            
        Raises:
            DuplicateOfferError: При дублировании предложений
            ExcelProcessingError: При ошибках создания записей
        """
        self.logger.debug(f"Создаем предложения для компании '{data['company_name']}'")
        
        created_offers = []
        
        try:
            with transaction.atomic():
                for year_data in data['years']:
                    # Проверяем на дублирование
                    existing_offer = InsuranceOffer.objects.filter(
                        summary=summary,
                        company_name=data['company_name'],
                        insurance_year=year_data['year']
                    ).first()
                    
                    if existing_offer:
                        raise DuplicateOfferError(data['company_name'], year_data['year'])
                    
                    # Создаем предложение
                    offer = self._create_single_offer(data, year_data, summary)
                    created_offers.append(offer)
                    
                    self.logger.debug(f"Создано предложение для {year_data['year']} года")
                
                # Обновляем счетчик предложений в своде
                summary.update_total_offers_count()
                
            self.logger.info(f"Успешно создано {len(created_offers)} предложений для компании '{data['company_name']}'")
            return created_offers
            
        except DuplicateOfferError:
            raise
        except Exception as e:
            error_msg = f"Ошибка при создании предложений: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise ExcelProcessingError(error_msg) from e
    
    def _create_single_offer(self, company_data: Dict[str, Any], year_data: Dict[str, Any], summary: InsuranceSummary) -> InsuranceOffer:
        """
        Создает одно предложение
        
        Args:
            company_data: Данные компании
            year_data: Данные года
            summary: Свод предложений
            
        Returns:
            Созданное предложение
        """
        # Определяем параметры рассрочки для основного варианта
        installment_available = year_data['installment'] > 1
        payments_per_year = year_data['installment']
        
        # Определяем параметры рассрочки для дополнительного варианта (если есть)
        installment_2_available = False
        payments_per_year_2 = 1
        if year_data.get('installment_2') is not None:
            installment_2_available = year_data['installment_2'] > 1
            payments_per_year_2 = year_data['installment_2']
        
        # Подготавливаем данные для создания предложения
        offer_data = {
            'summary': summary,
            'company_name': company_data['company_name'],
            'insurance_year': year_data['year'],
            'insurance_sum': year_data['insurance_sum'],
            'franchise_1': year_data['franchise'],
            'premium_with_franchise_1': year_data['premium'],
            'installment_variant_1': installment_available,
            'payments_per_year_variant_1': payments_per_year,
            # Для обратной совместимости
            'installment_available': installment_available,
            'payments_per_year': payments_per_year,
            'is_valid': True
        }
        
        # Добавляем дополнительные поля, если они присутствуют
        if year_data.get('premium_2') is not None:
            offer_data['premium_with_franchise_2'] = year_data['premium_2']
        
        if year_data.get('franchise_2') is not None:
            offer_data['franchise_2'] = year_data['franchise_2']
        
        if year_data.get('installment_2') is not None:
            offer_data['installment_variant_2'] = installment_2_available
            offer_data['payments_per_year_variant_2'] = payments_per_year_2
        
        # Добавляем примечания, если они есть
        if year_data.get('notes'):
            offer_data['notes'] = year_data['notes']
        
        offer = InsuranceOffer.objects.create(**offer_data)
        
        return offer


def get_excel_export_service() -> ExcelExportService:
    """
    Фабричная функция для создания экземпляра ExcelExportService
    с настройками по умолчанию
    
    Returns:
        ExcelExportService: Настроенный экземпляр сервиса
        
    Raises:
        ExcelExportServiceError: Если не удается создать сервис
    """
    try:
        # Получаем путь к шаблону из настроек или используем значение по умолчанию
        template_path = getattr(
            settings, 
            'SUMMARY_TEMPLATE_PATH', 
            settings.BASE_DIR / 'templates' / 'summary_template.xlsx'
        )
        
        return ExcelExportService(str(template_path))
        
    except Exception as e:
        error_msg = f"Ошибка при создании ExcelExportService: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise ExcelExportServiceError(error_msg) from e


def get_excel_response_processor() -> ExcelResponseProcessor:
    """
    Фабричная функция для создания экземпляра ExcelResponseProcessor
    
    Returns:
        ExcelResponseProcessor: Экземпляр процессора
    """
    return ExcelResponseProcessor()