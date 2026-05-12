"""
Представления для работы со сводами предложений
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from decimal import Decimal, InvalidOperation
from django.utils import timezone
import logging
import os

from .models import InsuranceSummary, InsuranceOffer, SummaryTemplate
from insurance_requests.models import InsuranceRequest
from insurance_requests.decorators import user_required, admin_required
from .forms import OfferForm, SummaryForm, AddOfferToSummaryForm, DealListFilterForm
from .exceptions import DuplicateOfferError
from .services.analytics_insurance_companies import (
    DATE_MODE_CHOICES,
    DATE_MODE_SUMMARY_CREATED,
    build_analytics_insurance_companies_payload,
)
from .services import analytics_managers as analytics_managers_service

logger = logging.getLogger(__name__)


@user_required
def summary_list(request):
    """Список всех сводов с фильтрацией и сортировкой"""
    from .forms import SummaryFilterForm
    
    # Получаем базовый queryset с оптимизацией запросов
    summaries = InsuranceSummary.objects.select_related('request', 'request__created_by')
    
    # Получаем список доступных филиалов из сводов
    available_branches = summaries.values_list('request__branch', flat=True).distinct().exclude(
        request__branch__isnull=True
    ).exclude(request__branch='').order_by('request__branch')

    available_insurance_types = list(
        summaries.exclude(
            request__insurance_type__isnull=True
        ).exclude(
            request__insurance_type=''
        ).values_list('request__insurance_type', flat=True).distinct().order_by('request__insurance_type')
    )
    manager_rows = summaries.exclude(
        request__created_by__isnull=True
    ).order_by().values(
        'request__created_by_id',
        'request__created_by__username',
        'request__created_by__first_name',
        'request__created_by__last_name',
    ).distinct()
    available_managers = []
    for manager_row in manager_rows:
        first_name = (manager_row.get('request__created_by__first_name') or '').strip()
        last_name = (manager_row.get('request__created_by__last_name') or '').strip()
        username = (manager_row.get('request__created_by__username') or '').strip()
        display_name = f"{first_name} {last_name}".strip() or username
        available_managers.append((str(manager_row.get('request__created_by_id')), display_name))
    available_managers = sorted(available_managers, key=lambda item: item[1].lower())
    
    # Применяем фильтры
    filter_form = SummaryFilterForm(
        request.GET or None,
        insurance_type_choices=available_insurance_types,
        manager_choices=available_managers,
    )
    is_filter_form_valid = filter_form.is_valid()
    cleaned_data = filter_form.cleaned_data if is_filter_form_valid else {}
    current_branch = request.GET.get('branch')

    def apply_summary_filters(queryset, include_branch=True):
        if include_branch and current_branch:
            queryset = queryset.filter(request__branch=current_branch)

        if not is_filter_form_valid:
            return queryset

        status = cleaned_data.get('status')
        if status:
            queryset = queryset.filter(status=status)

        search = cleaned_data.get('search')
        if search:
            queryset = queryset.filter(
                Q(request__dfa_number__icontains=search)
                | Q(request__client_name__icontains=search)
                | Q(request__inn__icontains=search)
                | Q(selected_company__icontains=search)
            )

        start_date = cleaned_data.get('start_date')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)

        end_date = cleaned_data.get('end_date')
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        insurance_type = cleaned_data.get('insurance_type')
        if insurance_type:
            queryset = queryset.filter(request__insurance_type=insurance_type)

        manager = cleaned_data.get('manager')
        if manager:
            queryset = queryset.filter(request__created_by_id=int(manager))

        return queryset

    summaries = apply_summary_filters(summaries, include_branch=True)
    
    # Подсчет количества сводов для каждого филиала (только если выбран конкретный филиал)
    branch_counts = {}
    if available_branches and current_branch:
        branch_summaries = apply_summary_filters(
            InsuranceSummary.objects.select_related('request', 'request__created_by'),
            include_branch=True,
        )
        branch_counts[current_branch] = branch_summaries.count()
    
    # Подсчет общего количества сводов (только если не выбран конкретный филиал)
    total_summaries_count = 0
    if not current_branch:
        total_summaries_queryset = apply_summary_filters(
            InsuranceSummary.objects.select_related('request', 'request__created_by'),
            include_branch=False,
        )
        total_summaries_count = total_summaries_queryset.count()
    
    # Сортировка по умолчанию
    sort_by = request.GET.get('sort', '-created_at')
    valid_sorts = ['-created_at', 'created_at', '-total_offers', 'total_offers', 'status']
    if sort_by in valid_sorts:
        summaries = summaries.order_by(sort_by)
    else:
        summaries = summaries.order_by('-created_at')
    
    # Реализация пагинации (требование 5.1, 5.2)
    paginator = Paginator(summaries, 30)  # 30 сводов на страницу
    page = request.GET.get('page')
    
    try:
        summaries = paginator.page(page)
    except PageNotAnInteger:
        # Если номер страницы не является целым числом, показываем первую страницу
        summaries = paginator.page(1)
    except EmptyPage:
        # Если номер страницы превышает максимальный, показываем последнюю страницу
        summaries = paginator.page(paginator.num_pages)
    
    # Обновляем контекст шаблона для передачи информации о филиале (требование 4.1, 4.2, 4.3)
    return render(request, 'summaries/summary_list.html', {
        'summaries': summaries,
        'paginator': paginator,
        'filter_form': filter_form,
        'available_branches': available_branches,
        'branch_counts': branch_counts,
        'current_branch': current_branch,
        'total_summaries_count': total_summaries_count,
        'current_sort': sort_by,
        'show_branch_counts': bool(current_branch),  # Показывать счетчики только для активного филиала
        'show_total_count': not current_branch,      # Показывать общий счетчик только для "Все своды"
    })


def _resolve_manager_online_name(insurance_request):
    """Возвращает отображаемое имя менеджера Онлайна из пользователя-создателя заявки."""
    created_by = getattr(insurance_request, 'created_by', None)
    if not created_by:
        return None

    if created_by.last_name and created_by.first_name:
        return f"{created_by.last_name} {created_by.first_name}"
    if created_by.last_name:
        return created_by.last_name
    if created_by.first_name:
        return created_by.first_name
    return created_by.username


def _safe_decimal_or_none(value):
    """Безопасно приводит значение к Decimal или возвращает None."""
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _get_summary_required_variants(summary):
    """Определяет, какие варианты премии нужны для компактной аналитики свода."""
    insurance_request = getattr(summary, 'request', None)
    franchise_type = getattr(insurance_request, 'franchise_type', '') if insurance_request else ''

    if franchise_type == 'both_variants':
        return [1, 2], 'оба варианта'
    if franchise_type == 'with_franchise':
        return [1], 'с франшизой'
    return [1], 'без франшизы'


def _get_offer_premium_for_variant(offer, variant):
    if variant == 2:
        return _safe_decimal_or_none(offer.premium_with_franchise_2)
    return _safe_decimal_or_none(offer.premium_with_franchise_1)


def _summary_offer_has_premium(offer, variant):
    premium = _get_offer_premium_for_variant(offer, variant)
    return premium is not None and premium > 0


def _summary_offer_supports_installment(offer, variant):
    if variant == 2:
        return offer.installment_variant_2 and offer.payments_per_year_variant_2 > 1
    return offer.installment_variant_1 and offer.payments_per_year_variant_1 > 1


def _summary_offer_payments_per_year(offer, variant):
    if variant == 2:
        return offer.payments_per_year_variant_2 if offer.installment_variant_2 else 1
    return offer.payments_per_year_variant_1 if offer.installment_variant_1 else 1


def _format_summary_years_label(years):
    sorted_years = sorted(years)
    if not sorted_years:
        return '—'
    if len(sorted_years) == 1:
        return str(sorted_years[0])
    if sorted_years == list(range(sorted_years[0], sorted_years[-1] + 1)):
        return f'{sorted_years[0]}-{sorted_years[-1]}'
    return ', '.join(str(year) for year in sorted_years)


def _sum_summary_company_variant_total(company_offers, variant, years):
    offers_by_year = {offer.insurance_year: offer for offer in company_offers}
    total = Decimal('0')

    for year in sorted(years):
        offer = offers_by_year.get(year)
        if offer is None:
            return None
        premium = _get_offer_premium_for_variant(offer, variant)
        if premium is None or premium <= 0:
            return None
        total += premium

    return total


def _build_summary_variant_price_range(offers_by_company, variant, years):
    company_totals = []
    for company_name, company_offers in offers_by_company.items():
        total = _sum_summary_company_variant_total(company_offers, variant, years)
        if total is not None:
            company_totals.append((company_name, total))

    if not company_totals:
        return None

    company_totals.sort(key=lambda item: (item[1], item[0]))
    min_total = company_totals[0][1]
    max_total = company_totals[-1][1]
    spread_abs = max_total - min_total
    spread_pct = ((spread_abs / min_total) * Decimal('100')) if min_total > 0 else None
    best_company_names = [
        company_name for company_name, total in company_totals
        if total == min_total
    ]

    return {
        'variant': variant,
        'label': f'Вариант {variant}',
        'min_total': min_total,
        'max_total': max_total,
        'spread_abs': spread_abs,
        'spread_pct': spread_pct,
        'best_company_name': ', '.join(best_company_names),
        'comparable_count': len(company_totals),
    }


def _build_summary_compact_analytics(summary, offers):
    """Собирает компактную аналитику для блока сводной информации."""
    offers_list = list(offers)
    offers_by_company = {}
    for offer in offers_list:
        offers_by_company.setdefault(offer.company_name, []).append(offer)

    sorted_companies = sorted(offers_by_company.keys())
    years = {offer.insurance_year for offer in offers_list}
    required_variants, required_variants_label = _get_summary_required_variants(summary)

    complete_companies = []
    incomplete_companies = []
    for company_name, company_offers in offers_by_company.items():
        offers_by_year = {offer.insurance_year: offer for offer in company_offers}
        company_years = set(offers_by_year.keys())
        has_all_years = bool(years) and company_years == years
        has_required_premiums = has_all_years and all(
            _summary_offer_has_premium(offers_by_year[year], variant)
            for year in years
            for variant in required_variants
        )

        if has_required_premiums:
            complete_companies.append(company_name)
        else:
            incomplete_companies.append(company_name)

    price_ranges = []
    for variant in required_variants:
        price_range = _build_summary_variant_price_range(offers_by_company, variant, years)
        if price_range:
            price_ranges.append(price_range)

    installment_companies = set()
    max_payments_per_year = 1
    for company_name, company_offers in offers_by_company.items():
        company_has_installment = False
        for offer in company_offers:
            for variant in required_variants:
                if _summary_offer_supports_installment(offer, variant):
                    company_has_installment = True
                    max_payments_per_year = max(
                        max_payments_per_year,
                        _summary_offer_payments_per_year(offer, variant),
                    )
        if company_has_installment:
            installment_companies.add(company_name)

    insurance_request = getattr(summary, 'request', None)
    show_installment = bool(
        getattr(insurance_request, 'has_installment', False)
        or installment_companies
    )

    selected = None
    selected_company = (summary.selected_company or '').strip()
    if selected_company:
        selected_variant = summary.selected_franchise_variant
        if selected_variant not in (1, 2):
            selected_variant = required_variants[0] if len(required_variants) == 1 else 1

        selected_offers = offers_by_company.get(selected_company, [])
        selected_years = {offer.insurance_year for offer in selected_offers}
        selected_total = _sum_summary_company_variant_total(
            selected_offers,
            selected_variant,
            selected_years,
        ) if selected_years else None

        price_row = _build_deal_price_row(
            summary,
            comparison_mode='selected_variant',
            require_full_coverage=True,
        )
        if price_row and selected_total is None:
            selected_total = price_row.get('selected_total')

        selected = {
            'company_name': selected_company,
            'variant': selected_variant,
            'variant_label': f'Вариант {selected_variant}',
            'total': selected_total,
            'rank': price_row.get('selected_rank') if price_row else None,
            'comparable_count': price_row.get('comparable_companies_count') if price_row else None,
            'delta_to_min_abs': price_row.get('delta_to_min_abs') if price_row else None,
            'delta_to_min_pct': price_row.get('delta_to_min_pct') if price_row else None,
            'is_min_selected': price_row.get('is_min_selected') if price_row else False,
        }

    preview_limit = 3
    return {
        'has_offers': bool(offers_list),
        'companies_count': len(sorted_companies),
        'years_count': len(years),
        'years_label': _format_summary_years_label(years),
        'required_variants': required_variants,
        'required_variants_label': required_variants_label,
        'complete_companies_count': len(complete_companies),
        'incomplete_companies_count': len(incomplete_companies),
        'incomplete_companies': sorted(incomplete_companies),
        'price_ranges': price_ranges,
        'installment': {
            'show': show_installment,
            'companies_count': len(installment_companies),
            'max_payments_per_year': max_payments_per_year,
        },
        'selected': selected,
        'company_preview': sorted_companies[:preview_limit],
        'company_preview_more_count': max(len(sorted_companies) - preview_limit, 0),
        'companies_title': ', '.join(sorted_companies),
    }


def _build_deal_list_row(summary):
    """Готовит строку списка сделок на основе завершенного свода."""
    selected_company = (summary.selected_company or '').strip()
    if not selected_company:
        return None

    selected_variant = summary.selected_franchise_variant if summary.selected_franchise_variant in (1, 2) else 1
    selected_variant_display = (
        summary.get_selected_franchise_variant_display()
        if summary.selected_franchise_variant
        else 'Вариант 1 (по умолчанию)'
    )

    valid_offers = [offer for offer in summary.offers.all() if offer.is_valid]
    companies_count = len({offer.company_name for offer in valid_offers})

    selected_offers = sorted(
        [offer for offer in valid_offers if offer.company_name == selected_company],
        key=lambda offer: offer.insurance_year,
    )

    selected_total = Decimal('0')
    has_complete_selected_totals = bool(selected_offers)
    for offer in selected_offers:
        premium_raw = (
            offer.premium_with_franchise_2
            if selected_variant == 2
            else offer.premium_with_franchise_1
        )
        premium_value = _safe_decimal_or_none(premium_raw)
        if premium_value is None or premium_value <= 0:
            has_complete_selected_totals = False
            continue
        selected_total += premium_value

    if not has_complete_selected_totals:
        selected_total = None

    price_range = None
    price_row = _build_deal_price_row(
        summary,
        comparison_mode='selected_variant',
        require_full_coverage=True,
    )
    if price_row:
        selected_point = next(
            (point for point in price_row['points'] if point.get('is_selected')),
            None
        )
        price_range = {
            'points': price_row['points'],
            'min_total': price_row['min_total'],
            'max_total': price_row['max_total'],
            'selected_rank': price_row['selected_rank'],
            'comparable_companies_count': price_row['comparable_companies_count'],
            'delta_to_min_abs': price_row['delta_to_min_abs'],
            'delta_to_min_pct': price_row['delta_to_min_pct'],
            'best_company_name': price_row['best_company_name'],
            'is_min_selected': price_row['is_min_selected'],
            'selected_position_pct': selected_point['position_pct'] if selected_point else None,
        }

    request_obj = summary.request
    return {
        'summary': summary,
        'request': request_obj,
        'closed_at': summary.deal_closed_at,
        'selected_company': selected_company,
        'selected_variant': selected_variant,
        'selected_variant_display': selected_variant_display,
        'selected_total': selected_total,
        'total_years': len(selected_offers),
        'companies_count': companies_count,
        'has_data_warning': not has_complete_selected_totals,
        'price_range': price_range,
    }


def _sort_deal_rows(rows, sort_value):
    """Сортирует строки списка сделок по выбранному полю."""
    if sort_value == 'closed_at':
        return sorted(rows, key=lambda row: row['closed_at'] or timezone.now())

    if sort_value == '-closed_at':
        return sorted(rows, key=lambda row: row['closed_at'] or timezone.now(), reverse=True)

    if sort_value in {'total_premium', '-total_premium'}:
        rows_with_premium = [row for row in rows if row['selected_total'] is not None]
        rows_without_premium = [row for row in rows if row['selected_total'] is None]
        rows_with_premium.sort(
            key=lambda row: (row['selected_total'], row['closed_at'] or timezone.now()),
            reverse=(sort_value == '-total_premium'),
        )
        return rows_with_premium + rows_without_premium

    if sort_value == 'client_name':
        return sorted(
            rows,
            key=lambda row: (
                (row['request'].client_name or '').lower(),
                row['closed_at'] or timezone.now(),
            ),
        )

    if sort_value == '-client_name':
        return sorted(
            rows,
            key=lambda row: (
                (row['request'].client_name or '').lower(),
                row['closed_at'] or timezone.now(),
            ),
            reverse=True,
        )

    return sorted(rows, key=lambda row: row['closed_at'] or timezone.now(), reverse=True)


def _build_deal_list_kpi(rows):
    """Собирает KPI для страницы списка сделок."""
    total_deals = len(rows)
    rows_with_totals = [row for row in rows if row['selected_total'] is not None]
    total_premium_sum = sum((row['selected_total'] for row in rows_with_totals), Decimal('0'))
    avg_premium = (
        total_premium_sum / Decimal(len(rows_with_totals))
        if rows_with_totals else None
    )
    avg_years = (
        sum(row['total_years'] for row in rows) / total_deals
        if total_deals > 0 else 0
    )

    return {
        'total_deals': total_deals,
        'total_premium_sum': total_premium_sum if rows_with_totals else None,
        'avg_premium': avg_premium,
        'avg_years': avg_years,
        'rows_with_warning': sum(1 for row in rows if row['has_data_warning']),
    }


@user_required
def deal_list(request):
    """Список реально заключенных сделок."""
    base_queryset = InsuranceSummary.objects.select_related('request', 'request__created_by').annotate(
        deal_closed_at_value=Coalesce('completed_at', 'updated_at')
    ).filter(
        status='completed_accepted'
    ).exclude(
        selected_company__isnull=True
    ).exclude(
        selected_company=''
    )

    available_branches = list(
        base_queryset.exclude(
            request__branch__isnull=True
        ).exclude(
            request__branch=''
        ).values_list('request__branch', flat=True).distinct().order_by('request__branch')
    )
    available_insurance_types = list(
        base_queryset.exclude(
            request__insurance_type__isnull=True
        ).exclude(
            request__insurance_type=''
        ).values_list('request__insurance_type', flat=True).distinct().order_by('request__insurance_type')
    )
    available_companies = list(
        base_queryset.exclude(
            selected_company__isnull=True
        ).exclude(
            selected_company=''
        ).values_list('selected_company', flat=True).distinct().order_by('selected_company')
    )
    manager_rows = base_queryset.exclude(
        request__created_by__isnull=True
    ).order_by().values(
        'request__created_by_id',
        'request__created_by__username',
        'request__created_by__first_name',
        'request__created_by__last_name',
    ).distinct()
    available_managers = []
    for manager_row in manager_rows:
        first_name = (manager_row.get('request__created_by__first_name') or '').strip()
        last_name = (manager_row.get('request__created_by__last_name') or '').strip()
        username = (manager_row.get('request__created_by__username') or '').strip()
        display_name = f"{first_name} {last_name}".strip() or username
        available_managers.append((str(manager_row.get('request__created_by_id')), display_name))
    available_managers = sorted(available_managers, key=lambda item: item[1].lower())

    filter_form = DealListFilterForm(
        request.GET or None,
        branch_choices=available_branches,
        insurance_type_choices=available_insurance_types,
        manager_choices=available_managers,
        company_choices=available_companies,
    )

    deals_queryset = base_queryset
    current_sort = '-closed_at'

    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        if search:
            deals_queryset = deals_queryset.filter(
                Q(request__dfa_number__icontains=search)
                | Q(request__client_name__icontains=search)
                | Q(request__inn__icontains=search)
                | Q(selected_company__icontains=search)
            )

        branch = filter_form.cleaned_data.get('branch')
        if branch:
            deals_queryset = deals_queryset.filter(request__branch=branch)

        insurance_type = filter_form.cleaned_data.get('insurance_type')
        if insurance_type:
            deals_queryset = deals_queryset.filter(request__insurance_type=insurance_type)

        manager_id = filter_form.cleaned_data.get('manager')
        if manager_id:
            deals_queryset = deals_queryset.filter(request__created_by_id=int(manager_id))

        selected_company = filter_form.cleaned_data.get('selected_company')
        if selected_company:
            deals_queryset = deals_queryset.filter(selected_company=selected_company)

        start_date = filter_form.cleaned_data.get('applied_start_date')
        end_date = filter_form.cleaned_data.get('applied_end_date')
        if start_date:
            deals_queryset = deals_queryset.filter(deal_closed_at_value__date__gte=start_date)
        if end_date:
            deals_queryset = deals_queryset.filter(deal_closed_at_value__date__lte=end_date)

        current_sort = filter_form.cleaned_data.get('sort') or '-closed_at'

    deals_queryset = deals_queryset.prefetch_related('offers')
    rows = []
    for summary in deals_queryset:
        row = _build_deal_list_row(summary)
        if row is not None:
            rows.append(row)

    rows = _sort_deal_rows(rows, current_sort)
    kpi = _build_deal_list_kpi(rows)

    paginator = Paginator(rows, 25)
    page = request.GET.get('page')
    try:
        rows_page = paginator.page(page)
    except PageNotAnInteger:
        rows_page = paginator.page(1)
    except EmptyPage:
        rows_page = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring_without_page = query_params.urlencode()

    return render(request, 'summaries/deal_list.html', {
        'deals': rows_page,
        'paginator': paginator,
        'filter_form': filter_form,
        'kpi': kpi,
        'current_sort': current_sort,
        'querystring_without_page': querystring_without_page,
    })


@user_required
def summary_detail(request, pk):
    """Детальная информация о своде"""
    summary = get_object_or_404(
        InsuranceSummary.objects.select_related('request').prefetch_related('offers'), 
        pk=pk
    )
    
    # Получаем предложения с правильной сортировкой по новой структуре
    offers = summary.offers.filter(is_valid=True).order_by('company_name', 'insurance_year')
    
    # Группируем предложения по компаниям (company-first grouping)
    companies_with_offers = summary.get_offers_grouped_by_company()
    
    # Получаем структурированные данные для шаблона (company-year matrix)
    company_year_matrix = summary.get_company_year_matrix()
    
    # Сортируем компании по алфавиту
    sorted_companies = sorted(companies_with_offers.keys())
    
    # Получаем корректное количество уникальных компаний
    unique_companies_count = summary.get_unique_companies_count()
    
    # Получаем данные о компаниях с количеством лет для отображения тегов
    companies_with_year_counts = summary.get_companies_with_year_counts()
    
    # Получаем данные об итоговых суммах по компаниям для многолетних предложений
    company_totals = summary.get_company_totals()
    
    # Получаем комментарии, сгруппированные по компаниям
    company_notes = summary.get_company_notes()

    # Компактная аналитика для бокового блока сводной информации
    summary_analytics = _build_summary_compact_analytics(summary, offers)

    # Данные о доступных франшизных вариантах по компаниям для UI статусов
    company_variant_requirements = {}
    for company_name in sorted_companies:
        company_variant_requirements[company_name] = {
            'available_variants': summary.get_company_available_variants(company_name),
            'requires_choice': summary.requires_variant_choice(company_name),
            'default_variant': summary.get_default_variant(company_name),
        }
    
    return render(request, 'summaries/summary_detail.html', {
        'summary': summary,
        'offers': offers,
        'companies_with_offers': companies_with_offers,
        'company_year_matrix': company_year_matrix,
        'sorted_companies': sorted_companies,
        'unique_companies_count': unique_companies_count,
        'companies_with_year_counts': companies_with_year_counts,
        'company_totals': company_totals,
        'company_notes': company_notes,
        'summary_analytics': summary_analytics,
        'company_variant_requirements': company_variant_requirements,
    })


@user_required
def deal_summary(request, summary_id):
    """Страница резюме по страховой сделке (технический лист)"""
    summary = get_object_or_404(
        InsuranceSummary.objects.select_related('request').prefetch_related('offers'), 
        pk=summary_id
    )
    
    # Проверяем, что свод завершен с акцептом и выбрана компания
    if summary.status != 'completed_accepted' or not summary.selected_company:
        messages.warning(request, 'Резюме по сделке доступно только для завершенных сводов с выбранной страховой компанией')
        return redirect('summaries:summary_detail', pk=summary_id)
    
    # Получаем заявку
    insurance_request = summary.request
    
    # Получаем предложения выбранной компании
    selected_offers = list(
        summary.offers.filter(
            company_name=summary.selected_company,
            is_valid=True
        ).order_by('insurance_year')
    )

    # Собираем общий комментарий по выбранным предложениям (без дублей, с сохранением порядка)
    selected_offer_notes = []
    for offer in selected_offers:
        note = (offer.notes or '').strip()
        if note and note not in selected_offer_notes:
            selected_offer_notes.append(note)

    selected_offer_notes_summary = ' | '.join(selected_offer_notes)

    # Менеджер Онлайна — форматируем имя
    created_by = insurance_request.created_by
    if created_by:
        if created_by.last_name and created_by.first_name:
            manager_online = f"{created_by.last_name} {created_by.first_name}"
        elif created_by.last_name:
            manager_online = created_by.last_name
        elif created_by.first_name:
            manager_online = created_by.first_name
        else:
            manager_online = created_by.username
    else:
        manager_online = None

    is_casco_type = insurance_request.insurance_type in ['КАСКО', 'страхование спецтехники']
    is_property_type = insurance_request.insurance_type == 'страхование имущества'

    context = {
        'summary': summary,
        'request': insurance_request,
        'selected_company': summary.selected_company,
        'selected_franchise_variant': summary.selected_franchise_variant,
        'selected_franchise_variant_display': summary.get_selected_franchise_variant_display() if summary.selected_franchise_variant else None,
        'selected_offers': selected_offers,
        'selected_offer_notes_summary': selected_offer_notes_summary,
        'total_years': len(selected_offers),

        # Основные данные заявки
        'request_number': insurance_request.dfa_number,
        'client_name': insurance_request.client_name,
        'client_inn': insurance_request.inn,
        'branch': insurance_request.get_branch_display() if hasattr(insurance_request, 'get_branch_display') else insurance_request.branch,
        'insurance_type': insurance_request.get_insurance_type_display(),
        'vehicle_info': insurance_request.vehicle_info,
        'manufacturing_year': insurance_request.manufacturing_year,
        'asset_status': insurance_request.asset_status,
        'insurance_period': getattr(insurance_request, 'insurance_period', ''),

        # Условия договора
        'creditor_bank': insurance_request.creditor_bank,
        'usage_purposes': insurance_request.usage_purposes,
        'has_franchise': insurance_request.has_franchise,
        'franchise_type': 'Да' if insurance_request.has_franchise else 'Нет',
        'has_installment': insurance_request.has_installment,
        'has_casco_ce': insurance_request.has_casco_ce if is_casco_type else None,
        'deal_status_display': insurance_request.get_deal_status_display() if hasattr(insurance_request, 'get_deal_status_display') else '',

        # Менеджеры
        'manager_name': insurance_request.manager_name,
        'manager_online': manager_online,

        # Параметры для КАСКО/спецтехники
        'has_autostart': insurance_request.has_autostart if is_casco_type else None,
        'key_completeness': insurance_request.key_completeness if is_casco_type else None,
        'pts_psm': insurance_request.pts_psm if is_casco_type else None,
        'telematics_complex': insurance_request.telematics_complex if is_casco_type else None,

        # Параметры для страхования имущества
        'insurance_territory': insurance_request.insurance_territory if is_property_type else None,
        'has_transportation': insurance_request.has_transportation if is_property_type else None,
        'has_construction_work': insurance_request.has_construction_work if is_property_type else None,
    }

    return render(request, 'summaries/deal_summary.html', context)


@user_required
def create_summary(request, request_id):
    """
    Создание свода для заявки с улучшенной обработкой ошибок и проверками доступа.
    
    Требования: 1.2, 1.3, 1.4
    """
    insurance_request = get_object_or_404(InsuranceRequest, pk=request_id)
    
    # Дополнительные проверки прав доступа (требование 1.2)
    if not request.user.is_authenticated:
        messages.error(request, 'Необходимо войти в систему для создания свода')
        return redirect('insurance_requests:login')
    
    # Проверяем права пользователя на создание сводов
    from insurance_requests.decorators import has_user_access
    if not has_user_access(request.user):
        messages.error(request, 'У вас недостаточно прав для создания сводов')
        return redirect('insurance_requests:request_detail', pk=request_id)
    
    # Проверяем, можно ли создать свод для этой заявки (требование 1.2)
    if hasattr(insurance_request, 'summary'):
        messages.info(request, f'Свод для заявки {insurance_request.get_display_name()} уже существует')
        return redirect('summaries:summary_detail', pk=insurance_request.summary.pk)
    
    # Проверяем статус заявки - можно создавать свод только для определенных статусов
    # Важно: список синхронизирован с InsuranceRequest.STATUS_CHOICES
    allowed_statuses = ['uploaded', 'email_generated', 'emails_sent']
    if insurance_request.status not in allowed_statuses:
        messages.error(request, 
                      f'Нельзя создать свод для заявки со статусом "{insurance_request.get_status_display()}". '
                      f'Свод можно создать только для заявок со статусами: '
                      f'{", ".join([dict(insurance_request.STATUS_CHOICES).get(s, s) for s in allowed_statuses])}')
        return redirect('insurance_requests:request_detail', pk=request_id)
    
    # Проверяем обязательные поля заявки
    validation_errors = []
    if not insurance_request.client_name or not insurance_request.client_name.strip():
        validation_errors.append('Не указано имя клиента')
    
    if not insurance_request.inn or not insurance_request.inn.strip():
        validation_errors.append('Не указан ИНН клиента')
    
    if not insurance_request.insurance_type:
        validation_errors.append('Не указан тип страхования')
    
    if validation_errors:
        messages.error(request, 
                      f'Невозможно создать свод: {"; ".join(validation_errors)}. '
                      'Пожалуйста, дополните информацию в заявке.')
        return redirect('insurance_requests:request_detail', pk=request_id)
    
    # Создание свода с улучшенной обработкой ошибок (требование 1.3)
    try:
        with transaction.atomic():
            # Создаем свод
            summary = InsuranceSummary.objects.create(
                request=insurance_request,
                status='collecting'
            )
            
            # Обновляем статус заявки, если необходимо
            if insurance_request.status == 'uploaded':
                insurance_request.status = 'email_generated'
                insurance_request.save(update_fields=['status', 'updated_at'])
            
            logger.info(f"Summary created successfully for request {request_id} by user {request.user.username}")
            
            # Улучшенное сообщение об успехе (требование 1.3)
            messages.success(request, 
                           f'Свод предложений для заявки {insurance_request.get_display_name()} '
                           f'({insurance_request.client_name}) успешно создан')
            
            # Корректное перенаправление после создания свода (требование 1.4)
            return redirect('summaries:summary_detail', pk=summary.pk)
            
    except Exception as e:
        # Улучшенная обработка ошибок (требование 1.3)
        error_message = str(e)
        logger.error(f"Error creating summary for request {request_id} by user {request.user.username}: {error_message}", 
                    exc_info=True)
        
        # Более информативные сообщения об ошибках для пользователя
        if 'UNIQUE constraint failed' in error_message or 'duplicate key' in error_message.lower():
            messages.error(request, 
                          f'Свод для заявки {insurance_request.get_display_name()} уже существует. '
                          'Обновите страницу и попробуйте снова.')
        elif 'permission' in error_message.lower() or 'access' in error_message.lower():
            messages.error(request, 
                          'Недостаточно прав для создания свода. '
                          'Обратитесь к администратору системы.')
        elif 'database' in error_message.lower() or 'connection' in error_message.lower():
            messages.error(request, 
                          'Временная ошибка базы данных. '
                          'Пожалуйста, попробуйте создать свод через несколько минут.')
        else:
            messages.error(request, 
                          f'Произошла ошибка при создании свода: {error_message}. '
                          'Если ошибка повторяется, обратитесь к администратору.')
        
        # Корректное перенаправление при ошибке (требование 1.4)
        return redirect('insurance_requests:request_detail', pk=request_id)


@user_required
def add_offer(request, summary_id):
    """Добавление предложения к своду"""
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    
    if request.method == 'POST':
        form = AddOfferToSummaryForm(request.POST)
        
        # Логируем данные формы для отладки
        logger.info(f"Form data received for summary {summary_id}: {request.POST}")
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    offer = form.save(commit=False)
                    offer.summary = summary
                    offer.save()
                    
                    # Обновляем счетчик предложений в своде
                    summary.update_total_offers_count()
                    
                    logger.info(f"Offer saved successfully: {offer.company_name} ({offer.get_insurance_year_display()}) for summary {summary_id}")
                    
                    messages.success(request, f'Предложение от {offer.company_name} ({offer.get_insurance_year_display()}) успешно добавлено')
                    return redirect('summaries:summary_detail', pk=summary_id)
                    
            except IntegrityError as e:
                # Специальная обработка ошибок дублирования предложений
                error_str = str(e)
                if ('UNIQUE constraint failed' in error_str or 
                    'duplicate key value violates unique constraint' in error_str):
                    # Извлекаем информацию о дублирующемся предложении из данных формы
                    company_name = form.cleaned_data.get('company_name', 'неизвестная компания')
                    insurance_year = form.cleaned_data.get('insurance_year', 'неизвестный год')
                    
                    # Создаем кастомное исключение для лучшей обработки
                    duplicate_error = DuplicateOfferError(company_name, insurance_year)
                    messages.error(request, duplicate_error.get_user_message())
                    logger.warning(f"Duplicate offer attempt for summary {summary_id}: {company_name} year {insurance_year}")
                else:
                    # Обработка других ошибок целостности данных
                    logger.error(f"IntegrityError adding offer to summary {summary_id}: {str(e)}", exc_info=True)
                    messages.error(request, f'Ошибка целостности данных при сохранении предложения: {str(e)}')
            except Exception as e:
                # Обработка всех остальных ошибок
                logger.error(f"Error adding offer to summary {summary_id}: {str(e)}", exc_info=True)
                messages.error(request, f'Ошибка при сохранении предложения: {str(e)}')
        else:
            # Логируем ошибки валидации
            logger.warning(f"Form validation failed for summary {summary_id}. Errors: {form.errors}")
            
            # Добавляем общее сообщение об ошибке валидации
            error_messages = []
            for field, errors in form.errors.items():
                if field == '__all__':
                    error_messages.extend(errors)
                else:
                    field_label = form.fields[field].label if field in form.fields else field
                    for error in errors:
                        error_messages.append(f"{field_label}: {error}")
            
            if error_messages:
                messages.error(request, f'Ошибки в форме: {"; ".join(error_messages)}')
            else:
                messages.error(request, 'Проверьте правильность заполнения всех полей')
    else:
        form = AddOfferToSummaryForm()
    
    return render(request, 'summaries/add_offer.html', {
        'form': form,
        'summary': summary
    })


@user_required
def generate_summary_file(request, summary_id):
    """Генерация Excel файла свода (полная версия с техническим листом)"""
    from datetime import datetime
    from urllib.parse import quote
    import re
    from .services import get_excel_export_service, ExcelExportServiceError, InvalidSummaryDataError, TemplateNotFoundError
    
    summary = get_object_or_404(InsuranceSummary.objects.select_related('request'), pk=summary_id)
    
    # Проверка статуса свода - требование 1.1
    if summary.status != 'ready':
        logger.warning(f"Attempt to generate Excel for summary {summary_id} with status '{summary.status}'")
        return JsonResponse({
            'error': 'Файл можно генерировать только для сводов в статусе "Готов к отправке"'
        }, status=400)
    
    try:
        # Создание сервиса для генерации Excel - требование 1.2
        service = get_excel_export_service()
        
        # Генерация Excel файла (полная версия) - требование 1.2, 1.3
        excel_file = service.generate_summary_excel(summary, is_client_version=False)
        
        # Формирование имени файла - требование 3.2, 15.1-15.8
        # Обработка номера ДФА: извлечение только цифр
        dfa_number_digits_only = re.sub(r'[^\d]', '', summary.request.dfa_number)
        # Изменение формата даты на день_месяц_год
        date_formatted = datetime.now().strftime('%d_%m_%Y')
        filename = f"full_svod_{dfa_number_digits_only}_{date_formatted}.xlsx"
        
        # Создание HTTP response с Excel файлом - требование 3.1
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Excel file successfully generated for summary {summary_id}: {filename}")
        return response
        
    except InvalidSummaryDataError as e:
        # Обработка ошибок валидации данных - требование 4.1, 4.2
        logger.error(f"Invalid summary data for {summary_id}: {str(e)}")
        return JsonResponse({
            'error': f'Ошибка в данных свода: {str(e)}'
        }, status=400)
        
    except TemplateNotFoundError as e:
        # Обработка ошибок недоступности шаблона - требование 4.3
        logger.error(f"Template not found for summary {summary_id}: {str(e)}")
        return JsonResponse({
            'error': 'Шаблон Excel-файла недоступен. Обратитесь к администратору.'
        }, status=500)
        
    except ExcelExportServiceError as e:
        # Обработка других ошибок сервиса - требование 3.3
        logger.error(f"Excel export service error for summary {summary_id}: {str(e)}")
        return JsonResponse({
            'error': f'Ошибка при генерации файла: {str(e)}'
        }, status=500)
        
    except Exception as e:
        # Обработка неожиданных ошибок - требование 3.3
        logger.error(f"Unexpected error generating summary file for {summary_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': 'Произошла неожиданная ошибка при генерации файла. Обратитесь к администратору.'
        }, status=500)


@user_required
def generate_client_summary_file(request, summary_id):
    """Генерация клиентского Excel файла свода (сокращенная версия без технического листа)"""
    from datetime import datetime
    from urllib.parse import quote
    import re
    from .services import get_excel_export_service, ExcelExportServiceError, InvalidSummaryDataError, TemplateNotFoundError
    
    summary = get_object_or_404(InsuranceSummary.objects.select_related('request'), pk=summary_id)
    
    # Проверка статуса свода
    if summary.status != 'ready':
        logger.warning(f"Attempt to generate client Excel for summary {summary_id} with status '{summary.status}'")
        return JsonResponse({
            'error': 'Файл можно генерировать только для сводов в статусе "Готов к отправке"'
        }, status=400)
    
    try:
        # Создание сервиса для генерации Excel
        service = get_excel_export_service()
        
        # Генерация клиентского Excel файла (без технического листа)
        excel_file = service.generate_summary_excel(summary, is_client_version=True)
        
        # Формирование имени файла с префиксом "client_"
        dfa_number_digits_only = re.sub(r'[^\d]', '', summary.request.dfa_number)
        date_formatted = datetime.now().strftime('%d_%m_%Y')
        filename = f"client_svod_{dfa_number_digits_only}_{date_formatted}.xlsx"
        
        # Создание HTTP response с Excel файлом
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Client Excel file successfully generated for summary {summary_id}: {filename}")
        return response
        
    except InvalidSummaryDataError as e:
        logger.error(f"Invalid summary data for client file {summary_id}: {str(e)}")
        return JsonResponse({
            'error': f'Ошибка в данных свода: {str(e)}'
        }, status=400)
        
    except TemplateNotFoundError as e:
        logger.error(f"Client template not found for summary {summary_id}: {str(e)}")
        return JsonResponse({
            'error': 'Клиентский шаблон Excel-файла недоступен. Обратитесь к администратору.'
        }, status=500)
        
    except ExcelExportServiceError as e:
        logger.error(f"Excel export service error for client file {summary_id}: {str(e)}")
        return JsonResponse({
            'error': f'Ошибка при генерации клиентского файла: {str(e)}'
        }, status=500)
        
    except Exception as e:
        logger.error(f"Unexpected error generating client summary file for {summary_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': 'Произошла неожиданная ошибка при генерации клиентского файла. Обратитесь к администратору.'
        }, status=500)


@require_http_methods(["POST"])
@user_required
def send_summary_to_client(request, summary_id):
    """Отправка свода клиенту без автоматического изменения статуса"""
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    
    try:
        # Логируем действие без изменения статуса
        logger.info(f"Summary {summary_id} manual send action by user {request.user.username}")
        
        # Возвращаем успешный результат без изменения статуса
        return JsonResponse({
            'success': True, 
            'message': 'Свод отправлен в Альянс. Для изменения статуса используйте блок "Управление статусом".'
        })
        
    except Exception as e:
        logger.error(f"Error sending summary {summary_id} to client: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Ошибка при отправке свода: {str(e)}'})


@user_required
def copy_offer(request, offer_id):
    """Копирование предложения"""
    original_offer = get_object_or_404(InsuranceOffer, pk=offer_id)
    
    if request.method == 'POST':
        form = OfferForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Создаем новое предложение на основе данных формы
                    new_offer = form.save(commit=False)
                    new_offer.summary = original_offer.summary
                    new_offer.save()
                    
                    # Обновляем счетчик предложений в своде
                    original_offer.summary.update_total_offers_count()
                    
                    logger.info(f"Offer copied: {original_offer.id} -> {new_offer.id} by user {request.user.username}")
                    
                    messages.success(request, f'Предложение от {new_offer.company_name} ({new_offer.get_insurance_year_display()}) успешно скопировано')
                    return redirect('summaries:summary_detail', pk=new_offer.summary.pk)
                    
            except IntegrityError as e:
                # Специальная обработка ошибок дублирования предложений
                error_str = str(e)
                if ('UNIQUE constraint failed' in error_str or 
                    'duplicate key value violates unique constraint' in error_str):
                    # Извлекаем информацию о дублирующемся предложении из данных формы
                    company_name = form.cleaned_data.get('company_name', 'неизвестная компания')
                    insurance_year = form.cleaned_data.get('insurance_year', 'неизвестный год')
                    
                    # Создаем кастомное исключение для лучшей обработки
                    duplicate_error = DuplicateOfferError(company_name, insurance_year)
                    messages.error(request, duplicate_error.get_user_message())
                    logger.warning(f"Duplicate offer attempt during copy for summary {original_offer.summary.id}: {company_name} year {insurance_year}")
                else:
                    # Обработка других ошибок целостности данных
                    logger.error(f"IntegrityError copying offer {offer_id}: {str(e)}", exc_info=True)
                    messages.error(request, f'Ошибка целостности данных при копировании предложения: {str(e)}')
            except Exception as e:
                # Обработка всех остальных ошибок
                logger.error(f"Error copying offer {offer_id}: {str(e)}", exc_info=True)
                messages.error(request, f'Ошибка при копировании предложения: {str(e)}')
        else:
            # Логируем ошибки валидации
            logger.warning(f"Form validation failed for copying offer {offer_id}. Errors: {form.errors}")
            
            # Добавляем общее сообщение об ошибке валидации
            error_messages = []
            for field, errors in form.errors.items():
                if field == '__all__':
                    error_messages.extend(errors)
                else:
                    field_label = form.fields[field].label if field in form.fields else field
                    for error in errors:
                        error_messages.append(f"{field_label}: {error}")
            
            if error_messages:
                messages.error(request, f'Ошибки в форме: {"; ".join(error_messages)}')
            else:
                messages.error(request, 'Проверьте правильность заполнения всех полей')
    else:
        # Создаем форму с данными из оригинального предложения
        initial_data = {
            'company_name': original_offer.company_name,
            'insurance_year': original_offer.insurance_year,
            'insurance_sum': original_offer.insurance_sum,
            'franchise_1': original_offer.franchise_1,
            'premium_with_franchise_1': original_offer.premium_with_franchise_1,
            'franchise_2': original_offer.franchise_2,
            'premium_with_franchise_2': original_offer.premium_with_franchise_2,
            'installment_variant_1': original_offer.installment_variant_1,
            'payments_per_year_variant_1': original_offer.payments_per_year_variant_1,
            'installment_variant_2': original_offer.installment_variant_2,
            'payments_per_year_variant_2': original_offer.payments_per_year_variant_2,
            'notes': original_offer.notes,
        }
        form = OfferForm(initial=initial_data)
    
    return render(request, 'summaries/copy_offer.html', {
        'form': form,
        'original_offer': original_offer
    })


@user_required
def edit_offer(request, offer_id):
    """Редактирование предложения"""
    offer = get_object_or_404(InsuranceOffer, pk=offer_id)
    
    if request.method == 'POST':
        form = OfferForm(request.POST, request.FILES, instance=offer)
        if form.is_valid():
            try:
                with transaction.atomic():
                    updated_offer = form.save()
                    

                    
                    messages.success(request, f'Предложение от {updated_offer.company_name} ({updated_offer.get_insurance_year_display()}) обновлено')
                    return redirect('summaries:summary_detail', pk=updated_offer.summary.pk)
                    
            except Exception as e:
                logger.error(f"Error updating offer {offer_id}: {str(e)}")
                messages.error(request, f'Ошибка при обновлении предложения: {str(e)}')
        else:
            # Добавляем обработку ошибок валидации для лучшего UX
            error_messages = []
            for field, errors in form.errors.items():
                if field == '__all__':
                    error_messages.extend(errors)
                else:
                    field_label = form.fields[field].label if field in form.fields else field
                    for error in errors:
                        error_messages.append(f"{field_label}: {error}")
            
            if error_messages:
                messages.error(request, f'Ошибки в форме: {"; ".join(error_messages)}')
            else:
                messages.error(request, 'Проверьте правильность заполнения всех полей')
    else:
        form = OfferForm(instance=offer)
    
    return render(request, 'summaries/edit_offer.html', {
        'form': form,
        'offer': offer
    })


@require_http_methods(["POST"])
@user_required
def delete_offer(request, offer_id):
    """Удаление предложения"""
    offer = get_object_or_404(InsuranceOffer, pk=offer_id)
    summary = offer.summary
    summary_id = summary.pk
    company_name = offer.company_name
    insurance_year_display = offer.get_insurance_year_display()
    
    try:
        with transaction.atomic():
            # Удаляем предложение
            offer.delete()
            
            # Обновляем счетчик предложений в своде
            summary.update_total_offers_count()
            
            logger.info(f"Offer deleted: {company_name} ({insurance_year_display}) from summary {summary_id}")
            
            return JsonResponse({
                'success': True, 
                'message': f'Предложение от {company_name} ({insurance_year_display}) удалено',
                'new_companies_count': summary.get_unique_companies_count()
            })
        
    except Exception as e:
        logger.error(f"Error deleting offer {offer_id}: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': f'Не удалось удалить предложение: {str(e)}'
        })


@require_http_methods(["POST"])
@user_required
def change_summary_status(request, summary_id):
    """Изменение статуса свода"""
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    new_status = request.POST.get('status')
    selected_company = request.POST.get('selected_company', '').strip()
    selected_franchise_variant_raw = request.POST.get('selected_franchise_variant', '').strip()
    selected_franchise_variant = None
    
    # Валидация статуса "Завершен: акцепт/распоряжение"
    if new_status == 'completed_accepted':
        if not selected_company:
            return JsonResponse({
                'success': False, 
                'error': 'Необходимо выбрать страховую компанию для статуса "Завершен: акцепт/распоряжение"'
            })
        
        # Проверяем, что выбранная компания есть в предложениях свода
        available_companies = summary.get_unique_companies_list()
        if selected_company not in available_companies:
            return JsonResponse({
                'success': False, 
                'error': 'Выбранная компания не найдена в предложениях свода'
            })

        available_variants = summary.get_company_available_variants(selected_company)
        if not available_variants:
            return JsonResponse({
                'success': False,
                'error': 'Для выбранной компании не найдено валидных вариантов премии'
            })

        if selected_franchise_variant_raw:
            try:
                selected_franchise_variant = int(selected_franchise_variant_raw)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Передан некорректный вариант франшизы'
                })

            if selected_franchise_variant not in available_variants:
                return JsonResponse({
                    'success': False,
                    'error': 'Выбранный вариант франшизы недоступен для выбранной компании'
                })

        # Если вариантов два, выбор обязателен. Если один - выбираем автоматически.
        if len(available_variants) > 1:
            if not selected_franchise_variant_raw:
                return JsonResponse({
                    'success': False,
                    'error': 'Необходимо выбрать вариант франшизы (1 или 2)'
                })
        elif not selected_franchise_variant:
            selected_franchise_variant = available_variants[0]
    
    if new_status in dict(InsuranceSummary.STATUS_CHOICES):
        old_status = summary.status
        summary.status = new_status
        
        # Сохраняем выбранную компанию для статуса "Завершен: акцепт/распоряжение"
        if new_status == 'completed_accepted':
            summary.selected_company = selected_company
            summary.selected_franchise_variant = selected_franchise_variant
            if old_status != 'completed_accepted' or not summary.completed_at:
                summary.completed_at = timezone.now()
        else:
            summary.selected_company = None
            summary.selected_franchise_variant = None
            summary.completed_at = None
        
        # Если статус изменен на "Отправлен в Альянс", устанавливаем время отправки
        if new_status == 'sent':
            summary.sent_to_client_at = timezone.now()
        
        summary.save()
        
        logger.info(f"Summary {summary_id} status changed from '{old_status}' to '{new_status}'" + 
                   (
                       f" with selected company '{selected_company}' and franchise variant '{selected_franchise_variant}'"
                       if new_status == 'completed_accepted' else ""
                   ))
        
        message = f'Статус изменен на "{summary.get_status_display()}"'
        if new_status == 'completed_accepted' and selected_company:
            message += f' (СК: {selected_company}, вариант: {summary.get_selected_franchise_variant_display()})'
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'new_status': new_status,
            'new_status_display': summary.get_status_display(),
            'selected_company': selected_company if new_status == 'completed_accepted' else None,
            'selected_franchise_variant': selected_franchise_variant if new_status == 'completed_accepted' else None
        })
    
    return JsonResponse({
        'success': False, 
        'error': 'Недопустимый статус'
    })


@require_http_methods(["POST"])
@user_required
def update_summary_notes(request, summary_id):
    """Обновление примечания к своду"""
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    notes = request.POST.get('notes', '').strip()
    
    # Обновляем примечание
    summary.notes = notes
    summary.save(update_fields=['notes'])
    
    logger.info(f"Summary {summary_id} notes updated by user {request.user.username}")
    
    return JsonResponse({
        'success': True,
        'message': 'Примечание успешно обновлено' if notes else 'Примечание удалено',
        'notes': notes
    })


# Удалена дублирующаяся функция upload_multiple_company_responses - используется версия ниже

@require_http_methods(["POST"])
@user_required
def upload_company_response(request, summary_id):
    """
    Загрузка ответа страховой компании через Excel файл
    
    Требования: 1.3, 2.1, 4.3, 5.1, 5.2, 5.3, 5.4, 5.5
    """
    from .services import get_excel_response_processor
    from .services import ExcelProcessingError, InvalidFileFormatError, MissingDataError, InvalidDataError, RowProcessingError
    from .forms import CompanyResponseUploadForm
    
    # Получаем свод с проверкой доступа
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    
    # Проверка аутентификации и прав пользователя (требование 1.3, 2.1)
    if not request.user.is_authenticated:
        logger.warning(f"Unauthenticated user attempted to upload company response for summary {summary_id}")
        return JsonResponse({
            'success': False,
            'error': 'Необходимо войти в систему для загрузки ответов компаний'
        }, status=401)
    
    # Проверяем права пользователя на изменение сводов
    from insurance_requests.decorators import has_user_access
    if not has_user_access(request.user):
        logger.warning(f"User {request.user.username} without sufficient permissions attempted to upload company response for summary {summary_id}")
        return JsonResponse({
            'success': False,
            'error': 'У вас недостаточно прав для загрузки ответов компаний'
        }, status=403)
    
    # Валидация статуса свода (только "Сбор предложений") (требование 2.1)
    if summary.status != 'collecting':
        logger.warning(f"Attempt to upload company response for summary {summary_id} with status '{summary.status}' by user {request.user.username}")
        return JsonResponse({
            'success': False,
            'error': 'Загрузка ответов компаний доступна только при статусе "Сбор предложений"'
        }, status=400)
    
    # Обработка POST запросов с файлами (требование 2.1)
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Метод не поддерживается. Используйте POST для загрузки файлов.'
        }, status=405)
    
    # Валидация формы загрузки файла
    form = CompanyResponseUploadForm(request.POST, request.FILES)
    
    if not form.is_valid():
        # Обработка ошибок валидации файлов (требование 5.1, 5.3, 5.5)
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        error_message = '; '.join(error_messages) if error_messages else 'Ошибка валидации файла'
        
        logger.warning(f"File validation failed for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'field_errors': form.errors
        }, status=400)
    
    # Получаем загруженный файл
    excel_file = form.cleaned_data['excel_file']
    
    try:
        # Интеграция с сервисом ExcelResponseProcessor (требование 4.3, 5.2, 5.4)
        processor = get_excel_response_processor()
        
        # Обработка успешной загрузки с созданием предложений (требование 4.3, 5.2, 5.4)
        result = processor.process_excel_file(excel_file, summary)
        
        # Логирование успешной операции
        logger.info(f"Company response uploaded successfully for summary {summary_id} by user {request.user.username}: "
                   f"Company '{result['company_name']}', {result['offers_created']} offers created for years {result['years']}")
        
        # Формируем улучшенное сообщение об успешной обработке (требование 3.1, 3.2)
        years_processed = result.get('years', [])
        skipped_rows = result.get('skipped_rows', [])
        processed_rows = result.get('processed_rows', [])
        
        # Основное сообщение с количеством обработанных лет
        if len(years_processed) == 1:
            base_message = f'Предложение от компании "{result["company_name"]}" успешно загружено для {years_processed[0]} года страхования'
        else:
            years_str = ', '.join(map(str, sorted(years_processed)))
            base_message = f'Предложение от компании "{result["company_name"]}" успешно загружено для {len(years_processed)} лет страхования ({years_str})'
        
        # Проверяем информацию о сопоставлении компании
        matching_info = result.get('company_matching_info', {})
        additional_messages = []
        
        # Информация о пропущенных строках (требование 3.2)
        if skipped_rows:
            if len(processed_rows) > 0:
                additional_messages.append(
                    f'Обработано строк: {len(processed_rows)} ({", ".join(map(str, processed_rows))}). '
                    f'Пропущено строк: {len(skipped_rows)} ({", ".join(map(str, skipped_rows))})'
                )
            else:
                additional_messages.append(
                    f'Пропущено строк с пустыми данными: {len(skipped_rows)} ({", ".join(map(str, skipped_rows))})'
                )
        elif processed_rows:
            additional_messages.append(
                f'Все данные успешно обработаны из {len(processed_rows)} строк ({", ".join(map(str, processed_rows))})'
            )
        
        # Информация о сопоставлении названия компании
        if matching_info.get('assigned_other'):
            additional_messages.append(
                f'Внимание: Название компании "{matching_info["original_name"]}" не найдено в списке, '
                f'поэтому было присвоено значение "Другое". Проверьте правильность названия в файле.'
            )
        elif matching_info.get('was_matched') and not matching_info.get('assigned_other'):
            additional_messages.append(
                f'Название компании автоматически сопоставлено: "{matching_info["original_name"]}" → "{matching_info["standardized_name"]}"'
            )
        
        # Возврат JSON ответов с результатами обработки (требование 5.2, 5.4)
        return JsonResponse({
            'success': True,
            'message': base_message,
            'additional_messages': additional_messages,
            'details': {
                'company_name': result['company_name'],
                'offers_created': result['offers_created'],
                'years': result['years'],
                'processed_rows': processed_rows,
                'skipped_rows': skipped_rows,
                'matching_info': matching_info
            }
        })
        
    except InvalidFileFormatError as e:
        # Улучшенная обработка ошибок формата файла (требование 3.4, 4.3)
        error_details = str(e)
        error_message = f"Ошибка формата файла: {error_details}"
        
        # Добавляем инструкции по исправлению
        correction_instructions = []
        if 'расширение' in error_details.lower() or '.xlsx' in error_details:
            correction_instructions.append("Убедитесь, что файл имеет расширение .xlsx")
            correction_instructions.append("Сохраните файл в формате Excel (.xlsx) если он в другом формате")
        elif 'лист' in error_details.lower():
            correction_instructions.append("Убедитесь, что файл содержит хотя бы один рабочий лист")
        elif 'поврежден' in error_details.lower() or 'corrupt' in error_details.lower():
            correction_instructions.append("Файл может быть поврежден - попробуйте пересохранить его")
        else:
            correction_instructions.append("Убедитесь, что загружаете корректный Excel файл (.xlsx)")
            correction_instructions.append("Проверьте, что файл не поврежден и открывается в Excel")
        
        if correction_instructions:
            error_message += f". Рекомендации: {'; '.join(correction_instructions)}"
        
        logger.warning(f"Invalid file format for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'file_format',
            'correction_instructions': correction_instructions
        }, status=400)
        
    except MissingDataError as e:
        # Улучшенная обработка ошибок отсутствующих данных (требование 3.4, 4.3)
        missing_cells = getattr(e, 'missing_cells', [])
        
        if missing_cells:
            cells_str = ', '.join(missing_cells)
            error_message = f"Отсутствуют обязательные данные в ячейках: {cells_str}"
            
            # Добавляем инструкции по исправлению для конкретных ячеек
            correction_instructions = []
            for cell in missing_cells:
                if cell == 'B2':
                    correction_instructions.append("Ячейка B2: укажите название страховой компании")
                elif cell.startswith('A') and cell[1:].isdigit():
                    row = cell[1:]
                    correction_instructions.append(f"Ячейка A{row}: укажите номер года страхования (1, 2, 3 и т.д.)")
                elif cell.startswith('B') and cell[1:].isdigit():
                    row = cell[1:]
                    correction_instructions.append(f"Ячейка B{row}: укажите страховую сумму")
                elif cell.startswith('D') and cell[1:].isdigit():
                    row = cell[1:]
                    correction_instructions.append(f"Ячейка D{row}: укажите размер премии")
                elif cell.startswith('E') and cell[1:].isdigit():
                    row = cell[1:]
                    correction_instructions.append(f"Ячейка E{row}: укажите размер франшизы (может быть 0)")
                elif cell.startswith('F') and cell[1:].isdigit():
                    row = cell[1:]
                    correction_instructions.append(f"Ячейка F{row}: укажите тип рассрочки (1, 2, 3, 4 или 12)")
            
            if correction_instructions:
                error_message += f". Необходимо заполнить: {'; '.join(correction_instructions)}"
        else:
            error_message = f"Отсутствуют обязательные данные: {str(e)}"
            correction_instructions = ["Проверьте, что файл содержит данные в правильном формате"]
        
        logger.warning(f"Missing data in file for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'missing_data',
            'missing_cells': missing_cells,
            'correction_instructions': correction_instructions
        }, status=400)
        
    except InvalidDataError as e:
        # Улучшенная обработка ошибок некорректных данных (требование 3.4, 4.3)
        field_name = getattr(e, 'field_name', 'неизвестное поле')
        field_value = getattr(e, 'value', 'неизвестное значение')
        expected_format = getattr(e, 'expected_format', 'корректный формат')
        
        error_message = f"Некорректные данные в поле '{field_name}': значение '{field_value}' не соответствует ожидаемому формату ({expected_format})"
        
        # Добавляем инструкции по исправлению
        correction_instructions = []
        if 'год' in field_name.lower():
            correction_instructions.append("Убедитесь, что год указан числом от 1 до 10")
        elif 'сумма' in field_name.lower() or 'премия' in field_name.lower() or 'франшиза' in field_name.lower():
            correction_instructions.append("Убедитесь, что сумма указана числом без пробелов и специальных символов")
        elif 'рассрочка' in field_name.lower():
            correction_instructions.append("Рассрочка должна быть одним из значений: 1, 2, 3, 4, 12")
        
        if correction_instructions:
            error_message += f". {' '.join(correction_instructions)}"
        
        logger.warning(f"Invalid data in file for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'invalid_data',
            'field_name': field_name,
            'field_value': field_value,
            'expected_format': expected_format,
            'correction_instructions': correction_instructions
        }, status=400)
        
    except RowProcessingError as e:
        # Обработка ошибок конкретных строк (требование 3.4, 4.3)
        row_number = getattr(e, 'row_number', 'неизвестная')
        field_name = getattr(e, 'field_name', 'неизвестное поле')
        cell_address = getattr(e, 'cell_address', None)
        error_details = getattr(e, 'error_message', str(e))
        
        if cell_address:
            error_message = f"Ошибка в строке {row_number}, ячейка {cell_address} (поле '{field_name}'): {error_details}"
        else:
            error_message = f"Ошибка в строке {row_number} (поле '{field_name}'): {error_details}"
        
        # Добавляем инструкции по исправлению в зависимости от типа ошибки
        correction_instructions = []
        if 'год' in field_name.lower():
            correction_instructions.append(f"Проверьте ячейку {cell_address or f'A{row_number}'} - год должен быть числом от 1 до 10")
        elif 'сумма' in field_name.lower():
            correction_instructions.append(f"Проверьте ячейку {cell_address or f'B{row_number}'} - страховая сумма должна быть положительным числом")
        elif 'премия' in field_name.lower():
            correction_instructions.append(f"Проверьте ячейку {cell_address or f'D{row_number}'} - премия должна быть положительным числом")
        elif 'франшиза' in field_name.lower():
            correction_instructions.append(f"Проверьте ячейку {cell_address or f'E{row_number}'} - франшиза должна быть числом (может быть 0)")
        elif 'рассрочка' in field_name.lower():
            correction_instructions.append(f"Проверьте ячейку {cell_address or f'F{row_number}'} - рассрочка должна быть одним из значений: 1, 2, 3, 4, 12")
        
        if correction_instructions:
            error_message += f". {' '.join(correction_instructions)}"
        
        logger.warning(f"Row processing error for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'row_processing_error',
            'row_number': row_number,
            'field_name': field_name,
            'cell_address': cell_address,
            'correction_instructions': correction_instructions
        }, status=400)
        
    except DuplicateOfferError as e:
        # Обработка дублирования предложений (IntegrityError) (требование 5.1, 5.3, 5.5)
        error_message = f"Дублирование предложения: {str(e)}"
        logger.warning(f"Duplicate offer attempt for summary {summary_id} by user {request.user.username}: {error_message}")
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'duplicate_offer',
            'company_name': getattr(e, 'company_name', None),
            'year': getattr(e, 'year', None)
        }, status=409)
        
    except ExcelProcessingError as e:
        # Обработка других ошибок обработки Excel (требование 5.1, 5.3, 5.5)
        error_message = f"Ошибка обработки Excel файла: {str(e)}"
        logger.error(f"Excel processing error for summary {summary_id} by user {request.user.username}: {error_message}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': error_message,
            'error_type': 'processing_error'
        }, status=500)
        
    except Exception as e:
        # Логирование всех операций и ошибок (требование 5.1, 5.3, 5.5)
        error_message = f"Неожиданная ошибка при загрузке ответа компании: {str(e)}"
        logger.error(f"Unexpected error uploading company response for summary {summary_id} by user {request.user.username}: {error_message}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Произошла неожиданная ошибка при обработке файла. Обратитесь к администратору.',
            'error_type': 'unexpected_error'
        }, status=500)


@require_http_methods(["POST"])
@user_required
def upload_multiple_company_responses(request, summary_id):
    """
    Загрузка множественных ответов страховых компаний через Excel файлы
    
    Требования: 1.1, 1.2, 2.1, 2.2, 3.1, 3.2, 4.1, 4.2
    """
    import time
    from .services.multiple_file_processor import MultipleFileProcessor
    from .forms import MultipleCompanyResponseUploadForm
    
    # Получаем свод с проверкой доступа
    summary = get_object_or_404(InsuranceSummary, pk=summary_id)
    
    # Проверка аутентификации и прав пользователя
    if not request.user.is_authenticated:
        logger.warning(
            f"UPLOAD_MULTIPLE_AUTH_ERROR - Неавторизованная попытка загрузки | "
            f"summary_id={summary_id} | ip={request.META.get('REMOTE_ADDR', 'unknown')}"
        )
        return JsonResponse({
            'success': False,
            'error': 'Необходимо войти в систему для загрузки ответов компаний'
        }, status=401)
    
    # Проверяем права пользователя на изменение сводов
    from insurance_requests.decorators import has_user_access
    if not has_user_access(request.user):
        logger.warning(
            f"UPLOAD_MULTIPLE_PERMISSION_ERROR - Недостаточно прав | "
            f"user={request.user.username} | user_id={request.user.id} | summary_id={summary_id}"
        )
        return JsonResponse({
            'success': False,
            'error': 'У вас недостаточно прав для загрузки ответов компаний'
        }, status=403)
    
    # Валидация статуса свода (только "Сбор предложений")
    if summary.status != 'collecting':
        logger.warning(
            f"UPLOAD_MULTIPLE_STATUS_ERROR - Неверный статус свода | "
            f"user={request.user.username} | summary_id={summary_id} | "
            f"current_status={summary.status} | required_status=collecting"
        )
        return JsonResponse({
            'success': False,
            'error': 'Загрузка ответов компаний доступна только при статусе "Сбор предложений"'
        }, status=400)
    
    # Обработка POST запросов с файлами
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Метод не поддерживается. Используйте POST для загрузки файлов.'
        }, status=405)
    
    # Валидация входящих файлов через единую форму
    form = MultipleCompanyResponseUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                for error in errors:
                    error_messages.append(f"{field}: {error}")

        error_message = '; '.join(error_messages) if error_messages else 'Ошибка валидации загруженных файлов'
        logger.warning(
            f"UPLOAD_MULTIPLE_FORM_VALIDATION_ERROR - Ошибка валидации формы | "
            f"user={request.user.username} | summary_id={summary_id} | "
            f"error_message={error_message} | form_errors={form.errors}"
        )
        return JsonResponse({
            'success': False,
            'error': error_message,
            'field_errors': form.errors,
        }, status=400)

    excel_files = form.cleaned_data['excel_files']
    
    try:
        # Засекаем время начала обработки
        start_time = time.time()
        total_size_mb = sum(file.size for file in excel_files) / (1024 * 1024)
        
        # Логирование начала операции
        logger.info(
            f"UPLOAD_MULTIPLE_START - Начало загрузки множественных файлов | "
            f"user={request.user.username} | user_id={request.user.id} | "
            f"summary_id={summary_id} | files_count={len(excel_files)} | "
            f"total_size_mb={total_size_mb:.2f} | "
            f"files=[{', '.join(f.name for f in excel_files)}]"
        )
        
        # Создаем процессор для множественных файлов
        processor = MultipleFileProcessor(summary)
        
        # Обрабатываем файлы
        results = processor.process_files(excel_files)
        
        # Вычисляем время обработки
        processing_time = time.time() - start_time
        
        # Подсчитываем статистику
        successful_files = sum(1 for result in results if result['success'])
        failed_files = len(results) - successful_files
        total_offers_created = sum(result.get('offers_created', 0) for result in results if result['success'])
        
        # Детальное логирование результатов
        logger.info(
            f"UPLOAD_MULTIPLE_COMPLETE - Загрузка множественных файлов завершена | "
            f"user={request.user.username} | summary_id={summary_id} | "
            f"total_files={len(results)} | successful={successful_files} | failed={failed_files} | "
            f"success_rate={successful_files/len(results)*100:.1f}% | "
            f"total_offers_created={total_offers_created} | processing_time={processing_time:.2f}s"
        )
        
        # Логирование деталей каждого файла
        for result in results:
            if result['success']:
                logger.debug(
                    f"UPLOAD_MULTIPLE_FILE_SUCCESS - Файл успешно обработан | "
                    f"filename={result['file_name']} | company={result.get('company_name', 'N/A')} | "
                    f"offers_created={result.get('offers_created', 0)}"
                )
            else:
                logger.debug(
                    f"UPLOAD_MULTIPLE_FILE_ERROR - Ошибка обработки файла | "
                    f"filename={result['file_name']} | error_type={result.get('error_type', 'unknown')} | "
                    f"error_message={result.get('error_message', 'N/A')}"
                )
        
        # Формируем ответ
        response_data = {
            'success': True,
            'total_files': len(results),
            'successful_files': successful_files,
            'failed_files': failed_files,
            'processing_time': processing_time,
            'results': results
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        processing_time = time.time() - start_time if 'start_time' in locals() else 0
        error_message = f"Неожиданная ошибка при загрузке множественных ответов компаний: {str(e)}"
        
        logger.error(
            f"UPLOAD_MULTIPLE_EXCEPTION - Неожиданная ошибка | "
            f"user={request.user.username} | summary_id={summary_id} | "
            f"files_count={len(excel_files)} | processing_time={processing_time:.2f}s | "
            f"error={error_message}",
            exc_info=True
        )
        
        return JsonResponse({
            'success': False,
            'error': 'Произошла неожиданная ошибка при обработке файлов. Обратитесь к администратору.',
            'error_type': 'unexpected_error'
        }, status=500)


@user_required
def offer_search(request):
    """Поиск и фильтрация предложений по различным критериям"""
    from .forms import CompanyOfferSearchForm
    from django.db.models import Q
    
    offers = InsuranceOffer.objects.filter(is_valid=True).select_related('summary', 'summary__request')
    search_form = CompanyOfferSearchForm(request.GET)
    
    if search_form.is_valid():
        # Фильтр по названию компании
        if search_form.cleaned_data.get('company_name'):
            offers = offers.filter(
                company_name__icontains=search_form.cleaned_data['company_name']
            )
        
        # Фильтр по минимальной премии
        if search_form.cleaned_data.get('min_premium'):
            offers = offers.filter(
                Q(premium_with_franchise_1__gte=search_form.cleaned_data['min_premium']) |
                Q(premium_with_franchise_2__gte=search_form.cleaned_data['min_premium'])
            )
        
        # Фильтр по максимальной премии
        if search_form.cleaned_data.get('max_premium'):
            offers = offers.filter(
                Q(premium_with_franchise_1__lte=search_form.cleaned_data['max_premium']) |
                Q(premium_with_franchise_2__lte=search_form.cleaned_data['max_premium'])
            )
        
        # Фильтр только предложения с рассрочкой
        if search_form.cleaned_data.get('installment_only'):
            offers = offers.filter(
                Q(installment_variant_1=True, payments_per_year_variant_1__gt=1) |
                Q(installment_variant_2=True, payments_per_year_variant_2__gt=1) |
                # Legacy fallback для старых записей
                Q(installment_available=True, payments_per_year__gt=1)
            )
    
    # Сортировка
    sort_by = request.GET.get('sort', 'premium_with_franchise_1')
    valid_sorts = [
        'premium_with_franchise_1', '-premium_with_franchise_1',
        'premium_with_franchise_2', '-premium_with_franchise_2',
        'company_name', '-company_name',
        'insurance_year', '-insurance_year',
        'payments_per_year', '-payments_per_year',
        'payments_per_year_variant_1', '-payments_per_year_variant_1',
        'payments_per_year_variant_2', '-payments_per_year_variant_2',
    ]
    if sort_by in valid_sorts:
        offers = offers.order_by(sort_by)
    else:
        offers = offers.order_by('premium_with_franchise_1')
    
    # Группировка предложений по компаниям для лучшего отображения
    companies_data = {}
    for offer in offers:
        if offer.company_name not in companies_data:
            companies_data[offer.company_name] = []
        companies_data[offer.company_name].append(offer)
    
    return render(request, 'summaries/offer_search.html', {
        'offers': offers,
        'search_form': search_form,
        'companies_data': companies_data,
        'current_sort': sort_by
    })



def get_russian_month_name(date):
    """Возвращает название месяца на русском языке"""
    months = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    return f"{months[date.month]} {date.year}"


def _parse_statistics_filters(request):
    """Парсинг и валидация фильтров периода для страницы статистики."""
    from datetime import datetime, timedelta
    from django.utils import timezone

    period = request.GET.get('period', 'all').strip().lower() or 'all'
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()

    start_date = None
    end_date = None
    errors = []

    def _parse_date(value, field_label):
        if not value:
            return None
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f'Некорректная дата в поле "{field_label}"')
            return None

    # Явные даты имеют приоритет над быстрым периодом.
    if start_date_str or end_date_str:
        start_date = _parse_date(start_date_str, 'Дата с')
        end_date = _parse_date(end_date_str, 'Дата по')
        period = 'custom'
    elif period in {'30', '90', '365'}:
        days = int(period)
        end_date = timezone.localdate()
        start_date = end_date - timedelta(days=days - 1)
    else:
        period = 'all'

    if start_date and end_date and start_date > end_date:
        errors.append('Дата начала периода не может быть позже даты окончания')
        start_date, end_date = None, None
        period = 'all'

    return {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_str': start_date.isoformat() if start_date else '',
        'end_date_str': end_date.isoformat() if end_date else '',
        'errors': errors,
    }


def _parse_company_analytics_filters(request):
    """Парсинг фильтров для аналитики по страховым компаниям."""
    filters = _parse_statistics_filters(request)

    date_mode = (request.GET.get('date_mode') or DATE_MODE_SUMMARY_CREATED).strip()
    if date_mode not in DATE_MODE_CHOICES:
        date_mode = DATE_MODE_SUMMARY_CREATED

    comparison_mode = (request.GET.get('comparison_mode') or 'selected_variant').strip()
    if comparison_mode not in {'selected_variant', 'best_available'}:
        comparison_mode = 'selected_variant'

    full_coverage_raw = (request.GET.get('full_coverage') or '1').strip().lower()
    require_full_coverage = full_coverage_raw not in {'0', 'false', 'no'}

    filters.update({
        'date_mode': date_mode,
        'branch': (request.GET.get('branch') or '').strip(),
        'insurance_type': (request.GET.get('insurance_type') or '').strip(),
        'manager_online': (request.GET.get('manager_online') or '').strip(),
        'manager_alliance': (request.GET.get('manager_alliance') or '').strip(),
        'selected_company': (request.GET.get('selected_company') or '').strip(),
        'deal_status': (request.GET.get('deal_status') or '').strip(),
        'comparison_mode': comparison_mode,
        'require_full_coverage': require_full_coverage,
    })
    return filters


def _apply_summary_date_filters(queryset, start_date=None, end_date=None):
    """Применяет фильтрацию периода к queryset сводов."""
    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)
    return queryset


def _build_statistics_payload(start_date=None, end_date=None):
    """Строит данные статистики для шаблона и экспортов."""
    from datetime import timedelta
    from django.db.models import Count, Avg, Sum
    from django.utils import timezone

    summaries_qs = InsuranceSummary.objects.select_related('request', 'request__created_by')
    summaries_qs = _apply_summary_date_filters(summaries_qs, start_date=start_date, end_date=end_date)
    offers_qs = InsuranceOffer.objects.filter(is_valid=True, summary__in=summaries_qs).select_related('summary__request')

    # Основная статистика
    stats = {
        'total_summaries': summaries_qs.count(),
        'collecting': summaries_qs.filter(status='collecting').count(),
        'ready': summaries_qs.filter(status='ready').count(),
        'sent': summaries_qs.filter(status='sent').count(),
        'completed_accepted': summaries_qs.filter(status='completed_accepted').count(),
        'completed_rejected': summaries_qs.filter(status='completed_rejected').count(),
        'completed': summaries_qs.filter(status__in=['completed_accepted', 'completed_rejected']).count(),
        'avg_offers_per_summary': summaries_qs.aggregate(avg=Avg('total_offers'))['avg'] or 0,
        'total_offers': offers_qs.count(),
    }

    # Дельта за 30 дней (в пределах выбранного периода)
    last_month = timezone.now() - timedelta(days=30)
    summaries_last_month = summaries_qs.filter(created_at__gte=last_month)
    offers_last_month = offers_qs.filter(received_at__gte=last_month)
    stats['summaries_last_month'] = summaries_last_month.count()
    stats['offers_last_month'] = offers_last_month.count()

    # Статистика по компаниям с разбивкой по типам страхования
    company_stats_detailed = {}
    for offer in offers_qs:
        company = offer.company_name
        summary_id = offer.summary.id
        insurance_type = offer.summary.request.insurance_type or 'Не указан'

        if company not in company_stats_detailed:
            company_stats_detailed[company] = {'summaries': set(), 'types': {}}

        company_stats_detailed[company]['summaries'].add(summary_id)
        company_stats_detailed[company]['types'].setdefault(insurance_type, set()).add(summary_id)

    for company_name in company_stats_detailed:
        company_stats_detailed[company_name]['total'] = len(company_stats_detailed[company_name]['summaries'])
        for insurance_type in list(company_stats_detailed[company_name]['types'].keys()):
            company_stats_detailed[company_name]['types'][insurance_type] = len(
                company_stats_detailed[company_name]['types'][insurance_type]
            )

    company_stats_detailed = dict(
        sorted(company_stats_detailed.items(), key=lambda x: x[1]['total'], reverse=True)
    )
    for company_name in company_stats_detailed:
        company_stats_detailed[company_name]['types'] = dict(
            sorted(company_stats_detailed[company_name]['types'].items(), key=lambda x: x[1], reverse=True)
        )

    # Итого по типам страхования для выбранного периода
    total_summaries_count = summaries_qs.count()
    summaries_by_type = {}
    for summary in summaries_qs:
        insurance_type = summary.request.insurance_type or 'Не указан'
        summaries_by_type[insurance_type] = summaries_by_type.get(insurance_type, 0) + 1

    company_totals = {
        'total': total_summaries_count,
        'kasko': summaries_by_type.get('КАСКО', 0),
        'spec': summaries_by_type.get('страхование спецтехники', 0),
        'property': summaries_by_type.get('страхование имущества', 0),
        'other': summaries_by_type.get('другое', 0),
    }

    # Статистика по годам страхования
    year_stats = list(
        offers_qs.values('insurance_year').annotate(
            count=Count('id'),
            avg_premium=Avg('premium_with_franchise_1'),
            total_premium=Sum('premium_with_franchise_1')
        ).order_by('insurance_year')
    )

    # Статистика по месяцам создания сводов
    monthly_summaries_stats = {}
    for summary in summaries_qs:
        month_key = summary.created_at.strftime('%Y-%m')
        month_display = get_russian_month_name(summary.created_at)
        monthly_summaries_stats.setdefault(month_key, {'display': month_display, 'count': 0})
        monthly_summaries_stats[month_key]['count'] += 1

    # Сортируем месяцы от новых к старым для таблиц
    monthly_summaries_stats = dict(
        sorted(monthly_summaries_stats.items(), key=lambda x: x[0], reverse=True)
    )

    # Статистика по филиалам
    branch_stats_detailed = {}
    summaries_with_branch = summaries_qs.filter(
        request__branch__isnull=False
    ).exclude(
        request__branch=''
    )

    for summary in summaries_with_branch:
        branch = summary.request.branch
        insurance_type = summary.request.insurance_type or 'Не указан'
        branch_stats_detailed.setdefault(branch, {'total': 0, 'types': {}})
        branch_stats_detailed[branch]['total'] += 1
        branch_stats_detailed[branch]['types'][insurance_type] = (
            branch_stats_detailed[branch]['types'].get(insurance_type, 0) + 1
        )

    branch_stats_detailed = dict(
        sorted(branch_stats_detailed.items(), key=lambda x: x[1]['total'], reverse=True)
    )
    for branch in branch_stats_detailed:
        branch_stats_detailed[branch]['types'] = dict(
            sorted(branch_stats_detailed[branch]['types'].items(), key=lambda x: x[1], reverse=True)
        )
    branch_stats_detailed = dict(list(branch_stats_detailed.items())[:10])

    branch_totals = {
        'total': 0,
        'kasko': 0,
        'spec': 0,
        'property': 0,
        'other': 0,
    }
    for branch_data in branch_stats_detailed.values():
        branch_totals['total'] += branch_data['total']
        branch_totals['kasko'] += branch_data['types'].get('КАСКО', 0)
        branch_totals['spec'] += branch_data['types'].get('страхование спецтехники', 0)
        branch_totals['property'] += branch_data['types'].get('страхование имущества', 0)
        branch_totals['other'] += branch_data['types'].get('другое', 0)

    # Статистика по пользователям Django
    user_stats = {}
    user_monthly_stats = {}
    offers_count_by_summary = dict(
        offers_qs.values('summary_id').annotate(count=Count('id')).values_list('summary_id', 'count')
    )

    for summary in summaries_qs:
        user = summary.request.created_by
        if not user:
            continue

        user_display = f"{user.first_name} {user.last_name}".strip() or user.username
        if user_display not in user_stats:
            user_stats[user_display] = {
                'count': 0,
                'offers_count': 0,
                'accepted': 0,
                'rejected': 0,
            }

        user_stats[user_display]['count'] += 1
        user_stats[user_display]['offers_count'] += offers_count_by_summary.get(summary.id, 0)
        if summary.status == 'completed_accepted':
            user_stats[user_display]['accepted'] += 1
        elif summary.status == 'completed_rejected':
            user_stats[user_display]['rejected'] += 1

        month_key = summary.created_at.strftime('%Y-%m')
        month_display = get_russian_month_name(summary.created_at)
        user_monthly_stats.setdefault(user_display, {})
        user_monthly_stats[user_display].setdefault(month_key, {'display': month_display, 'count': 0})
        user_monthly_stats[user_display][month_key]['count'] += 1

    user_priority_order = ['grigoriigrachev', 'test_user', 'testuser']
    priority_users = [u for u in user_stats.keys() if u in user_priority_order]
    other_users = [u for u in user_stats.keys() if u not in user_priority_order]
    priority_users.sort(key=lambda x: user_priority_order.index(x))
    other_users.sort()
    ordered_users = priority_users + other_users
    user_stats = {user: user_stats[user] for user in ordered_users if user in user_stats}
    user_monthly_stats = {user: user_monthly_stats[user] for user in ordered_users if user in user_monthly_stats}

    for user_display in user_monthly_stats:
        user_monthly_stats[user_display] = dict(
            sorted(user_monthly_stats[user_display].items(), key=lambda x: x[0], reverse=True)
        )

    # Данные для мини-графиков
    monthly_chart_items = sorted(monthly_summaries_stats.items(), key=lambda x: x[0])
    top_companies_items = list(company_stats_detailed.items())[:8]
    chart_data = {
        'statuses': {
            'labels': ['Сбор', 'Готов', 'Отправлен', 'Акцепт', 'Не будет'],
            'values': [
                stats['collecting'],
                stats['ready'],
                stats['sent'],
                stats['completed_accepted'],
                stats['completed_rejected'],
            ],
            'colors': ['#f59e0b', '#06b6d4', '#3b82f6', '#16a34a', '#64748b'],
        },
        'monthly': {
            'labels': [item[1]['display'] for item in monthly_chart_items],
            'values': [item[1]['count'] for item in monthly_chart_items],
            'color': '#0b7a75',
        },
        'top_companies': {
            'labels': [item[0] for item in top_companies_items],
            'values': [item[1]['total'] for item in top_companies_items],
            'color': '#7c3aed',
        },
    }

    return {
        'stats': stats,
        'company_stats_detailed': company_stats_detailed,
        'company_totals': company_totals,
        'year_stats': year_stats,
        'monthly_summaries_stats': monthly_summaries_stats,
        'branch_stats_detailed': branch_stats_detailed,
        'branch_totals': branch_totals,
        'user_stats': user_stats,
        'user_monthly_stats': user_monthly_stats,
        'chart_data': chart_data,
    }


@admin_required
def summary_statistics(request):
    """Статистика по сводам"""
    filters = _parse_statistics_filters(request)
    for error_message in filters['errors']:
        messages.warning(request, error_message)

    payload = _build_statistics_payload(
        start_date=filters['start_date'],
        end_date=filters['end_date']
    )

    context = {
        **payload,
        'filters': {
            'period': filters['period'],
            'start_date': filters['start_date_str'],
            'end_date': filters['end_date_str'],
        },
    }
    return render(request, 'summaries/statistics.html', context)


@admin_required
def export_statistics_widget(request):
    """Экспорт виджетов статистики в XLSX."""
    from io import BytesIO
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font

    filters = _parse_statistics_filters(request)
    payload = _build_statistics_payload(
        start_date=filters['start_date'],
        end_date=filters['end_date']
    )
    widget = request.GET.get('widget', 'overview').strip().lower() or 'overview'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Статистика'

    title_map = {
        'overview': 'Сводка по статистике',
        'statuses': 'Статусы сводов',
        'monthly': 'Динамика по месяцам',
        'companies': 'Компании',
        'branches': 'Филиалы',
        'users': 'Пользователи',
    }
    worksheet['A1'] = title_map.get(widget, title_map['overview'])
    worksheet['A1'].font = Font(bold=True, size=14)

    period_display = 'Все время'
    if filters['start_date'] and filters['end_date']:
        period_display = f"{filters['start_date_str']} - {filters['end_date_str']}"
    elif filters['start_date']:
        period_display = f"с {filters['start_date_str']}"
    elif filters['end_date']:
        period_display = f"по {filters['end_date_str']}"

    worksheet['A2'] = f"Период: {period_display}"
    worksheet['A3'] = f"Сформировано: {timezone.localtime().strftime('%d.%m.%Y %H:%M')}"

    row = 5

    def write_headers(*headers):
        nonlocal row
        for idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=idx, value=header)
            cell.font = Font(bold=True)
        row += 1

    def write_row(*values):
        nonlocal row
        for idx, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=idx, value=value)
        row += 1

    if widget == 'statuses':
        write_headers('Статус', 'Количество')
        status_labels = payload['chart_data']['statuses']['labels']
        status_values = payload['chart_data']['statuses']['values']
        for label, value in zip(status_labels, status_values):
            write_row(label, value)
    elif widget == 'monthly':
        write_headers('Месяц', 'Количество сводов')
        for month_key, month_data in sorted(payload['monthly_summaries_stats'].items(), key=lambda x: x[0]):
            write_row(month_data['display'], month_data['count'])
    elif widget == 'companies':
        write_headers('Компания', 'В сводах, шт.', 'КАСКО', 'Спецтехника', 'Имущество', 'Другое')
        for company_name, company_data in payload['company_stats_detailed'].items():
            write_row(
                company_name,
                company_data.get('total', 0),
                company_data['types'].get('КАСКО', 0),
                company_data['types'].get('страхование спецтехники', 0),
                company_data['types'].get('страхование имущества', 0),
                company_data['types'].get('другое', 0),
            )
    elif widget == 'branches':
        write_headers('Филиал', 'Всего, шт.', 'КАСКО', 'Спецтехника', 'Имущество', 'Другое')
        for branch_name, branch_data in payload['branch_stats_detailed'].items():
            write_row(
                branch_name,
                branch_data.get('total', 0),
                branch_data['types'].get('КАСКО', 0),
                branch_data['types'].get('страхование спецтехники', 0),
                branch_data['types'].get('страхование имущества', 0),
                branch_data['types'].get('другое', 0),
            )
    elif widget == 'users':
        write_headers('Пользователь', 'Сводов', 'Предложений', 'Акцептов', 'Не будет')
        for user_display, user_data in payload['user_stats'].items():
            write_row(
                user_display,
                user_data.get('count', 0),
                user_data.get('offers_count', 0),
                user_data.get('accepted', 0),
                user_data.get('rejected', 0),
            )
    else:
        write_headers('Метрика', 'Значение')
        write_row('Всего сводов', payload['stats']['total_summaries'])
        write_row('Всего предложений', payload['stats']['total_offers'])
        write_row('Среднее предложений на свод', float(payload['stats']['avg_offers_per_summary'] or 0))
        write_row('Сбор предложений', payload['stats']['collecting'])
        write_row('Готов к отправке', payload['stats']['ready'])
        write_row('Отправлен в Альянс', payload['stats']['sent'])
        write_row('Завершен: акцепт/распоряжение', payload['stats']['completed_accepted'])
        write_row('Завершен: не будет', payload['stats']['completed_rejected'])

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    today = timezone.localtime().strftime('%d_%m_%Y')
    filename = f"stats_{widget}_{today}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _safe_decimal(value):
    """Безопасно приводит значение к Decimal."""
    if value is None:
        return None

    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _median_decimal(values):
    """Возвращает медиану для списка Decimal значений."""
    if not values:
        return None

    sorted_values = sorted(values)
    total_count = len(sorted_values)
    middle_index = total_count // 2

    if total_count % 2 == 1:
        return sorted_values[middle_index]

    return (sorted_values[middle_index - 1] + sorted_values[middle_index]) / Decimal('2')


def _sum_company_premium_for_variant(offers, variant, years_for_comparison):
    """
    Суммирует итоговую премию компании по выбранному варианту франшизы.

    Возвращает None, если по любому из годов нет валидной премии.
    """
    offers_by_year = {offer.insurance_year: offer for offer in offers}
    total = Decimal('0')

    for year in sorted(years_for_comparison):
        offer = offers_by_year.get(year)
        if offer is None:
            return None

        premium_raw = (
            offer.premium_with_franchise_1
            if variant == 1
            else offer.premium_with_franchise_2
        )
        premium_value = _safe_decimal(premium_raw)
        if premium_value is None or premium_value <= 0:
            return None

        total += premium_value

    return total


def _build_deal_price_row(summary, comparison_mode='selected_variant', require_full_coverage=True):
    """Формирует аналитическую строку по одной завершенной сделке."""
    selected_company = (summary.selected_company or '').strip()
    if not selected_company:
        return None

    selected_variant = summary.selected_franchise_variant or 1
    if selected_variant not in (1, 2):
        selected_variant = 1

    prefetched_valid_offers = getattr(summary, 'valid_offers_prefetched', None)
    offers_iterable = prefetched_valid_offers if prefetched_valid_offers is not None else summary.offers.all()

    offers_by_company = {}
    for offer in offers_iterable:
        if prefetched_valid_offers is None and not offer.is_valid:
            continue
        offers_by_company.setdefault(offer.company_name, []).append(offer)

    selected_offers = offers_by_company.get(selected_company, [])
    if not selected_offers:
        return None

    selected_years = {offer.insurance_year for offer in selected_offers}
    if not selected_years:
        return None

    company_totals = {}
    company_variants = {}

    for company_name, company_offers in offers_by_company.items():
        company_years = {offer.insurance_year for offer in company_offers}
        years_for_comparison = selected_years if require_full_coverage else company_years

        if require_full_coverage and company_years != selected_years:
            continue

        if company_name == selected_company:
            selected_total = _sum_company_premium_for_variant(
                company_offers,
                selected_variant,
                years_for_comparison,
            )
            if selected_total is None:
                continue

            company_totals[company_name] = selected_total
            company_variants[company_name] = selected_variant
            continue

        if comparison_mode == 'best_available':
            total_variant_1 = _sum_company_premium_for_variant(
                company_offers,
                1,
                years_for_comparison,
            )
            total_variant_2 = _sum_company_premium_for_variant(
                company_offers,
                2,
                years_for_comparison,
            )

            candidates = []
            if total_variant_1 is not None:
                candidates.append((1, total_variant_1))
            if total_variant_2 is not None:
                candidates.append((2, total_variant_2))

            if not candidates:
                continue

            best_variant, best_total = min(candidates, key=lambda item: item[1])
            company_totals[company_name] = best_total
            company_variants[company_name] = best_variant
        else:
            total_same_variant = _sum_company_premium_for_variant(
                company_offers,
                selected_variant,
                years_for_comparison,
            )
            if total_same_variant is None:
                continue

            company_totals[company_name] = total_same_variant
            company_variants[company_name] = selected_variant

    if selected_company not in company_totals:
        return None

    if len(company_totals) < 2:
        return None

    sorted_companies = sorted(company_totals.items(), key=lambda item: (item[1], item[0]))
    selected_total = company_totals[selected_company]
    min_total = sorted_companies[0][1]
    max_total = sorted_companies[-1][1]
    spread_abs = max_total - min_total
    spread_pct = ((spread_abs / min_total) * Decimal('100')) if min_total > 0 else None

    selected_rank = 1 + sum(
        1 for _, total_value in company_totals.items()
        if total_value < selected_total
    )
    delta_to_min_abs = selected_total - min_total
    delta_to_min_pct = ((delta_to_min_abs / min_total) * Decimal('100')) if min_total > 0 else None
    is_min_selected = selected_total == min_total

    best_company_names = [
        company_name for company_name, total_value in sorted_companies
        if total_value == min_total
    ]
    best_company_name = ', '.join(best_company_names)

    points = []
    for company_name, total_value in sorted_companies:
        if spread_abs > 0:
            position_pct = float(((total_value - min_total) / spread_abs) * Decimal('100'))
        else:
            position_pct = 50.0

        points.append({
            'company_name': company_name,
            'total': total_value,
            'variant_used': company_variants.get(company_name),
            'is_selected': company_name == selected_company,
            'position_pct': round(position_pct, 2),
        })

    competitor_points = [point for point in points if not point['is_selected']]
    top_competitors = competitor_points[:3]

    return {
        'summary': summary,
        'request': summary.request,
        'selected_company': selected_company,
        'selected_variant': selected_variant,
        'selected_total': selected_total,
        'min_total': min_total,
        'max_total': max_total,
        'spread_abs': spread_abs,
        'spread_pct': spread_pct,
        'delta_to_min_abs': delta_to_min_abs,
        'delta_to_min_pct': delta_to_min_pct,
        'selected_rank': selected_rank,
        'is_min_selected': is_min_selected,
        'best_company_name': best_company_name,
        'comparable_companies_count': len(company_totals),
        'total_companies_count': len(offers_by_company),
        'years_count': len(selected_years),
        'points': points,
        'top_competitors': top_competitors,
    }


@admin_required
def analytics_insurance_offers(request):
    """MVP аналитики по выбору страховых предложений."""
    filters = _parse_statistics_filters(request)
    for error_message in filters['errors']:
        messages.warning(request, error_message)

    comparison_mode = (request.GET.get('comparison_mode') or 'selected_variant').strip()
    if comparison_mode not in {'selected_variant', 'best_available'}:
        comparison_mode = 'selected_variant'

    full_coverage_raw = (request.GET.get('full_coverage') or '1').strip().lower()
    require_full_coverage = full_coverage_raw not in {'0', 'false', 'no'}

    selected_branch = (request.GET.get('branch') or '').strip()
    selected_insurance_type = (request.GET.get('insurance_type') or '').strip()
    selected_manager_id = (request.GET.get('manager') or '').strip()

    filter_base_qs = InsuranceSummary.objects.select_related('request', 'request__created_by').filter(
        status='completed_accepted'
    ).exclude(
        selected_company__isnull=True
    ).exclude(
        selected_company=''
    )
    filter_base_qs = _apply_summary_date_filters(
        filter_base_qs,
        start_date=filters['start_date'],
        end_date=filters['end_date']
    )

    available_branches = list(
        filter_base_qs.exclude(
            request__branch__isnull=True
        ).exclude(
            request__branch=''
        ).values_list('request__branch', flat=True).distinct().order_by('request__branch')
    )

    available_insurance_types = list(
        filter_base_qs.exclude(
            request__insurance_type__isnull=True
        ).exclude(
            request__insurance_type=''
        ).values_list('request__insurance_type', flat=True).distinct().order_by('request__insurance_type')
    )

    manager_rows = filter_base_qs.exclude(
        request__created_by__isnull=True
    ).order_by().values(
        'request__created_by_id',
        'request__created_by__username',
        'request__created_by__first_name',
        'request__created_by__last_name'
    ).distinct()

    available_managers = []
    for manager_row in manager_rows:
        first_name = (manager_row.get('request__created_by__first_name') or '').strip()
        last_name = (manager_row.get('request__created_by__last_name') or '').strip()
        username = (manager_row.get('request__created_by__username') or '').strip()
        display_name = f"{first_name} {last_name}".strip() or username

        available_managers.append({
            'id': str(manager_row.get('request__created_by_id')),
            'name': display_name,
        })
    available_managers = sorted(available_managers, key=lambda item: item['name'].lower())

    summaries_qs = filter_base_qs
    if selected_branch:
        summaries_qs = summaries_qs.filter(request__branch=selected_branch)
    if selected_insurance_type:
        summaries_qs = summaries_qs.filter(request__insurance_type=selected_insurance_type)
    if selected_manager_id:
        try:
            summaries_qs = summaries_qs.filter(request__created_by_id=int(selected_manager_id))
        except ValueError:
            messages.warning(request, 'Некорректный фильтр менеджера был сброшен')
            selected_manager_id = ''

    summaries_qs = summaries_qs.prefetch_related('offers').order_by('-created_at')

    rows = []
    for summary in summaries_qs:
        row_data = _build_deal_price_row(
            summary,
            comparison_mode=comparison_mode,
            require_full_coverage=require_full_coverage,
        )
        if row_data is not None:
            rows.append(row_data)

    total_deals = len(rows)
    min_selected_count = sum(1 for row in rows if row['is_min_selected'])
    non_min_selected_count = total_deals - min_selected_count

    delta_abs_values = [row['delta_to_min_abs'] for row in rows if row['delta_to_min_abs'] is not None]
    delta_pct_values = [row['delta_to_min_pct'] for row in rows if row['delta_to_min_pct'] is not None]
    comparable_counts = [row['comparable_companies_count'] for row in rows]

    median_delta_abs = _median_decimal(delta_abs_values)
    median_delta_pct = _median_decimal(delta_pct_values)
    avg_companies = (sum(comparable_counts) / total_deals) if total_deals > 0 else 0
    min_selected_rate = (
        (Decimal(min_selected_count) / Decimal(total_deals) * Decimal('100'))
        if total_deals > 0 else Decimal('0')
    )

    paginator = Paginator(rows, 25)
    page = request.GET.get('page')
    try:
        rows_page = paginator.page(page)
    except PageNotAnInteger:
        rows_page = paginator.page(1)
    except EmptyPage:
        rows_page = paginator.page(paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring_without_page = query_params.urlencode()

    context = {
        'rows': rows_page,
        'paginator': paginator,
        'analytics_kpi': {
            'total_deals': total_deals,
            'min_selected_count': min_selected_count,
            'non_min_selected_count': non_min_selected_count,
            'min_selected_rate': min_selected_rate,
            'median_delta_abs': median_delta_abs,
            'median_delta_pct': median_delta_pct,
            'avg_companies': avg_companies,
        },
        'filters': {
            'period': filters['period'],
            'start_date': filters['start_date_str'],
            'end_date': filters['end_date_str'],
            'branch': selected_branch,
            'insurance_type': selected_insurance_type,
            'manager': selected_manager_id,
            'comparison_mode': comparison_mode,
            'full_coverage': require_full_coverage,
        },
        'available_branches': available_branches,
        'available_insurance_types': available_insurance_types,
        'available_managers': available_managers,
        'querystring_without_page': querystring_without_page,
    }
    return render(request, 'summaries/analytics_insurance_offers.html', context)


@admin_required
def analytics_insurance_companies(request):
    """Аналитика по страховым компаниям."""
    filters = _parse_company_analytics_filters(request)
    for error_message in filters['errors']:
        messages.warning(request, error_message)

    payload = build_analytics_insurance_companies_payload(
        start_date=filters['start_date'],
        end_date=filters['end_date'],
        date_mode=filters['date_mode'],
        branch=filters['branch'],
        insurance_type=filters['insurance_type'],
        manager_online=filters['manager_online'],
        manager_alliance=filters['manager_alliance'],
        selected_company=filters['selected_company'],
        deal_status=filters['deal_status'],
        comparison_mode=filters['comparison_mode'],
        require_full_coverage=filters['require_full_coverage'],
        price_row_builder=_build_deal_price_row,
    )
    for error_message in payload.get('filter_errors', []):
        messages.warning(request, error_message)

    context = {
        **payload,
        'filters': {
            'period': filters['period'],
            'start_date': filters['start_date_str'],
            'end_date': filters['end_date_str'],
            'date_mode': filters['date_mode'],
            'branch': filters['branch'],
            'insurance_type': filters['insurance_type'],
            'manager_online': filters['manager_online'],
            'manager_alliance': filters['manager_alliance'],
            'selected_company': filters['selected_company'],
            'deal_status': filters['deal_status'],
            'comparison_mode': filters['comparison_mode'],
            'require_full_coverage': filters['require_full_coverage'],
        },
        'date_mode_choices': [
            {'value': value, 'label': label}
            for value, label in DATE_MODE_CHOICES.items()
        ],
    }
    return render(request, 'summaries/analytics_insurance_companies.html', context)


@admin_required
def export_analytics_insurance_companies_widget(request):
    """Экспорт виджетов аналитики по страховым компаниям в XLSX."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    filters = _parse_company_analytics_filters(request)
    payload = build_analytics_insurance_companies_payload(
        start_date=filters['start_date'],
        end_date=filters['end_date'],
        date_mode=filters['date_mode'],
        branch=filters['branch'],
        insurance_type=filters['insurance_type'],
        manager_online=filters['manager_online'],
        manager_alliance=filters['manager_alliance'],
        selected_company=filters['selected_company'],
        deal_status=filters['deal_status'],
        comparison_mode=filters['comparison_mode'],
        require_full_coverage=filters['require_full_coverage'],
        price_row_builder=_build_deal_price_row,
    )

    export_payload = payload['export_payload']
    widget = request.GET.get('widget', 'overview').strip().lower() or 'overview'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Аналитика СК'

    title_map = {
        'overview': 'Сводка KPI',
        'rating': 'Рейтинг страховых компаний',
        'competitiveness': 'Ценовая конкурентность',
        'conversion': 'Конверсия в выбор',
        'slice_branch': 'Разрез СК x Филиал',
        'slice_manager_alliance': 'Разрез СК x Менеджер Альянса',
        'slice_insurance_type': 'Разрез СК x Тип страхования',
        'slice_deal_status': 'Разрез СК x Статус сделки',
        'dynamics': 'Динамика по времени',
        'data_quality': 'Data Quality',
    }
    worksheet['A1'] = title_map.get(widget, title_map['overview'])
    worksheet['A1'].font = Font(bold=True, size=14)

    period_display = 'Все время'
    if filters['start_date'] and filters['end_date']:
        period_display = f"{filters['start_date_str']} - {filters['end_date_str']}"
    elif filters['start_date']:
        period_display = f"с {filters['start_date_str']}"
    elif filters['end_date']:
        period_display = f"по {filters['end_date_str']}"

    worksheet['A2'] = f"Период: {period_display}"
    worksheet['A3'] = f"Режим даты: {DATE_MODE_CHOICES.get(filters['date_mode'], DATE_MODE_CHOICES[DATE_MODE_SUMMARY_CREATED])}"
    worksheet['A4'] = f"Сформировано: {timezone.localtime().strftime('%d.%m.%Y %H:%M')}"

    row = 6

    def write_headers(*headers):
        nonlocal row
        for idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=idx, value=header)
            cell.font = Font(bold=True)
        row += 1

    def write_row(*values):
        nonlocal row
        for idx, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=idx, value=value)
        row += 1

    if widget == 'rating':
        write_headers(
            'Позиция',
            'СК',
            'Участвовала в сделках, шт.',
            'Побед, шт.',
            'Покрытие, %',
            'Win rate при участии, %',
            'Доля побед, %',
            'Сумма премии побед, ₽',
            'Средняя премия побед, ₽',
            'Средний ранг при выборе',
            'Медиана Δ к минимуму, ₽',
            'Доля min-selected при выборе, %',
        )
        for rating_row in export_payload['rating_rows']:
            write_row(
                rating_row['position'],
                rating_row['company_name'],
                rating_row['offered_in_deals_count'],
                rating_row['selected_wins_count'],
                float(rating_row['coverage_pct']),
                float(rating_row['win_rate_when_offered_pct']),
                float(rating_row['win_share_pct']),
                float(rating_row['selected_premium_sum']),
                float(rating_row['selected_premium_avg']) if rating_row['selected_premium_avg'] is not None else None,
                rating_row['avg_rank_when_selected'],
                float(rating_row['median_delta_abs_when_selected']) if rating_row['median_delta_abs_when_selected'] is not None else None,
                float(rating_row['min_selected_rate_when_selected']),
            )
    elif widget == 'competitiveness':
        write_headers(
            'СК',
            'Выбрана в сделках, шт.',
            'Сопоставимых сделок, шт.',
            'Средний ранг',
            'Медиана Δ к минимуму, ₽',
            'Медиана Δ к минимуму, %',
            'Доля выбора минимума, %',
            'Среднее число конкурентов',
        )
        for row_data in export_payload['competitiveness_rows']:
            write_row(
                row_data['company_name'],
                row_data['selected_deals_count'],
                row_data['comparable_selected_deals_count'],
                row_data['avg_rank'],
                float(row_data['median_delta_abs']) if row_data['median_delta_abs'] is not None else None,
                float(row_data['median_delta_pct']) if row_data['median_delta_pct'] is not None else None,
                float(row_data['min_selected_rate']),
                row_data['avg_competitors'],
            )
    elif widget == 'conversion':
        write_headers('СК', 'Участвовала в сделках, шт.', 'Выбрана, шт.', 'Конверсия в выбор, %', 'Доля побед, %')
        for row_data in export_payload['conversion_rows']:
            write_row(
                row_data['company_name'],
                row_data['offered_in_deals_count'],
                row_data['selected_wins_count'],
                float(row_data['conversion_pct']),
                float(row_data['win_share_pct']),
            )
    elif widget in {'slice_branch', 'slice_manager_alliance', 'slice_insurance_type', 'slice_deal_status'}:
        slice_key_map = {
            'slice_branch': 'branch',
            'slice_manager_alliance': 'manager_alliance',
            'slice_insurance_type': 'insurance_type',
            'slice_deal_status': 'deal_status',
        }
        slice_rows = export_payload['slices'][slice_key_map[widget]]
        write_headers('СК', 'Значение разреза', 'Участвовала в сделках, шт.', 'Выбрана, шт.', 'Win rate при участии, %')
        for row_data in slice_rows:
            write_row(
                row_data['company_name'],
                row_data['dimension_value'],
                row_data['offered_in_deals_count'],
                row_data['selected_wins_count'],
                float(row_data['win_rate_when_offered_pct']),
            )
    elif widget == 'dynamics':
        write_headers(
            'Месяц',
            'Сделок, шт.',
            'Сопоставимых сделок, шт.',
            'Выбор минимума, шт.',
            'Доля выбора минимума, %',
            'Средняя премия выбора, ₽',
            'Уникальных выбранных СК',
        )
        for row_data in export_payload['dynamics_rows']:
            write_row(
                row_data['month_label'],
                row_data['total_deals'],
                row_data['comparable_deals'],
                row_data['min_selected_count'],
                float(row_data['min_selected_rate']),
                float(row_data['selected_premium_avg']) if row_data['selected_premium_avg'] is not None else None,
                row_data['distinct_selected_companies'],
            )
    elif widget == 'data_quality':
        write_headers('Индикатор', 'Количество', 'Доля, %')
        for quality_row in export_payload['data_quality_rows']:
            write_row(
                quality_row['label'],
                quality_row['count'],
                float(quality_row['rate']),
            )
    else:
        kpi = export_payload['kpi']
        write_headers('Метрика', 'Значение')
        write_row('Сделок в выборке', kpi['total_deals'])
        write_row('Сопоставимых сделок', kpi['comparable_deals'])
        write_row('Уникальных СК (участвовали)', kpi['distinct_companies_offered'])
        write_row('Уникальных СК (выбраны)', kpi['distinct_companies_selected'])
        write_row('Выбран минимум, шт.', kpi['min_selected_count'])
        write_row('Выбран минимум, %', float(kpi['min_selected_rate']))
        write_row('Среднее число конкурентов', kpi['avg_competitors'])
        write_row('Средняя выбранная премия, ₽', float(kpi['avg_selected_premium']) if kpi['avg_selected_premium'] is not None else None)
        write_row('Медиана Δ к минимуму, ₽', float(kpi['median_delta_abs']) if kpi['median_delta_abs'] is not None else None)
        write_row('Медиана Δ к минимуму, %', float(kpi['median_delta_pct']) if kpi['median_delta_pct'] is not None else None)
        write_row('Средний ранг выбора', kpi['avg_rank'])
        write_row('Доля многолетних сделок, %', float(kpi['multiyear_rate']))
        write_row('Доля сделок с рассрочкой, %', float(kpi['installment_rate']))
        write_row('Конкурентные сделки (>=3 СК), шт.', kpi['competitive_deals_count'])
        write_row('Конкурентные сделки (>=3 СК), %', float(kpi['competitive_deals_rate']))
        write_row('Среднее число СК на сделку', kpi['avg_offered_companies_per_deal'])
        write_row('Среднее время request -> summary, ч', kpi['avg_hours_request_to_summary'])
        write_row('Среднее время summary -> close, ч', kpi['avg_hours_summary_to_close'])

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 54)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    today = timezone.localtime().strftime('%d_%m_%Y')
    filename = f"analytics_companies_{widget}_{today}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


@admin_required
def analytics_placeholder(request):
    """Индексный экран раздела аналитики со списком вложенных отчетов."""
    return render(request, 'summaries/analytics_placeholder.html')


@admin_required
def analytics_managers(request):
    """Аналитика по сотрудникам — обзор."""
    filters = analytics_managers_service.parse_filters(request.GET)
    payload = analytics_managers_service.build_overview_payload(filters)
    payload['alerts'] = analytics_managers_service.build_alerts(filters)
    return render(request, 'summaries/analytics_managers.html', payload)


@admin_required
def analytics_manager_detail(request, user_id):
    """Аналитика по сотрудникам — досье одного сотрудника."""
    filters = analytics_managers_service.parse_filters(request.GET)
    payload = analytics_managers_service.build_manager_profile_payload(user_id, filters)
    return render(request, 'summaries/analytics_manager_detail.html', payload)


@admin_required
def analytics_managers_compare(request):
    """Side-by-side сравнение нескольких сотрудников."""
    filters = analytics_managers_service.parse_filters(request.GET)
    raw_ids = request.GET.get('ids', '')
    user_ids: list[int] = []
    for token in raw_ids.split(','):
        token = token.strip()
        if not token:
            continue
        try:
            user_ids.append(int(token))
        except ValueError:
            filters.errors.append(f'Некорректный id: {token!r}')
    payload = analytics_managers_service.build_compare_payload(user_ids, filters)
    return render(request, 'summaries/analytics_managers_compare.html', payload)


@admin_required
def analytics_managers_leaderboard(request):
    """Леденборд сотрудников по composite efficiency-index (admin-only)."""
    filters = analytics_managers_service.parse_filters(request.GET)
    payload = analytics_managers_service.build_leaderboard_payload(filters)
    return render(request, 'summaries/analytics_managers_leaderboard.html', payload)


@admin_required
def export_analytics_managers_widget(request, user_id=None):
    """XLSX-экспорт: общий обзор или досье одного сотрудника."""
    filters = analytics_managers_service.parse_filters(request.GET)
    today = timezone.localdate().strftime('%Y-%m-%d')
    if user_id is None:
        output = analytics_managers_service.export_overview_xlsx(filters)
        filename = f'employee_analytics_{today}.xlsx'
    else:
        output = analytics_managers_service.export_manager_dossier_xlsx(user_id, filters)
        filename = f'employee_dossier_{user_id}_{today}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@user_required
def help_page(request):
    """
    Справочная страница для модуля сводов
    
    Отображает справочную информацию о работе со сводами предложений,
    включая процессы загрузки ответов страховщиков, выгрузки сводов,
    рабочий процесс и примеры использования.
    
    Требования: 2.1, 2.2, 2.3, 2.4, 2.5, 3.1, 3.2, 3.3, 5.1, 5.2, 5.4, 5.5
    
    Args:
        request: HTTP запрос от пользователя
        
    Returns:
        HttpResponse: Отрендеренная страница справки или редирект при ошибке
        
    Raises:
        Exception: Любые ошибки логируются и обрабатываются gracefully
    """
    try:
        # Подготавливаем контекст для шаблона справки
        context = {
            'title': 'Справка по работе со сводами',
            'sections': [
                'upload_responses',    # Раздел о загрузке ответов страховщиков
                'export_summaries',    # Раздел о выгрузке сводов
                'examples'            # Раздел с примерами и образцами
            ]
        }
        
        # Отображаем шаблон справки с подготовленным контекстом
        return render(request, 'summaries/help.html', context)
        
    except Exception as e:
        # Логируем ошибку с полной трассировкой для отладки
        logger.error(f"Error loading help page: {str(e)}", exc_info=True)
        
        # Показываем пользователю понятное сообщение об ошибке
        messages.error(request, 'Произошла ошибка при загрузке справочной страницы. Обратитесь к администратору.')
        
        # Выполняем graceful fallback - перенаправляем к списку сводов
        return redirect('summaries:summary_list')
