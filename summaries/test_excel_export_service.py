"""
Unit тесты для ExcelExportService
"""

import os
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock

from django.test import TestCase
from django.conf import settings
from openpyxl import Workbook
from openpyxl.workbook import Workbook as OpenpyxlWorkbook

from insurance_requests.models import InsuranceRequest
from summaries.models import InsuranceSummary
from summaries.services import (
    ExcelExportService,
    ExcelExportServiceError,
    TemplateNotFoundError,
    InvalidSummaryDataError
)


class ExcelExportServiceTests(TestCase):
    """Тесты для класса ExcelExportService"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона для тестов
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        # Создаем простой Excel файл для тестов
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        workbook.save(self.template_path)
        
        # Создаем тестовые данные
        self.insurance_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            insurance_type='КАСКО',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН',
            branch='Тестовый филиал'
        )
        
        self.insurance_summary = InsuranceSummary.objects.create(
            request=self.insurance_request,
            status='ready'
        )
    
    def tearDown(self):
        """Очистка после тестов"""
        # Удаляем временные файлы
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)
    
    def test_init_with_valid_template(self):
        """Тест инициализации сервиса с валидным шаблоном"""
        service = ExcelExportService(self.template_path)
        self.assertEqual(service.template_path, Path(self.template_path))
    
    def test_init_with_nonexistent_template(self):
        """Тест инициализации сервиса с несуществующим шаблоном"""
        nonexistent_path = '/path/to/nonexistent/template.xlsx'
        
        with self.assertRaises(TemplateNotFoundError) as context:
            ExcelExportService(nonexistent_path)
        
        self.assertIn('Шаблон Excel не найден по пути', str(context.exception))
    
    def test_init_with_directory_instead_of_file(self):
        """Тест инициализации сервиса с директорией вместо файла"""
        with self.assertRaises(TemplateNotFoundError) as context:
            ExcelExportService(self.temp_dir)
        
        self.assertIn('Путь к шаблону не является файлом', str(context.exception))


class ExcelExportServiceValidationTests(TestCase):
    """Тесты валидации данных в ExcelExportService"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        workbook = Workbook()
        workbook.save(self.template_path)
        
        self.service = ExcelExportService(self.template_path)
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)
    
    def test_validate_summary_data_valid_data(self):
        """Тест валидации корректных данных свода"""
        # Создаем заявку с полными данными
        request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        # Валидация должна пройти без исключений
        try:
            self.service._validate_summary_data(summary)
        except InvalidSummaryDataError:
            self.fail("Валидация корректных данных не должна вызывать исключение")
    
    def test_validate_summary_data_missing_request(self):
        """Тест валидации свода без связанной заявки"""
        # Создаем мок объекта свода без связанной заявки
        summary = Mock()
        summary.request = None
        summary.id = 999
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        self.assertIn('связанная заявка', str(context.exception))
    
    def test_validate_summary_data_missing_dfa_number(self):
        """Тест валидации заявки без номера ДФА"""
        request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number=''  # Пустой номер ДФА
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        self.assertIn('номер заявки (dfa_number)', str(context.exception))
    
    def test_validate_summary_data_missing_vehicle_info(self):
        """Тест валидации заявки без информации о предмете лизинга"""
        request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='',  # Пустая информация о предмете лизинга
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        self.assertIn('информация о предмете лизинга (vehicle_info)', str(context.exception))
    
    def test_validate_summary_data_missing_client_name(self):
        """Тест валидации заявки без названия клиента"""
        request = InsuranceRequest.objects.create(
            client_name='',  # Пустое название клиента
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        self.assertIn('название клиента (client_name)', str(context.exception))
    
    def test_validate_summary_data_multiple_missing_fields(self):
        """Тест валидации заявки с несколькими отсутствующими полями"""
        request = InsuranceRequest.objects.create(
            client_name='',  # Пустое название клиента
            inn='1234567890',
            vehicle_info='',  # Пустая информация о предмете лизинга
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        error_message = str(context.exception)
        self.assertIn('название клиента (client_name)', error_message)
        self.assertIn('информация о предмете лизинга (vehicle_info)', error_message)
    
    def test_validate_summary_data_whitespace_only_fields(self):
        """Тест валидации заявки с полями, содержащими только пробелы"""
        request = InsuranceRequest.objects.create(
            client_name='   ',  # Только пробелы
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        summary = InsuranceSummary.objects.create(
            request=request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError) as context:
            self.service._validate_summary_data(summary)
        
        self.assertIn('название клиента (client_name)', str(context.exception))


class ExcelExportServiceExcelOperationsTests(TestCase):
    """Тесты операций с Excel в ExcelExportService"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        # Создаем Excel файл с именованным листом
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        workbook.save(self.template_path)
        
        self.service = ExcelExportService(self.template_path)
        
        # Создаем тестовые данные
        self.insurance_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        self.insurance_summary = InsuranceSummary.objects.create(
            request=self.insurance_request,
            status='ready'
        )
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)
    
    @patch('summaries.services.load_workbook')
    def test_load_template_success(self, mock_load_workbook):
        """Тест успешной загрузки шаблона"""
        mock_workbook = Mock()
        mock_load_workbook.return_value = mock_workbook
        
        result = self.service._load_template()
        
        mock_load_workbook.assert_called_once_with(Path(self.template_path))
        self.assertEqual(result, mock_workbook)
    
    @patch('summaries.services.load_workbook')
    def test_load_template_error(self, mock_load_workbook):
        """Тест ошибки при загрузке шаблона"""
        mock_load_workbook.side_effect = Exception("Ошибка загрузки файла")
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._load_template()
        
        self.assertIn('Ошибка при загрузке шаблона Excel', str(context.exception))
    
    def test_get_target_worksheet_named_sheet(self):
        """Тест получения целевого листа по имени"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        
        result = self.service._get_target_worksheet(workbook)
        
        self.assertEqual(result.title, 'summary_template_sheet')
    
    def test_get_target_worksheet_first_sheet(self):
        """Тест получения первого листа, если именованный не найден"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'other_sheet'
        
        result = self.service._get_target_worksheet(workbook)
        
        self.assertEqual(result.title, 'other_sheet')
    
    def test_get_target_worksheet_no_sheets(self):
        """Тест ошибки при отсутствии листов в книге"""
        workbook = Mock()
        workbook.sheetnames = []
        workbook.worksheets = []
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._get_target_worksheet(workbook)
        
        self.assertIn('не найдено ни одного рабочего листа', str(context.exception))
    
    def test_set_merged_cell_value_success(self):
        """Тест успешной записи значения в ячейку"""
        workbook = Workbook()
        worksheet = workbook.active
        
        self.service._set_merged_cell_value(worksheet, 'A1', 'Тестовое значение')
        
        self.assertEqual(worksheet['A1'].value, 'Тестовое значение')
    
    def test_set_merged_cell_value_error(self):
        """Тест ошибки при записи в ячейку"""
        # Создаем мок worksheet, который вызывает исключение при обращении к ячейке
        worksheet = Mock()
        # Настраиваем мок для вызова исключения при обращении к ячейке
        mock_cell = Mock()
        worksheet.__getitem__ = Mock(side_effect=Exception("Ошибка доступа к ячейке"))
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._set_merged_cell_value(worksheet, 'A1', 'Тестовое значение')
        
        self.assertIn('Ошибка при записи в ячейку A1', str(context.exception))
    
    def test_fill_template_data_success(self):
        """Тест успешного заполнения шаблона данными"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        
        self.service._fill_template_data(workbook, self.insurance_summary)
        
        # Проверяем, что данные записались в правильные ячейки
        self.assertEqual(worksheet['C1'].value, 'ТС-19827-ГА-ВН')
        self.assertEqual(worksheet['C2'].value, 'б/у грузовой тягач седельный SCANIA R440A4X2NA')
        self.assertEqual(worksheet['C3'].value, 'ООО "ТЕСТОВАЯ КОМПАНИЯ"')
    
    @patch.object(ExcelExportService, '_get_target_worksheet')
    def test_fill_template_data_worksheet_error(self, mock_get_worksheet):
        """Тест ошибки при получении рабочего листа"""
        mock_get_worksheet.side_effect = Exception("Ошибка получения листа")
        
        workbook = Mock()
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._fill_template_data(workbook, self.insurance_summary)
        
        self.assertIn('Ошибка при заполнении данных в Excel', str(context.exception))


class ExcelExportServiceIntegrationTests(TestCase):
    """Интеграционные тесты для ExcelExportService"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        # Создаем Excel файл с именованным листом
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        workbook.save(self.template_path)
        
        self.service = ExcelExportService(self.template_path)
        
        # Создаем тестовые данные
        self.insurance_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        self.insurance_summary = InsuranceSummary.objects.create(
            request=self.insurance_request,
            status='ready'
        )
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)
    
    def test_generate_summary_excel_success(self):
        """Тест успешной генерации Excel файла"""
        result = self.service.generate_summary_excel(self.insurance_summary)
        
        # Проверяем, что результат - это BytesIO
        self.assertIsInstance(result, BytesIO)
        
        # Проверяем, что файл не пустой
        self.assertGreater(len(result.getvalue()), 0)
        
        # Проверяем, что позиция указателя установлена в начало
        self.assertEqual(result.tell(), 0)
    
    def test_generate_summary_excel_invalid_data(self):
        """Тест генерации Excel файла с невалидными данными"""
        # Создаем свод с невалидными данными
        invalid_request = InsuranceRequest.objects.create(
            client_name='',  # Пустое название клиента
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        invalid_summary = InsuranceSummary.objects.create(
            request=invalid_request,
            status='ready'
        )
        
        with self.assertRaises(InvalidSummaryDataError):
            self.service.generate_summary_excel(invalid_summary)
    
    @patch.object(ExcelExportService, '_load_template')
    def test_generate_summary_excel_template_error(self, mock_load_template):
        """Тест обработки ошибки загрузки шаблона"""
        mock_load_template.side_effect = Exception("Ошибка загрузки шаблона")
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service.generate_summary_excel(self.insurance_summary)
        
        self.assertIn('Ошибка при генерации Excel-файла', str(context.exception))
    
    @patch.object(ExcelExportService, '_fill_template_data')
    def test_generate_summary_excel_fill_error(self, mock_fill_data):
        """Тест обработки ошибки заполнения данных"""
        mock_fill_data.side_effect = Exception("Ошибка заполнения данных")
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service.generate_summary_excel(self.insurance_summary)
        
        self.assertIn('Ошибка при генерации Excel-файла', str(context.exception))


class ExcelExportServiceFactoryTests(TestCase):
    """Тесты фабричной функции get_excel_export_service"""
    
    @patch('summaries.services.settings')
    def test_get_excel_export_service_with_custom_path(self, mock_settings):
        """Тест создания сервиса с пользовательским путем к шаблону"""
        # Создаем временный файл
        temp_dir = tempfile.mkdtemp()
        template_path = os.path.join(temp_dir, 'custom_template.xlsx')
        
        workbook = Workbook()
        workbook.save(template_path)
        
        try:
            # Настраиваем мок settings
            mock_settings.SUMMARY_TEMPLATE_PATH = template_path
            
            from summaries.services import get_excel_export_service
            service = get_excel_export_service()
            
            self.assertIsInstance(service, ExcelExportService)
            self.assertEqual(service.template_path, Path(template_path))
        
        finally:
            # Очистка
            if os.path.exists(template_path):
                os.remove(template_path)
            os.rmdir(temp_dir)
    
    @patch('summaries.services.settings')
    def test_get_excel_export_service_with_default_path(self, mock_settings):
        """Тест создания сервиса с путем по умолчанию"""
        # Создаем временный файл для пути по умолчанию
        temp_dir = tempfile.mkdtemp()
        default_template_path = os.path.join(temp_dir, 'summary_template.xlsx')
        
        workbook = Workbook()
        workbook.save(default_template_path)
        
        try:
            # Настраиваем мок settings без SUMMARY_TEMPLATE_PATH
            mock_settings.BASE_DIR = Path(temp_dir)
            del mock_settings.SUMMARY_TEMPLATE_PATH  # Удаляем атрибут, если он есть
            
            # Мокаем getattr для возврата пути по умолчанию
            with patch('summaries.services.getattr') as mock_getattr:
                mock_getattr.return_value = default_template_path
                
                from summaries.services import get_excel_export_service
                service = get_excel_export_service()
                
                self.assertIsInstance(service, ExcelExportService)
        
        finally:
            # Очистка
            if os.path.exists(default_template_path):
                os.remove(default_template_path)
            os.rmdir(temp_dir)
    
    @patch('summaries.services.ExcelExportService')
    def test_get_excel_export_service_creation_error(self, mock_service_class):
        """Тест обработки ошибки создания сервиса"""
        mock_service_class.side_effect = Exception("Ошибка создания сервиса")
        
        from summaries.services import get_excel_export_service, ExcelExportServiceError
        
        with self.assertRaises(ExcelExportServiceError) as context:
            get_excel_export_service()
        
        self.assertIn('Ошибка при создании ExcelExportService', str(context.exception))


class ExcelExportServiceErrorHandlingTests(TestCase):
    """Тесты обработки различных типов ошибок в ExcelExportService"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        workbook = Workbook()
        workbook.save(self.template_path)
        
        self.service = ExcelExportService(self.template_path)
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)
    
    def test_template_not_found_error_inheritance(self):
        """Тест наследования TemplateNotFoundError от ExcelExportServiceError"""
        self.assertTrue(issubclass(TemplateNotFoundError, ExcelExportServiceError))
    
    def test_invalid_summary_data_error_inheritance(self):
        """Тест наследования InvalidSummaryDataError от ExcelExportServiceError"""
        self.assertTrue(issubclass(InvalidSummaryDataError, ExcelExportServiceError))
    
    def test_excel_export_service_error_inheritance(self):
        """Тест наследования ExcelExportServiceError от Exception"""
        self.assertTrue(issubclass(ExcelExportServiceError, Exception))
    
    def test_error_messages_are_strings(self):
        """Тест что сообщения об ошибках являются строками"""
        try:
            ExcelExportService('/nonexistent/path')
        except TemplateNotFoundError as e:
            self.assertIsInstance(str(e), str)
            self.assertGreater(len(str(e)), 0)
        
        try:
            raise InvalidSummaryDataError("Тестовая ошибка")
        except InvalidSummaryDataError as e:
            self.assertIsInstance(str(e), str)
            self.assertEqual(str(e), "Тестовая ошибка")
        
        try:
            raise ExcelExportServiceError("Общая ошибка сервиса")
        except ExcelExportServiceError as e:
            self.assertIsInstance(str(e), str)
            self.assertEqual(str(e), "Общая ошибка сервиса")