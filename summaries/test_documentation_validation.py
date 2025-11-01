"""
Тесты для валидации актуальности справочной документации
"""
from django.test import TestCase
from unittest.mock import Mock, patch
import tempfile
import os
from openpyxl import Workbook
from core.excel_utils import ExcelReader


class DocumentationValidationTests(TestCase):
    """Тесты для проверки соответствия документации реальному коду"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный Excel файл для тестирования
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()
        
        # Создаем тестовый Excel файл
        workbook = Workbook()
        sheet = workbook.active
        
        # Заполняем тестовыми данными
        sheet['D7'] = 'Тестовый клиент'
        sheet['D9'] = '1234567890'
        sheet['C45'] = 'КАСКО C/E данные'  # Для тестирования КАСКО кат. C/E
        sheet['C44'] = 'Перевозка данные'  # Для тестирования перевозки
        sheet['C48'] = 'СМР данные'       # Для тестирования СМР
        
        # Данные для франшизы (юр.лицо)
        sheet['D29'] = 'Без франшизы'
        sheet['E29'] = ''
        sheet['F29'] = ''
        
        workbook.save(self.temp_file.name)
        workbook.close()
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
    
    def test_casco_ce_row_45_validation(self):
        """Тест валидации строки 45 для КАСКО кат. C/E"""
        # Тест для юр.лица
        reader = ExcelReader(
            self.temp_file.name, 
            application_type='legal_entity',
            application_format='casco_equipment'
        )
        
        # Проверяем, что метод _determine_casco_ce_openpyxl использует строку 45
        with patch.object(reader, '_get_cell_with_adjustment_openpyxl') as mock_get_cell:
            mock_get_cell.return_value = 'КАСКО C/E данные'
            
            # Создаем мок листа
            mock_sheet = Mock()
            result = reader._determine_casco_ce_openpyxl(mock_sheet)
            
            # Проверяем, что метод вызывался с правильными параметрами
            calls = mock_get_cell.call_args_list
            row_45_calls = [call for call in calls if call[0][2] == 45]
            
            self.assertTrue(len(row_45_calls) > 0, "Метод должен проверять строку 45")
            self.assertTrue(result, "Должен вернуть True при наличии данных в строке 45")
    
    def test_transportation_parameter_c44_validation(self):
        """Тест валидации ячейки C44 для параметра перевозки"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='legal_entity',
            application_format='property'
        )
        
        # Проверяем, что метод использует ячейку C44
        with patch.object(reader, '_get_cell_with_adjustment_openpyxl') as mock_get_cell:
            mock_get_cell.return_value = 'Перевозка данные'
            
            mock_sheet = Mock()
            result = reader._detect_transportation_parameter_openpyxl(mock_sheet)
            
            # Проверяем, что метод вызывался с правильными параметрами
            mock_get_cell.assert_called_with(mock_sheet, 'C', 44)
            self.assertTrue(result, "Должен вернуть True при наличии данных в C44")
    
    def test_construction_work_parameter_c48_validation(self):
        """Тест валидации ячейки C48 для параметра СМР"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='legal_entity',
            application_format='property'
        )
        
        # Проверяем, что метод использует ячейку C48
        with patch.object(reader, '_get_cell_with_adjustment_openpyxl') as mock_get_cell:
            mock_get_cell.return_value = 'СМР данные'
            
            mock_sheet = Mock()
            result = reader._detect_construction_work_parameter_openpyxl(mock_sheet)
            
            # Проверяем, что метод вызывался с правильными параметрами
            mock_get_cell.assert_called_with(mock_sheet, 'C', 48)
            self.assertTrue(result, "Должен вернуть True при наличии данных в C48")
    
    def test_ip_row_offset_logic_validation(self):
        """Тест валидации логики смещения строк для ИП"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='individual_entrepreneur',
            application_format='casco_equipment'
        )
        
        # Тестируем смещение для строк > 8
        adjusted_row_45 = reader._get_adjusted_row(45)
        self.assertEqual(adjusted_row_45, 46, "Строка 45 должна смещаться на +1 для ИП")
        
        adjusted_row_44 = reader._get_adjusted_row(44)
        self.assertEqual(adjusted_row_44, 45, "Строка 44 должна смещаться на +1 для ИП")
        
        adjusted_row_48 = reader._get_adjusted_row(48)
        self.assertEqual(adjusted_row_48, 49, "Строка 48 должна смещаться на +1 для ИП")
        
        # Тестируем отсутствие смещения для строк <= 8
        adjusted_row_7 = reader._get_adjusted_row(7)
        self.assertEqual(adjusted_row_7, 7, "Строка 7 не должна смещаться для ИП")
    
    def test_franchise_type_values_validation(self):
        """Тест валидации значений типов франшизы"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='legal_entity',
            application_format='casco_equipment'
        )
        
        # Тестируем возможные значения франшизы
        with patch.object(reader, '_get_cell_with_adjustment_openpyxl') as mock_get_cell:
            mock_sheet = Mock()
            
            # Тест 'none' - только D29 заполнена
            mock_get_cell.side_effect = lambda sheet, col, row: {
                ('D', 29): 'Без франшизы',
                ('E', 29): None,
                ('F', 29): None
            }.get((col, row))
            
            result = reader._determine_franchise_type(mock_sheet, is_openpyxl=True)
            self.assertEqual(result, 'none', "Должен вернуть 'none' когда только D29 заполнена")
            
            # Тест 'with_franchise' - только E29 или F29 заполнены
            mock_get_cell.side_effect = lambda sheet, col, row: {
                ('D', 29): None,
                ('E', 29): 'С франшизой',
                ('F', 29): None
            }.get((col, row))
            
            result = reader._determine_franchise_type(mock_sheet, is_openpyxl=True)
            self.assertEqual(result, 'with_franchise', "Должен вернуть 'with_franchise' когда только E29 заполнена")
            
            # Тест 'both_variants' - D29 и (E29 или F29) заполнены
            mock_get_cell.side_effect = lambda sheet, col, row: {
                ('D', 29): 'Без франшизы',
                ('E', 29): 'С франшизой',
                ('F', 29): None
            }.get((col, row))
            
            result = reader._determine_franchise_type(mock_sheet, is_openpyxl=True)
            self.assertEqual(result, 'both_variants', "Должен вернуть 'both_variants' когда D29 и E29 заполнены")
    
    def test_additional_parameters_cells_validation(self):
        """Тест валидации ячеек для дополнительных параметров КАСКО"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='legal_entity',
            application_format='casco_equipment'
        )
        
        # Проверяем, что метод извлечения дополнительных параметров существует
        self.assertTrue(
            hasattr(reader, '_extract_casco_additional_parameters_openpyxl'),
            "Должен существовать метод извлечения дополнительных параметров КАСКО"
        )
        
        # Проверяем, что метод возвращает словарь с ожидаемыми ключами
        with patch.object(reader, '_get_cell_with_adjustment_openpyxl') as mock_get_cell:
            mock_get_cell.return_value = 'Тестовое значение'
            mock_sheet = Mock()
            
            result = reader._extract_casco_additional_parameters_openpyxl(mock_sheet)
            
            # Проверяем наличие ожидаемых полей
            expected_fields = [
                'key_completeness', 'pts_psm', 'bank_creditor',
                'usage_purposes', 'telematics_complex', 'insurance_territory'
            ]
            
            for field in expected_fields:
                self.assertIn(field, result, f"Поле '{field}' должно присутствовать в дополнительных параметрах")
    
    def test_property_format_specific_parameters(self):
        """Тест специфичных параметров для формата 'имущество'"""
        reader = ExcelReader(
            self.temp_file.name,
            application_type='legal_entity',
            application_format='property'
        )
        
        # Проверяем, что для формата 'имущество' автозапуск всегда False
        data = reader.read_insurance_request()
        self.assertFalse(data['has_autostart'], "Автозапуск должен быть False для формата 'имущество'")
        
        # Проверяем, что КАСКО кат. C/E всегда False для формата 'имущество'
        self.assertFalse(data['has_casco_ce'], "КАСКО кат. C/E должен быть False для формата 'имущество'")
        
        # Проверяем, что параметры перевозки и СМР могут быть определены автоматически
        self.assertIn('has_transportation', data, "Должен присутствовать параметр перевозки")
        self.assertIn('has_construction_work', data, "Должен присутствовать параметр СМР")


class TechnicalAccuracyTests(TestCase):
    """Тесты технической точности документированной информации"""
    
    def test_documented_cell_mappings_accuracy(self):
        """Тест точности документированных соответствий ячеек"""
        # Создаем временный файл для тестирования
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            
            # Заполняем ячейки согласно документации
            sheet['D7'] = 'Тестовый клиент'  # Название клиента
            sheet['D9'] = '1234567890'       # ИНН
            sheet['M24'] = 'да'              # Автозапуск
            sheet['C45'] = 'КАСКО C/E'       # КАСКО кат. C/E (строка 45)
            sheet['C44'] = 'Перевозка'       # Перевозка (C44)
            sheet['C48'] = 'СМР'             # СМР (C48)
            
            workbook.save(temp_file.name)
            workbook.close()
            
            # Тестируем извлечение данных
            reader = ExcelReader(
                temp_file.name,
                application_type='legal_entity',
                application_format='casco_equipment'
            )
            
            data = reader.read_insurance_request()
            
            # Проверяем правильность извлечения
            self.assertEqual(data['client_name'], 'Тестовый клиент')
            self.assertEqual(data['inn'], '1234567890')
            self.assertTrue(data['has_autostart'])  # M24 = 'да'
            self.assertTrue(data['has_casco_ce'])   # C45 содержит данные
            
        finally:
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
    
    def test_ip_vs_legal_entity_differences(self):
        """Тест различий в обработке заявок от ИП и юр.лиц"""
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            
            # Заполняем данные для тестирования смещения
            sheet['D7'] = 'Клиент'
            sheet['D30'] = 'Франшиза ИП'  # Для ИП (смещение +1 от D29)
            sheet['C45'] = 'КАСКО ИП'     # Для ИП (смещение +1 от C44)
            
            workbook.save(temp_file.name)
            workbook.close()
            
            # Тестируем для ИП
            reader_ip = ExcelReader(
                temp_file.name,
                application_type='individual_entrepreneur',
                application_format='casco_equipment'
            )
            
            # Тестируем для юр.лица
            reader_legal = ExcelReader(
                temp_file.name,
                application_type='legal_entity',
                application_format='casco_equipment'
            )
            
            # Проверяем смещение строк
            self.assertEqual(reader_ip._get_adjusted_row(29), 30, "Строка 29 должна смещаться на +1 для ИП")
            self.assertEqual(reader_legal._get_adjusted_row(29), 29, "Строка 29 не должна смещаться для юр.лица")
            
        finally:
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
    
    def test_examples_correctness(self):
        """Тест корректности примеров в документации"""
        # Проверяем, что примеры соответствуют реальной логике
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            
            # Пример из документации: заполнение ячеек
            sheet['D7'] = 'ООО "Тестовая компания"'
            sheet['D9'] = '7701234567'
            sheet['D29'] = '0'        # Без франшизы
            sheet['E29'] = ''         # Пустая
            sheet['F29'] = ''         # Пустая
            sheet['C44'] = 'да'       # Перевозка
            sheet['C48'] = ''         # СМР отсутствует
            
            workbook.save(temp_file.name)
            workbook.close()
            
            reader = ExcelReader(
                temp_file.name,
                application_type='legal_entity',
                application_format='property'
            )
            
            data = reader.read_insurance_request()
            
            # Проверяем соответствие примеру
            self.assertEqual(data['client_name'], 'ООО "Тестовая компания"')
            self.assertEqual(data['inn'], '7701234567')
            self.assertEqual(data['franchise_type'], 'none')  # Только D29 заполнена
            self.assertTrue(data['has_transportation'])       # C44 = 'да'
            self.assertFalse(data['has_construction_work'])   # C48 пустая
            
        finally:
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)