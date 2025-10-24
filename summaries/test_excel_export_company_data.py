"""
Unit тесты для новых методов ExcelExportService, связанных с заполнением данных компаний
"""

import os
import tempfile
from decimal import Decimal
from unittest.mock import Mock, patch, MagicMock, call
from copy import copy

from django.test import TestCase
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Fill, Alignment, Protection

from insurance_requests.models import InsuranceRequest
from summaries.models import InsuranceSummary, InsuranceOffer
from summaries.services.excel_services import (
    ExcelExportService,
    ExcelExportServiceError,
    InvalidSummaryDataError
)


class ExcelExportServiceCompanyDataTests(TestCase):
    """Тесты для методов работы с данными компаний"""
    
    def setUp(self):
        """Настройка тестовых данных"""
        # Создаем временный файл шаблона
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        workbook.save(self.template_path)
        
        self.service = ExcelExportService(self.template_path)
        
        # Создаем тестовые данные
        self.insurance_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='1234567890',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН'
        )
        
        self.insurance_summary = InsuranceSummary.objects.create(
            request=self.insurance_request,
            status='ready'
        )
        
        # Создаем тестовые предложения
        self.offer1 = InsuranceOffer.objects.create(
            summary=self.insurance_summary,
            company_name='Абсолют',
            insurance_year=1,
            insurance_sum=Decimal('1000000'),
            premium_with_franchise_1=Decimal('50000'),
            franchise_1=Decimal('10000'),
            payments_per_year_variant_1=1,
            installment_variant_1=False
        )
        
        self.offer2 = InsuranceOffer.objects.create(
            summary=self.insurance_summary,
            company_name='Абсолют',
            insurance_year=2,
            insurance_sum=Decimal('800000'),
            premium_with_franchise_1=Decimal('40000'),
            franchise_1=Decimal('8000'),
            premium_with_franchise_2=Decimal('45000'),
            franchise_2=Decimal('5000'),
            payments_per_year_variant_1=4,
            installment_variant_1=True,
            payments_per_year_variant_2=2,
            installment_variant_2=True
        )
        
        self.offer3 = InsuranceOffer.objects.create(
            summary=self.insurance_summary,
            company_name='ВСК',
            insurance_year=1,
            insurance_sum=Decimal('1000000'),
            premium_with_franchise_1=Decimal('55000'),
            franchise_1=Decimal('15000'),
            payments_per_year_variant_1=12,
            installment_variant_1=True,
            notes='Специальные условия'
        )
    
    def tearDown(self):
        """Очистка после тестов"""
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)


class GetCompaniesSortedDataTests(ExcelExportServiceCompanyDataTests):
    """Тесты для метода _get_companies_sorted_data"""
    
    def test_get_companies_sorted_data_success(self):
        """Тест успешного получения отсортированных данных компаний"""
        # Мокаем метод модели
        expected_data = {
            'Абсолют': [self.offer1, self.offer2],
            'ВСК': [self.offer3]
        }
        
        with patch.object(self.insurance_summary, 'get_offers_grouped_by_company', return_value=expected_data):
            result = self.service._get_companies_sorted_data(self.insurance_summary)
        
        self.assertEqual(result, expected_data)
    
    def test_get_companies_sorted_data_empty_result(self):
        """Тест получения пустых данных компаний"""
        with patch.object(self.insurance_summary, 'get_offers_grouped_by_company', return_value={}):
            result = self.service._get_companies_sorted_data(self.insurance_summary)
        
        self.assertEqual(result, {})
    
    def test_get_companies_sorted_data_exception_handling(self):
        """Тест обработки исключений при получении данных компаний"""
        with patch.object(self.insurance_summary, 'get_offers_grouped_by_company', side_effect=Exception("Database error")):
            with self.assertRaises(ExcelExportServiceError) as context:
                self.service._get_companies_sorted_data(self.insurance_summary)
        
        self.assertIn('Ошибка при получении данных компаний', str(context.exception))


class FillCompanyYearRowTests(ExcelExportServiceCompanyDataTests):
    """Тесты для метода _fill_company_year_row"""
    
    def setUp(self):
        super().setUp()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Настраиваем мок для методов форматирования
        self.format_patches = [
            patch.object(self.service, '_copy_row_styles'),
            patch.object(self.service, '_format_insurance_sum', return_value=Decimal('1000000')),
            patch.object(self.service, '_format_premium', return_value=Decimal('50000')),
            patch.object(self.service, '_format_franchise', return_value=Decimal('10000')),
            patch.object(self.service, '_format_installment_payments', return_value=1),
            patch.object(self.service, '_fill_insurance_rate'),
            patch.object(self.service, '_validate_numeric_data_before_write', return_value=True)
        ]
        
        self.mocks = [p.start() for p in self.format_patches]
    
    def tearDown(self):
        super().tearDown()
        for p in self.format_patches:
            p.stop()
    
    def test_fill_company_year_row_basic_data(self):
        """Тест заполнения основных данных строки"""
        self.service._fill_company_year_row(
            self.worksheet, 11, 'Абсолют', self.offer1, '1 год'
        )
        
        # Проверяем основные данные
        self.assertEqual(self.worksheet['A11'].value, 'Абсолют')
        self.assertEqual(self.worksheet['B11'].value, '1 год')
        self.assertEqual(self.worksheet['C11'].value, Decimal('1000000'))
        self.assertEqual(self.worksheet['F11'].value, Decimal('50000'))
        self.assertEqual(self.worksheet['G11'].value, Decimal('10000'))
        self.assertEqual(self.worksheet['H11'].value, 1)
    
    def test_fill_company_year_row_with_second_variant(self):
        """Тест заполнения данных с вторым вариантом предложения"""
        # Настраиваем мок для has_second_franchise_variant
        with patch.object(self.offer2, 'has_second_franchise_variant', return_value=True):
            self.service._fill_company_year_row(
                self.worksheet, 11, 'Абсолют', self.offer2, '2 год'
            )
        
        # Проверяем, что вызваны методы для второго варианта
        format_premium_calls = self.mocks[2].call_args_list
        self.assertEqual(len(format_premium_calls), 2)  # Вызовы для варианта 1 и 2
        self.assertEqual(format_premium_calls[0], call(self.offer2, 1))
        self.assertEqual(format_premium_calls[1], call(self.offer2, 2))
    
    def test_fill_company_year_row_with_notes(self):
        """Тест заполнения примечаний"""
        self.service._fill_company_year_row(
            self.worksheet, 11, 'ВСК', self.offer3, '1 год'
        )
        
        # Проверяем примечания
        self.assertEqual(self.worksheet['Q11'].value, 'Специальные условия')
    
    def test_fill_company_year_row_copies_styles_for_non_first_row(self):
        """Тест копирования стилей для строк, отличных от первой"""
        self.service._fill_company_year_row(
            self.worksheet, 11, 'Абсолют', self.offer1, '1 год'
        )
        
        # Проверяем, что _copy_row_styles был вызван
        self.mocks[0].assert_called_once_with(self.worksheet, 10, 11)
    
    def test_fill_company_year_row_no_style_copy_for_first_row(self):
        """Тест отсутствия копирования стилей для первой строки данных"""
        self.service._fill_company_year_row(
            self.worksheet, 10, 'Абсолют', self.offer1, '1 год'
        )
        
        # Проверяем, что _copy_row_styles НЕ был вызван
        self.mocks[0].assert_not_called()
    
    def test_fill_company_year_row_long_notes_truncation(self):
        """Тест обрезки длинных примечаний"""
        long_notes = 'А' * 1500  # Превышает MAX_NOTES_LENGTH (1000)
        offer_with_long_notes = Mock()
        offer_with_long_notes.notes = long_notes
        offer_with_long_notes.has_second_franchise_variant.return_value = False
        offer_with_long_notes.insurance_year = 1
        
        self.service._fill_company_year_row(
            self.worksheet, 11, 'Тест', offer_with_long_notes, '1 год'
        )
        
        # Проверяем, что примечания обрезаны
        result_notes = self.worksheet['Q11'].value
        self.assertIsNotNone(result_notes)
        self.assertLessEqual(len(result_notes), 1003)  # 1000 + "..."
        self.assertTrue(result_notes.endswith('...'))
    
    def test_fill_company_year_row_exception_handling(self):
        """Тест обработки исключений при заполнении строки"""
        # Настраиваем мок для вызова исключения
        self.mocks[1].side_effect = Exception("Format error")
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._fill_company_year_row(
                self.worksheet, 11, 'Абсолют', self.offer1, '1 год'
            )
        
        self.assertIn('Ошибка при заполнении строки 11', str(context.exception))


class FormatInstallmentPaymentsTests(ExcelExportServiceCompanyDataTests):
    """Тесты для метода _format_installment_payments"""
    
    def test_format_installment_payments_variant_1_no_installment(self):
        """Тест форматирования рассрочки для варианта 1 без рассрочки"""
        offer = Mock()
        offer.payments_per_year_variant_1 = 1
        offer.installment_variant_1 = False
        
        result = self.service._format_installment_payments(offer, 1)
        
        self.assertEqual(result, 1)
    
    def test_format_installment_payments_variant_1_with_installment(self):
        """Тест форматирования рассрочки для варианта 1 с рассрочкой"""
        offer = Mock()
        offer.payments_per_year_variant_1 = 4
        offer.installment_variant_1 = True
        
        result = self.service._format_installment_payments(offer, 1)
        
        self.assertEqual(result, 4)
    
    def test_format_installment_payments_variant_2_with_installment(self):
        """Тест форматирования рассрочки для варианта 2 с рассрочкой"""
        offer = Mock()
        offer.payments_per_year_variant_2 = 12
        offer.installment_variant_2 = True
        
        result = self.service._format_installment_payments(offer, 2)
        
        self.assertEqual(result, 12)
    
    def test_format_installment_payments_none_payments(self):
        """Тест обработки None в количестве платежей"""
        offer = Mock()
        offer.payments_per_year_variant_1 = None
        offer.installment_variant_1 = True
        
        result = self.service._format_installment_payments(offer, 1)
        
        self.assertEqual(result, 1)
    
    def test_format_installment_payments_invalid_variant(self):
        """Тест обработки некорректного номера варианта"""
        offer = Mock()
        
        with self.assertRaises(ValueError) as context:
            self.service._format_installment_payments(offer, 3)
        
        self.assertIn('Номер варианта должен быть 1 или 2', str(context.exception))
    
    def test_format_installment_payments_exception_handling(self):
        """Тест обработки исключений при форматировании рассрочки"""
        offer = Mock()
        offer.payments_per_year_variant_1 = Mock(side_effect=Exception("Attribute error"))
        
        result = self.service._format_installment_payments(offer, 1)
        
        # При ошибке должен возвращаться 1 (единовременный платеж)
        self.assertEqual(result, 1)


class CopySeparatorRowTests(ExcelExportServiceCompanyDataTests):
    """Тесты для метода _copy_separator_row"""
    
    def setUp(self):
        super().setUp()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Настраиваем исходную строку с данными
        self.worksheet['A9'].value = 'Separator'
        self.worksheet['B9'].value = 'Data'
        
        # Мокаем _copy_row_styles
        self.copy_styles_patcher = patch.object(self.service, '_copy_row_styles')
        self.mock_copy_styles = self.copy_styles_patcher.start()
    
    def tearDown(self):
        super().tearDown()
        self.copy_styles_patcher.stop()
    
    def test_copy_separator_row_success(self):
        """Тест успешного копирования строки-разделителя"""
        self.service._copy_separator_row(self.worksheet, 9, 12)
        
        # Проверяем, что стили скопированы
        self.mock_copy_styles.assert_called_once_with(self.worksheet, 9, 12)
        
        # Проверяем, что содержимое скопировано
        self.assertEqual(self.worksheet['A12'].value, 'Separator')
        self.assertEqual(self.worksheet['B12'].value, 'Data')
    
    def test_copy_separator_row_with_comments(self):
        """Тест копирования строки с комментариями"""
        # Добавляем комментарий к исходной ячейке
        from openpyxl.comments import Comment
        self.worksheet['A9'].comment = Comment('Test comment', 'Author')
        
        self.service._copy_separator_row(self.worksheet, 9, 12)
        
        # Проверяем, что комментарий скопирован
        self.assertIsNotNone(self.worksheet['A12'].comment)
    
    def test_copy_separator_row_exception_handling(self):
        """Тест обработки исключений при копировании строки-разделителя"""
        # Настраиваем мок для вызова исключения
        self.mock_copy_styles.side_effect = Exception("Style copy error")
        
        with self.assertRaises(ExcelExportServiceError) as context:
            self.service._copy_separator_row(self.worksheet, 9, 12)
        
        self.assertIn('Ошибка при копировании строки-разделителя', str(context.exception))


class CopyRowStylesTests(ExcelExportServiceCompanyDataTests):
    """Тесты для метода _copy_row_styles"""
    
    def setUp(self):
        super().setUp()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Настраиваем исходную строку со стилями
        source_cell = self.worksheet['A10']
        source_cell.font = Font(name='Arial', size=12, bold=True)
        source_cell.border = Border()
        source_cell.fill = Fill()
        source_cell.alignment = Alignment(horizontal='center')
        source_cell.protection = Protection(locked=True)
        source_cell.number_format = '0.00'
    
    def test_copy_row_styles_success(self):
        """Тест успешного копирования стилей строки"""
        self.service._copy_row_styles(self.worksheet, 10, 11)
        
        # Проверяем, что стили скопированы
        target_cell = self.worksheet['A11']
        self.assertEqual(target_cell.font.name, 'Arial')
        self.assertEqual(target_cell.font.size, 12)
        self.assertTrue(target_cell.font.bold)
        self.assertEqual(target_cell.alignment.horizontal, 'center')
        self.assertTrue(target_cell.protection.locked)
        self.assertEqual(target_cell.number_format, '0.00')
    
    def test_copy_row_styles_row_dimensions(self):
        """Тест копирования свойств строки (высота, скрытость)"""
        # Настраиваем свойства исходной строки
        self.worksheet.row_dimensions[10].height = 25
        self.worksheet.row_dimensions[10].hidden = True
        self.worksheet.row_dimensions[10].outline_level = 1
        
        self.service._copy_row_styles(self.worksheet, 10, 11)
        
        # Проверяем, что свойства строки скопированы
        target_row = self.worksheet.row_dimensions[11]
        self.assertEqual(target_row.height, 25)
        self.assertTrue(target_row.hidden)
        self.assertEqual(target_row.outline_level, 1)
    
    def test_copy_row_styles_no_styles(self):
        """Тест копирования из ячейки без стилей"""
        # Создаем ячейку без стилей
        source_cell = self.worksheet['B10']
        
        # Метод не должен вызывать исключений
        try:
            self.service._copy_row_styles(self.worksheet, 10, 11)
        except Exception as e:
            self.fail(f"Копирование стилей не должно вызывать исключение: {e}")
    
    def test_copy_row_styles_partial_failure(self):
        """Тест частичного сбоя при копировании стилей"""
        # Создаем мок ячейки, которая вызывает исключение при копировании стилей
        with patch.object(self.worksheet, '__getitem__') as mock_getitem:
            # Настраиваем мок для возврата проблемной ячейки для определенных координат
            def side_effect(key):
                if key == 'B10' or key == 'B11':
                    mock_cell = Mock()
                    mock_cell.has_style = True
                    mock_cell.font = Mock()
                    mock_cell.font.copy.side_effect = Exception("Font copy error")
                    return mock_cell
                else:
                    return self.worksheet.cell(row=int(key[1:]), column=ord(key[0]) - ord('A') + 1)
            
            mock_getitem.side_effect = side_effect
            
            # Метод должен продолжить работу несмотря на ошибки
            try:
                self.service._copy_row_styles(self.worksheet, 10, 11)
            except Exception as e:
                self.fail(f"Метод должен обрабатывать частичные сбои: {e}")


class CopyFormulaToRowTests(ExcelExportServiceCompanyDataTests):
    """Тесты для методов копирования формул и вычисления тарифов"""
    
    def setUp(self):
        super().setUp()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Настраиваем исходную ячейку с формулой
        self.worksheet['E10'].value = '=IF(AND(C10<>0,F10<>0),F10/C10*100,"")'
        
        # Мокаем _create_rate_formula
        self.create_formula_patcher = patch.object(self.service, '_create_rate_formula')
        self.mock_create_formula = self.create_formula_patcher.start()
    
    def tearDown(self):
        super().tearDown()
        self.create_formula_patcher.stop()
    
    def test_copy_formula_to_row_with_formula(self):
        """Тест копирования формулы с адаптацией к новой строке"""
        # Настраиваем мок для data_type
        source_cell = self.worksheet['E10']
        source_cell.data_type = 'f'  # formula
        
        self.service._copy_formula_to_row(self.worksheet, 10, 11, 'E')
        
        # Проверяем, что формула адаптирована
        target_cell = self.worksheet['E11']
        expected_formula = '=IF(AND(C11<>0,F11<>0),F11/C11*100,"")'
        self.assertEqual(target_cell.value, expected_formula)
    
    def test_copy_formula_to_row_no_formula(self):
        """Тест обработки ячейки без формулы"""
        # Настраиваем ячейку без формулы
        source_cell = self.worksheet['E10']
        source_cell.data_type = 'n'  # number
        source_cell.value = 5.5
        
        self.service._copy_formula_to_row(self.worksheet, 10, 11, 'E')
        
        # Проверяем, что вызван метод создания формулы
        self.mock_create_formula.assert_called_once_with(self.worksheet, 11, 'E')
    
    def test_copy_formula_to_row_graceful_degradation(self):
        """Тест корректной обработки ошибок при копировании формулы"""
        # Метод должен обрабатывать исключения и не падать
        # Тестируем, что метод не выбрасывает исключения при проблемах с ячейками
        try:
            self.service._copy_formula_to_row(self.worksheet, 10, 11, 'E')
        except Exception as e:
            self.fail(f"Метод не должен выбрасывать исключение: {e}")
    
    def test_create_rate_formula_variant_1(self):
        """Тест создания формулы для тарифа варианта 1"""
        self.create_formula_patcher.stop()  # Останавливаем мок для этого теста
        
        self.service._create_rate_formula(self.worksheet, 11, 'E')
        
        expected_formula = '=IF(AND(C11<>0,F11<>0),F11/C11*100,"")'
        self.assertEqual(self.worksheet['E11'].value, expected_formula)
        
        self.create_formula_patcher.start()  # Возобновляем мок
    
    def test_create_rate_formula_variant_2(self):
        """Тест создания формулы для тарифа варианта 2"""
        self.create_formula_patcher.stop()  # Останавливаем мок для этого теста
        
        self.service._create_rate_formula(self.worksheet, 11, 'K')
        
        expected_formula = '=IF(AND(C11<>0,L11<>0),L11/C11*100,"")'
        self.assertEqual(self.worksheet['K11'].value, expected_formula)
        
        self.create_formula_patcher.start()  # Возобновляем мок
    
    def test_create_rate_formula_invalid_column(self):
        """Тест обработки некорректной колонки тарифа"""
        self.create_formula_patcher.stop()  # Останавливаем мок для этого теста
        
        with self.assertRaises(ValueError) as context:
            self.service._create_rate_formula(self.worksheet, 11, 'X')
        
        self.assertIn('Неизвестная колонка тарифа: X', str(context.exception))
        
        self.create_formula_patcher.start()  # Возобновляем мок
    
    def test_fill_insurance_rate_first_row_skip(self):
        """Тест пропуска заполнения тарифа для первой строки данных"""
        with patch.object(self.service, '_copy_formula_to_row') as mock_copy:
            self.service._fill_insurance_rate(self.worksheet, 10, 'E', 1)
        
        # Проверяем, что копирование НЕ было вызвано для строки 10
        mock_copy.assert_not_called()
    
    def test_fill_insurance_rate_other_rows(self):
        """Тест заполнения тарифа для других строк"""
        with patch.object(self.service, '_copy_formula_to_row') as mock_copy:
            self.service._fill_insurance_rate(self.worksheet, 11, 'E', 1)
        
        # Проверяем, что копирование было вызвано
        mock_copy.assert_called_once_with(self.worksheet, 10, 11, 'E')


class EdgeCaseHandlingTests(ExcelExportServiceCompanyDataTests):
    """Тесты для обработки граничных случаев"""
    
    def test_validate_companies_data_empty_input(self):
        """Тест обработки пустых данных компаний"""
        result = self.service._validate_companies_data({})
        
        self.assertEqual(result, {})
    
    def test_validate_companies_data_company_limit(self):
        """Тест ограничения количества компаний"""
        # Создаем данные с превышением лимита
        companies_data = {}
        for i in range(self.service.MAX_COMPANIES_LIMIT + 10):
            company_name = f'Компания_{i:03d}'
            companies_data[company_name] = [Mock()]
        
        result = self.service._validate_companies_data(companies_data)
        
        # Проверяем, что количество ограничено
        self.assertLessEqual(len(result), self.service.MAX_COMPANIES_LIMIT)
    
    def test_validate_companies_data_years_limit(self):
        """Тест ограничения количества лет на компанию"""
        # Создаем компанию с превышением лимита лет
        offers = []
        for i in range(self.service.MAX_YEARS_PER_COMPANY + 5):
            offer = Mock()
            offer.insurance_year = i + 1
            offers.append(offer)
        
        companies_data = {'Тест Компания': offers}
        
        with patch.object(self.service, '_validate_offer_data', return_value=True):
            result = self.service._validate_companies_data(companies_data)
        
        # Проверяем, что количество лет ограничено
        self.assertLessEqual(len(result['Тест Компания']), self.service.MAX_YEARS_PER_COMPANY)
    
    def test_validate_offer_data_missing_year(self):
        """Тест валидации предложения без года страхования"""
        offer = Mock()
        offer.insurance_year = None
        
        result = self.service._validate_offer_data(offer, 'Тест Компания')
        
        self.assertFalse(result)
    
    def test_validate_offer_data_invalid_year_range(self):
        """Тест валидации предложения с некорректным годом"""
        offer = Mock()
        offer.insurance_year = 25  # Превышает допустимый диапазон
        
        result = self.service._validate_offer_data(offer, 'Тест Компания')
        
        self.assertFalse(result)
    
    def test_validate_offer_data_missing_insurance_sum(self):
        """Тест валидации предложения без страховой суммы"""
        offer = Mock()
        offer.insurance_year = 1
        offer.insurance_sum = None
        
        result = self.service._validate_offer_data(offer, 'Тест Компания')
        
        self.assertFalse(result)
    
    def test_validate_offer_data_missing_premium(self):
        """Тест валидации предложения без премии"""
        offer = Mock()
        offer.insurance_year = 1
        offer.insurance_sum = Decimal('1000000')
        offer.premium_with_franchise_1 = None
        
        result = self.service._validate_offer_data(offer, 'Тест Компания')
        
        self.assertFalse(result)
    
    def test_validate_offer_data_valid_offer(self):
        """Тест валидации корректного предложения"""
        offer = Mock()
        offer.insurance_year = 1
        offer.insurance_sum = Decimal('1000000')
        offer.premium_with_franchise_1 = Decimal('50000')
        offer.franchise_1 = Decimal('10000')
        offer.has_second_franchise_variant.return_value = False
        
        result = self.service._validate_offer_data(offer, 'Тест Компания')
        
        self.assertTrue(result)
    
    def test_validate_numeric_data_before_write_valid_decimal(self):
        """Тест валидации корректного числового значения"""
        result = self.service._validate_numeric_data_before_write(
            Decimal('50000'), 'премия', 'Тест Компания', 1
        )
        
        self.assertTrue(result)
    
    def test_validate_numeric_data_before_write_none_value(self):
        """Тест валидации None значения"""
        result = self.service._validate_numeric_data_before_write(
            None, 'франшиза', 'Тест Компания', 1
        )
        
        self.assertTrue(result)  # None допустимо
    
    def test_validate_numeric_data_before_write_negative_value(self):
        """Тест валидации отрицательного значения"""
        result = self.service._validate_numeric_data_before_write(
            Decimal('-1000'), 'премия', 'Тест Компания', 1
        )
        
        self.assertFalse(result)
    
    def test_validate_numeric_data_before_write_excessive_value(self):
        """Тест валидации чрезмерно большого значения"""
        excessive_value = Decimal('2000000000')  # Превышает MAX_INSURANCE_SUM
        
        result = self.service._validate_numeric_data_before_write(
            excessive_value, 'страховая сумма', 'Тест Компания', 1
        )
        
        self.assertFalse(result)
    
    def test_validate_numeric_data_before_write_non_numeric(self):
        """Тест валидации нечислового значения"""
        result = self.service._validate_numeric_data_before_write(
            'не число', 'премия', 'Тест Компания', 1
        )
        
        self.assertFalse(result)


class NumericFormattingTests(ExcelExportServiceCompanyDataTests):
    """Тесты для методов форматирования числовых данных"""
    
    def test_format_numeric_value_valid_decimal(self):
        """Тест форматирования корректного Decimal значения"""
        result = self.service._format_numeric_value(Decimal('50000.123'), 'премия')
        
        self.assertEqual(result, Decimal('50000.12'))  # Округлено до 2 знаков
    
    def test_format_numeric_value_string_input(self):
        """Тест форматирования строкового числового значения"""
        result = self.service._format_numeric_value('50 000.50', 'премия')
        
        self.assertEqual(result, Decimal('50000.50'))
    
    def test_format_numeric_value_none_input(self):
        """Тест форматирования None значения"""
        result = self.service._format_numeric_value(None, 'премия')
        
        self.assertIsNone(result)
    
    def test_format_numeric_value_empty_string(self):
        """Тест форматирования пустой строки"""
        result = self.service._format_numeric_value('', 'премия')
        
        self.assertIsNone(result)
    
    def test_format_numeric_value_negative_to_zero(self):
        """Тест преобразования отрицательного значения в ноль"""
        result = self.service._format_numeric_value(Decimal('-1000'), 'премия')
        
        self.assertEqual(result, Decimal('0'))
    
    def test_format_numeric_value_excessive_to_max(self):
        """Тест ограничения чрезмерно большого значения"""
        excessive_value = Decimal('2000000000')
        result = self.service._format_numeric_value(excessive_value, 'премия')
        
        self.assertEqual(result, self.service.MAX_INSURANCE_SUM)
    
    def test_format_numeric_value_below_minimum(self):
        """Тест обработки значения ниже минимума"""
        small_value = Decimal('0.5')
        result = self.service._format_numeric_value(small_value, 'премия')
        
        self.assertEqual(result, self.service.MIN_INSURANCE_SUM)
    
    def test_format_insurance_sum(self):
        """Тест форматирования страховой суммы"""
        offer = Mock()
        offer.insurance_sum = Decimal('1000000')
        
        result = self.service._format_insurance_sum(offer)
        
        self.assertEqual(result, Decimal('1000000.00'))
    
    def test_format_premium_variant_1(self):
        """Тест форматирования премии для варианта 1"""
        offer = Mock()
        offer.premium_with_franchise_1 = Decimal('50000')
        
        result = self.service._format_premium(offer, 1)
        
        self.assertEqual(result, Decimal('50000.00'))
    
    def test_format_premium_variant_2(self):
        """Тест форматирования премии для варианта 2"""
        offer = Mock()
        offer.premium_with_franchise_2 = Decimal('45000')
        
        result = self.service._format_premium(offer, 2)
        
        self.assertEqual(result, Decimal('45000.00'))
    
    def test_format_franchise_variant_1_none_to_zero(self):
        """Тест форматирования франшизы варианта 1: None -> 0"""
        offer = Mock()
        offer.franchise_1 = None
        
        result = self.service._format_franchise(offer, 1)
        
        self.assertEqual(result, Decimal('0'))
    
    def test_format_franchise_variant_2_none_stays_none(self):
        """Тест форматирования франшизы варианта 2: None остается None"""
        offer = Mock()
        offer.franchise_2 = None
        
        result = self.service._format_franchise(offer, 2)
        
        self.assertIsNone(result)
    
    def test_format_franchise_valid_value(self):
        """Тест форматирования корректного значения франшизы"""
        offer = Mock()
        offer.franchise_1 = Decimal('10000')
        
        result = self.service._format_franchise(offer, 1)
        
        self.assertEqual(result, Decimal('10000.00'))


class IntegrationTests(ExcelExportServiceCompanyDataTests):
    """Интеграционные тесты для полного цикла заполнения данных компаний"""
    
    def test_fill_company_data_complete_flow(self):
        """Тест полного цикла заполнения данных компаний"""
        workbook = Workbook()
        worksheet = workbook.active
        
        # Мокаем все необходимые методы
        companies_data = {
            'Абсолют': [self.offer1, self.offer2],
            'ВСК': [self.offer3]
        }
        
        with patch.object(self.service, '_get_companies_sorted_data', return_value=companies_data), \
             patch.object(self.service, '_validate_companies_data', return_value=companies_data), \
             patch.object(self.service, '_fill_company_year_row') as mock_fill_row, \
             patch.object(self.service, '_copy_separator_row') as mock_separator:
            
            self.service._fill_company_data(workbook, self.insurance_summary)
        
        # Проверяем количество вызовов заполнения строк
        self.assertEqual(mock_fill_row.call_count, 3)  # 3 предложения
        
        # Проверяем количество вызовов разделителей
        self.assertEqual(mock_separator.call_count, 1)  # 1 разделитель между компаниями
    
    def test_fill_company_data_no_offers(self):
        """Тест обработки отсутствия предложений"""
        workbook = Workbook()
        
        with patch.object(self.service, '_get_companies_sorted_data', return_value={}), \
             patch.object(self.service, '_validate_companies_data', return_value={}), \
             patch.object(self.service, '_fill_company_year_row') as mock_fill_row:
            
            # Не должно вызывать исключений
            self.service._fill_company_data(workbook, self.insurance_summary)
        
        # Проверяем, что заполнение строк не вызывалось
        mock_fill_row.assert_not_called()
    
    def test_fill_company_data_row_limit_reached(self):
        """Тест достижения лимита строк"""
        workbook = Workbook()
        
        # Создаем много предложений для превышения лимита
        many_offers = []
        for i in range(50):
            offer = Mock()
            offer.insurance_year = i + 1
            many_offers.append(offer)
        
        companies_data = {'Тест Компания': many_offers}
        
        # Устанавливаем низкий лимит для теста
        original_limit = self.service.MAX_ROWS_LIMIT
        self.service.MAX_ROWS_LIMIT = 15
        
        try:
            with patch.object(self.service, '_get_companies_sorted_data', return_value=companies_data), \
                 patch.object(self.service, '_validate_companies_data', return_value=companies_data), \
                 patch.object(self.service, '_fill_company_year_row') as mock_fill_row:
                
                self.service._fill_company_data(workbook, self.insurance_summary)
            
            # Проверяем, что количество вызовов ограничено
            self.assertLess(mock_fill_row.call_count, 50)
        
        finally:
            # Восстанавливаем исходный лимит
            self.service.MAX_ROWS_LIMIT = original_limit
    
    def test_fill_company_data_exception_handling(self):
        """Тест обработки исключений в _fill_company_data"""
        workbook = Workbook()
        
        with patch.object(self.service, '_get_companies_sorted_data', side_effect=Exception("Data error")):
            with self.assertRaises(ExcelExportServiceError) as context:
                self.service._fill_company_data(workbook, self.insurance_summary)
        
        self.assertIn('Ошибка при заполнении данных компаний', str(context.exception))