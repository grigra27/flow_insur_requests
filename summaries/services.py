"""
Сервисы для работы со сводами предложений
"""

import logging
from io import BytesIO
from pathlib import Path
from typing import Optional
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from django.conf import settings

from .models import InsuranceSummary
from insurance_requests.models import InsuranceRequest


logger = logging.getLogger(__name__)


class ExcelExportServiceError(Exception):
    """Базовое исключение для ошибок ExcelExportService"""
    pass


class TemplateNotFoundError(ExcelExportServiceError):
    """Исключение для случаев, когда шаблон Excel не найден"""
    pass


class InvalidSummaryDataError(ExcelExportServiceError):
    """Исключение для случаев, когда данные свода некорректны"""
    pass


class ExcelExportService:
    """Сервис для генерации Excel-файлов сводов предложений"""
    
    def __init__(self, template_path: str):
        """
        Инициализация сервиса
        
        Args:
            template_path: Путь к файлу шаблона Excel
        """
        self.template_path = Path(template_path)
        self._validate_template()
    
    def _validate_template(self) -> None:
        """Проверяет доступность шаблона Excel"""
        if not self.template_path.exists():
            error_msg = f"Шаблон Excel не найден по пути: {self.template_path}"
            logger.error(error_msg)
            raise TemplateNotFoundError(error_msg)
        
        if not self.template_path.is_file():
            error_msg = f"Путь к шаблону не является файлом: {self.template_path}"
            logger.error(error_msg)
            raise TemplateNotFoundError(error_msg)
    
    def generate_summary_excel(self, summary: InsuranceSummary) -> BytesIO:
        """
        Генерирует Excel-файл для свода предложений
        
        Args:
            summary: Объект свода предложений
            
        Returns:
            BytesIO: Сгенерированный Excel-файл в памяти
            
        Raises:
            InvalidSummaryDataError: Если данные свода некорректны
            ExcelExportServiceError: При ошибках работы с Excel
        """
        logger.info(f"Начинаем генерацию Excel-файла для свода ID: {summary.id}")
        
        try:
            # Валидация данных свода
            self._validate_summary_data(summary)
            
            # Загрузка шаблона
            workbook = self._load_template()
            
            # Заполнение данными
            self._fill_template_data(workbook, summary)
            
            # Сохранение в память
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Excel-файл успешно сгенерирован для свода ID: {summary.id}")
            return excel_buffer
            
        except (InvalidSummaryDataError, TemplateNotFoundError):
            # Переброс известных исключений
            raise
        except Exception as e:
            error_msg = f"Ошибка при генерации Excel-файла для свода ID {summary.id}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _validate_summary_data(self, summary: InsuranceSummary) -> None:
        """
        Валидирует данные свода перед генерацией
        
        Args:
            summary: Объект свода предложений
            
        Raises:
            InvalidSummaryDataError: Если данные некорректны
        """
        missing_fields = []
        
        # Проверка наличия связанной заявки
        if not summary.request:
            missing_fields.append("связанная заявка")
        else:
            request = summary.request
            
            # Проверка обязательных полей заявки
            if not request.dfa_number or request.dfa_number.strip() == '':
                missing_fields.append("номер заявки (dfa_number)")
            
            if not request.vehicle_info or request.vehicle_info.strip() == '':
                missing_fields.append("информация о предмете лизинга (vehicle_info)")
            
            if not request.client_name or request.client_name.strip() == '':
                missing_fields.append("название клиента (client_name)")
        
        if missing_fields:
            error_msg = f"Отсутствуют обязательные данные: {', '.join(missing_fields)}"
            logger.error(f"Валидация данных свода ID {summary.id} не пройдена: {error_msg}")
            raise InvalidSummaryDataError(error_msg)
        
        logger.debug(f"Валидация данных свода ID {summary.id} успешно пройдена")
    
    def _load_template(self) -> Workbook:
        """
        Загружает шаблон Excel из файла
        
        Returns:
            Workbook: Загруженная книга Excel
            
        Raises:
            ExcelExportServiceError: При ошибках загрузки шаблона
        """
        try:
            logger.debug(f"Загружаем шаблон из файла: {self.template_path}")
            workbook = load_workbook(self.template_path)
            logger.debug("Шаблон успешно загружен")
            return workbook
        except Exception as e:
            error_msg = f"Ошибка при загрузке шаблона Excel: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _fill_template_data(self, workbook: Workbook, summary: InsuranceSummary) -> None:
        """
        Заполняет шаблон данными из свода
        
        Args:
            workbook: Книга Excel для заполнения
            summary: Объект свода предложений
            
        Raises:
            ExcelExportServiceError: При ошибках заполнения данных
        """
        try:
            logger.debug(f"Начинаем заполнение данных для свода ID: {summary.id}")
            
            # Получаем рабочий лист (используем первый лист или ищем по имени)
            worksheet = self._get_target_worksheet(workbook)
            
            request = summary.request
            
            # Заполнение данных согласно требованиям:
            # CDE1 - номер заявки
            self._set_merged_cell_value(worksheet, 'C1', request.dfa_number)
            logger.debug(f"Записан номер заявки в C1: {request.dfa_number}")
            
            # CDE2 - информация о предмете лизинга
            self._set_merged_cell_value(worksheet, 'C2', request.vehicle_info)
            logger.debug(f"Записана информация о предмете лизинга в C2: {request.vehicle_info[:50]}...")
            
            # CDE3 - название клиента
            self._set_merged_cell_value(worksheet, 'C3', request.client_name)
            logger.debug(f"Записано название клиента в C3: {request.client_name}")
            
            logger.info(f"Данные успешно заполнены для свода ID: {summary.id}")
            
        except Exception as e:
            error_msg = f"Ошибка при заполнении данных в Excel: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _get_target_worksheet(self, workbook: Workbook):
        """
        Получает целевой рабочий лист для заполнения
        
        Args:
            workbook: Книга Excel
            
        Returns:
            Worksheet: Рабочий лист для заполнения
            
        Raises:
            ExcelExportServiceError: Если не удается найти подходящий лист
        """
        try:
            # Сначала пытаемся найти лист с именем "summary_template_sheet"
            if 'summary_template_sheet' in workbook.sheetnames:
                logger.debug("Найден лист 'summary_template_sheet'")
                return workbook['summary_template_sheet']
            
            # Если не найден, используем первый лист
            if workbook.worksheets:
                worksheet = workbook.worksheets[0]
                logger.debug(f"Используем первый лист: {worksheet.title}")
                return worksheet
            
            # Если нет листов вообще
            raise ExcelExportServiceError("В шаблоне Excel не найдено ни одного рабочего листа")
            
        except Exception as e:
            error_msg = f"Ошибка при получении рабочего листа: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e
    
    def _set_merged_cell_value(self, worksheet, cell_address: str, value: str) -> None:
        """
        Устанавливает значение в ячейку (может быть объединенной)
        
        Args:
            worksheet: Рабочий лист
            cell_address: Адрес ячейки (например, 'C1')
            value: Значение для записи
        """
        try:
            cell = worksheet[cell_address]
            cell.value = value
            logger.debug(f"Установлено значение в ячейку {cell_address}: {value}")
        except Exception as e:
            error_msg = f"Ошибка при записи в ячейку {cell_address}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ExcelExportServiceError(error_msg) from e


def get_excel_export_service() -> ExcelExportService:
    """
    Фабричная функция для создания экземпляра ExcelExportService
    с настройками по умолчанию
    
    Returns:
        ExcelExportService: Настроенный экземпляр сервиса
        
    Raises:
        ExcelExportServiceError: Если не удается создать сервис
    """
    try:
        # Получаем путь к шаблону из настроек или используем значение по умолчанию
        template_path = getattr(
            settings, 
            'SUMMARY_TEMPLATE_PATH', 
            settings.BASE_DIR / 'templates' / 'summary_template.xlsx'
        )
        
        return ExcelExportService(str(template_path))
        
    except Exception as e:
        error_msg = f"Ошибка при создании ExcelExportService: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise ExcelExportServiceError(error_msg) from e