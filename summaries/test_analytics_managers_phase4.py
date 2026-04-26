"""Phase 4 — досье сотрудника и XLSX-экспорты."""
from __future__ import annotations

from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from insurance_requests.models import InsuranceRequest

from .models import InsuranceCompany, InsuranceOffer, InsuranceSummary
from .services import analytics_managers


class ManagerProfilePayloadTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(
            username='alice', password='p',
            first_name='Алиса', last_name='Иванова',
            email='alice@example.com',
        )
        InsuranceCompany.objects.get_or_create(
            name='Альфа', defaults={'display_name': 'Альфа', 'sort_order': 10})
        now = timezone.now()
        cls.now = now

        # Active request (uploaded), no summary
        active_req = InsuranceRequest.objects.create(
            client_name='Active', inn='1234567890',
            insurance_type='КАСКО', branch='Москва',
            created_by=cls.user, status='uploaded',
        )
        InsuranceRequest.objects.filter(pk=active_req.pk).update(
            created_at=now - timedelta(days=2),
        )

        # Overdue request
        ov = InsuranceRequest.objects.create(
            client_name='Overdue', inn='1234567890',
            insurance_type='КАСКО', branch='Москва',
            created_by=cls.user, status='uploaded',
        )
        InsuranceRequest.objects.filter(pk=ov.pk).update(
            created_at=now - timedelta(days=1),
            response_deadline=now - timedelta(hours=4),
        )

        # Accepted summary with company Альфа, branch Казань
        ac_req = InsuranceRequest.objects.create(
            client_name='Accept', inn='1234567890',
            insurance_type='КАСКО', branch='Казань',
            created_by=cls.user, status='emails_sent',
        )
        InsuranceRequest.objects.filter(pk=ac_req.pk).update(
            created_at=now - timedelta(days=20),
        )
        ac_sum = InsuranceSummary.objects.create(
            request=ac_req, status='completed_accepted',
            selected_company='Альфа', selected_franchise_variant=1,
        )
        InsuranceSummary.objects.filter(pk=ac_sum.pk).update(
            created_at=now - timedelta(days=18),
            sent_to_client_at=now - timedelta(days=15),
            completed_at=now - timedelta(days=10),
        )
        InsuranceOffer.objects.create(
            summary=ac_sum, company_name='Альфа',
            insurance_sum=Decimal('500000'), insurance_year=1,
            franchise_1=Decimal('0'),
            premium_with_franchise_1=Decimal('25000'),
        )

    def _filters(self, **overrides):
        f = analytics_managers.parse_filters({'period': '365'})
        f.update(overrides)
        return f

    def test_profile_returns_user_and_row(self):
        payload = analytics_managers.build_manager_profile_payload(self.user.pk, self._filters())
        self.assertFalse(payload.get('not_found'))
        self.assertEqual(payload['user'].pk, self.user.pk)
        self.assertIsNotNone(payload['row'])
        self.assertEqual(payload['row']['accepted'], 1)

    def test_profile_active_overdue_lists(self):
        payload = analytics_managers.build_manager_profile_payload(self.user.pk, self._filters())
        active_kinds = {item['kind'] for item in payload['active_items']}
        self.assertIn('request', active_kinds)
        self.assertGreaterEqual(len(payload['overdue']), 1)

    def test_profile_top_companies_and_branches(self):
        payload = analytics_managers.build_manager_profile_payload(self.user.pk, self._filters())
        company_names = [c['name'] for c in payload['top_companies']]
        self.assertIn('Альфа', company_names)
        branch_names = [b['name'] for b in payload['top_branches']]
        self.assertIn('Москва', branch_names)
        self.assertIn('Казань', branch_names)

    def test_profile_self_benchmark_present_with_dates(self):
        payload = analytics_managers.build_manager_profile_payload(self.user.pk, self._filters())
        # period=365 → даты заполнены
        self.assertIsNotNone(payload['self_benchmark'])

    def test_profile_timeline_uses_status_events(self):
        payload = analytics_managers.build_manager_profile_payload(self.user.pk, self._filters())
        # Сигналы создали StatusEvent для каждой созданной заявки и свода
        self.assertGreater(len(payload['timeline']), 0)
        # Проверяем формат
        evt = payload['timeline'][0]
        for key in ('changed_at', 'kind', 'target_label', 'to_status'):
            self.assertIn(key, evt)

    def test_profile_not_found_for_unknown_user(self):
        payload = analytics_managers.build_manager_profile_payload(999999, self._filters())
        self.assertTrue(payload.get('not_found'))


class XlsxExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username='exp', password='p', first_name='Экс')
        InsuranceCompany.objects.get_or_create(
            name='Альфа', defaults={'display_name': 'Альфа', 'sort_order': 10})
        now = timezone.now()
        for i in range(3):
            r = InsuranceRequest.objects.create(
                client_name=f'X{i}', inn='1234567890',
                insurance_type='КАСКО', branch='Москва',
                created_by=cls.user, status='emails_sent',
            )
            InsuranceRequest.objects.filter(pk=r.pk).update(
                created_at=now - timedelta(days=i + 1),
            )

    def test_overview_xlsx_has_expected_sheets(self):
        filters = analytics_managers.parse_filters({'period': '365'})
        output = analytics_managers.export_overview_xlsx(filters)
        wb = load_workbook(output)
        names = wb.sheetnames
        for expected in ('KPI', 'Сотрудники', 'Качество данных', 'Funnel'):
            self.assertIn(expected, names, msg=f'Missing sheet: {expected}')

    def test_dossier_xlsx_has_expected_sheets(self):
        filters = analytics_managers.parse_filters({'period': '365'})
        output = analytics_managers.export_manager_dossier_xlsx(self.user.pk, filters)
        wb = load_workbook(output)
        names = wb.sheetnames
        for expected in ('KPI', 'Активные', 'Просроченные', 'Зависшие', 'Топ СК и Филиалы', 'Активность'):
            self.assertIn(expected, names, msg=f'Missing sheet: {expected}')


class ExportEndpointSmokeTests(TestCase):
    def setUp(self):
        admin_group, _ = Group.objects.get_or_create(name='Администраторы')
        self.admin = User.objects.create_user(username='adm', password='p')
        self.admin.groups.add(admin_group)
        self.client = Client()
        self.client.login(username='adm', password='p')

    def test_overview_export_returns_xlsx(self):
        url = reverse('summaries:export_analytics_managers_widget')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml.sheet', response['Content-Type'])
        # Проверим, что это валидный xlsx
        wb = load_workbook(BytesIO(response.content))
        self.assertIn('KPI', wb.sheetnames)

    def test_dossier_export_returns_xlsx(self):
        # Пользователь без заявок — экспорт должен всё равно вернуть xlsx с KPI листом
        target = User.objects.create_user(username='target', password='p')
        url = reverse('summaries:export_analytics_manager_detail', kwargs={'user_id': target.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        self.assertIn('KPI', wb.sheetnames)
