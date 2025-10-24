"""
Тесты для функционала объединения ячеек с суммированием премий
"""

import unittest
from unittest.mock import Mock, MagicMock, patch
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Alignment

from summaries.services.excel_services import ExcelExportService


class PremiumSummaryFunctionalityTests(unittest.TestCase):
    """Тесты для методов объединения ячеек с суммированием премий"""
    
    def setUp(self):
        """Настройка тестового окружения"""
        # Создаем сервис с мокированной валидацией шаблона
        with patch.object(ExcelExportService, '_validate_template'):
            self.service = ExcelExportService('/fake/template/path.xlsx')
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Создаем тестовые предложения
        self.offers = []
        for i in range(3):
            offer = Mock()
            offer.insurance_year = i + 1
            offer.premium_with_franchise_1 = Decimal(f'{1000 + i * 100}')  # 1000, 1100, 1200
            offer.premium_with_franchise_2 = Decimal(f'{500 + i * 50}') if i < 2 else None  # 500, 550, None
            offer.has_second_franchise_variant.return_value = i < 2  # Только первые два
            self.offers.append(offer)
    
    def test_calculate_premium_sum_type_1(self):
        """Тест вычисления суммы премий типа 1"""
        result = self.service._calculate_premium_sum(self.offers, 1)
        
        # Ожидаем сумму: 1000 + 1100 + 1200 = 3300
        expected = Decimal('3300')
        self.assertEqual(result, expected)
    
    def test_calculate_premium_sum_type_2(self):
        """Тест вычисления суммы премий типа 2"""
        result = self.service._calculate_premium_sum(self.offers, 2)
        
        # Ожидаем сумму: 500 + 550 = 1050 (третье предложение не имеет премии-2)
        expected = Decimal('1050')
        self.assertEqual(result, expected)
    
    def test_calculate_premium_sum_no_valid_premiums(self):
        """Тест вычисления суммы когда нет валидных премий"""
        # Создаем предложения без премий-2
        offers_no_premium_2 = []
        for i in range(2):
            offer = Mock()
            offer.premium_with_franchise_2 = None
            offer.has_second_franchise_variant.return_value = False
            offers_no_premium_2.append(offer)
        
        with patch.object(self.service, '_format_premium', return_value=None):
            result = self.service._calculate_premium_sum(offers_no_premium_2, 2)
        
        self.assertIsNone(result)
    
    def test_calculate_premium_sum_with_errors(self):
        """Тест вычисления суммы с ошибками в данных"""
        # Создаем предложения с ошибками
        offers_with_errors = []
        for i in range(3):
            offer = Mock()
            offer.insurance_year = i + 1
            offers_with_errors.append(offer)
        
        # Мокаем _format_premium чтобы возвращать ошибки для некоторых предложений
        def mock_format_premium(offer, variant):
            if offer.insurance_year == 2:  # Второе предложение с ошибкой
                raise Exception("Format error")
            return Decimal('1000') if offer.insurance_year in [1, 3] else None
        
        with patch.object(self.service, '_format_premium', side_effect=mock_format_premium):
            result = self.service._calculate_premium_sum(offers_with_errors, 1)
        
        # Ожидаем сумму только валидных предложений: 1000 + 1000 = 2000
        expected = Decimal('2000')
        self.assertEqual(result, expected)
    
    def test_apply_premium_cell_formatting(self):
        """Тест применения форматирования к ячейкам премий"""
        # Настраиваем шаблонную ячейку с реальными стилями
        from openpyxl.styles import Font, Border, PatternFill, Protection
        template_cell = self.worksheet['I10']
        template_cell.font = Font(name='Arial', size=12)
        template_cell.border = Border()
        template_cell.fill = PatternFill()
        template_cell.number_format = '#,##0.00'
        template_cell.protection = Protection()
        
        # Применяем форматирование
        self.service._apply_premium_cell_formatting(self.worksheet, 'I12', 10)
        
        # Проверяем, что форматирование применено
        target_cell = self.worksheet['I12']
        self.assertIsNotNone(target_cell.alignment)
        self.assertEqual(target_cell.alignment.vertical, 'center')
    
    def test_apply_premium_cell_formatting_no_template_style(self):
        """Тест применения форматирования когда у шаблона нет стилей"""
        # Шаблонная ячейка без стилей (по умолчанию)
        template_cell = self.worksheet['I10']
        # Не устанавливаем никаких стилей
        
        # Применяем форматирование
        self.service._apply_premium_cell_formatting(self.worksheet, 'I12', 10)
        
        # Проверяем, что вертикальное выравнивание все равно применено
        target_cell = self.worksheet['I12']
        self.assertIsNotNone(target_cell.alignment)
        self.assertEqual(target_cell.alignment.vertical, 'center')
    
    def test_merge_premium_summary_cells_success(self):
        """Тест успешного объединения ячеек с суммами премий"""
        company_name = "Тест Компания"
        start_row = 10
        end_row = 12
        
        # Мокаем методы вычисления сумм
        with patch.object(self.service, '_calculate_premium_sum') as mock_calc, \
             patch.object(self.service, '_apply_premium_cell_formatting') as mock_format:
            
            mock_calc.side_effect = [Decimal('3000'), Decimal('1500')]  # премии-1 и премии-2
            
            self.service._merge_premium_summary_cells(
                self.worksheet, start_row, end_row, company_name, self.offers
            )
            
            # Проверяем, что методы были вызваны
            self.assertEqual(mock_calc.call_count, 2)
            mock_calc.assert_any_call(self.offers, 1)
            mock_calc.assert_any_call(self.offers, 2)
            
            # Проверяем, что форматирование применено к обеим ячейкам
            self.assertEqual(mock_format.call_count, 2)
            mock_format.assert_any_call(self.worksheet, 'I10', 10)
            mock_format.assert_any_call(self.worksheet, 'O10', 10)
            
            # Проверяем значения в ячейках
            self.assertEqual(self.worksheet['I10'].value, Decimal('3000'))
            self.assertEqual(self.worksheet['O10'].value, Decimal('1500'))
    
    def test_merge_premium_summary_cells_no_premium_2(self):
        """Тест объединения ячеек когда нет премий-2"""
        company_name = "Тест Компания"
        start_row = 10
        end_row = 12
        
        with patch.object(self.service, '_calculate_premium_sum') as mock_calc, \
             patch.object(self.service, '_apply_premium_cell_formatting') as mock_format:
            
            mock_calc.side_effect = [Decimal('3000'), None]  # есть премии-1, нет премий-2
            
            self.service._merge_premium_summary_cells(
                self.worksheet, start_row, end_row, company_name, self.offers
            )
            
            # Проверяем значения в ячейках
            self.assertEqual(self.worksheet['I10'].value, Decimal('3000'))
            self.assertIsNone(self.worksheet['O10'].value)  # Пустая ячейка для премий-2
            
            # Проверяем, что форматирование все равно применено
            self.assertEqual(mock_format.call_count, 2)
    
    def test_merge_premium_summary_cells_zero_premium_2(self):
        """Тест объединения ячеек когда премии-2 равны нулю"""
        company_name = "Тест Компания"
        start_row = 10
        end_row = 12
        
        with patch.object(self.service, '_calculate_premium_sum') as mock_calc, \
             patch.object(self.service, '_apply_premium_cell_formatting') as mock_format:
            
            mock_calc.side_effect = [Decimal('3000'), Decimal('0')]  # есть премии-1, премии-2 = 0
            
            self.service._merge_premium_summary_cells(
                self.worksheet, start_row, end_row, company_name, self.offers
            )
            
            # Проверяем, что ячейка O остается пустой при нулевой сумме
            self.assertEqual(self.worksheet['I10'].value, Decimal('3000'))
            self.assertIsNone(self.worksheet['O10'].value)
    
    def test_merge_premium_summary_cells_exception_handling(self):
        """Тест обработки исключений при объединении ячеек"""
        company_name = "Тест Компания"
        start_row = 10
        end_row = 12
        
        # Мокаем исключение при объединении ячеек
        with patch.object(self.worksheet, 'merge_cells', side_effect=Exception("Merge error")), \
             patch.object(self.service, '_fill_premium_summary_fallback') as mock_fallback:
            
            self.service._merge_premium_summary_cells(
                self.worksheet, start_row, end_row, company_name, self.offers
            )
            
            # Проверяем, что вызван fallback метод
            mock_fallback.assert_called_once_with(self.worksheet, start_row, end_row, self.offers)
    
    def test_fill_premium_summary_fallback(self):
        """Тест fallback метода для заполнения премий в отдельные ячейки"""
        start_row = 10
        end_row = 12
        
        with patch.object(self.service, '_format_premium') as mock_format:
            # Настраиваем возвращаемые значения для премий
            mock_format.side_effect = lambda offer, variant: {
                (self.offers[0], 1): Decimal('1000'),
                (self.offers[0], 2): Decimal('500'),
                (self.offers[1], 1): Decimal('1100'),
                (self.offers[1], 2): Decimal('550'),
                (self.offers[2], 1): Decimal('1200'),
                (self.offers[2], 2): None,
            }.get((offer, variant))
            
            self.service._fill_premium_summary_fallback(
                self.worksheet, start_row, end_row, self.offers
            )
            
            # Проверяем, что значения записаны в отдельные ячейки
            self.assertEqual(self.worksheet['I10'].value, Decimal('1000'))  # Первая строка, премия-1
            self.assertEqual(self.worksheet['O10'].value, Decimal('500'))   # Первая строка, премия-2
            self.assertEqual(self.worksheet['I11'].value, Decimal('1100'))  # Вторая строка, премия-1
            self.assertEqual(self.worksheet['O11'].value, Decimal('550'))   # Вторая строка, премия-2
            self.assertEqual(self.worksheet['I12'].value, Decimal('1200'))  # Третья строка, премия-1
            self.assertIsNone(self.worksheet['O12'].value)                  # Третья строка, нет премии-2
    
    def test_fill_premium_summary_fallback_with_unmerge(self):
        """Тест fallback метода с разъединением ячеек"""
        start_row = 10
        end_row = 12
        
        # Предварительно объединяем ячейки
        self.worksheet.merge_cells('I10:I12')
        self.worksheet.merge_cells('O10:O12')
        
        with patch.object(self.service, '_format_premium', return_value=Decimal('1000')):
            self.service._fill_premium_summary_fallback(
                self.worksheet, start_row, end_row, self.offers
            )
            
            # Проверяем, что ячейки были разъединены и заполнены
            # (точная проверка разъединения сложна в тестах, но метод должен работать без ошибок)
            self.assertEqual(self.worksheet['I10'].value, Decimal('1000'))


class PremiumSummaryIntegrationTests(unittest.TestCase):
    """Интеграционные тесты для функционала суммирования премий"""
    
    def setUp(self):
        """Настройка тестового окружения"""
        # Создаем сервис с мокированной валидацией шаблона
        with patch.object(ExcelExportService, '_validate_template'):
            self.service = ExcelExportService('/fake/template/path.xlsx')
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
    
    def test_integration_with_fill_company_data(self):
        """Тест интеграции с методом _fill_company_data"""
        # Создаем мок свода с компаниями
        summary = Mock()
        
        # Создаем тестовые предложения для компании с несколькими годами
        offers_multi_year = []
        for i in range(3):
            offer = Mock()
            offer.insurance_year = i + 1
            offer.premium_with_franchise_1 = Decimal(f'{1000 + i * 100}')
            offer.premium_with_franchise_2 = Decimal(f'{500 + i * 50}') if i < 2 else None
            offer.has_second_franchise_variant.return_value = i < 2
            offers_multi_year.append(offer)
        
        # Создаем тестовые предложения для компании с одним годом
        offer_single_year = Mock()
        offer_single_year.insurance_year = 1
        offer_single_year.premium_with_franchise_1 = Decimal('2000')
        offer_single_year.premium_with_franchise_2 = Decimal('1000')
        offer_single_year.has_second_franchise_variant.return_value = True
        
        companies_data = {
            'Компания А': offers_multi_year,  # Несколько лет - должно объединяться
            'Компания Б': [offer_single_year]  # Один год - не должно объединяться
        }
        
        with patch.object(self.service, '_get_companies_sorted_data', return_value=companies_data), \
             patch.object(self.service, '_validate_companies_data', return_value=companies_data), \
             patch.object(self.service, '_get_target_worksheet', return_value=self.worksheet), \
             patch.object(self.service, '_fill_company_year_row'), \
             patch.object(self.service, '_merge_company_name_cells'), \
             patch.object(self.service, '_merge_premium_summary_cells') as mock_merge_premium, \
             patch.object(self.service, '_copy_separator_row'):
            
            self.service._fill_company_data(self.workbook, summary)
            
            # Проверяем, что объединение премий вызвано только для компании с несколькими годами
            mock_merge_premium.assert_called_once()
            call_args = mock_merge_premium.call_args[0]
            self.assertEqual(call_args[3], 'Компания А')  # Название компании
            self.assertEqual(call_args[4], offers_multi_year)  # Предложения компании
    
    def test_constants_updated(self):
        """Тест что константы COMPANY_DATA_COLUMNS содержат новые столбцы"""
        columns = self.service.COMPANY_DATA_COLUMNS
        
        # Проверяем наличие новых столбцов для сумм премий
        self.assertIn('premium_1_summary', columns)
        self.assertIn('premium_2_summary', columns)
        self.assertEqual(columns['premium_1_summary'], 'I')
        self.assertEqual(columns['premium_2_summary'], 'O')


if __name__ == '__main__':
    unittest.main()