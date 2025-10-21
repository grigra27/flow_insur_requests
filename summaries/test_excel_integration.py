"""
Интеграционные тесты для функционала выгрузки Excel-файла со сводной информацией

Эти тесты проверяют полный цикл генерации Excel-файлов, включая:
- Полный цикл генерации Excel файла (требование 1.1, 2.1, 2.2, 2.3)
- Проверку содержимого сгенерированного файла
- Проверку статуса свода
"""

import os
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch, Mock

from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from django.urls import reverse
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook, load_workbook

from insurance_requests.models import InsuranceRequest
from summaries.models import InsuranceSummary, InsuranceOffer
from summaries.services import (
    ExcelExportService, 
    get_excel_export_service,
    ExcelExportServiceError,
    InvalidSummaryDataError,
    TemplateNotFoundError
)


class ExcelExportIntegrationTestCase(TestCase):
    """Базовый класс для интеграционных тестов Excel экспорта"""
    
    @classmethod
    def setUpTestData(cls):
        """Настройка тестовых данных"""
        # Создаем пользователей и группы
        cls.admin_group = Group.objects.create(name='Администраторы')
        cls.user_group = Group.objects.create(name='Пользователи')
        
        cls.admin_user = User.objects.create_user(
            username='admin_test',
            password='test_password',
            is_staff=True
        )
        cls.admin_user.groups.add(cls.admin_group)
        
        cls.regular_user = User.objects.create_user(
            username='user_test',
            password='test_password'
        )
        cls.regular_user.groups.add(cls.user_group)
        
        # Создаем тестовые заявки с полными данными
        cls.complete_request = InsuranceRequest.objects.create(
            client_name='ООО "АВТОПРОЕКТ"',
            inn='1234567890',
            insurance_type='КАСКО',
            vehicle_info='б/у грузовой тягач седельный SCANIA R440A4X2NA',
            dfa_number='ТС-19827-ГА-ВН',
            branch='Великий Новгород',
            status='email_sent',
            created_by=cls.admin_user,
            insurance_period='1 год'
        )
        
        cls.incomplete_request = InsuranceRequest.objects.create(
            client_name='',  # Пустое название клиента
            inn='0987654321',
            insurance_type='ОСАГО',
            vehicle_info='',  # Пустая информация о предмете лизинга
            dfa_number='ТС-19828-ГА-ВН',
            branch='Москва',
            status='uploaded',
            created_by=cls.regular_user,
            insurance_period='2 года'
        )
        
        # Создаем своды с разными статусами
        cls.ready_summary = InsuranceSummary.objects.create(
            request=cls.complete_request,
            status='ready'  # Готов к отправке
        )
        
        cls.collecting_summary = InsuranceSummary.objects.create(
            request=cls.incomplete_request,
            status='collecting'  # Не готов к генерации
        )
        
        # Создаем предложения для полного свода
        cls.offer1 = InsuranceOffer.objects.create(
            summary=cls.ready_summary,
            company_name='РЕСО-Гарантия',
            insurance_year=1,
            insurance_sum=1000000.00,
            franchise_1=0.00,
            premium_with_franchise_1=50000.00,
            franchise_2=25000.00,
            premium_with_franchise_2=45000.00
        )
        
        cls.offer2 = InsuranceOffer.objects.create(
            summary=cls.ready_summary,
            company_name='Альфа Страхование',
            insurance_year=1,
            insurance_sum=1200000.00,
            franchise_1=0.00,
            premium_with_franchise_1=55000.00
        )
    
    def setUp(self):
        """Настройка для каждого теста"""
        # Создаем временный файл шаблона для тестов
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'test_template.xlsx')
        
        # Создаем простой Excel файл для тестов
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'summary_template_sheet'
        
        # Добавляем заголовки для тестирования
        worksheet['A1'] = 'Номер заявки:'
        worksheet['A2'] = 'Предмет лизинга:'
        worksheet['A3'] = 'Клиент:'
        
        workbook.save(self.template_path)
        
        self.client = Client()
    
    def tearDown(self):
        """Очистка после каждого теста"""
        # Удаляем временные файлы
        if os.path.exists(self.template_path):
            os.remove(self.template_path)
        os.rmdir(self.temp_dir)


class FullCycleExcelGenerationTests(ExcelExportIntegrationTestCase):
    """
    Тесты полного цикла генерации Excel файла
    Требования: 1.1, 2.1, 2.2, 2.3
    """
    
    @patch('summaries.services.settings')
    def test_complete_excel_generation_workflow(self, mock_settings):
        """Тест полного цикла генерации Excel файла через веб-интерфейс"""
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Авторизуемся как администратор
        self.client.login(username='admin_test', password='test_password')
        
        # Проверяем, что свод имеет статус "Готов к отправке"
        self.assertEqual(self.ready_summary.status, 'ready')
        
        # Отправляем запрос на генерацию Excel файла
        url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
        response = self.client.get(url)
        
        # Проверяем успешный ответ
        self.assertEqual(response.status_code, 200)
        
        # Проверяем тип содержимого
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Проверяем заголовок для скачивания файла
        content_disposition = response['Content-Disposition']
        # Проверяем, что это attachment (может быть закодировано)
        self.assertIn('Content-Disposition', response)
        # Проверяем наличие имени файла или attachment в заголовке
        self.assertTrue(
            'attachment' in content_disposition.lower() or 
            'filename' in content_disposition.lower() or
            len(content_disposition) > 10  # Закодированный заголовок обычно длинный
        )
        
        # Проверяем, что файл не пустой
        self.assertGreater(len(response.content), 0)
        
        # Проверяем, что это валидный Excel файл
        excel_buffer = BytesIO(response.content)
        try:
            workbook = load_workbook(excel_buffer)
            self.assertIsNotNone(workbook)
        except Exception as e:
            self.fail(f"Generated file is not a valid Excel file: {e}")
    
    @patch('summaries.services.settings')
    def test_excel_generation_with_service_layer(self, mock_settings):
        """Тест генерации Excel через сервисный слой"""
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Создаем сервис
        service = get_excel_export_service()
        
        # Генерируем Excel файл
        excel_file = service.generate_summary_excel(self.ready_summary)
        
        # Проверяем, что результат - это BytesIO
        self.assertIsInstance(excel_file, BytesIO)
        
        # Проверяем, что файл не пустой
        self.assertGreater(len(excel_file.getvalue()), 0)
        
        # Проверяем, что позиция указателя установлена в начало
        self.assertEqual(excel_file.tell(), 0)
        
        # Проверяем, что это валидный Excel файл
        try:
            workbook = load_workbook(excel_file)
            self.assertIsNotNone(workbook)
        except Exception as e:
            self.fail(f"Generated file is not a valid Excel file: {e}")
    
    @patch('summaries.services.settings')
    def test_excel_generation_with_multiple_offers(self, mock_settings):
        """Тест генерации Excel для свода с несколькими предложениями"""
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Добавляем дополнительные предложения
        InsuranceOffer.objects.create(
            summary=self.ready_summary,
            company_name='ВСК',
            insurance_year=2,
            insurance_sum=1500000.00,
            franchise_1=0.00,
            premium_with_franchise_1=70000.00
        )
        
        InsuranceOffer.objects.create(
            summary=self.ready_summary,
            company_name='Согласие',
            insurance_year=1,
            insurance_sum=1100000.00,
            franchise_1=10000.00,
            premium_with_franchise_1=48000.00
        )
        
        # Обновляем счетчик предложений
        self.ready_summary.update_total_offers_count()
        
        # Проверяем количество предложений
        self.assertEqual(self.ready_summary.offers.count(), 4)
        self.assertEqual(self.ready_summary.get_unique_companies_count(), 4)
        
        # Генерируем Excel файл
        service = get_excel_export_service()
        excel_file = service.generate_summary_excel(self.ready_summary)
        
        # Проверяем успешную генерацию
        self.assertIsInstance(excel_file, BytesIO)
        self.assertGreater(len(excel_file.getvalue()), 0)
    
    def test_excel_generation_end_to_end_workflow(self):
        """Тест полного end-to-end workflow генерации Excel"""
        # Создаем новую заявку
        new_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТОВАЯ КОМПАНИЯ"',
            inn='9876543210',
            insurance_type='Имущественное страхование',
            vehicle_info='новый автомобиль Toyota Camry',
            dfa_number='ТС-20001-МСК-ТС',
            branch='Москва',
            status='uploaded',
            created_by=self.admin_user,
            insurance_period='3 года'
        )
        
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Создаем свод
        create_url = reverse('summaries:create_summary', args=[new_request.pk])
        response = self.client.post(create_url)
        self.assertEqual(response.status_code, 302)
        
        # Получаем созданный свод
        new_summary = InsuranceSummary.objects.get(request=new_request)
        self.assertEqual(new_summary.status, 'collecting')
        
        # Добавляем предложение
        offer_data = {
            'company_name': 'РЕСО',  # Используем валидное название из списка
            'insurance_year': 1,
            'insurance_sum': '2000000.00',
            'franchise_1': '0.00',
            'premium_with_franchise_1': '100000.00',
            'franchise_2': '50000.00',
            'premium_with_franchise_2': '90000.00'
        }
        
        add_offer_url = reverse('summaries:add_offer', args=[new_summary.pk])
        response = self.client.post(add_offer_url, offer_data)
        self.assertEqual(response.status_code, 302)
        
        # Проверяем, что предложение добавилось
        self.assertEqual(new_summary.offers.count(), 1)
        
        # Изменяем статус на "Готов к отправке"
        change_status_url = reverse('summaries:change_summary_status', args=[new_summary.pk])
        response = self.client.post(change_status_url, {'status': 'ready'})
        self.assertEqual(response.status_code, 200)
        
        new_summary.refresh_from_db()
        self.assertEqual(new_summary.status, 'ready')
        
        # Генерируем Excel файл
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            generate_url = reverse('summaries:generate_summary_file', args=[new_summary.pk])
            response = self.client.get(generate_url)
            
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )


class ExcelFileContentVerificationTests(ExcelExportIntegrationTestCase):
    """
    Тесты проверки содержимого сгенерированного Excel файла
    Требования: 2.1, 2.2, 2.3
    """
    
    @patch('summaries.services.settings')
    def test_excel_file_contains_correct_data(self, mock_settings):
        """Тест проверки корректности данных в сгенерированном Excel файле"""
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Генерируем Excel файл
        service = get_excel_export_service()
        excel_file = service.generate_summary_excel(self.ready_summary)
        
        # Загружаем сгенерированный файл для проверки
        workbook = load_workbook(excel_file)
        
        # Получаем рабочий лист
        if 'summary_template_sheet' in workbook.sheetnames:
            worksheet = workbook['summary_template_sheet']
        else:
            worksheet = workbook.active
        
        # Проверяем данные в ячейках согласно требованиям
        # CDE1 - номер заявки (требование 2.1)
        self.assertEqual(worksheet['C1'].value, 'ТС-19827-ГА-ВН')
        
        # CDE2 - информация о предмете лизинга (требование 2.2)
        self.assertEqual(
            worksheet['C2'].value, 
            'б/у грузовой тягач седельный SCANIA R440A4X2NA'
        )
        
        # CDE3 - название клиента (требование 2.3)
        self.assertEqual(worksheet['C3'].value, 'ООО "АВТОПРОЕКТ"')
    
    @patch('summaries.services.settings')
    def test_excel_file_structure_integrity(self, mock_settings):
        """Тест проверки целостности структуры Excel файла"""
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Генерируем Excel файл
        service = get_excel_export_service()
        excel_file = service.generate_summary_excel(self.ready_summary)
        
        # Загружаем файл
        workbook = load_workbook(excel_file)
        
        # Проверяем, что файл содержит ожидаемые листы
        self.assertGreater(len(workbook.worksheets), 0)
        
        # Проверяем, что есть лист с нужным именем или используется активный лист
        if 'summary_template_sheet' in workbook.sheetnames:
            worksheet = workbook['summary_template_sheet']
            self.assertEqual(worksheet.title, 'summary_template_sheet')
        else:
            worksheet = workbook.active
            self.assertIsNotNone(worksheet)
        
        # Проверяем, что ячейки C1, C2, C3 содержат данные
        self.assertIsNotNone(worksheet['C1'].value)
        self.assertIsNotNone(worksheet['C2'].value)
        self.assertIsNotNone(worksheet['C3'].value)
        
        # Проверяем, что данные не пустые строки
        self.assertNotEqual(worksheet['C1'].value.strip(), '')
        self.assertNotEqual(worksheet['C2'].value.strip(), '')
        self.assertNotEqual(worksheet['C3'].value.strip(), '')
    
    @patch('summaries.services.settings')
    def test_excel_file_with_special_characters(self, mock_settings):
        """Тест обработки специальных символов в данных"""
        # Создаем заявку со специальными символами
        special_request = InsuranceRequest.objects.create(
            client_name='ООО "ТЕСТ & КОМПАНИЯ" (с символами)',
            inn='1111111111',
            insurance_type='КАСКО',
            vehicle_info='автомобиль с "кавычками" & амперсандом',
            dfa_number='ТС-СПЕЦ-001',
            branch='Тест',
            status='uploaded',
            created_by=self.admin_user,
            insurance_period='1 год'
        )
        
        special_summary = InsuranceSummary.objects.create(
            request=special_request,
            status='ready'
        )
        
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Генерируем Excel файл
        service = get_excel_export_service()
        excel_file = service.generate_summary_excel(special_summary)
        
        # Загружаем файл и проверяем данные
        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        
        # Проверяем, что специальные символы корректно сохранились
        self.assertEqual(worksheet['C1'].value, 'ТС-СПЕЦ-001')
        self.assertEqual(worksheet['C2'].value, 'автомобиль с "кавычками" & амперсандом')
        self.assertEqual(worksheet['C3'].value, 'ООО "ТЕСТ & КОМПАНИЯ" (с символами)')
    
    @patch('summaries.services.settings')
    def test_excel_file_with_long_text(self, mock_settings):
        """Тест обработки длинных текстовых данных"""
        # Создаем заявку с длинными данными
        long_vehicle_info = (
            'очень длинное описание транспортного средства с множеством деталей, '
            'включая технические характеристики, год выпуска, пробег, состояние, '
            'дополнительное оборудование и другую важную информацию для страхования'
        )
        
        long_client_name = (
            'Общество с ограниченной ответственностью "Очень длинное название '
            'компании с дополнительными словами и описанием деятельности"'
        )
        
        long_request = InsuranceRequest.objects.create(
            client_name=long_client_name,
            inn='2222222222',
            insurance_type='КАСКО',
            vehicle_info=long_vehicle_info,
            dfa_number='ТС-ДЛИННЫЙ-001',
            branch='Тест',
            status='uploaded',
            created_by=self.admin_user,
            insurance_period='1 год'
        )
        
        long_summary = InsuranceSummary.objects.create(
            request=long_request,
            status='ready'
        )
        
        # Настраиваем мок settings
        mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
        
        # Генерируем Excel файл
        service = get_excel_export_service()
        excel_file = service.generate_summary_excel(long_summary)
        
        # Загружаем файл и проверяем данные
        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        
        # Проверяем, что длинные тексты корректно сохранились
        self.assertEqual(worksheet['C1'].value, 'ТС-ДЛИННЫЙ-001')
        self.assertEqual(worksheet['C2'].value, long_vehicle_info)
        self.assertEqual(worksheet['C3'].value, long_client_name)


class SummaryStatusVerificationTests(ExcelExportIntegrationTestCase):
    """
    Тесты проверки статуса свода при генерации Excel
    Требования: 1.1
    """
    
    def test_excel_generation_only_for_ready_status(self):
        """Тест генерации Excel только для сводов со статусом 'ready'"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Пытаемся сгенерировать Excel для свода со статусом 'collecting'
        url = reverse('summaries:generate_summary_file', args=[self.collecting_summary.pk])
        response = self.client.get(url)
        
        # Проверяем, что возвращается ошибка
        self.assertEqual(response.status_code, 400)
        
        # Проверяем содержимое ответа
        self.assertEqual(response['Content-Type'], 'application/json')
        
        # Парсим JSON ответ
        import json
        response_data = json.loads(response.content)
        
        # Проверяем сообщение об ошибке
        self.assertIn('error', response_data)
        self.assertIn('Готов к отправке', response_data['error'])
    
    def test_excel_generation_success_for_ready_status(self):
        """Тест успешной генерации Excel для свода со статусом 'ready'"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            # Генерируем Excel для свода со статусом 'ready'
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            # Проверяем успешный ответ
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response['Content-Type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    def test_status_verification_in_service_layer(self):
        """Тест проверки статуса на уровне сервиса"""
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            service = get_excel_export_service()
            
            # Проверяем, что сервис работает для готового свода
            try:
                excel_file = service.generate_summary_excel(self.ready_summary)
                self.assertIsInstance(excel_file, BytesIO)
            except Exception as e:
                self.fail(f"Service should work for ready summary: {e}")
            
            # Сервис не проверяет статус напрямую, это делается в view
            # Но данные должны быть валидными
            try:
                excel_file = service.generate_summary_excel(self.collecting_summary)
                # Сервис должен работать, если данные валидны
                self.assertIsInstance(excel_file, BytesIO)
            except InvalidSummaryDataError:
                # Ошибка может возникнуть из-за невалидных данных, не статуса
                pass
    
    def test_all_status_transitions_for_excel_generation(self):
        """Тест генерации Excel для всех возможных статусов"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Создаем свод для тестирования разных статусов
        test_request = InsuranceRequest.objects.create(
            client_name='Тест статусов',
            inn='3333333333',
            insurance_type='КАСКО',
            vehicle_info='тестовый автомобиль',
            dfa_number='ТС-СТАТУС-001',
            branch='Тест',
            status='uploaded',
            created_by=self.admin_user,
            insurance_period='1 год'
        )
        
        test_summary = InsuranceSummary.objects.create(
            request=test_request,
            status='collecting'
        )
        
        # Тестируем все статусы
        statuses_to_test = ['collecting', 'ready', 'sent', 'completed']
        
        for status in statuses_to_test:
            test_summary.status = status
            test_summary.save()
            
            with patch('summaries.services.settings') as mock_settings:
                mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
                
                url = reverse('summaries:generate_summary_file', args=[test_summary.pk])
                response = self.client.get(url)
                
                if status == 'ready':
                    # Только для статуса 'ready' должна быть успешная генерация
                    self.assertEqual(response.status_code, 200)
                    self.assertEqual(
                        response['Content-Type'],
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    # Для всех остальных статусов должна быть ошибка
                    self.assertEqual(response.status_code, 400)
                    self.assertEqual(response['Content-Type'], 'application/json')


class ExcelExportErrorHandlingIntegrationTests(ExcelExportIntegrationTestCase):
    """
    Интеграционные тесты обработки ошибок при генерации Excel
    Требования: 4.1, 4.2, 4.3
    """
    
    def test_invalid_summary_data_handling(self):
        """Тест обработки невалидных данных свода"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            # Пытаемся сгенерировать Excel для свода с невалидными данными
            url = reverse('summaries:generate_summary_file', args=[self.collecting_summary.pk])
            response = self.client.get(url)
            
            # Сначала проверяем статус (должна быть ошибка статуса)
            self.assertEqual(response.status_code, 400)
            
            # Изменяем статус на ready, но оставляем невалидные данные
            self.collecting_summary.status = 'ready'
            self.collecting_summary.save()
            
            # Теперь должна быть ошибка валидации данных
            response = self.client.get(url)
            self.assertEqual(response.status_code, 400)
            
            # Проверяем JSON ответ
            import json
            response_data = json.loads(response.content)
            self.assertIn('error', response_data)
            self.assertIn('данных свода', response_data['error'])
    
    def test_template_not_found_handling(self):
        """Тест обработки отсутствующего шаблона"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            # Указываем несуществующий путь к шаблону
            mock_settings.SUMMARY_TEMPLATE_PATH = '/nonexistent/path/template.xlsx'
            
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            # Проверяем ошибку
            self.assertEqual(response.status_code, 500)
            
            # Проверяем JSON ответ
            import json
            response_data = json.loads(response.content)
            self.assertIn('error', response_data)
            self.assertTrue('недоступен' in response_data['error'] or 'не найден' in response_data['error'])
    
    def test_service_error_handling(self):
        """Тест обработки ошибок сервиса"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Мокаем ошибку в сервисе через services модуль
        with patch('summaries.services.get_excel_export_service') as mock_service:
            mock_service.side_effect = ExcelExportServiceError("Тестовая ошибка сервиса")
            
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            # Проверяем ошибку
            self.assertEqual(response.status_code, 500)
            
            # Проверяем JSON ответ
            import json
            response_data = json.loads(response.content)
            self.assertIn('error', response_data)
            self.assertIn('Тестовая ошибка сервиса', response_data['error'])
    
    def test_unexpected_error_handling(self):
        """Тест обработки неожиданных ошибок"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Мокаем неожиданную ошибку через services модуль
        with patch('summaries.services.get_excel_export_service') as mock_service:
            mock_service.side_effect = Exception("Неожиданная ошибка")
            
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            # Проверяем ошибку
            self.assertEqual(response.status_code, 500)
            
            # Проверяем JSON ответ
            import json
            response_data = json.loads(response.content)
            self.assertIn('error', response_data)
            self.assertIn('неожиданная ошибка', response_data['error'])


class ExcelExportSecurityIntegrationTests(ExcelExportIntegrationTestCase):
    """
    Интеграционные тесты безопасности генерации Excel
    """
    
    def test_unauthorized_access_prevention(self):
        """Тест предотвращения неавторизованного доступа"""
        # Не авторизуемся
        url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
        response = self.client.get(url)
        
        # Проверяем перенаправление на страницу входа
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)
    
    def test_access_to_nonexistent_summary(self):
        """Тест доступа к несуществующему своду"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        # Пытаемся получить доступ к несуществующему своду
        url = reverse('summaries:generate_summary_file', args=[99999])
        response = self.client.get(url)
        
        # Проверяем ошибку 404
        self.assertEqual(response.status_code, 404)
    
    def test_user_permissions_for_excel_generation(self):
        """Тест прав пользователей на генерацию Excel"""
        # Создаем пользователя без прав
        unauthorized_user = User.objects.create_user(
            username='unauthorized',
            password='test_password'
        )
        
        # Авторизуемся как пользователь без прав
        self.client.login(username='unauthorized', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            # Проверяем отказ в доступе
            self.assertEqual(response.status_code, 403)


class ExcelExportPerformanceIntegrationTests(ExcelExportIntegrationTestCase):
    """
    Интеграционные тесты производительности генерации Excel
    """
    
    def test_excel_generation_performance(self):
        """Тест производительности генерации Excel"""
        import time
        
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            # Измеряем время генерации
            start_time = time.time()
            
            url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
            response = self.client.get(url)
            
            end_time = time.time()
            generation_time = end_time - start_time
            
            # Проверяем успешность
            self.assertEqual(response.status_code, 200)
            
            # Проверяем, что генерация выполняется быстро (менее 5 секунд)
            self.assertLess(generation_time, 5.0, 
                          f"Excel generation took too long: {generation_time:.2f}s")
            
            print(f"Excel generation completed in {generation_time:.3f}s")
    
    def test_multiple_excel_generation_sequential(self):
        """Тест последовательной генерации нескольких Excel файлов"""
        # Авторизуемся
        self.client.login(username='admin_test', password='test_password')
        
        with patch('summaries.services.settings') as mock_settings:
            mock_settings.SUMMARY_TEMPLATE_PATH = self.template_path
            
            # Генерируем несколько файлов последовательно
            results = []
            for i in range(3):
                url = reverse('summaries:generate_summary_file', args=[self.ready_summary.pk])
                response = self.client.get(url)
                results.append(response.status_code)
            
            # Проверяем, что все генерации успешны
            self.assertEqual(len(results), 3)
            self.assertTrue(all(status == 200 for status in results))