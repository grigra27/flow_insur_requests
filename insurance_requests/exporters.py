from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytz
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


MOSCOW_TZ = pytz.timezone('Europe/Moscow')

SECTION_TITLES = OrderedDict([
    ('system', 'Системные данные'),
    ('client', 'Клиент и сделка'),
    ('object', 'Объект страхования'),
    ('insurance', 'Условия страхования'),
    ('customer', 'Реквизиты страхователя'),
    ('communication', 'Письмо и комментарии'),
    ('technical', 'Технические данные'),
    ('attachments', 'Вложения'),
    ('other', 'Прочее'),
])

FIELD_SECTION_MAP = {
    'id': 'system',
    'created_at': 'system',
    'updated_at': 'system',
    'created_by': 'system',
    'status': 'system',
    'client_name': 'client',
    'inn': 'client',
    'dfa_number': 'client',
    'branch': 'client',
    'manager_name': 'client',
    'deal_status': 'client',
    'submission_date': 'client',
    'vehicle_info': 'object',
    'brand': 'object',
    'model': 'object',
    'condition': 'object',
    'equipment_type': 'object',
    'power_or_capacity': 'object',
    'acquisition_cost_value': 'object',
    'acquisition_cost_currency': 'object',
    'source_batch_id': 'object',
    'item_no': 'object',
    'item_count': 'object',
    'source_object_count': 'object',
    'manufacturing_year': 'object',
    'asset_status': 'object',
    'usage_purposes': 'object',
    'insurance_type': 'insurance',
    'insurance_period': 'insurance',
    'franchise_type': 'insurance',
    'has_installment': 'insurance',
    'premium_frequency': 'insurance',
    'has_autostart': 'insurance',
    'has_casco_ce': 'insurance',
    'has_transportation': 'insurance',
    'transportation_departure': 'insurance',
    'transportation_destination': 'insurance',
    'transportation_days': 'insurance',
    'has_construction_work': 'insurance',
    'response_deadline': 'insurance',
    'insured_party': 'insurance',
    'insured_sum_type': 'insurance',
    'creditor_bank': 'insurance',
    'key_completeness': 'insurance',
    'pts_psm': 'insurance',
    'telematics_complex': 'insurance',
    'insurance_territory': 'insurance',
    'guard_conditions': 'insurance',
    'property_location_right_holder': 'insurance',
    'legal_address': 'customer',
    'postal_address': 'customer',
    'business_activity': 'customer',
    'birth_date': 'customer',
    'email_subject': 'communication',
    'email_body': 'communication',
    'notes': 'communication',
    'additional_data': 'technical',
}

FIELD_ORDER = [
    'id',
    'created_at',
    'updated_at',
    'created_by',
    'status',
    'client_name',
    'inn',
    'dfa_number',
    'branch',
    'manager_name',
    'deal_status',
    'submission_date',
    'vehicle_info',
    'brand',
    'model',
    'condition',
    'equipment_type',
    'power_or_capacity',
    'acquisition_cost_value',
    'acquisition_cost_currency',
    'source_batch_id',
    'item_no',
    'item_count',
    'source_object_count',
    'manufacturing_year',
    'asset_status',
    'usage_purposes',
    'insurance_type',
    'insurance_period',
    'franchise_type',
    'has_installment',
    'premium_frequency',
    'has_autostart',
    'has_casco_ce',
    'has_transportation',
    'transportation_departure',
    'transportation_destination',
    'transportation_days',
    'has_construction_work',
    'response_deadline',
    'insured_party',
    'insured_sum_type',
    'creditor_bank',
    'key_completeness',
    'pts_psm',
    'telematics_complex',
    'insurance_territory',
    'guard_conditions',
    'property_location_right_holder',
    'legal_address',
    'postal_address',
    'business_activity',
    'birth_date',
    'email_subject',
    'email_body',
    'notes',
]


def _format_decimal(value: Decimal) -> str:
    normalized = format(value, 'f')
    if '.' in normalized:
        normalized = normalized.rstrip('0').rstrip('.')
    return normalized


def _format_datetime(value: datetime) -> str:
    if timezone.is_naive(value):
        value = timezone.make_aware(value, timezone.get_default_timezone())
    return timezone.localtime(value, MOSCOW_TZ).strftime('%d.%m.%Y %H:%M:%S')


def _format_scalar(value) -> str:
    if value is None or value == '':
        return '—'
    if isinstance(value, bool):
        return 'Да' if value else 'Нет'
    if isinstance(value, datetime):
        return _format_datetime(value)
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    if isinstance(value, Decimal):
        return _format_decimal(value)
    return str(value)


def _format_model_field_value(insurance_request, field) -> str:
    value = getattr(insurance_request, field.name)
    if value is None or value == '':
        return '—'

    if field.many_to_one:
        full_name = ' '.join(
            part for part in [getattr(value, 'last_name', ''), getattr(value, 'first_name', '')] if part
        )
        username = getattr(value, 'username', '')
        if username and full_name:
            return f'{username} ({full_name}) [id={value.pk}]'
        if username:
            return f'{username} [id={value.pk}]'
        return f'id={value.pk}'

    display_getter = getattr(insurance_request, f'get_{field.name}_display', None)
    if callable(display_getter) and field.choices:
        display_value = display_getter()
        if display_value and str(display_value) != str(value):
            return f'{display_value} [{value}]'

    return _format_scalar(value)


def _append_flat_rows(rows, prefix: str, value) -> None:
    if isinstance(value, dict):
        if not value:
            rows.append((prefix, '{}'))
            return
        for nested_key, nested_value in value.items():
            next_prefix = f'{prefix}.{nested_key}' if prefix else str(nested_key)
            _append_flat_rows(rows, next_prefix, nested_value)
        return

    if isinstance(value, list):
        if not value:
            rows.append((prefix, '[]'))
            return
        for index, nested_value in enumerate(value, start=1):
            next_prefix = f'{prefix}[{index}]'
            _append_flat_rows(rows, next_prefix, nested_value)
        return

    rows.append((prefix, _format_scalar(value)))


def _build_request_rows(insurance_request):
    rows_by_section = OrderedDict((section_key, []) for section_key in SECTION_TITLES)
    model_fields = {field.name: field for field in insurance_request._meta.fields}
    handled_fields = set()

    for field_name in FIELD_ORDER:
        field = model_fields.get(field_name)
        if field is None:
            continue
        handled_fields.add(field_name)
        section_key = FIELD_SECTION_MAP.get(field_name, 'other')
        key = f'{field.verbose_name} [{field.name}]'
        rows_by_section[section_key].append((key, _format_model_field_value(insurance_request, field)))

    for field in insurance_request._meta.fields:
        if field.name in handled_fields or field.name == 'additional_data':
            continue
        section_key = FIELD_SECTION_MAP.get(field.name, 'other')
        key = f'{field.verbose_name} [{field.name}]'
        rows_by_section[section_key].append((key, _format_model_field_value(insurance_request, field)))

    additional_data = insurance_request.additional_data if isinstance(insurance_request.additional_data, dict) else {}
    flat_additional_rows = []
    _append_flat_rows(flat_additional_rows, 'additional_data', additional_data)
    if not flat_additional_rows:
        flat_additional_rows.append(('additional_data', '{}'))
    rows_by_section['technical'].extend(flat_additional_rows)

    attachments = list(insurance_request.attachments.order_by('uploaded_at', 'pk'))
    if attachments:
        rows_by_section['attachments'].append(('attachments.count', str(len(attachments))))
        for index, attachment in enumerate(attachments, start=1):
            base_key = f'attachments[{index}]'
            rows_by_section['attachments'].extend([
                (f'{base_key}.id', str(attachment.pk)),
                (f'{base_key}.original_filename', attachment.original_filename or '—'),
                (f'{base_key}.file_type', attachment.file_type or '—'),
                (f'{base_key}.file_name', attachment.file.name or '—'),
                (f'{base_key}.uploaded_at', _format_datetime(attachment.uploaded_at)),
            ])
    else:
        rows_by_section['attachments'].append(('attachments.count', '0'))

    return rows_by_section


def build_request_export_workbook(insurance_request) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Карточка заявки'
    rows_by_section = _build_request_rows(insurance_request)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    section_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')

    worksheet.append(['Раздел', 'Ключ', 'Значение'])
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for section_key, section_title in SECTION_TITLES.items():
        section_rows = rows_by_section.get(section_key, [])
        if not section_rows:
            continue

        worksheet.append([section_title, '', ''])
        for cell in worksheet[worksheet.max_row]:
            cell.font = bold_font
            cell.fill = section_fill
            cell.alignment = Alignment(vertical='center')

        for key, value in section_rows:
            worksheet.append(['', key, value])

    for column_letter, width in {'A': 26, 'B': 42, 'C': 90}.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_request_export_filename(insurance_request) -> str:
    base_name = insurance_request.dfa_number or f'request_{insurance_request.pk}'
    safe_name = get_valid_filename(base_name) or f'request_{insurance_request.pk}'
    safe_name = safe_name[:80]
    timestamp = timezone.localtime(timezone.now(), MOSCOW_TZ).strftime('%Y%m%d_%H%M')
    return f'request_card_{safe_name}_{timestamp}.xlsx'
