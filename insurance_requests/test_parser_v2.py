from datetime import date
from decimal import Decimal
from io import BytesIO
import os
import shutil
import tempfile

from django import forms as forms_module
from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from .forms import DEFAULT_BRANCH, ParserV2PreviewForm
from .models import InsuranceRequest, RequestAttachment
from .parsers.excel_v2 import ExcelRequestParserV2
from .parsers.excel_v2.parser import (
    GridCell,
    classify_equipment_or_power,
    group_identical_objects,
    normalize_condition,
    normalize_currency,
    normalize_insured_party,
    normalize_insured_sum_type,
    normalize_premium_frequency_label,
    normalize_property_location_right_holder,
    parse_cost_value,
    parse_date_value,
    split_brand_model,
)


class ObjectRowHelpersTests(TestCase):
    """Stage 3.1: low-level extractors used while parsing an object row."""

    def test_normalize_currency_recognises_common_variants(self):
        self.assertEqual(normalize_currency('руб'), 'RUB')
        self.assertEqual(normalize_currency('руб.'), 'RUB')
        self.assertEqual(normalize_currency('РУБ'), 'RUB')
        self.assertEqual(normalize_currency('рублей'), 'RUB')
        self.assertEqual(normalize_currency('₽'), 'RUB')
        self.assertEqual(normalize_currency('USD'), 'USD')
        self.assertEqual(normalize_currency('$'), 'USD')
        self.assertEqual(normalize_currency('долл'), 'USD')
        self.assertEqual(normalize_currency('EUR'), 'EUR')
        self.assertEqual(normalize_currency('евро'), 'EUR')
        self.assertEqual(normalize_currency('€'), 'EUR')
        self.assertIsNone(normalize_currency('XXX'))
        self.assertIsNone(normalize_currency(''))
        self.assertIsNone(normalize_currency(None))

    def test_normalize_condition_recognises_common_variants(self):
        self.assertEqual(normalize_condition('новое'), 'new')
        self.assertEqual(normalize_condition('Новый'), 'new')
        self.assertEqual(normalize_condition('НОВАЯ'), 'new')
        self.assertEqual(normalize_condition('б/у'), 'used')
        self.assertEqual(normalize_condition('БУ'), 'used')
        self.assertEqual(normalize_condition('б-у'), 'used')
        self.assertIsNone(normalize_condition('неизвестно'))
        self.assertIsNone(normalize_condition(None))

    def test_parse_cost_value_strips_separators(self):
        self.assertEqual(parse_cost_value('1490000'), Decimal('1490000'))
        self.assertEqual(parse_cost_value('1 490 000'), Decimal('1490000'))
        self.assertEqual(parse_cost_value('1490000,00'), Decimal('1490000.00'))
        self.assertEqual(parse_cost_value('147.51'), Decimal('147.51'))
        self.assertIsNone(parse_cost_value('abc'))
        self.assertIsNone(parse_cost_value(''))
        self.assertIsNone(parse_cost_value(None))

    def test_classify_equipment_or_power_separates_text_from_number(self):
        self.assertEqual(classify_equipment_or_power('колесная'), ('колесная', None))
        self.assertEqual(classify_equipment_or_power('гусеничная'), ('гусеничная', None))
        self.assertEqual(classify_equipment_or_power('78.05'), (None, '78.05'))
        self.assertEqual(classify_equipment_or_power('147,51'), (None, '147,51'))
        self.assertEqual(classify_equipment_or_power('8 700'), (None, '8 700'))
        self.assertEqual(classify_equipment_or_power(''), (None, None))
        self.assertEqual(classify_equipment_or_power(None), (None, None))

    def test_split_brand_model_preserves_short_model_numbers(self):
        self.assertEqual(
            split_brand_model('LADA Largus KS045L 2024 б/у 78.05 1490000 руб', '2024'),
            ('LADA', 'Largus KS045L'),
        )
        # Short numerics inside model name (Mercedes G 400, BMW X5, Audi A4)
        # must survive — only prices (4+ digits) and fractional numbers are dropped.
        self.assertEqual(split_brand_model('Mercedes-Benz G 400 2024', '2024'),
                         ('Mercedes-Benz', 'G 400'))
        self.assertEqual(split_brand_model('Toyota Land Cruiser Prado 250 2024', '2024'),
                         ('Toyota', 'Land Cruiser Prado 250'))
        self.assertEqual(split_brand_model('Haval H7 2025 новое', '2025'),
                         ('Haval', 'H7'))
        # Single token → brand=None, all in model.
        self.assertEqual(split_brand_model('Mining', None), (None, 'Mining'))
        self.assertEqual(split_brand_model('', None), (None, None))
        self.assertEqual(split_brand_model(None, None), (None, None))

    def test_split_brand_model_ignores_generic_object_type_prefixes(self):
        self.assertEqual(
            split_brand_model('Автомобиль JAC T9 2026 новое 163,15 3919000 руб', '2026'),
            ('JAC', 'T9'),
        )
        self.assertEqual(
            split_brand_model('б/у Автомобиль Mercedes-Benz CLA 200 4MATIC 2021 б/у 149,56 3750000 руб', '2021'),
            ('Mercedes-Benz', 'CLA 200 4MATIC'),
        )
        self.assertEqual(
            split_brand_model('Трактор LS R36iHT 2022 новое 3050000 руб', '2022'),
            ('LS', 'R36iHT'),
        )
        self.assertEqual(
            split_brand_model('Погрузчик с бортовым поворотом LIUGONG CLG385B 2026 новое 4780000 руб', '2026'),
            ('LIUGONG', 'CLG385B'),
        )
        self.assertEqual(
            split_brand_model('экскаватор-погрузчик LOVOL FB878X-LMD 2025 новое 9320000 руб', '2025'),
            ('LOVOL', 'FB878X-LMD'),
        )
        self.assertEqual(
            split_brand_model('А/м ГАЗель БИЗНЕС 330232 2026 новое 106,73 2687000 руб', '2026'),
            ('ГАЗель', 'БИЗНЕС 330232'),
        )
        self.assertEqual(
            split_brand_model('автобус НЕФАЗ 5299-0000040-52 2026 новое', '2026'),
            ('НЕФАЗ', '5299-0000040-52'),
        )
        self.assertEqual(
            split_brand_model('Автотопливозаправщик ГРАЗ 36139-0000011 2018 б/у', '2018'),
            ('ГРАЗ', '36139-0000011'),
        )
        self.assertEqual(
            split_brand_model('Б/у асфальтоукладчик VOGELE SUPER 1300-3 2017 б/у', '2017'),
            ('VOGELE', 'SUPER 1300-3'),
        )
        self.assertEqual(
            split_brand_model('полуприцеп-самосвал GRUNWALD 9453-0000011-60 2025 новое 4025000 руб', '2025'),
            ('GRUNWALD', '9453-0000011-60'),
        )
        self.assertEqual(
            split_brand_model('экскаватор LOVOL FR225E2-N 2025 новое', '2025'),
            ('LOVOL', 'FR225E2-N'),
        )


class AutostartExtractionTests(TestCase):
    """`_extract_autostart` must require an explicit «да» — empty cell means «нет»."""

    @staticmethod
    def _coord(row: int, col: int) -> str:
        return f"{chr(64 + col)}{row}"

    def _run(self, cells_spec):
        cells = [
            GridCell(row=r, col=c, coordinate=self._coord(r, c), value=v)
            for r, c, v in cells_spec
        ]
        rows = {}
        for cell in cells:
            rows.setdefault(cell.row, []).append(cell)
        return ExcelRequestParserV2()._extract_autostart(cells, rows)

    def test_returns_true_when_value_cell_is_da(self):
        self.assertTrue(self._run([(24, 12, 'Автозапуск'), (24, 13, 'да')]))

    def test_returns_false_when_value_cell_is_net(self):
        self.assertFalse(self._run([(24, 12, 'Автозапуск'), (24, 13, 'нет')]))

    def test_returns_false_when_value_cell_is_empty(self):
        # Regression: standard «Автозапуск» row with no value must NOT default to True.
        self.assertFalse(self._run([(24, 12, 'Автозапуск')]))

    def test_returns_false_for_da_net_header_without_value(self):
        self.assertFalse(self._run([(24, 12, 'Автозапуск (да/нет):')]))

    def test_returns_true_for_inline_label_with_da(self):
        self.assertTrue(self._run([(24, 12, 'Автозапуск: да')]))

    def test_returns_true_for_da_net_header_with_explicit_da_value(self):
        self.assertTrue(self._run([(24, 12, 'Автозапуск (да/нет):'), (24, 13, 'да')]))

    def test_returns_false_when_label_missing(self):
        self.assertFalse(self._run([(24, 12, 'Что-то другое'), (24, 13, 'да')]))


class CustomerDealHelpersTests(TestCase):
    """Stages 3.2 / 3.3: low-level extractors for customer and deal fields."""

    def test_parse_date_handles_common_formats(self):
        self.assertEqual(parse_date_value('17.04.2026'), date(2026, 4, 17))
        self.assertEqual(parse_date_value('12.06.80'), date(1980, 6, 12))
        self.assertEqual(parse_date_value('1982-12-12'), date(1982, 12, 12))
        self.assertEqual(parse_date_value('1982-12-12 00:00:00'), date(1982, 12, 12))
        self.assertEqual(parse_date_value(date(2024, 1, 1)), date(2024, 1, 1))
        self.assertIsNone(parse_date_value(''))
        self.assertIsNone(parse_date_value(None))
        self.assertIsNone(parse_date_value('not a date'))

    def test_normalize_insured_party(self):
        self.assertEqual(normalize_insured_party('ЛизингоДАТЕЛЬ'), 'lessor')
        self.assertEqual(normalize_insured_party('ЛИЗИНГОПОЛУЧАТЕЛЬ'), 'lessee')
        self.assertEqual(normalize_insured_party('Лизингополучатель'), 'lessee')
        self.assertIsNone(normalize_insured_party('Оба'))
        self.assertIsNone(normalize_insured_party('что-то другое'))
        self.assertIsNone(normalize_insured_party(None))

    def test_normalize_insured_sum_type_handles_substring_collision(self):
        # «неагрегатная» содержит «агрегатная» как подстроку — нельзя путать.
        self.assertEqual(normalize_insured_sum_type('Неагрегатная'), 'non_aggregate')
        self.assertEqual(normalize_insured_sum_type('агрегатная'), 'aggregate')
        self.assertIsNone(normalize_insured_sum_type('что-то'))

    def test_normalize_property_location_right_holder(self):
        self.assertEqual(
            normalize_property_location_right_holder('собственность лизингополучателя'),
            'lessee_owner',
        )
        self.assertEqual(
            normalize_property_location_right_holder('Стороннее лицо'),
            'third_party_owner',
        )
        self.assertIsNone(normalize_property_location_right_holder('—'))

    def test_normalize_premium_frequency_label(self):
        self.assertEqual(normalize_premium_frequency_label('Единовременно'), 'single')
        self.assertEqual(normalize_premium_frequency_label('ежеквартально'), 'quarterly')
        self.assertEqual(normalize_premium_frequency_label('Поквартально'), 'quarterly')
        self.assertEqual(normalize_premium_frequency_label('2 раза в год'), 'biannual')
        self.assertEqual(normalize_premium_frequency_label('полугодовой'), 'biannual')
        self.assertEqual(normalize_premium_frequency_label('ежегодно'), 'annual')
        # «прочее» — вне нашего enum.
        self.assertIsNone(normalize_premium_frequency_label('прочее (укажите)'))
        self.assertIsNone(normalize_premium_frequency_label(None))


class ObjectRowPayloadIntegrationTests(TestCase):
    """Stage 3.1: insured_objects[] entries must carry structured fields."""

    def _build_minimal_workbook_with_object_row(self):
        wb = Workbook()
        sheet = wb.active
        # Minimal headers so that V2 detects an object table starting at row 41.
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '7707083893'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        # Object table header.
        sheet['C41'] = 'Наименование и описание имущества'
        sheet['J41'] = 'Год выпуска'
        sheet['L41'] = 'Мощность'
        sheet['M41'] = 'Стоимость на момент приобретения'
        sheet['N41'] = 'Валюта'
        # Object row.
        sheet['C43'] = 'LADA Largus KS045L'
        sheet['J43'] = 2024
        sheet['K43'] = 'б/у'
        sheet['L43'] = '78.05'
        sheet['M43'] = 1490000
        sheet['N43'] = 'руб'
        return wb

    def _parse_workbook(self, workbook):
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            temp_file.close()
            workbook.save(temp_file.name)
            return ExcelRequestParserV2().parse(temp_file.name, original_filename='object-row.xlsx')
        finally:
            os.unlink(temp_file.name)

    def test_object_payload_contains_structured_fields(self):
        result = self._parse_workbook(self._build_minimal_workbook_with_object_row())
        objects = result.data['parser_v2_payload']['insured_objects']
        self.assertEqual(len(objects), 1)
        obj = objects[0]
        self.assertEqual(obj['brand'], 'LADA')
        self.assertEqual(obj['model'], 'Largus KS045L')
        self.assertEqual(obj['year'], '2024')
        self.assertEqual(obj['condition'], 'used')
        self.assertEqual(obj['power_or_capacity'], '78.05')
        self.assertEqual(obj['acquisition_cost_value'], '1490000')
        self.assertEqual(obj['acquisition_cost_currency'], 'RUB')

    def test_object_payload_normalises_alternative_currency_synonyms(self):
        wb = self._build_minimal_workbook_with_object_row()
        wb.active['N43'] = 'USD'
        wb.active['M43'] = 25000
        result = self._parse_workbook(wb)
        obj = result.data['parser_v2_payload']['insured_objects'][0]
        self.assertEqual(obj['acquisition_cost_currency'], 'USD')
        self.assertEqual(obj['acquisition_cost_value'], '25000')

    def test_casco_ce_uses_object_section_not_empty_template_header(self):
        wb = self._build_minimal_workbook_with_object_row()
        sheet = wb.active
        sheet['B42'] = 'Транспортные средства категории B'
        sheet['B44'] = 'Транспортные средства категории C'

        result = self._parse_workbook(wb)
        obj = result.data['parser_v2_payload']['insured_objects'][0]
        self.assertEqual(obj.get('vehicle_category'), 'B')
        self.assertEqual(obj.get('vehicle_category_source'), 'B42:B42')
        self.assertEqual(obj.get('equipment_type'), 'Категория B')
        self.assertFalse(result.data.get('has_casco_ce'))
        self.assertNotIn('has_casco_ce', result.source_map)

    def test_casco_ce_is_selected_for_object_under_category_c(self):
        wb = self._build_minimal_workbook_with_object_row()
        sheet = wb.active
        sheet['B42'] = 'Транспортные средства категории C'

        result = self._parse_workbook(wb)
        obj = result.data['parser_v2_payload']['insured_objects'][0]
        self.assertEqual(obj.get('vehicle_category'), 'C')
        self.assertEqual(obj.get('equipment_type'), 'Категория C')
        self.assertTrue(result.data.get('has_casco_ce'))
        self.assertEqual(result.source_map.get('has_casco_ce'), 'B42:B42')

    def test_casco_ce_is_selected_when_any_object_is_under_category_c(self):
        wb = self._build_minimal_workbook_with_object_row()
        sheet = wb.active
        sheet['B42'] = 'Транспортные средства категории B'
        sheet['B44'] = 'Транспортные средства категории C'
        sheet['B45'] = 2
        sheet['C45'] = 'Грузовой автомобиль MAN TGS'
        sheet['J45'] = 2024
        sheet['K45'] = 'новое'
        sheet['L45'] = '12'
        sheet['M45'] = 9800000
        sheet['N45'] = 'руб'

        result = self._parse_workbook(wb)
        objects = result.data['parser_v2_payload']['insured_objects']
        self.assertEqual([obj.get('vehicle_category') for obj in objects], ['B', 'C'])
        self.assertTrue(result.data.get('has_casco_ce'))
        self.assertEqual(result.source_map.get('has_casco_ce'), 'B44:B44')

    def test_object_section_category_fills_empty_equipment_type(self):
        wb = self._build_minimal_workbook_with_object_row()
        sheet = wb.active
        sheet['B42'] = 'Транспортные средства категории D'

        result = self._parse_workbook(wb)
        obj = result.data['parser_v2_payload']['insured_objects'][0]
        self.assertEqual(obj.get('vehicle_category'), 'D')
        self.assertEqual(obj.get('equipment_type'), 'Категория D')
        self.assertEqual(obj.get('power_or_capacity'), '78.05')

    def test_special_equipment_keeps_specific_equipment_type(self):
        wb = self._build_minimal_workbook_with_object_row()
        sheet = wb.active
        sheet['B42'] = 'Специальная техника'
        sheet['C43'] = 'Трактор LS R36iHT'
        sheet['L43'] = 'колесная'

        result = self._parse_workbook(wb)
        obj = result.data['parser_v2_payload']['insured_objects'][0]
        self.assertEqual(obj.get('vehicle_category'), 'special_equipment')
        self.assertEqual(obj.get('equipment_type'), 'колесная')
        self.assertIsNone(obj.get('power_or_capacity'))


class IdenticalObjectGroupingTests(TestCase):
    """parser_v2_identical_object_multiplicity: fully identical object rows
    collapse into one object that records how many source rows it represents."""

    @staticmethod
    def _obj(**overrides):
        base = {
            'description': 'LADA Largus KS045L',
            'brand': 'LADA',
            'model': 'Largus KS045L',
            'year': '2024',
            'condition': 'used',
            'equipment_type': None,
            'power_or_capacity': '78.05',
            'acquisition_cost_value': '1490000',
            'acquisition_cost_currency': 'RUB',
            'vehicle_category': None,
            'source': 'C43:N43',
            'source_object_count': 1,
            'duplicate_sources': ['C43:N43'],
        }
        base.update(overrides)
        return base

    def test_group_mixed_file_sums_multiplicity_in_first_appearance_order(self):
        a = self._obj(brand='LADA', model='Largus KS045L')
        b = self._obj(brand='Toyota', model='Camry XV70', acquisition_cost_value='4500000')
        c = self._obj(brand='Haval', model='H7', acquisition_cost_value='3649000')
        grouped, meta = group_identical_objects([a, a, a, b, c, c])

        self.assertEqual([o['brand'] for o in grouped], ['LADA', 'Toyota', 'Haval'])
        self.assertEqual([o['source_object_count'] for o in grouped], [3, 1, 2])
        self.assertEqual(meta['raw_object_count'], 6)
        self.assertEqual(meta['unique_object_count'], 3)
        self.assertEqual([g['source_object_count'] for g in meta['groups']], [3, 1, 2])

    def test_group_collapses_cost_format_variants(self):
        a = self._obj(acquisition_cost_value='1490000')
        b = self._obj(acquisition_cost_value='1490000.00')
        grouped, meta = group_identical_objects([a, b])

        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0]['source_object_count'], 2)
        self.assertEqual(meta['raw_object_count'], 2)
        self.assertEqual(meta['unique_object_count'], 1)

    def test_group_keeps_objects_that_differ_in_one_business_field(self):
        a = self._obj(year='2024')
        b = self._obj(year='2023')
        grouped, _ = group_identical_objects([a, b])

        self.assertEqual(len(grouped), 2)
        self.assertEqual([o['source_object_count'] for o in grouped], [1, 1])

    def test_group_does_not_merge_poorly_parsed_rows_with_empty_key(self):
        # Both rows have an all-empty structural key; only the free-text
        # description differs. They must stay separate (fallback rule).
        bare = {
            'description': 'Объект А', 'brand': None, 'model': None, 'year': '',
            'condition': None, 'equipment_type': None, 'power_or_capacity': None,
            'acquisition_cost_value': None, 'acquisition_cost_currency': None,
            'vehicle_category': None, 'source': 'B25:B25',
            'source_object_count': 1, 'duplicate_sources': ['B25:B25'],
        }
        other = dict(bare, description='Объект Б', source='B26:B26', duplicate_sources=['B26:B26'])
        grouped, meta = group_identical_objects([bare, other])

        self.assertEqual(len(grouped), 2)
        self.assertEqual(meta['unique_object_count'], 2)

    def _build_workbook_with_object_rows(self, rows):
        wb = Workbook()
        sheet = wb.active
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '7707083893'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['C41'] = 'Наименование и описание имущества'
        sheet['J41'] = 'Год выпуска'
        sheet['L41'] = 'Мощность'
        sheet['M41'] = 'Стоимость на момент приобретения'
        sheet['N41'] = 'Валюта'
        for idx, (desc, year, cond, power, cost) in enumerate(rows):
            row = 43 + idx * 2
            sheet[f'C{row}'] = desc
            sheet[f'J{row}'] = year
            sheet[f'K{row}'] = cond
            sheet[f'L{row}'] = power
            sheet[f'M{row}'] = cost
            sheet[f'N{row}'] = 'руб'
        return wb

    def _parse_workbook(self, workbook):
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            temp_file.close()
            workbook.save(temp_file.name)
            return ExcelRequestParserV2().parse(temp_file.name, original_filename='grouping.xlsx')
        finally:
            os.unlink(temp_file.name)

    def test_three_identical_rows_collapse_to_one_object_through_parse(self):
        identical = ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000)
        result = self._parse_workbook(
            self._build_workbook_with_object_rows([identical, identical, identical])
        )
        payload = result.data['parser_v2_payload']
        objects = payload['insured_objects']
        self.assertEqual(len(objects), 1)
        self.assertEqual(objects[0]['source_object_count'], 3)
        self.assertEqual(len(objects[0]['duplicate_sources']), 3)

        grouping = payload['object_grouping']
        self.assertEqual(grouping['raw_object_count'], 3)
        self.assertEqual(grouping['unique_object_count'], 1)

    def test_distinct_rows_stay_separate_through_parse(self):
        result = self._parse_workbook(self._build_workbook_with_object_rows([
            ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000),
            ('Toyota Camry XV70', 2023, 'новое', '249', 4500000),
        ]))
        payload = result.data['parser_v2_payload']
        self.assertEqual(len(payload['insured_objects']), 2)
        self.assertEqual([o['source_object_count'] for o in payload['insured_objects']], [1, 1])
        self.assertEqual(payload['object_grouping']['raw_object_count'], 2)
        self.assertEqual(payload['object_grouping']['unique_object_count'], 2)


class CustomerDealPayloadIntegrationTests(TestCase):
    """Stages 3.2 / 3.3: customer + deal fields must surface in parse() result."""

    def _build_workbook(self):
        wb = Workbook()
        sheet = wb.active
        # Header / customer block. Submission date sits in M2 with a «дата
        # подачи» label one cell to the left, just like the real corpus.
        sheet['L2'] = 'дата подачи'
        sheet['M2'] = '17.04.2026'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['B8'] = 'Юридический адрес:'
        sheet['D8'] = '194354, Санкт-Петербург г, Северный пр-кт, дом № 11'
        sheet['D9'] = '7707083893'
        sheet['B10'] = 'Почтовый адрес:'
        sheet['D10'] = '194354, Санкт-Петербург г, Северный пр-кт, дом № 11'
        sheet['B11'] = 'Основной вид деятельности:'
        sheet['D11'] = 'Строительство автомобильных дорог'
        # Deal / insurance block. Two label cells + an «Х» under one of them
        # is the canonical layout in the real corpus.
        sheet['B14'] = 'Страхователь'
        sheet['D14'] = 'ЛизингоДАТЕЛЬ'
        sheet['E14'] = 'ЛизингоПОЛУЧАТЕЛЬ'
        sheet['D15'] = 'Х'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['D24'] = 'страховая сумма'
        sheet['E24'] = 'Неагрегатная'
        sheet['D26'] = 'условия по охране (день/ночь)'
        sheet['E26'] = 'без ограничений'
        # Property-location right holder block: two labels in one row + «Х»
        # under one of them (same crosshair layout as the «Страхователь» block).
        sheet['B27'] = 'Правообладатель места расположения'
        sheet['D27'] = 'собственность лизингополучателя'
        sheet['F27'] = 'собственность третьего лица'
        sheet['D28'] = 'Х'
        # Premium frequency block: «Х» next to «ежеквартально».
        sheet['B31'] = 'Порядок уплаты страховой премии'
        sheet['D31'] = 'Единовременно'
        sheet['E31'] = 'В рассрочку'
        sheet['E32'] = 'ежеквартально'
        sheet['F32'] = 'Х'
        sheet['E33'] = '2 раза в год'
        sheet['E34'] = 'ежегодно'
        return wb

    def _parse_workbook(self, workbook):
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            temp_file.close()
            workbook.save(temp_file.name)
            return ExcelRequestParserV2().parse(temp_file.name, original_filename='customer-deal.xlsx')
        finally:
            os.unlink(temp_file.name)

    def test_customer_fields_surface_in_parse_result(self):
        result = self._parse_workbook(self._build_workbook())
        self.assertEqual(result.data.get('legal_address'),
                         '194354, Санкт-Петербург г, Северный пр-кт, дом № 11')
        self.assertEqual(result.data.get('postal_address'),
                         '194354, Санкт-Петербург г, Северный пр-кт, дом № 11')
        self.assertEqual(result.data.get('business_activity'),
                         'Строительство автомобильных дорог')
        self.assertEqual(result.data.get('submission_date'), '2026-04-17')

    def test_deal_fields_surface_in_parse_result(self):
        result = self._parse_workbook(self._build_workbook())
        self.assertEqual(result.data.get('insured_party'), 'lessor')
        self.assertEqual(result.data.get('insured_sum_type'), 'non_aggregate')
        self.assertEqual(result.data.get('guard_conditions'), 'без ограничений')
        self.assertEqual(result.data.get('property_location_right_holder'), 'lessee_owner')

    def test_transportation_ignores_usage_purpose_text(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['C37'] = 'Цели использования'
        sheet['D37'] = 'Перевозка сотрудников, инструментов, материалов к месту проведения работ'
        result = self._parse_workbook(wb)
        self.assertFalse(result.data.get('has_transportation'))
        self.assertNotIn('has_transportation', result.source_map)

    def test_transportation_ignores_business_activity_text(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D11'] = 'Деятельность автомобильного грузового транспорта и услуги по перевозкам'
        result = self._parse_workbook(wb)
        self.assertFalse(result.data.get('has_transportation'))
        self.assertNotIn('has_transportation', result.source_map)

    def test_transportation_template_without_mark_or_details_is_not_selected(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['B44'] = 'Дополнительные виды страхования оборудования(отметьте знаком "Х")'
        sheet['D44'] = 'Перевозка (с погрузкой, выгрузкой) от поставщика к лизингополучателю'
        sheet['C45'] = 'Пункт отправления'
        sheet['C46'] = 'Пункт назначения'
        sheet['C47'] = 'Ориентировочный срок перевозки, в днях'
        result = self._parse_workbook(wb)
        self.assertFalse(result.data.get('has_transportation'))
        self.assertNotIn('has_transportation', result.source_map)

    def test_transportation_mark_in_additional_insurance_block_selects_option(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['B44'] = 'Дополнительные виды страхования оборудования(отметьте знаком "Х")'
        sheet['D44'] = 'Перевозка (с погрузкой, выгрузкой) от поставщика к лизингополучателю'
        sheet['E44'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertTrue(result.data.get('has_transportation'))
        self.assertEqual(result.source_map.get('has_transportation'), 'E44')

    def test_transportation_filled_details_select_option(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['B44'] = 'Дополнительные виды страхования оборудования(отметьте знаком "Х")'
        sheet['D44'] = 'Перевозка (с погрузкой, выгрузкой) от поставщика к лизингополучателю'
        sheet['C45'] = 'Пункт отправления'
        sheet['D45'] = 'Москва'
        result = self._parse_workbook(wb)
        self.assertTrue(result.data.get('has_transportation'))
        self.assertEqual(result.source_map.get('has_transportation'), 'D45')

    def test_insurance_type_uses_d21_mark_for_casco(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D21'] = 'Х'
        sheet['D22'] = None
        sheet['B21'] = 'КАСКО'
        sheet['B22'] = 'страхование спецтехники'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('insurance_type'), 'КАСКО')
        self.assertEqual(result.source_map.get('insurance_type'), 'D21')

    def test_insurance_type_uses_d22_mark_for_equipment(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D21'] = None
        sheet['D22'] = 'Х'
        sheet['B21'] = 'КАСКО'
        sheet['B22'] = 'страхование спецтехники'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('insurance_type'), 'страхование спецтехники')
        self.assertEqual(result.source_map.get('insurance_type'), 'D22')

    def test_insurance_type_uses_d23_mark_for_shifted_ip_layout(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['B21'] = 'Вид страхования (отметьте знаком "Х")'
        sheet['B22'] = 'КАСКО'
        sheet['B23'] = 'страхование спецтехники'
        sheet['D21'] = None
        sheet['D22'] = None
        sheet['D23'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('insurance_type'), 'страхование спецтехники')
        self.assertEqual(result.source_map.get('insurance_type'), 'D23')

    def test_insurance_type_not_forced_to_property_by_object_header_text(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D21'] = 'Х'
        sheet['D22'] = None
        sheet['B21'] = 'КАСКО'
        sheet['B22'] = 'страхование спецтехники'
        # This phrase appears in object-table headers and used to trigger a
        # false positive for property insurance in the old full-sheet scan.
        sheet['C41'] = 'Наименование и описание имущества'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('insurance_type'), 'КАСКО')

    def test_insured_party_x_marker_picks_lessee_when_marked(self):
        wb = self._build_workbook()
        sheet = wb.active
        # Move the «Х» from under lessor (D15) to under lessee (E15).
        sheet['D15'] = None
        sheet['E15'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('insured_party'), 'lessee')

    def test_insured_party_warns_when_no_mark_is_present(self):
        wb = self._build_workbook()
        sheet = wb.active
        # Strip every X-mark candidate cell below the label row.
        for coord in ('D15', 'E15'):
            sheet[coord] = None
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('insured_party'))
        self.assertTrue(
            any(
                w.get('field') == 'insured_party' and w.get('level') == 'manual_required'
                for w in result.warnings
            ),
            f"Expected a manual_required insured_party warning, got: {result.warnings}",
        )

    def test_insured_party_ambiguous_when_both_columns_marked(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D15'] = 'Х'
        sheet['E15'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('insured_party'))
        self.assertEqual(result.data.get('premium_frequency'), 'quarterly')

    def test_plrh_x_marker_picks_third_party_when_marked(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D28'] = None
        sheet['F28'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('property_location_right_holder'), 'third_party_owner')

    def test_plrh_warns_only_for_property_insurance(self):
        # No mark + insurance_type = «страхование имущества» → warning fires.
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D21'] = 'страхование имущества'
        sheet['D28'] = None
        sheet['F28'] = None
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('property_location_right_holder'))
        self.assertTrue(
            any(
                w.get('field') == 'property_location_right_holder'
                and w.get('level') == 'manual_required'
                for w in result.warnings
            ),
            f"Expected a manual_required PLRH warning, got: {result.warnings}",
        )

    def test_plrh_warning_suppressed_for_casco(self):
        # No mark + default КАСКО → no PLRH warning (field is irrelevant for CASCO).
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D28'] = None
        sheet['F28'] = None
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('property_location_right_holder'))
        self.assertFalse(
            any(w.get('field') == 'property_location_right_holder' for w in result.warnings),
            f"Did not expect a PLRH warning for CASCO, got: {result.warnings}",
        )

    def test_plrh_ambiguous_when_both_columns_marked(self):
        wb = self._build_workbook()
        sheet = wb.active
        sheet['D28'] = 'Х'
        sheet['F28'] = 'Х'
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('property_location_right_holder'))

    def test_has_installment_derived_from_premium_frequency(self):
        # «Рассрочка» в нашей семантике = только внутригодовые платежи
        # (quarterly / biannual). annual (один платёж за год) — НЕ рассрочка.
        cases = {
            'quarterly': ('F32', True),
            'biannual':  ('F33', True),
            'annual':    ('F34', False),
        }
        for expected_freq, (mark_cell, expected_installment) in cases.items():
            wb = self._build_workbook()
            sheet = wb.active
            # Clear the default «Х» under «ежеквартально».
            sheet['F32'] = None
            sheet[mark_cell] = 'Х'
            result = self._parse_workbook(wb)
            self.assertEqual(
                result.data.get('premium_frequency'), expected_freq,
                f"premium_frequency for mark {mark_cell}",
            )
            self.assertEqual(
                result.data.get('has_installment'), expected_installment,
                f"has_installment for premium_frequency={expected_freq}",
            )

    def test_has_installment_false_when_premium_frequency_missing(self):
        wb = self._build_workbook()
        sheet = wb.active
        # Strip every X-marker in the premium-frequency block.
        for coord in ('F32', 'F33', 'F34'):
            sheet[coord] = None
        result = self._parse_workbook(wb)
        self.assertIsNone(result.data.get('premium_frequency'))
        self.assertFalse(result.data.get('has_installment'))

    def test_empty_template_does_not_pick_neighbour_labels(self):
        # Build a "template only" workbook — labels present, values empty.
        wb = Workbook()
        sheet = wb.active
        sheet['L2'] = 'дата подачи'
        sheet['M2'] = '17.04.2026'
        sheet['B8'] = 'Юридический адрес:'
        sheet['B10'] = 'Почтовый адрес:'
        sheet['B11'] = 'Основной вид деятельности:'
        sheet['B12'] = 'ПАРАМЕТРЫ СТРАХОВОЙ СДЕЛКИ'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '7707083893'
        result = self._parse_workbook(wb)
        # No value -> field stays absent (or None), never the next label.
        for field in ('legal_address', 'postal_address', 'business_activity'):
            value = result.data.get(field)
            self.assertNotIn('Почтовый адрес', value or '')
            self.assertNotIn('Основной вид деятельности', value or '')
            self.assertNotIn('ПАРАМЕТРЫ', value or '')

    def test_empty_creditor_bank_does_not_borrow_neighbour_value(self):
        # «Банк-кредитор» label with an empty value column, followed a few
        # columns over by the next field's label. The empty creditor must NOT
        # capture «Необходимый период страхования».
        wb = Workbook()
        sheet = wb.active
        sheet['B17'] = 'Банк-кредитор'
        # B18..G17 left empty (creditor unknown yet)
        sheet['H17'] = 'Необходимый период страхования'
        sheet['L17'] = '1 год'
        # The cell below the label belongs to another attribute and must not be
        # taken as the creditor value.
        sheet['B18'] = 'Вид страхования'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('creditor_bank', ''), '')

    def test_filled_creditor_bank_is_still_extracted(self):
        wb = Workbook()
        sheet = wb.active
        sheet['B17'] = 'Банк-кредитор'
        sheet['C17'] = 'ВТБ'
        result = self._parse_workbook(wb)
        self.assertEqual(result.data.get('creditor_bank', ''), 'ВТБ')


class ParserV2UploadTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.media_root = tempfile.mkdtemp()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()

        admin_group, _ = Group.objects.get_or_create(name='Администраторы')
        user_group, _ = Group.objects.get_or_create(name='Пользователи')

        self.superuser = User.objects.create_superuser(
            username='parser_v2_root',
            email='root@example.com',
            password='pwd',
        )
        self.superuser.groups.add(admin_group)

        self.regular_user = User.objects.create_user(
            username='parser_v2_user',
            email='user@example.com',
            password='pwd',
        )
        self.regular_user.groups.add(user_group)

        self.admin_user = User.objects.create_user(
            username='parser_v2_admin',
            email='admin@example.com',
            password='pwd',
        )
        self.admin_user.groups.add(admin_group)

    def tearDown(self):
        self.settings_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def _xlsx_upload(self, filename='заявка 20213-ЛТ-КЗ.xlsx'):
        workbook = Workbook()
        sheet = workbook.active
        sheet['C4'] = 'Казанский филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D6'] = '20213-ЛТ-КЗ'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['B14'] = 'Страхователь'
        sheet['D14'] = 'ЛизингоДАТЕЛЬ'
        sheet['E14'] = 'ЛизингоПОЛУЧАТЕЛЬ'
        sheet['D15'] = 'Х'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['B24'] = 'Предмет лизинга'
        sheet['B25'] = 'Мини-погрузчик Sunward SWL 4028, 2024 г.в.'
        sheet['D29'] = 'Без франшизы'

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            filename,
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _xlsx_upload_with_template_object_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['C4'] = 'Казанский филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['A39'] = 'СВЕДЕНИЯ ОБ ОБЪЕКТЕ СТРАХОВАНИЯ'
        sheet['A41'] = '№ п/п'
        sheet['B41'] = 'Наименование и описание имущества  (марка модель комплектация)'
        sheet['C41'] = 'Год выпуска'
        sheet['A42'] = 'Транспортные средства категории B'
        sheet['A43'] = '1'
        sheet['B43'] = 'Lixiang L9 (пробег 13 000 км)'
        sheet['C43'] = '2024'
        sheet['D43'] = 'б/у'
        sheet['E43'] = '9 000 000'
        sheet['A45'] = 'Противоугонные системы и оборудование (отметьте знаком "Х")'
        sheet['B45'] = 'Штатная'
        sheet['C45'] = 'Установленная дополнительно'
        sheet['D45'] = 'название, модель'
        sheet['A46'] = 'Сигнализация'
        sheet['A47'] = 'Иммобилайзер'

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'заявка object-template.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _xlsx_upload_with_unknown_branch(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['C4'] = 'Неизвестный филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['B24'] = 'Предмет лизинга'
        sheet['B25'] = 'Мини-погрузчик Sunward SWL 4028, 2024 г.в.'

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'заявка unknown-branch.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _xlsx_upload_with_telematics_template_header(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['C4'] = 'Казанский филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['B24'] = 'Предмет лизинга'
        sheet['B25'] = 'Мини-погрузчик Sunward SWL 4028, 2024 г.в.'
        sheet['A53'] = 'Телематический комплекс'
        sheet['B53'] = 'Наименование'
        sheet['C53'] = 'StarLine M96'

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'заявка telematics-template.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _parse_workbook_v2(self, workbook):
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            temp_file.close()
            workbook.save(temp_file.name)
            return ExcelRequestParserV2().parse(temp_file.name, original_filename='franchise-test.xlsx')
        finally:
            os.unlink(temp_file.name)

    def _post_data_from_preview(self, source):
        """Build POST data that mirrors what the operator would submit.

        Accepts either an HttpResponse from the upload step (preferred —
        also collects the object formset) or a bare ParserV2PreviewForm
        for back-compat with tests that don't deal with the formset.
        """
        if hasattr(source, 'context') and source.context is not None:
            form = source.context['form']
            object_formset = source.context.get('object_formset')
        else:
            form = source
            object_formset = None

        data = {}
        checkbox_fields = {
            'has_installment',
            'has_autostart',
            'has_casco_ce',
            'has_transportation',
            'has_construction_work',
        }
        for field in form.fields:
            if field in checkbox_fields:
                if form.initial.get(field):
                    data[field] = 'on'
            else:
                value = form.initial.get(field, '')
                data[field] = '' if value is None else value

        if object_formset is not None:
            prefix = object_formset.prefix or 'form'
            data[f'{prefix}-TOTAL_FORMS'] = str(object_formset.total_form_count())
            data[f'{prefix}-INITIAL_FORMS'] = str(object_formset.initial_form_count())
            data[f'{prefix}-MIN_NUM_FORMS'] = '0'
            data[f'{prefix}-MAX_NUM_FORMS'] = '1000'
            for index, obj_form in enumerate(object_formset.forms):
                for field_name, field in obj_form.fields.items():
                    initial = obj_form.initial.get(field_name, '')
                    key = f'{prefix}-{index}-{field_name}'
                    if isinstance(field, forms_module.BooleanField):
                        if initial:
                            data[key] = 'on'
                    else:
                        data[key] = '' if initial is None else initial
        return data

    def test_primary_upload_route_uses_new_parser_for_all_v1_users(self):
        url = reverse('insurance_requests:upload_excel')

        anonymous_response = self.client.get(url)
        self.assertEqual(anonymous_response.status_code, 302)

        self.client.login(username='parser_v2_user', password='pwd')
        regular_response = self.client.get(url)
        self.assertEqual(regular_response.status_code, 200)
        self.assertTemplateUsed(regular_response, 'insurance_requests/upload_excel_v2.html')

        self.client.logout()
        self.client.login(username='parser_v2_admin', password='pwd')
        admin_response = self.client.get(url)
        self.assertEqual(admin_response.status_code, 200)
        self.assertTemplateUsed(admin_response, 'insurance_requests/upload_excel_v2.html')

        self.client.logout()
        self.client.login(username='parser_v2_root', password='pwd')
        superuser_response = self.client.get(url)
        self.assertEqual(superuser_response.status_code, 200)
        self.assertTemplateUsed(superuser_response, 'insurance_requests/upload_excel_v2.html')

    def test_old_loader_has_dedicated_fallback_route(self):
        self.client.login(username='parser_v2_user', password='pwd')

        response = self.client.get(reverse('insurance_requests:upload_excel_legacy'))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'insurance_requests/upload_excel.html')
        self.assertContains(response, 'Старый загрузчик')

    def test_request_list_uses_primary_upload_route_without_extra_parser_v2_button(self):
        request_list_url = reverse('insurance_requests:request_list')
        primary_upload_url = reverse('insurance_requests:upload_excel')
        parser_v2_url = reverse('insurance_requests:upload_excel_v2')

        self.client.login(username='parser_v2_user', password='pwd')
        regular_response = self.client.get(request_list_url)
        self.assertEqual(regular_response.status_code, 200)
        self.assertContains(regular_response, primary_upload_url)
        self.assertNotContains(regular_response, parser_v2_url)

        self.client.logout()
        self.client.login(username='parser_v2_admin', password='pwd')
        admin_response = self.client.get(request_list_url)
        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, primary_upload_url)
        self.assertNotContains(admin_response, parser_v2_url)

        self.client.logout()
        self.client.login(username='parser_v2_root', password='pwd')
        superuser_response = self.client.get(request_list_url)
        self.assertEqual(superuser_response.status_code, 200)
        self.assertContains(superuser_response, primary_upload_url)
        self.assertNotContains(superuser_response, parser_v2_url)

    def test_parser_v2_upload_renders_editable_preview(self):
        self.client.login(username='parser_v2_user', password='pwd')

        response = self.client.post(
            reverse('insurance_requests:upload_excel'),
            {'excel_file': self._xlsx_upload()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'insurance_requests/upload_excel_v2_preview.html')
        self.assertContains(response, 'ООО Ромашка')
        self.assertContains(response, 'Мини-погрузчик Sunward SWL 4028')
        self.assertIn('draft_id', response.context)
        self.assertEqual(response.context['form'].initial['client_name'], 'ООО Ромашка')
        self.assertEqual(response.context['form'].initial['manager_name'], 'Иванов Иван')

    def test_primary_upload_page_links_to_old_loader(self):
        self.client.login(username='parser_v2_user', password='pwd')

        response = self.client.get(reverse('insurance_requests:upload_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('insurance_requests:upload_excel_legacy'))

    def test_primary_upload_page_uses_magic_context_label(self):
        self.client.login(username='parser_v2_user', password='pwd')

        response = self.client.get(reverse('insurance_requests:upload_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['app_navigation']['current_context_label'],
            'Заявки / Alla Borisovna Magic Parser',
        )

    def test_parser_v2_unrecognized_branch_preserves_raw_value_and_warns(self):
        self.client.login(username='parser_v2_root', password='pwd')

        response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_unknown_branch()},
        )

        form = response.context['form']
        self.assertEqual(form.initial['branch'], 'Неизвестный филиал')
        self.assertEqual(form.fields['branch'].widget.__class__.__name__, 'Select')
        warnings = response.context['warnings']
        self.assertTrue(
            any(w.get('field') == 'branch' and w.get('level') == 'manual_required' for w in warnings),
            f"Expected a manual_required branch warning, got: {warnings}",
        )

    def test_parser_v2_branch_accepts_only_known_choices(self):
        form = ParserV2PreviewForm(data={'draft_id': 'draft', 'branch': 'Неизвестный филиал'})

        self.assertFalse(form.is_valid())
        self.assertIn('branch', form.errors)

    def test_parser_v2_does_not_take_template_rows_as_vehicle_info(self):
        self.client.login(username='parser_v2_root', password='pwd')

        response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_template_object_rows()},
        )

        vehicle_info = response.context['form'].initial['vehicle_info']
        self.assertIn('Lixiang L9', vehicle_info)
        self.assertNotIn('Транспортные средства категории B', vehicle_info)
        self.assertNotIn('Противоугонные системы', vehicle_info)
        self.assertNotIn('Сигнализация', vehicle_info)

    def test_parser_v2_keeps_raw_object_description_for_persistence(self):
        self.client.login(username='parser_v2_root', password='pwd')

        response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_template_object_rows()},
        )

        insured_object = response.context['insured_objects'][0]
        self.assertEqual(
            insured_object['object_description'],
            insured_object['description'],
        )
        self.assertIn('Lixiang L9', insured_object['object_description'])

    def test_parser_v2_does_not_take_name_header_as_telematics_complex(self):
        self.client.login(username='parser_v2_root', password='pwd')

        response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_telematics_template_header()},
        )

        telematics_complex = response.context['form'].initial['telematics_complex']
        self.assertEqual(telematics_complex, 'StarLine M96')
        self.assertNotEqual(telematics_complex, 'Наименование')

    def test_parser_v2_franchise_uses_marked_column_not_all_template_labels(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['D28'] = 'Без франшизы'
        sheet['E28'] = 'Франшиза в процентах от страховой суммы'
        sheet['F28'] = 'Абсолютная сумма'
        sheet['D29'] = 'X'

        result = self._parse_workbook_v2(workbook)

        self.assertEqual(result.data['franchise_type'], 'none')
        self.assertIn('D29', result.source_map['franchise_type'])
        self.assertNotEqual(result.data['franchise_type'], 'both_variants')

    def test_parser_v2_franchise_ignores_unmarked_template_labels(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['D28'] = 'Без франшизы'
        sheet['E28'] = 'Франшиза в процентах от страховой суммы'
        sheet['F28'] = 'Абсолютная сумма'

        result = self._parse_workbook_v2(workbook)

        self.assertEqual(result.data['franchise_type'], 'none')
        self.assertNotIn('franchise_type', result.source_map)

    def test_parser_v2_franchise_detects_percent_and_absolute_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['D28'] = 'Без франшизы'
        sheet['E28'] = 'Франшиза в процентах от страховой суммы'
        sheet['F28'] = 'Абсолютная сумма'
        sheet['E29'] = 'X'

        percent_result = self._parse_workbook_v2(workbook)
        self.assertEqual(percent_result.data['franchise_type'], 'with_franchise')
        self.assertIn('E29', percent_result.source_map['franchise_type'])

        sheet['E29'] = None
        sheet['F29'] = '50 000'

        absolute_result = self._parse_workbook_v2(workbook)
        self.assertEqual(absolute_result.data['franchise_type'], 'with_franchise')
        self.assertIn('F29', absolute_result.source_map['franchise_type'])

    def test_parser_v2_franchise_detects_both_variants(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['D28'] = 'Без франшизы'
        sheet['F28'] = 'Абсолютная сумма'
        sheet['D29'] = 'X'
        sheet['F29'] = '50 000'

        result = self._parse_workbook_v2(workbook)

        self.assertEqual(result.data['franchise_type'], 'both_variants')
        self.assertIn('D29', result.source_map['franchise_type'])
        self.assertIn('F29', result.source_map['franchise_type'])

    def test_parser_v2_franchise_uses_shifted_row_for_individual_entrepreneur(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Заявка от ИП Иванов И.И.'
        sheet['D29'] = 'Без франшизы'
        sheet['E29'] = 'Франшиза в процентах от страховой суммы'
        sheet['F29'] = 'Абсолютная сумма'
        sheet['E30'] = 'X'

        result = self._parse_workbook_v2(workbook)

        self.assertEqual(result.data['franchise_type'], 'with_franchise')
        self.assertIn('E30', result.source_map['franchise_type'])
        self.assertEqual(result.data['parser_v2_payload']['franchise_details']['value_row'], 30)

    def test_parser_v2_creates_request_from_preview_and_keeps_original_attachment(self):
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Ромашка Проверено'

        create_response = self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created_request = InsuranceRequest.objects.get()
        self.assertRedirects(
            create_response,
            reverse('insurance_requests:request_detail', kwargs={'pk': created_request.pk}),
        )
        self.assertEqual(created_request.client_name, 'ООО Ромашка Проверено')
        self.assertEqual(created_request.manager_name, 'Иванов Иван')
        self.assertEqual(created_request.branch, 'Казань')
        self.assertEqual(created_request.additional_data['parser_version'], 'v2')
        self.assertEqual(created_request.additional_data['parser_v2']['version'], '2.0.0')
        self.assertTrue(RequestAttachment.objects.filter(request=created_request).exists())

    def test_parser_v2_tracks_scalar_edit_and_stores_original_snapshot(self):
        """Фаза 1: ручная правка основного поля попадает в tracking, полный
        снимок распознанных данных сохраняется, а блок виден в карточке."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Лютик'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        # Полный снимок «до» сохранён.
        self.assertEqual(
            created.additional_data['parser_v2']['original_data']['client_name'],
            'ООО Ромашка',
        )
        # Правка отслежена.
        edits = created.parser_v2_field_edits
        client_edits = [e for e in edits if e['field'] == 'client_name']
        self.assertEqual(len(client_edits), 1)
        self.assertEqual(client_edits[0]['original'], 'ООО Ромашка')
        self.assertEqual(client_edits[0]['modified'], 'ООО Лютик')
        self.assertEqual(client_edits[0]['edit_type'], 'changed')
        self.assertGreaterEqual(created.parser_v2_edit_count, 1)

        # Блок «Ручные правки оператора» рендерится в карточке заявки.
        detail = self.client.get(
            reverse('insurance_requests:request_detail', kwargs={'pk': created.pk})
        )
        self.assertContains(detail, 'Ручные правки оператора')
        self.assertContains(detail, 'ООО Лютик')

    def test_parser_v2_no_false_positive_edits_when_unchanged(self):
        """Фаза 1: если оператор ничего не правил, tracking пуст — плейсхолдеры
        to_request_fields() не создают ложных правок."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertEqual(
            created.parser_v2_edit_count, 0,
            msg=f'Неожиданные правки: {created.parser_v2_all_edits}',
        )

    def test_parser_v2_tracks_object_edit_per_sibling(self):
        """Фаза 1: объектная правка привязана к нужной заявке партии по item_no."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=2)},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['objects-0-brand'] = 'LADA (исправлено)'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        first = InsuranceRequest.objects.get(item_no=1)
        second = InsuranceRequest.objects.get(item_no=2)
        first_brand_edits = [e for e in first.parser_v2_object_edits if e['field'] == 'brand']
        self.assertEqual(len(first_brand_edits), 1)
        self.assertEqual(first_brand_edits[0]['modified'], 'LADA (исправлено)')
        # Вторая сестра не правилась.
        self.assertEqual(second.parser_v2_object_edits, [])

    def test_parser_v2_comparison_page_renders_three_columns(self):
        """Страница сравнения показывает три состояния поля и исходный Excel."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Лютик'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)
        created = InsuranceRequest.objects.get()

        response = self.client.get(
            reverse('insurance_requests:request_comparison', kwargs={'pk': created.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'insurance_requests/request_comparison.html')
        # Видны все три колонки и значения распознанного/текущего.
        self.assertContains(response, 'Распознано из Excel')
        self.assertContains(response, 'Оператор при создании')
        self.assertContains(response, 'Текущее значение')
        self.assertContains(response, 'ООО Ромашка')
        self.assertContains(response, 'ООО Лютик')
        # Ссылка на исходный Excel присутствует.
        self.assertContains(response, 'Исходный Excel')
        self.assertEqual(response.context['changed_count'], 1)
        self.assertEqual(response.context['changed_after_count'], 0)

    def test_parser_v2_comparison_redirects_for_non_v2_request(self):
        """Фаза 2: для не-V2 заявки страница сравнения уводит на карточку."""
        self.client.login(username='parser_v2_root', password='pwd')
        legacy = InsuranceRequest.objects.create(
            client_name='Старая заявка', inn='1', additional_data={}
        )
        response = self.client.get(
            reverse('insurance_requests:request_comparison', kwargs={'pk': legacy.pk})
        )
        self.assertRedirects(
            response,
            reverse('insurance_requests:request_detail', kwargs={'pk': legacy.pk}),
        )

    def test_parser_v2_sets_manual_edits_count_per_sibling(self):
        """Фаза 3: денормализованный счётчик правок выставляется на каждую заявку."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=2)},
        )
        post_data = self._post_data_from_preview(upload_response)
        # Одна общая правка + одна объектная правка у первой сестры.
        post_data['client_name'] = 'ООО Лютик'
        post_data['objects-0-brand'] = 'LADA (исправлено)'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        first = InsuranceRequest.objects.get(item_no=1)
        second = InsuranceRequest.objects.get(item_no=2)
        # Первая: общая (client_name) + объектная (brand) = 2.
        self.assertEqual(first.manual_edits_count, 2)
        self.assertEqual(first.manual_edits_count, first.parser_v2_edit_count)
        # Вторая: только общая правка = 1.
        self.assertEqual(second.manual_edits_count, 1)

    def test_parser_v2_zero_edits_count_when_unchanged(self):
        """Фаза 3: без правок счётчик 0."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertEqual(created.manual_edits_count, 0)

    def test_request_list_shows_manual_edits_badge(self):
        """Фаза 3: бейдж правок в списке ведёт на страницу сравнения."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Лютик'
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)
        created = InsuranceRequest.objects.get()

        response = self.client.get(reverse('insurance_requests:request_list'))
        self.assertContains(
            response,
            reverse('insurance_requests:request_comparison', kwargs={'pk': created.pk}),
        )

    def test_parser_v2_records_field_edit_rows(self):
        """Фаза 4: строки RequestFieldEdit пишутся (общие — раз на партию)."""
        from .models import RequestFieldEdit
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=2)},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Лютик'           # общая правка
        post_data['objects-0-brand'] = 'LADA (исправлено)'  # объектная правка первой сестры

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        first = InsuranceRequest.objects.get(item_no=1)
        second = InsuranceRequest.objects.get(item_no=2)
        # Общая правка записана один раз — к первой заявке партии.
        common = RequestFieldEdit.objects.filter(scope='common')
        self.assertEqual(common.count(), 1)
        self.assertEqual(common.first().request_id, first.pk)
        self.assertEqual(common.first().field_name, 'client_name')
        # Объектная правка — у своей сестры.
        object_rows = RequestFieldEdit.objects.filter(scope='object')
        self.assertEqual(object_rows.count(), 1)
        self.assertEqual(object_rows.first().request_id, first.pk)
        self.assertEqual(object_rows.first().field_name, 'brand')
        # У второй сестры собственных правок нет.
        self.assertFalse(RequestFieldEdit.objects.filter(request=second).exists())

    def test_parser_v2_sets_parser_confidence(self):
        """Фаза 4: денормализованная уверенность парсера сохраняется на заявке."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertIsNotNone(created.parser_confidence)
        self.assertGreaterEqual(created.parser_confidence, 0.0)
        self.assertLessEqual(created.parser_confidence, 1.0)

    def test_admin_shows_edits_and_filter(self):
        """Фаза 3: админка фильтрует по правкам и показывает их таблицу."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload()},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['client_name'] = 'ООО Лютик'
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)
        created = InsuranceRequest.objects.get()

        # Read-only таблица правок на странице заявки в админке.
        change_url = reverse('admin:insurance_requests_insurancerequest_change', args=[created.pk])
        change_response = self.client.get(change_url)
        self.assertEqual(change_response.status_code, 200)
        self.assertContains(change_response, 'ООО Лютик')

        # Фильтр «С правками» оставляет заявку, «Без правок» — скрывает.
        changelist_url = reverse('admin:insurance_requests_insurancerequest_changelist')
        with_edits = self.client.get(changelist_url, {'has_manual_edits': 'yes'})
        self.assertContains(with_edits, created.get_display_name())
        without_edits = self.client.get(changelist_url, {'has_manual_edits': 'no'})
        self.assertNotContains(without_edits, 'ООО Лютик')

    def _xlsx_upload_with_multiple_objects(self, object_count=3, filename='partia.xlsx'):
        """Build a synthetic xlsx with several object rows so the parser
        finds an N-object batch."""
        wb = Workbook()
        sheet = wb.active
        sheet['C4'] = 'Казанский филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['B14'] = 'Страхователь'
        sheet['D14'] = 'ЛизингоДАТЕЛЬ'
        sheet['E14'] = 'ЛизингоПОЛУЧАТЕЛЬ'
        sheet['D15'] = 'Х'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        # Object table header.
        sheet['C41'] = 'Наименование и описание имущества'
        sheet['J41'] = 'Год выпуска'
        sheet['L41'] = 'Мощность'
        sheet['M41'] = 'Стоимость на момент приобретения'
        sheet['N41'] = 'Валюта'
        # Object rows at 43, 45, 47, 49 — stop at the requested count.
        descriptions = [
            ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000),
            ('Toyota Camry XV70', 2023, 'новое', '249', 4500000),
            ('Haval H7', 2025, 'новое', '170', 3649000),
            ('Mercedes-Benz Sprinter', 2022, 'б/у', '170', 5200000),
        ]
        for idx, (desc, year, cond, power, cost) in enumerate(descriptions[:object_count]):
            row = 43 + idx * 2
            sheet[f'C{row}'] = desc
            sheet[f'J{row}'] = year
            sheet[f'K{row}'] = cond
            sheet[f'L{row}'] = power
            sheet[f'M{row}'] = cost
            sheet[f'N{row}'] = 'руб'
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            filename,
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_parser_v2_splits_multi_object_file_into_sibling_requests(self):
        """Stage 4.1: 3 objects → 3 requests sharing one source_batch_id."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=3)},
        )

        # Preview must announce a batch and the submit button must reflect it.
        self.assertContains(upload_response, 'Партия объектов')
        self.assertContains(upload_response, 'Создать партию из 3 заявок')

        post_data = self._post_data_from_preview(upload_response)

        create_response = self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        # Three sibling rows must be created.
        siblings = InsuranceRequest.objects.order_by('item_no')
        self.assertEqual(siblings.count(), 3)

        # Common fields are duplicated.
        for s in siblings:
            self.assertEqual(s.client_name, 'ООО Ромашка')
            self.assertEqual(s.inn, '1234567890')
            self.assertEqual(s.branch, 'Казань')
            self.assertEqual(s.insured_party, 'lessor')

        # Batch identity is consistent across siblings.
        batch_ids = {s.source_batch_id for s in siblings}
        self.assertEqual(len(batch_ids), 1)
        self.assertIsNotNone(siblings.first().source_batch_id)
        item_nos = [s.item_no for s in siblings]
        self.assertEqual(item_nos, [1, 2, 3])
        for s in siblings:
            self.assertEqual(s.item_count, 3)

        # Per-object fields differ.
        first, second, third = siblings
        self.assertEqual(first.brand, 'LADA')
        self.assertEqual(second.brand, 'Toyota')
        self.assertEqual(third.brand, 'Haval')
        self.assertEqual(first.acquisition_cost_value, Decimal('1490000'))
        self.assertEqual(second.acquisition_cost_value, Decimal('4500000'))
        self.assertEqual(third.acquisition_cost_value, Decimal('3649000'))
        for s in siblings:
            self.assertEqual(s.acquisition_cost_currency, 'RUB')

        # Original Excel is attached to every sibling.
        self.assertEqual(RequestAttachment.objects.count(), 3)
        for s in siblings:
            self.assertTrue(RequestAttachment.objects.filter(request=s).exists())

        # The view redirects to the first request of the batch.
        self.assertRedirects(
            create_response,
            reverse('insurance_requests:request_detail', kwargs={'pk': first.pk}),
        )

    def test_parser_v2_single_object_file_keeps_batch_fields_null(self):
        """Stage 4.1: 1 object → 1 request, source_batch_id stays NULL."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=1)},
        )
        # No "Партия объектов" block for single-object uploads.
        self.assertNotContains(upload_response, 'Партия объектов')
        self.assertContains(upload_response, 'Создать заявку')

        post_data = self._post_data_from_preview(upload_response)
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertIsNone(created.source_batch_id)
        self.assertIsNone(created.item_no)
        self.assertIsNone(created.item_count)
        # Object fields are still filled (single-object stage 4.1 still uses the payload).
        self.assertEqual(created.brand, 'LADA')
        self.assertEqual(created.acquisition_cost_value, Decimal('1490000'))
        self.assertEqual(RequestAttachment.objects.count(), 1)

    def test_parser_v2_can_skip_a_sibling_in_the_formset(self):
        """Stage 4.2: marking one card as skip removes it from the created batch."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=3)},
        )
        post_data = self._post_data_from_preview(upload_response)
        # Drop the middle sibling.
        post_data['objects-1-skip'] = 'on'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        siblings = list(InsuranceRequest.objects.order_by('item_no'))
        self.assertEqual(len(siblings), 2)
        # Both surviving siblings share the same batch and have item_count=2.
        self.assertEqual({s.source_batch_id for s in siblings}, {siblings[0].source_batch_id})
        self.assertIsNotNone(siblings[0].source_batch_id)
        self.assertEqual([s.item_no for s in siblings], [1, 2])
        for s in siblings:
            self.assertEqual(s.item_count, 2)
        # First and third objects survive, second (Toyota) was skipped.
        brands = [s.brand for s in siblings]
        self.assertIn('LADA', brands)
        self.assertIn('Haval', brands)
        self.assertNotIn('Toyota', brands)
        self.assertEqual(RequestAttachment.objects.count(), 2)

    def test_parser_v2_skipping_all_siblings_blocks_creation(self):
        """Stage 4.2: if every card is skipped, nothing is created."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=3)},
        )
        post_data = self._post_data_from_preview(upload_response)
        for index in range(3):
            post_data[f'objects-{index}-skip'] = 'on'

        response = self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        self.assertEqual(InsuranceRequest.objects.count(), 0)
        self.assertEqual(RequestAttachment.objects.count(), 0)
        # User is shown the preview again with an explicit error.
        self.assertContains(response, 'Все объекты партии отмечены к пропуску')

    def test_parser_v2_operator_edits_propagate_into_created_sibling(self):
        """Stage 4.2: edits in the formset overwrite parser payload defaults."""
        self.client.login(username='parser_v2_root', password='pwd')
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_multiple_objects(object_count=2)},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['objects-0-brand'] = 'LADA (исправлено)'
        post_data['objects-0-acquisition_cost_value'] = '2500000'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        first = InsuranceRequest.objects.get(item_no=1)
        self.assertEqual(first.brand, 'LADA (исправлено)')
        self.assertEqual(first.acquisition_cost_value, Decimal('2500000'))
        # Second sibling untouched.
        second = InsuranceRequest.objects.get(item_no=2)
        self.assertEqual(second.brand, 'Toyota')

    def _xlsx_upload_with_object_specs(self, specs, filename='multiplicity.xlsx'):
        """Build a synthetic xlsx whose object rows are given verbatim by
        `specs` (list of (desc, year, condition, power, cost) tuples), so the
        test can place fully identical rows next to distinct ones."""
        wb = Workbook()
        sheet = wb.active
        sheet['C4'] = 'Казанский филиал'
        sheet['C5'] = 'Иванов Иван'
        sheet['D7'] = 'ООО Ромашка'
        sheet['D9'] = '1234567890'
        sheet['D21'] = 'КАСКО'
        sheet['N17'] = '1 год'
        sheet['C41'] = 'Наименование и описание имущества'
        sheet['J41'] = 'Год выпуска'
        sheet['L41'] = 'Мощность'
        sheet['M41'] = 'Стоимость на момент приобретения'
        sheet['N41'] = 'Валюта'
        for idx, (desc, year, cond, power, cost) in enumerate(specs):
            row = 43 + idx * 2
            sheet[f'C{row}'] = desc
            sheet[f'J{row}'] = year
            sheet[f'K{row}'] = cond
            sheet[f'L{row}'] = power
            sheet[f'M{row}'] = cost
            sheet[f'N{row}'] = 'руб'
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            filename,
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_parser_v2_identical_rows_create_single_request_with_multiplicity(self):
        """Three identical rows → one request, source_object_count=3, no batch."""
        self.client.login(username='parser_v2_root', password='pwd')
        identical = ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000)
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_object_specs([identical, identical, identical])},
        )
        # Preview shows one card, not a 3-request batch.
        self.assertNotContains(upload_response, 'Партия объектов')
        self.assertContains(upload_response, 'Создать заявку')
        self.assertContains(upload_response, 'Одинаковых строк: 3')

        post_data = self._post_data_from_preview(upload_response)
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertEqual(created.source_object_count, 3)
        self.assertIsNone(created.source_batch_id)
        self.assertIsNone(created.item_no)
        self.assertIsNone(created.item_count)
        grouping = created.additional_data['parser_v2']['parsed_payload']['object_grouping']
        self.assertEqual(grouping['raw_object_count'], 3)
        self.assertEqual(grouping['unique_object_count'], 1)

    def test_parser_v2_mixed_file_groups_duplicates_into_sibling_multiplicities(self):
        """A, A, A, B, C, C → three siblings with multiplicities [3, 1, 2]."""
        self.client.login(username='parser_v2_root', password='pwd')
        a = ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000)
        b = ('Toyota Camry XV70', 2023, 'новое', '249', 4500000)
        c = ('Haval H7', 2025, 'новое', '170', 3649000)
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_object_specs([a, a, a, b, c, c])},
        )
        self.assertContains(upload_response, 'Создать партию из 3 заявок')

        post_data = self._post_data_from_preview(upload_response)
        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        siblings = list(InsuranceRequest.objects.order_by('item_no'))
        self.assertEqual(len(siblings), 3)
        self.assertEqual([s.brand for s in siblings], ['LADA', 'Toyota', 'Haval'])
        self.assertEqual([s.source_object_count for s in siblings], [3, 1, 2])
        self.assertEqual([s.item_no for s in siblings], [1, 2, 3])
        for s in siblings:
            self.assertEqual(s.item_count, 3)
        self.assertEqual(len({s.source_batch_id for s in siblings}), 1)

    def test_parser_v2_operator_can_override_multiplicity_in_preview(self):
        """Parser sets source_object_count=3; operator edits it to 2."""
        self.client.login(username='parser_v2_root', password='pwd')
        identical = ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000)
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_object_specs([identical, identical, identical])},
        )
        post_data = self._post_data_from_preview(upload_response)
        post_data['objects-0-source_object_count'] = '2'

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertEqual(created.source_object_count, 2)

    def test_parser_v2_skip_after_grouping_leaves_single_request(self):
        """A, A, B → two cards (A×2, B×1); skipping A leaves one request."""
        self.client.login(username='parser_v2_root', password='pwd')
        a = ('LADA Largus KS045L', 2024, 'б/у', '78.05', 1490000)
        b = ('Toyota Camry XV70', 2023, 'новое', '249', 4500000)
        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': self._xlsx_upload_with_object_specs([a, a, b])},
        )
        # Two cards: A×2 and B×1.
        self.assertContains(upload_response, 'Создать партию из 2 заявок')
        self.assertContains(upload_response, 'Одинаковых строк: 2')

        post_data = self._post_data_from_preview(upload_response)
        post_data['objects-0-skip'] = 'on'  # drop the A×2 card

        self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created = InsuranceRequest.objects.get()
        self.assertEqual(created.brand, 'Toyota')
        self.assertEqual(created.source_object_count, 1)
        self.assertIsNone(created.source_batch_id)
        self.assertIsNone(created.item_count)

    def test_parser_v2_can_create_minimal_request_after_unreadable_file(self):
        self.client.login(username='parser_v2_root', password='pwd')
        unreadable_file = SimpleUploadedFile(
            'broken.xlsx',
            b'not an excel file',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        upload_response = self.client.post(
            reverse('insurance_requests:upload_excel_v2'),
            {'excel_file': unreadable_file},
        )

        self.assertEqual(upload_response.status_code, 200)
        self.assertContains(upload_response, 'Файл не удалось прочитать')

        post_data = self._post_data_from_preview(upload_response.context['form'])
        create_response = self.client.post(reverse('insurance_requests:upload_excel_v2'), post_data)

        created_request = InsuranceRequest.objects.get()
        self.assertRedirects(
            create_response,
            reverse('insurance_requests:request_detail', kwargs={'pk': created_request.pk}),
        )
        self.assertEqual(created_request.client_name, 'Клиент не указан')
        self.assertEqual(created_request.dfa_number, 'Номер ДФА не указан')
        self.assertTrue(created_request.additional_data['parser_v2']['warnings'])
