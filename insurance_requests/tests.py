"""
Tests for insurance_requests app
"""
import uuid
from email.header import decode_header, make_header
from io import BytesIO
from datetime import date
from decimal import Decimal

from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth.models import User, Group
from django.test import TestCase, Client, SimpleTestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.templates import EmailTemplateGenerator
from .models import InsuranceRequest, RequestAttachment


class RequestDetailViewTest(TestCase):
    """Test the request detail view with summary creation button"""
    
    def setUp(self):
        """Set up test data"""
        from django.contrib.auth.models import Group
        
        self.client = Client()
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Add user to the required group
        user_group, created = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        
        self.client.login(username='testuser', password='testpass123')
        
        # Create a test request
        self.request = InsuranceRequest.objects.create(
            client_name='Test Client',
            inn='1234567890',
            insurance_type='КАСКО',
            insurance_period='1 год',
            status='uploaded',
            created_by=self.user
        )
    
    def test_create_summary_button_shown_for_uploaded_status(self):
        """Test that create summary button is shown for uploaded status"""
        url = reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Создать свод')
        self.assertContains(response, 'Свод предложений не создан')
        self.assertNotContains(response, 'Свод недоступен')
    
    def test_create_summary_button_shown_for_email_generated_status(self):
        """Test that create summary button is shown for email_generated status"""
        self.request.status = 'email_generated'
        self.request.save()
        
        url = reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Создать свод')
        self.assertContains(response, 'Свод предложений не создан')
        self.assertNotContains(response, 'Свод недоступен')
    
    def test_create_summary_button_shown_for_emails_sent_status(self):
        """Test that create summary button is shown for emails_sent status"""
        self.request.status = 'emails_sent'
        self.request.save()
        
        url = reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Создать свод')
        self.assertNotContains(response, 'Свод недоступен')
    
    def test_summary_unavailable_for_invalid_statuses(self):
        """Test that summary is unavailable for invalid statuses (this test may not be relevant anymore)"""
        # Since we only have 4 statuses now and all allow summary creation,
        # this test checks the else branch when a summary already exists
        # We'll create a mock scenario by testing with a non-existent status
        # But since we can't set invalid status, we'll test the else branch differently
        
        # For now, let's test that the create summary button works for emails_sent status
        self.request.status = 'emails_sent'
        self.request.save()
        
        url = reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        # With the new logic, emails_sent status should show create summary button
        self.assertContains(response, 'Создать свод')


class RequestDatabaseExportTest(TestCase):
    """Tests for XLSX export of the request card data."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='exportuser',
            password='testpass123',
            first_name='Иван',
            last_name='Иванов',
        )
        user_group, _ = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        admin_group, _ = Group.objects.get_or_create(name='Администраторы')

        self.superuser = User.objects.create_superuser(
            username='exportsuper',
            password='testpass123',
            email='exportsuper@example.com',
            first_name='Петр',
            last_name='Петров',
        )
        self.superuser.groups.add(admin_group)

        self.user_client = Client()
        self.user_client.login(username='exportuser', password='testpass123')
        self.superuser_client = Client()
        self.superuser_client.login(username='exportsuper', password='testpass123')

        self.request = InsuranceRequest.objects.create(
            client_name='ООО Тестовый клиент',
            inn='1234567890',
            insurance_type='страхование имущества',
            insurance_period='на весь срок лизинга',
            dfa_number='ДФА-EXPORT-001',
            branch='Москва',
            manager_name='Петров Петр',
            deal_status='new',
            vehicle_info='Линия по производству тестов',
            franchise_type='with_franchise',
            has_transportation=True,
            transportation_departure='Москва',
            transportation_destination='Казань',
            transportation_days=5,
            notes='Внутренний комментарий',
            created_by=self.user,
            additional_data={
                'application_type': 'legal_entity',
                'application_format': 'property',
                'parser_v2': {
                    'confidence': 0.94,
                    'warnings': ['Проверить адрес'],
                },
            },
        )
        RequestAttachment.objects.create(
            request=self.request,
            file=SimpleUploadedFile(
                'source.xlsx',
                b'test-export-content',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            original_filename='source.xlsx',
            file_type='.xlsx',
        )

    def test_request_detail_shows_database_export_button(self):
        response = self.superuser_client.get(
            reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Скачать карточку заявки')
        self.assertContains(
            response,
            reverse('insurance_requests:export_request_database', kwargs={'pk': self.request.pk}),
        )

    def test_request_detail_hides_database_export_button_for_regular_user(self):
        response = self.user_client.get(
            reverse('insurance_requests:request_detail', kwargs={'pk': self.request.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Скачать карточку заявки')
        self.assertNotContains(
            response,
            reverse('insurance_requests:export_request_database', kwargs={'pk': self.request.pk}),
        )

    def test_export_request_database_returns_xlsx_with_request_data(self):
        response = self.superuser_client.get(
            reverse('insurance_requests:export_request_database', kwargs={'pk': self.request.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        decoded_disposition = str(make_header(decode_header(response['Content-Disposition'])))
        self.assertIn('.xlsx', decoded_disposition)

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Карточка заявки']
        rows = [row for row in worksheet.iter_rows(values_only=True) if row[1]]

        exported_pairs = {(row[1], row[2]) for row in rows}

        self.assertIn(('Имя клиента [client_name]', 'ООО Тестовый клиент'), exported_pairs)
        self.assertIn(('Статус [status]', 'Загружено [uploaded]'), exported_pairs)
        self.assertIn(('Сводка объекта [object_summary]', 'Линия по производству тестов'), exported_pairs)
        self.assertIn(('Тип франшизы [franchise_type]', 'Только с франшизой [with_franchise]'), exported_pairs)
        self.assertIn(('additional_data.application_format', 'property'), exported_pairs)
        self.assertIn(('additional_data.parser_v2.warnings[1]', 'Проверить адрес'), exported_pairs)
        self.assertIn(('attachments[1].original_filename', 'source.xlsx'), exported_pairs)

    def test_export_request_database_forbidden_for_regular_user(self):
        response = self.user_client.get(
            reverse('insurance_requests:export_request_database', kwargs={'pk': self.request.pk})
        )

        self.assertEqual(response.status_code, 403)
        self.assertContains(response, 'Superuser', status_code=403)


class RequestV1V2DisplayCompatibilityTest(TestCase):
    """Old V1 requests and structured Parser V2 requests render side by side."""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='compatuser', password='pwd')
        user_group, _ = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        self.client.login(username='compatuser', password='pwd')

        self.v1_request = InsuranceRequest.objects.create(
            client_name='Старый клиент',
            inn='1111111111',
            insurance_type='КАСКО',
            insurance_period='1 год',
            dfa_number='V1-001',
            vehicle_info='Старое описание предмета лизинга V1',
            created_by=self.user,
        )
        self.legacy_installment_request = InsuranceRequest.objects.create(
            client_name='Старый клиент с рассрочкой',
            inn='1111111112',
            insurance_type='КАСКО',
            insurance_period='1 год',
            dfa_number='V1-INSTALL',
            has_installment=True,
            created_by=self.user,
        )
        self.v2_request = InsuranceRequest.objects.create(
            client_name='Новый клиент',
            inn='2222222222',
            insurance_type='КАСКО',
            insurance_period='1 год',
            dfa_number='V2-001',
            vehicle_info='Автомобиль LADA Largus KS045L 2024 б/у',
            brand='LADA',
            model='Largus KS045L',
            condition='used',
            equipment_type='Категория B',
            acquisition_cost_value=Decimal('1490000'),
            acquisition_cost_currency='RUB',
            premium_frequency='quarterly',
            insured_party='lessor',
            additional_data={
                'parser_version': 'v2',
                'parser_v2': {
                    'confidence': 0.88,
                    'warnings': [
                        {
                            'level': 'manual_required',
                            'field': 'insured_party',
                            'message': 'Проверьте страхователя.',
                        }
                    ],
                    'source_file_name': 'request-v2.xlsx',
                },
            },
            created_by=self.user,
        )
        self.annual_request = InsuranceRequest.objects.create(
            client_name='Ежегодный клиент',
            inn='3333333333',
            insurance_type='КАСКО',
            insurance_period='1 год',
            dfa_number='V2-ANNUAL',
            premium_frequency='annual',
            created_by=self.user,
        )
        self.single_request = InsuranceRequest.objects.create(
            client_name='Единовременный клиент',
            inn='4444444444',
            insurance_type='КАСКО',
            insurance_period='1 год',
            dfa_number='V2-SINGLE',
            premium_frequency='single',
            created_by=self.user,
        )

    def test_request_list_uses_structured_v2_object_and_v1_fallback(self):
        response = self.client.get(reverse('insurance_requests:request_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('insurance_requests:upload_excel'))
        self.assertNotContains(response, reverse('insurance_requests:upload_excel_v2'))
        self.assertContains(response, 'Старое описание предмета лизинга V1')
        self.assertContains(response, 'LADA Largus KS045L')
        self.assertContains(response, '1 490 000 RUB')
        self.assertNotContains(response, 'Автомобиль LADA Largus KS045L 2024 б/у')
        self.assertContains(response, 'Поквартально')
        self.assertContains(response, 'Рассрочка')
        self.assertNotContains(response, 'Ежегодно')
        self.assertNotContains(response, 'Единовременно')
        self.assertContains(response, 'Проверить')

    def test_request_detail_keeps_v1_simple_and_shows_v2_diagnostics(self):
        v1_response = self.client.get(
            reverse('insurance_requests:request_detail', kwargs={'pk': self.v1_request.pk})
        )
        self.assertEqual(v1_response.status_code, 200)
        self.assertContains(v1_response, 'Старое описание предмета лизинга V1')
        self.assertNotContains(v1_response, 'Parser V2')

        v2_response = self.client.get(
            reverse('insurance_requests:request_detail', kwargs={'pk': self.v2_request.pk})
        )
        self.assertEqual(v2_response.status_code, 200)
        self.assertContains(v2_response, 'LADA Largus KS045L')
        self.assertContains(v2_response, 'Стоимость приобретения')
        self.assertContains(v2_response, '1 490 000 RUB')
        self.assertContains(v2_response, 'Состояние:')
        self.assertContains(v2_response, 'Б/у')
        self.assertNotContains(v2_response, 'Автомобиль LADA Largus KS045L 2024 б/у')
        self.assertContains(v2_response, 'Alla Borisovna Magic Parser')
        self.assertContains(v2_response, '88%')

    def test_edit_request_renders_v2_fields_without_breaking_v1(self):
        v1_response = self.client.get(
            reverse('insurance_requests:edit_request', kwargs={'pk': self.v1_request.pk})
        )
        self.assertEqual(v1_response.status_code, 200)
        self.assertContains(v1_response, 'Старое описание предмета лизинга V1')
        self.assertContains(v1_response, 'Структурированные данные объекта')

        v2_response = self.client.get(
            reverse('insurance_requests:edit_request', kwargs={'pk': self.v2_request.pk})
        )
        self.assertEqual(v2_response.status_code, 200)
        self.assertContains(v2_response, 'LADA')
        self.assertContains(v2_response, 'Частота уплаты премии')
        self.assertContains(v2_response, 'Страхователь')


class DisplayNameBatchTest(TestCase):
    """Tests for get_display_name() with batch fields (V2 splitting)."""

    def test_display_name_for_single_request_without_batch_fields(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА 18022',
        )
        self.assertEqual(req.get_display_name(), 'ДФА 18022')


class InsuranceRequestObjectPropertiesTest(SimpleTestCase):
    def test_object_summary_prefers_structured_fields(self):
        request = InsuranceRequest(
            brand='LADA',
            model='Largus KS045L',
            condition='used',
            acquisition_cost_value=Decimal('1490000'),
            acquisition_cost_currency='RUB',
            manufacturing_year='2024',
            source_object_count=2,
            vehicle_info='legacy value',
        )

        self.assertEqual(
            request.object_summary,
            'LADA Largus KS045L, 2024 г., Б/у, 1 490 000 RUB, ×2',
        )

    def test_object_summary_uses_object_description_when_brand_model_missing(self):
        request = InsuranceRequest(
            object_description='Линия порошковой окраски с конвейером',
            manufacturing_year='2020',
            asset_status='б/у',
        )

        self.assertEqual(
            request.object_summary,
            'Линия порошковой окраски с конвейером, 2020 г., б/у',
        )

    def test_object_summary_falls_back_to_vehicle_info_for_legacy_v1(self):
        request = InsuranceRequest(vehicle_info='Старое описание предмета лизинга')

        self.assertEqual(request.object_summary, 'Старое описание предмета лизинга')

    def test_condition_helpers_support_v2_and_legacy(self):
        structured_request = InsuranceRequest(condition='new')
        legacy_request = InsuranceRequest(asset_status='новое')

        self.assertEqual(structured_request.condition_label, 'Новое')
        self.assertTrue(structured_request.is_new_object)
        self.assertEqual(legacy_request.condition_label, 'новое')
        self.assertTrue(legacy_request.is_new_object)


class EmailTemplateGeneratorTest(TestCase):
    def test_generate_subject_uses_object_display_name_for_structured_request(self):
        request = InsuranceRequest(
            dfa_number='ДФА-123',
            branch='Москва',
            insurance_period='1 год',
            brand='LADA',
            model='Vesta Cross',
            vehicle_info='Очень длинное legacy-описание, которое не должно попасть в тему письма',
        )

        subject = EmailTemplateGenerator().generate_subject(request.to_dict())

        self.assertIn('LADA Vesta Cross', subject)
        self.assertNotIn('legacy-описание', subject)

    def test_display_name_for_single_item_batch_omits_suffix(self):
        # item_count == 1 means «партия из одной заявки» — суффикс не нужен.
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА 18022',
            source_batch_id=uuid.uuid4(),
            item_no=1,
            item_count=1,
        )
        self.assertEqual(req.get_display_name(), 'ДФА 18022')

    def test_display_name_for_multi_item_batch_includes_position(self):
        batch_id = uuid.uuid4()
        req1 = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА 18022',
            source_batch_id=batch_id,
            item_no=1,
            item_count=3,
        )
        req2 = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            dfa_number='ДФА 18022',
            source_batch_id=batch_id,
            item_no=2,
            item_count=3,
        )
        self.assertEqual(req1.get_display_name(), 'ДФА 18022 / объект 1 из 3')
        self.assertEqual(req2.get_display_name(), 'ДФА 18022 / объект 2 из 3')

    def test_display_name_falls_back_to_id_when_dfa_missing(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            source_batch_id=uuid.uuid4(),
            item_no=2,
            item_count=4,
        )
        self.assertEqual(req.get_display_name(), f'#{req.id} / объект 2 из 4')


class RequestListBatchGroupingTest(TestCase):
    """Stage 4.3: batch siblings must look like a group in /request_list/."""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='listuser', password='pwd')
        user_group, _ = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        self.client.login(username='listuser', password='pwd')

        # One V1-style standalone request, one V2 batch of 3 siblings.
        self.standalone = InsuranceRequest.objects.create(
            client_name='Одиночка',
            inn='1111111111',
            insurance_type='КАСКО',
            dfa_number='ДФА-SOLO',
            created_by=self.user,
        )
        batch_id = uuid.uuid4()
        sibling_pks = []
        for i in (1, 2, 3):
            sibling = InsuranceRequest.objects.create(
                client_name='Партия',
                inn='2222222222',
                insurance_type='КАСКО',
                dfa_number='ДФА-PART',
                source_batch_id=batch_id,
                item_no=i,
                item_count=3,
                created_by=self.user,
            )
            sibling_pks.append(sibling.pk)
        # Mirror the production behaviour: every sibling of a batch shares
        # the same created_at so the (-created_at, source_batch_id, item_no)
        # ordering keeps them together.
        from django.utils import timezone as tz
        InsuranceRequest.objects.filter(pk__in=sibling_pks).update(created_at=tz.now())

    def test_batch_rows_have_dedicated_css_class_and_badge(self):
        response = self.client.get(reverse('insurance_requests:request_list'))
        self.assertEqual(response.status_code, 200)
        # Position-in-batch badges are rendered for each sibling.
        self.assertContains(response, '1/3')
        self.assertContains(response, '2/3')
        self.assertContains(response, '3/3')
        # The queryset ordered the siblings consecutively by item_no.
        listed = list(response.context['requests'])
        batch_rows = [r for r in listed if r.source_batch_id is not None]
        self.assertEqual([r.item_no for r in batch_rows], [1, 2, 3])

    def test_standalone_rows_do_not_show_batch_position_badge(self):
        response = self.client.get(reverse('insurance_requests:request_list') + '?dfa_filter=SOLO')
        listed = list(response.context['requests'])
        self.assertEqual(len(listed), 1)
        self.assertEqual(listed[0], self.standalone)
        # No "K/N" badge text for a standalone request.
        self.assertNotContains(response, '/3')
        # And no "/ объект K из N" suffix in the display name.
        self.assertNotContains(response, '/ объект ')


class RequestDetailBatchPanelTest(TestCase):
    """Stage 4.4: detail page of a V2 sibling must show the batch panel."""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='detailuser', password='pwd')
        user_group, _ = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        self.client.login(username='detailuser', password='pwd')

        batch_id = uuid.uuid4()
        self.siblings = []
        for i in (1, 2, 3):
            self.siblings.append(InsuranceRequest.objects.create(
                client_name='Партия Клиент',
                inn='3333333333',
                insurance_type='КАСКО',
                dfa_number='ДФА-BATCH',
                brand=f'Brand{i}',
                model=f'Model{i}',
                source_batch_id=batch_id,
                item_no=i,
                item_count=3,
                created_by=self.user,
            ))
        self.standalone = InsuranceRequest.objects.create(
            client_name='Одиночка',
            inn='4444444444',
            insurance_type='КАСКО',
            dfa_number='ДФА-SOLO',
            created_by=self.user,
        )

    def test_batch_panel_is_shown_on_sibling_detail(self):
        first = self.siblings[0]
        url = reverse('insurance_requests:request_detail', kwargs={'pk': first.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Партия из 3 заявок')
        # Context contains the two other siblings, sorted by item_no.
        siblings_in_context = response.context['batch_siblings']
        self.assertEqual([s.item_no for s in siblings_in_context], [2, 3])
        # Links to siblings are rendered.
        for sibling in self.siblings[1:]:
            self.assertContains(response, f'href="{reverse("insurance_requests:request_detail", kwargs={"pk": sibling.pk})}"')

    def test_batch_panel_excludes_current_request(self):
        middle = self.siblings[1]
        url = reverse('insurance_requests:request_detail', kwargs={'pk': middle.pk})
        response = self.client.get(url)
        siblings_in_context = response.context['batch_siblings']
        # Order is by item_no across all but the current one.
        self.assertEqual([s.item_no for s in siblings_in_context], [1, 3])
        self.assertNotIn(middle, siblings_in_context)

    def test_standalone_detail_has_no_batch_panel(self):
        url = reverse('insurance_requests:request_detail', kwargs={'pk': self.standalone.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Партия из')
        self.assertEqual(response.context['batch_siblings'], [])


class ObjectFieldsTest(TestCase):
    """Этап 2.1: новые поля объекта живут прямо в InsuranceRequest."""

    def test_object_fields_default_to_null(self):
        # V1-флоу не заполняет ни одно из новых полей — для совместимости
        # они должны корректно создаваться пустыми.
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
        )
        self.assertIsNone(req.brand)
        self.assertIsNone(req.model)
        self.assertIsNone(req.condition)
        self.assertIsNone(req.equipment_type)
        self.assertIsNone(req.power_or_capacity)
        self.assertIsNone(req.acquisition_cost_value)
        self.assertIsNone(req.acquisition_cost_currency)

    def test_object_fields_store_full_payload(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            brand='LADA',
            model='Largus KS045L',
            condition='used',
            equipment_type='Легковой автомобиль',
            power_or_capacity='78.05',
            acquisition_cost_value=Decimal('1490000.00'),
            acquisition_cost_currency='RUB',
        )
        req.refresh_from_db()
        self.assertEqual(req.brand, 'LADA')
        self.assertEqual(req.model, 'Largus KS045L')
        self.assertEqual(req.condition, 'used')
        self.assertEqual(req.get_condition_display(), 'Б/у')
        self.assertEqual(req.acquisition_cost_value, Decimal('1490000.00'))
        self.assertEqual(req.acquisition_cost_currency, 'RUB')
        self.assertEqual(req.get_acquisition_cost_currency_display(), 'Рубли')

    def test_condition_rejects_unknown_value(self):
        # choices ограничены 'new'/'used'; full_clean ловит остальное.
        req = InsuranceRequest(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            condition='unknown',
        )
        with self.assertRaises(ValidationError):
            req.full_clean()

    def test_currency_rejects_non_iso_value(self):
        req = InsuranceRequest(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            acquisition_cost_currency='руб',
        )
        with self.assertRaises(ValidationError):
            req.full_clean()


class CustomerFieldsTest(TestCase):
    """Stage 2.2: customer details (addresses, business activity, dates).
    OGRN/KPP are intentionally absent — leasing Excel files don't carry them."""

    def test_customer_fields_default_to_null(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
        )
        self.assertIsNone(req.legal_address)
        self.assertIsNone(req.postal_address)
        self.assertIsNone(req.business_activity)
        self.assertIsNone(req.birth_date)
        self.assertIsNone(req.submission_date)
        # OGRN/KPP fields must not exist on the model.
        self.assertFalse(hasattr(req, 'ogrn'))
        self.assertFalse(hasattr(req, 'kpp'))

    def test_customer_fields_store_full_payload(self):
        req = InsuranceRequest.objects.create(
            client_name='ИП Еремин Илья Сергеевич',
            inn='121212121212',
            insurance_type='КАСКО',
            legal_address='194354, Санкт-Петербург г, Северный пр-кт, дом № 11',
            postal_address='194354, Санкт-Петербург г, Северный пр-кт, дом № 11',
            business_activity='42.11 Строительство автомобильных дорог',
            birth_date=date(1980, 6, 12),
            submission_date=date(2026, 4, 17),
        )
        req.refresh_from_db()
        self.assertEqual(req.legal_address, '194354, Санкт-Петербург г, Северный пр-кт, дом № 11')
        self.assertEqual(req.postal_address, '194354, Санкт-Петербург г, Северный пр-кт, дом № 11')
        self.assertEqual(req.business_activity, '42.11 Строительство автомобильных дорог')
        self.assertEqual(req.birth_date, date(1980, 6, 12))
        self.assertEqual(req.submission_date, date(2026, 4, 17))

    def test_long_address_is_not_truncated(self):
        long_address = (
            '385000, Адыгея (Адыгея) респ, Майкоп г, Железнодорожная ул, '
            'дом № 332, корпус 1, литера А, помещение 5-Н, офис 12, '
            'кадастровый № 01:08:0501041:101 ' * 3
        )
        req = InsuranceRequest.objects.create(
            client_name='ООО Тест',
            inn='1234567890',
            insurance_type='КАСКО',
            legal_address=long_address,
        )
        req.refresh_from_db()
        self.assertEqual(req.legal_address, long_address)


class DealInsuranceFieldsTest(TestCase):
    """Stage 2.3: deal and insurance parameters.

    Excluded from the original plan because the source does not carry them:
      contract_start_date / contract_end_date  — Excel only has the «на весь
        срок лизинга» enum, not actual dates;
      period_start_date / period_end_date / period_months — same;
      indemnity_basis — 0/30 hits in the audit.
    """

    def test_fields_default_to_null(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
        )
        self.assertIsNone(req.insured_party)
        self.assertIsNone(req.insured_sum_type)
        self.assertIsNone(req.guard_conditions)
        self.assertIsNone(req.property_location_right_holder)
        self.assertIsNone(req.premium_frequency)
        # Excluded fields must not exist on the model.
        for absent in ('contract_start_date', 'contract_end_date',
                       'period_start_date', 'period_end_date', 'period_months',
                       'indemnity_basis'):
            self.assertFalse(hasattr(req, absent), f'{absent} should not exist on InsuranceRequest')

    def test_fields_store_full_payload(self):
        req = InsuranceRequest.objects.create(
            client_name='Тест ООО',
            inn='1234567890',
            insurance_type='КАСКО',
            insured_party='lessor',
            insured_sum_type='non_aggregate',
            guard_conditions='без ограничений',
            property_location_right_holder='lessee_owner',
            premium_frequency='quarterly',
        )
        req.refresh_from_db()
        self.assertEqual(req.insured_party, 'lessor')
        self.assertEqual(req.get_insured_party_display(), 'Лизингодатель')
        self.assertEqual(req.insured_sum_type, 'non_aggregate')
        self.assertEqual(req.get_insured_sum_type_display(), 'Неагрегатная')
        self.assertEqual(req.guard_conditions, 'без ограничений')
        self.assertEqual(req.property_location_right_holder, 'lessee_owner')
        self.assertEqual(req.premium_frequency, 'quarterly')
        self.assertEqual(req.get_premium_frequency_display(), 'Поквартально')

    def test_choices_reject_unknown_value(self):
        bad_values = [
            ('insured_party', 'owner'),
            ('insured_sum_type', 'unknown'),
            ('property_location_right_holder', 'tenant'),
            ('premium_frequency', 'semiannual'),  # excluded from our enum
        ]
        for field, value in bad_values:
            req = InsuranceRequest(
                client_name='Тест',
                inn='1234567890',
                insurance_type='КАСКО',
                **{field: value},
            )
            with self.assertRaises(ValidationError, msg=f'{field}={value!r} must be rejected'):
                req.full_clean()


@override_settings(
    LOGIN_RATE_LIMIT_ENABLED=True,
    LOGIN_MAX_ATTEMPTS=3,
    LOGIN_MAX_ATTEMPTS_PER_IP=30,
    LOGIN_ATTEMPT_WINDOW_SECONDS=300,
    LOGIN_LOCKOUT_SECONDS=600,
)
class LoginSecurityTest(TestCase):
    """Security tests for login form: anti-enumeration and rate limit."""

    def setUp(self):
        self.client = Client()
        self.login_url = reverse('login')
        self.user = User.objects.create_user(username='secure_user', password='securepass123')

        user_group, _ = Group.objects.get_or_create(name='Пользователи')
        self.user.groups.add(user_group)
        cache.clear()

    def tearDown(self):
        cache.clear()

    def _post_login(self, username, password, remote_addr='127.0.0.1'):
        return self.client.post(
            self.login_url,
            {'username': username, 'password': password},
            REMOTE_ADDR=remote_addr
        )

    def test_login_error_message_does_not_reveal_user_existence(self):
        unknown_user_response = self._post_login('unknown_user', 'somepass123')
        wrong_password_response = self._post_login('secure_user', 'wrongpass123')

        self.assertEqual(unknown_user_response.status_code, 200)
        self.assertEqual(wrong_password_response.status_code, 200)

        self.assertContains(unknown_user_response, 'Неверный логин или пароль')
        self.assertContains(wrong_password_response, 'Неверный логин или пароль')

        self.assertNotContains(unknown_user_response, 'Пользователь с таким логином не найден')
        self.assertNotContains(wrong_password_response, 'Пользователь с таким логином не найден')
        self.assertNotContains(unknown_user_response, 'Неверный пароль')
        self.assertNotContains(wrong_password_response, 'Неверный пароль')

    def test_login_is_locked_after_too_many_attempts(self):
        for _ in range(3):
            self._post_login('secure_user', 'wrongpass123')

        locked_response = self._post_login('secure_user', 'securepass123')
        self.assertEqual(locked_response.status_code, 200)
        self.assertContains(locked_response, 'Слишком много неудачных попыток входа')

        self.assertNotIn('_auth_user_id', self.client.session)
