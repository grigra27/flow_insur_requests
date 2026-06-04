"""Best-effort Excel parser for insurance request uploads V2."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
import logging
import os
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.excel_utils import AVAILABLE_BRANCHES, map_branch_name

logger = logging.getLogger(__name__)


PARSER_V2_VERSION = "2.0.0"
MISSING_CLIENT = "Клиент не указан"
MISSING_DFA = "Номер ДФА не указан"
MISSING_VEHICLE = "Предмет лизинга не указан"
CLIENT_COORDINATES = ("D7", "D8")
CLIENT_MAX_LABEL_ROW = 10
OBJECT_TEMPLATE_ROW_MARKERS = [
    "транспортные средства категории b",
    "транспортные средства категории c",
    "транспортные средства категории d",
    "специальная техника",
    "противоугонные системы и оборудование",
    "штатная",
    "установленная дополнительно",
    "название модель",
    "название модели",
    "модель сигнализации",
    "сигнализация",
    "иммобилайзер",
    "иммобилизатор",
    "мобилизатор",
    "механические противоугонные устройства",
    "спутниковая противоугонная система",
    "капот",
    "рычаг кпп",
    "прочее",
    "марка модель конфигурация",
]


@dataclass(frozen=True)
class GridCell:
    row: int
    col: int
    coordinate: str
    value: str

    @property
    def normalized(self) -> str:
        return normalize_text(self.value)


@dataclass
class ParserV2Result:
    data: Dict[str, Any]
    warnings: List[Dict[str, str]] = field(default_factory=list)
    source_map: Dict[str, str] = field(default_factory=dict)
    confidence: float = 0.0
    raw_debug: Dict[str, Any] = field(default_factory=dict)
    original_filename: str = ""
    parser_version: str = PARSER_V2_VERSION

    def to_session_dict(self) -> Dict[str, Any]:
        return {
            "parser_version": self.parser_version,
            "original_filename": self.original_filename,
            "data": self.data,
            "warnings": self.warnings,
            "source_map": self.source_map,
            "confidence": self.confidence,
            "raw_debug": self.raw_debug,
        }


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_text(value: Any) -> str:
    text = clean_value(value).lower().replace("ё", "е")
    text = text.replace("№", " no ")
    text = re.sub(r"[\n\r\t:;,.()]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# --- Object-row field extractors (stage 3.1) ---------------------------------
#
# These helpers turn a raw object row from the leasing-company Excel into the
# structured fields introduced by stage 2.1 (brand, model, vin, condition,
# acquisition_cost_*, …). They live at module level so they can be unit-tested
# in isolation and reused by tests via direct import.

_NUMERIC_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")

# Currency synonyms → ISO code. Keep keys in normalized lowercase form.
_CURRENCY_MAP = {
    "руб": "RUB",
    "рубли": "RUB",
    "рублей": "RUB",
    "руб ": "RUB",
    "rur": "RUB",
    "rub": "RUB",
    "₽": "RUB",
    "usd": "USD",
    "$": "USD",
    "долл": "USD",
    "доллар": "USD",
    "доллары": "USD",
    "долларов": "USD",
    "eur": "EUR",
    "€": "EUR",
    "евро": "EUR",
}

# Condition values (column K in the canonical CASCO/equipment layout).
_CONDITION_MAP = {
    "новое": "new",
    "новый": "new",
    "новая": "new",
    "new": "new",
    "б/у": "used",
    "бу": "used",
    "б\\у": "used",
    "б-у": "used",
    "used": "used",
}

_OBJECT_TYPE_PREFIXES = (
    ("погрузчик", "с", "бортовым", "поворотом"),
    ("легковой", "автомобиль"),
    ("грузовой", "автомобиль"),
    ("седельный", "тягач"),
    ("экскаватор-погрузчик",),
    ("полуприцеп-самосвал",),
    ("автотопливозаправщик",),
    ("асфальтоукладчик",),
    ("автомобиль",),
    ("автобус",),
    ("трактор",),
    ("тягач",),
    ("погрузчик",),
    ("экскаватор",),
    ("а/м",),
)


def normalize_currency(value: Any) -> Optional[str]:
    """Map a raw currency cell value to one of RUB/USD/EUR. Returns None if unknown."""
    if value is None:
        return None
    text = normalize_text(value).strip(" .")
    if not text:
        return None
    return _CURRENCY_MAP.get(text)


def normalize_condition(value: Any) -> Optional[str]:
    """Map a raw condition cell value to 'new'/'used'. Returns None if unknown."""
    if value is None:
        return None
    text = normalize_text(value).strip(" .")
    if not text:
        return None
    return _CONDITION_MAP.get(text)


def parse_cost_value(value: Any) -> Optional[Decimal]:
    """Convert a price-like cell ('1 490 000', '1490000,00', '147.51') to Decimal."""
    if value is None:
        return None
    text = clean_value(value)
    if not text:
        return None
    # Strip thousand separators (regular and non-breaking spaces), unify decimal mark.
    text = re.sub(r"[\s ]", "", text)
    text = text.replace(",", ".")
    if not _NUMERIC_RE.match(text):
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def classify_equipment_or_power(value: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Column L in the canonical layout is overloaded: it can hold engine power
    ('78.05', '147,51') or an equipment kind ('колесная', 'гусеничная').
    Distinguish by whether the cell parses as a single number.

    Returns (equipment_type, power_or_capacity).
    """
    if value is None:
        return None, None
    text = clean_value(value)
    if not text:
        return None, None
    # Strip spaces, normalize decimal mark before the numeric test.
    no_spaces = re.sub(r"[\s ]", "", text).replace(",", ".")
    if _NUMERIC_RE.match(no_spaces):
        return None, text
    return text, None


def split_brand_model(
    text: Any,
    year: Optional[str] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """
    Split a free-form object description into (brand, model).

    Strategy:
      - drop the year and any obvious price tokens (numbers with 4+ digits);
      - drop currency / condition markers and fractional numbers (engine power);
      - drop a generic object type prefix (e.g. 'Автомобиль', 'Трактор');
      - the first remaining token is the brand, the rest is the model;
      - if only one token survives, brand=None, model=that token (we won't guess).
    """
    if text is None:
        return None, None
    cleaned = clean_value(text)
    if not cleaned:
        return None, None
    if year:
        cleaned = re.sub(rf"\b{re.escape(year)}\b", " ", cleaned)
    # Drop tokens that are obviously not part of the brand/model:
    #   - currency / condition markers,
    #   - prices (4+ digit integers like 1490000),
    #   - fractional numbers (powers like 78.05, 147,51).
    # Short integers (G 400, X5, A4, Prado 250) stay — they belong to the model name.
    tokens: List[str] = []
    raw_tokens = cleaned.split()
    for index, raw in enumerate(raw_tokens):
        token_norm = normalize_text(raw).strip(" .")
        if not token_norm:
            continue
        if token_norm in _CURRENCY_MAP or token_norm in _CONDITION_MAP:
            continue
        normalized_number = raw.replace(",", ".")
        next_raw = raw_tokens[index + 1] if index + 1 < len(raw_tokens) else None
        if _looks_like_price_token(raw, next_raw):
            continue
        if re.fullmatch(r"\d+\.\d+", normalized_number):  # fractional = power
            continue
        tokens.append(raw)
    tokens = _strip_object_type_prefix(tokens)
    if not tokens:
        return None, None
    if len(tokens) == 1:
        return None, tokens[0]
    return tokens[0], " ".join(tokens[1:])


def _looks_like_price_token(raw: str, next_raw: Optional[str] = None) -> bool:
    if not re.fullmatch(r"\d{4,}", raw):
        return False
    if normalize_currency(next_raw):
        return True
    try:
        return Decimal(raw) >= Decimal("1000000")
    except (InvalidOperation, ValueError):
        return False


def _strip_object_type_prefix(tokens: List[str]) -> List[str]:
    normalized_tokens = [normalize_text(token).strip(" .") for token in tokens]
    for prefix in _OBJECT_TYPE_PREFIXES:
        if tuple(normalized_tokens[:len(prefix)]) == prefix:
            return tokens[len(prefix):]
    return tokens


# --- Identical-object grouping (parser_v2_identical_object_multiplicity) ------
#
# A second, smarter dedup layer on top of the exact-row_text counting done in
# `_extract_objects`. It collapses rows that are *business-equivalent* after
# normalization (e.g. differ only in whitespace or in how the cost is written:
# 1490000 / 1 490 000 / 1490000.00) into a single insured object, summing their
# multiplicity. The exact-row_text layer already counts byte-identical rows;
# this layer catches formatting variants that the byte comparison misses.

# Structural fields that define object identity. `description` (== row_text) is
# intentionally excluded: it is the raw row text, so including it would just
# reproduce the byte-level dedup and make this layer useless.
_CANONICAL_KEY_FIELDS = (
    "brand",
    "model",
    "year",
    "condition",
    "equipment_type",
    "power_or_capacity",
    "vehicle_category",
)


def _normalized_cost(value: Any) -> str:
    """Normalize an acquisition cost to a canonical string so that
    1490000, '1 490 000' and '1490000.00' compare equal. Empty → ''."""
    if value in (None, ""):
        return ""
    parsed = parse_cost_value(value)
    if parsed is None:
        return normalize_text(value)
    # Decimal.normalize() drops trailing zeros; format 'f' avoids exponent form.
    return format(parsed.normalize(), "f")


def _object_canonical_key(obj: Dict[str, Any]) -> Tuple[str, ...]:
    structural = tuple(
        normalize_text(obj.get(field)) if obj.get(field) else ""
        for field in _CANONICAL_KEY_FIELDS
    )
    cost = _normalized_cost(obj.get("acquisition_cost_value"))
    currency = (obj.get("acquisition_cost_currency") or "").upper()
    key = structural + (cost, currency)
    if any(part for part in key):
        return key
    # Fallback: the structural key is empty (a poorly parsed row with only a
    # free-text description). Use the normalized description so two genuinely
    # different unparsed rows are not merged by an all-empty key.
    return ("__desc__", normalize_text(obj.get("description")))


def group_identical_objects(
    insured_objects: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Collapse business-equivalent objects, summing their multiplicity.

    Returns (grouped_objects, grouping_meta). `grouped_objects` is the list the
    preview formset and request creation work with — one entry per unique
    object, each carrying `source_object_count` and `duplicate_sources`.
    `grouping_meta` is an audit trail stored in parser_v2_payload.object_grouping.
    Group order follows first appearance in the Excel (important for item_no).
    """
    grouped: List[Dict[str, Any]] = []
    groups_meta: List[Dict[str, Any]] = []
    index_by_key: Dict[Tuple[str, ...], int] = {}

    for obj in insured_objects:
        key = _object_canonical_key(obj)
        count = obj.get("source_object_count") or 1
        sources = list(obj.get("duplicate_sources") or ([obj["source"]] if obj.get("source") else []))
        if key in index_by_key:
            gi = index_by_key[key]
            target = grouped[gi]
            target["source_object_count"] = (target.get("source_object_count") or 1) + count
            target["duplicate_sources"] = (target.get("duplicate_sources") or []) + sources
            groups_meta[gi]["source_object_count"] = target["source_object_count"]
            groups_meta[gi]["sources"] = list(target["duplicate_sources"])
        else:
            index_by_key[key] = len(grouped)
            new_obj = dict(obj)
            new_obj["source_object_count"] = count
            new_obj["duplicate_sources"] = list(sources)
            grouped.append(new_obj)
            groups_meta.append({
                "group_index": len(grouped),
                "source_object_count": count,
                "sources": list(sources),
                "canonical_key": {
                    "brand": obj.get("brand"),
                    "model": obj.get("model"),
                    "year": obj.get("year"),
                    "acquisition_cost_value": _normalized_cost(obj.get("acquisition_cost_value")),
                    "acquisition_cost_currency": (obj.get("acquisition_cost_currency") or "").upper(),
                },
            })

    grouping_meta = {
        "raw_object_count": sum((o.get("source_object_count") or 1) for o in grouped),
        "unique_object_count": len(grouped),
        "groups": groups_meta,
    }
    return grouped, grouping_meta


# --- Customer/deal field extractors (stages 3.2 / 3.3) ----------------------
#
# Helpers used to fill the columns added by stages 2.2 and 2.3. They do not
# write to the database — they populate the parser_v2 result dict, which the
# preview view exposes for the operator to confirm. The actual InsuranceRequest
# fields will be filled during stage 4 (splitting).

_DATE_FORMATS = (
    "%d.%m.%Y",
    "%d.%m.%y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
)


def parse_date_value(value: Any) -> Optional[date]:
    """Convert a date-ish cell into a `date`. Returns None on unknown formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_value(value)
    if not text:
        return None
    # Excel sometimes hands us "1982-12-12 00:00:00" (timestamp string).
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except (ValueError, TypeError):
        return None


def normalize_insured_party(value: Any) -> Optional[str]:
    """Map a raw cell value to the lease.insured_party enum.

    Used only as a fallback when the canonical X-marker layout is missing
    (e.g. a free-text cell carrying just the role). The X-marker pickup
    lives in :meth:`ExcelRequestParserV2._extract_insured_party`.
    """
    if value is None:
        return None
    text = normalize_text(value)
    if not text:
        return None
    if "лизингодател" in text:
        return "lessor"
    if "лизингополучател" in text:
        return "lessee"
    return None


def normalize_insured_sum_type(value: Any) -> Optional[str]:
    """Map a raw cell value to the coverage_terms.insured_sum_type enum."""
    if value is None:
        return None
    text = normalize_text(value)
    if not text:
        return None
    # Order matters: «неагрегатная» contains «агрегатная» as a substring.
    if "неагрегатн" in text:
        return "non_aggregate"
    if "агрегатн" in text:
        return "aggregate"
    return None


def normalize_property_location_right_holder(value: Any) -> Optional[str]:
    """Map a raw cell value to property_location_right_holder enum."""
    if value is None:
        return None
    text = normalize_text(value)
    if not text:
        return None
    if "лизингополучател" in text or "собственн" in text and "лизинг" in text:
        return "lessee_owner"
    if "сторонн" in text or "трет" in text or "арендодател" in text:
        return "third_party_owner"
    return None


def normalize_premium_frequency_label(value: Any) -> Optional[str]:
    """Map a raw label cell (one of «Единовременно» / «ежеквартально» /
    «2 раза в год» / «ежегодно») to the premium_frequency enum.
    «прочее» is out of our enum and returns None.
    """
    if value is None:
        return None
    text = normalize_text(value)
    if not text:
        return None
    if "единовременн" in text:
        return "single"
    if "ежекварт" in text or "покварталь" in text:
        return "quarterly"
    if "2 раз" in text or "дваж" in text or "полугод" in text:
        return "biannual"
    if "ежегодн" in text:
        return "annual"
    return None


class ExcelRequestParserV2:
    """Label-based Parser V2 that never blocks request creation."""

    def parse(self, file_path: str, original_filename: str = "") -> ParserV2Result:
        warnings: List[Dict[str, str]] = []
        source_map: Dict[str, str] = {}

        try:
            cells, debug = self._read_cells(file_path)
        except Exception as exc:  # noqa: BLE001 - parser must return a useful draft
            logger.warning("Parser V2 could not read file %s: %s", file_path, exc, exc_info=True)
            data = self._default_data()
            data["notes"] = self._build_notes(
                [
                    {
                        "level": "error",
                        "field": "excel_file",
                        "message": "Файл не удалось прочитать. Создайте минимальную заявку и обработайте файл вручную.",
                        "source": original_filename or os.path.basename(file_path),
                    }
                ]
            )
            return ParserV2Result(
                data=data,
                warnings=[
                    {
                        "level": "error",
                        "field": "excel_file",
                        "message": f"Файл не удалось прочитать: {exc}",
                        "source": original_filename or os.path.basename(file_path),
                    }
                ],
                source_map={},
                confidence=0.0,
                raw_debug={
                    "reader": "failed",
                    "error": str(exc),
                    "file_name": original_filename or os.path.basename(file_path),
                },
                original_filename=original_filename,
            )

        data = self._default_data()
        rows = self._rows(cells)

        client_name, source = self._extract_client_name(cells, rows)
        if client_name:
            data["client_name"] = client_name
            source_map["client_name"] = source

        inn, source = self._extract_inn(cells, rows)
        if inn:
            data["inn"] = inn
            source_map["inn"] = source

        manager_name, source = self._extract_labeled_value(
            cells,
            rows,
            label_groups=[("менеджер",)],
            fallback_coordinate="C5",
        )
        if manager_name:
            data["manager_name"] = manager_name
            source_map["manager_name"] = source

        branch_raw, source = self._extract_branch(cells, rows)
        if branch_raw:
            data["branch"] = map_branch_name(branch_raw)
            source_map["branch"] = source

        dfa_number, source = self._extract_dfa_number(cells, original_filename)
        if dfa_number:
            data["dfa_number"] = dfa_number
            source_map["dfa_number"] = source

        data["insurance_type"], insurance_type_source = self._extract_insurance_type(cells, rows)
        if insurance_type_source:
            source_map["insurance_type"] = insurance_type_source

        data["insurance_period"], period_source = self._extract_insurance_period(cells)
        if period_source:
            source_map["insurance_period"] = period_source

        insured_objects = self._extract_objects(cells, rows)
        insured_objects, object_grouping = group_identical_objects(insured_objects)
        if insured_objects:
            data["vehicle_info"] = self._vehicle_summary(insured_objects)
            source_map["vehicle_info"] = insured_objects[0].get("source", "")
            source_map["insured_objects"] = ", ".join(
                obj["source"] for obj in insured_objects[:10] if obj.get("source")
            )

        application_type = self._detect_application_type(cells)

        franchise_type, franchise_details = self._extract_franchise_type(cells, application_type)
        data["franchise_type"] = franchise_type
        if franchise_details.get("source"):
            source_map["franchise_type"] = franchise_details["source"]
        data["has_autostart"] = self._extract_autostart(cells, rows)
        has_casco_ce, casco_ce_source = self._extract_casco_ce_from_objects(insured_objects)
        data["has_casco_ce"] = has_casco_ce
        if casco_ce_source:
            source_map["has_casco_ce"] = casco_ce_source
        has_transportation, transportation_source = self._extract_transportation_required(cells, rows)
        data["has_transportation"] = has_transportation
        if transportation_source:
            source_map["has_transportation"] = transportation_source
        data["has_construction_work"] = self._contains_any(cells, ["смр", "строительно монтаж"])

        manufacturing_year, source = self._extract_manufacturing_year(cells)
        if manufacturing_year:
            data["manufacturing_year"] = manufacturing_year
            source_map["manufacturing_year"] = source

        for field_name, label_groups in {
            "key_completeness": [("комплект", "ключ")],
            "pts_psm": [("птс",), ("псм",)],
            "creditor_bank": [("банк", "кредитор")],
            "usage_purposes": [("цель", "использ"), ("цели", "использ")],
            "insurance_territory": [("территор", "страх")],
        }.items():
            value, source = self._extract_labeled_value(cells, rows, label_groups=label_groups)
            if value:
                data[field_name] = value
                source_map[field_name] = source

        asset_value, asset_source = self._extract_asset_status(cells, rows, application_type)
        if asset_value:
            data["asset_status"] = asset_value
            source_map["asset_status"] = asset_source

        telematics_value, telematics_source = self._extract_telematics_complex(cells, rows)
        if telematics_value:
            data["telematics_complex"] = telematics_value
            source_map["telematics_complex"] = telematics_source

        # Stage 3.2 — customer details. The audit found these labels at fixed
        # anchors (R8/R10/R11 for юр.лицо), but we still go through the
        # label-based extractor so IP templates with a +1 row shift work too.
        for field_name, label_groups in {
            "legal_address": [("юридическ", "адрес"),
                              ("юр", "адрес")],
            "postal_address": [("почтов", "адрес"),
                               ("фактическ", "адрес")],
            "business_activity": [("вид", "деятельност"),
                                  ("оквэд",)],
        }.items():
            value, source = self._extract_labeled_value(cells, rows, label_groups=label_groups)
            if value:
                data[field_name] = value
                source_map[field_name] = source

        # Date-typed customer fields (3.2).
        birth_raw, birth_source = self._extract_labeled_value(
            cells, rows, label_groups=[("дата", "рождени")]
        )
        birth_value = parse_date_value(birth_raw)
        if birth_value:
            data["birth_date"] = birth_value.isoformat()
            source_map["birth_date"] = birth_source

        submission_raw, submission_source = self._extract_labeled_value(
            cells, rows, label_groups=[("дата", "подачи"),
                                       ("дата", "заявки"),
                                       ("дата", "составлен")]
        )
        submission_value = parse_date_value(submission_raw)
        if submission_value:
            data["submission_date"] = submission_value.isoformat()
            source_map["submission_date"] = submission_source

        # Stage 3.3 — deal / insurance parameters.
        party_value, party_source = self._extract_insured_party(cells, rows)
        if party_value:
            data["insured_party"] = party_value
            source_map["insured_party"] = party_source

        sum_raw, sum_source = self._extract_labeled_value(
            cells, rows, label_groups=[("страхов", "сумма")]
        )
        sum_value = normalize_insured_sum_type(sum_raw)
        if sum_value:
            data["insured_sum_type"] = sum_value
            source_map["insured_sum_type"] = sum_source

        guard_raw, guard_source = self._extract_labeled_value(
            cells, rows, label_groups=[("условия", "охран"),
                                       ("условия", "хранен")]
        )
        if guard_raw:
            data["guard_conditions"] = guard_raw
            source_map["guard_conditions"] = guard_source

        plrh_value, plrh_source = self._extract_property_location_right_holder(cells, rows)
        if plrh_value:
            data["property_location_right_holder"] = plrh_value
            source_map["property_location_right_holder"] = plrh_source

        freq_value, freq_source = self._extract_premium_frequency(cells, rows)
        if freq_value:
            data["premium_frequency"] = freq_value
            source_map["premium_frequency"] = freq_source

        # has_installment is a derived flag: True only for within-year
        # installment plans (quarterly / biannual). single (one upfront
        # payment) and annual (one payment per year on a multi-year policy)
        # are not considered installments by the insurers we mail.
        data["has_installment"] = data.get("premium_frequency") in ("quarterly", "biannual")

        parser_payload = {
            "insured_objects": insured_objects,
            "object_grouping": object_grouping,
            "raw_branch": branch_raw,
            "application_format": self._detect_application_format(data),
            "application_type": application_type,
            "franchise_details": franchise_details,
        }
        data["parser_v2_payload"] = parser_payload

        warnings.extend(self._build_warnings(data, insured_objects))
        data["notes"] = self._build_notes(warnings)

        confidence = self._calculate_confidence(data, insured_objects)
        debug.update(
            {
                "file_name": original_filename or os.path.basename(file_path),
                "object_count": len(insured_objects),
                "warning_count": len(warnings),
            }
        )

        return ParserV2Result(
            data=data,
            warnings=warnings,
            source_map=source_map,
            confidence=confidence,
            raw_debug=debug,
            original_filename=original_filename,
        )

    def _read_cells(self, file_path: str) -> Tuple[List[GridCell], Dict[str, Any]]:
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            cells: List[GridCell] = []
            for row in sheet.iter_rows():
                for cell in row:
                    value = clean_value(cell.value)
                    if value:
                        cells.append(GridCell(cell.row, cell.column, cell.coordinate, value))
            return cells, {
                "reader": "openpyxl",
                "sheet_name": sheet.title,
                "row_count": sheet.max_row,
                "column_count": sheet.max_column,
                "cell_count": len(cells),
            }
        except Exception as openpyxl_error:
            logger.info("Parser V2 openpyxl read failed, trying pandas: %s", openpyxl_error)

        df = pd.read_excel(file_path, sheet_name=0, header=None)
        cells = []
        for row_index, row in df.iterrows():
            for col_index, value in row.items():
                clean = clean_value(value)
                if clean:
                    row_number = int(row_index) + 1
                    col_number = int(col_index) + 1
                    cells.append(
                        GridCell(
                            row=row_number,
                            col=col_number,
                            coordinate=f"{get_column_letter(col_number)}{row_number}",
                            value=clean,
                        )
                    )
        return cells, {
            "reader": "pandas",
            "sheet_name": "0",
            "row_count": int(df.shape[0]),
            "column_count": int(df.shape[1]),
            "cell_count": len(cells),
        }

    def _default_data(self) -> Dict[str, Any]:
        deadline = timezone.localtime(timezone.now() + timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M")
        return {
            "client_name": MISSING_CLIENT,
            "inn": "",
            "insurance_type": "другое",
            "insurance_period": "",
            "vehicle_info": MISSING_VEHICLE,
            "dfa_number": MISSING_DFA,
            "branch": "",
            "manager_name": "",
            "deal_status": "new",
            "franchise_type": "none",
            "has_installment": False,
            "has_autostart": False,
            "has_casco_ce": False,
            "has_transportation": False,
            "has_construction_work": False,
            "manufacturing_year": "",
            "asset_status": "",
            "key_completeness": "",
            "pts_psm": "",
            "creditor_bank": "",
            "usage_purposes": "",
            "telematics_complex": "",
            "insurance_territory": "",
            "response_deadline": deadline,
            "notes": "",
            "parser_v2_payload": {},
        }

    def _rows(self, cells: Iterable[GridCell]) -> Dict[int, List[GridCell]]:
        rows: Dict[int, List[GridCell]] = {}
        for cell in cells:
            rows.setdefault(cell.row, []).append(cell)
        for row_cells in rows.values():
            row_cells.sort(key=lambda item: item.col)
        return rows

    def _extract_labeled_value(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
        label_groups: List[Tuple[str, ...]],
        fallback_coordinate: str = "",
    ) -> Tuple[str, str]:
        fallback_cell = self._cell_by_coordinate(cells, fallback_coordinate) if fallback_coordinate else None

        for cell in cells:
            if not self._matches_any_group(cell.normalized, label_groups):
                continue

            inline = self._inline_value_after_label(cell.value)
            if inline:
                return inline, cell.coordinate

            right = self._first_value_right(rows, cell)
            if right:
                return right.value, right.coordinate

            below = self._first_value_below(rows, cell)
            if below:
                return below.value, below.coordinate

        if fallback_cell and fallback_cell.value:
            return fallback_cell.value, fallback_cell.coordinate

        return "", ""

    def _extract_client_name(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> Tuple[str, str]:
        for coordinate in CLIENT_COORDINATES:
            cell = self._cell_by_coordinate(cells, coordinate)
            if cell and self._looks_like_client_name(cell.value):
                return cell.value, cell.coordinate

        for cell in cells:
            if cell.row > CLIENT_MAX_LABEL_ROW:
                continue
            if not self._matches_any_group(
                cell.normalized,
                [("наименование", "лизингополучател"), ("страхователь",), ("клиент",)],
            ):
                continue

            inline = self._inline_value_after_label(cell.value)
            if self._looks_like_client_name(inline):
                return inline, cell.coordinate

            right = self._first_value_right(rows, cell)
            if right and self._looks_like_client_name(right.value):
                return right.value, right.coordinate

            below = self._first_value_below(rows, cell)
            if below and self._looks_like_client_name(below.value):
                return below.value, below.coordinate

        return "", ""

    def _extract_inn(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> Tuple[str, str]:
        value, source = self._extract_labeled_value(cells, rows, label_groups=[("инн",)])
        inn = self._normalize_inn(value)
        if inn:
            return inn, source

        for cell in cells:
            match = re.search(r"\b\d{10}(\d{2})?\b", cell.value)
            if match:
                return match.group(0), cell.coordinate
        return "", ""

    def _extract_branch(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> Tuple[str, str]:
        for cell in cells:
            if "филиал" not in cell.normalized and "обособленное подразделение" not in cell.normalized:
                continue
            if cell.normalized in {"филиал", "филиал оп"}:
                continue
            return cell.value, cell.coordinate

        value, source = self._extract_labeled_value(cells, rows, label_groups=[("филиал",)])
        if value:
            return value, source
        return "", ""

    def _extract_dfa_number(self, cells: List[GridCell], original_filename: str) -> Tuple[str, str]:
        pattern = re.compile(r"\b\d{4,6}(?:[-/][0-9A-Za-zА-Яа-яЁё]+){1,5}\b")
        for cell in cells:
            match = pattern.search(cell.value)
            if match:
                return match.group(0), cell.coordinate
        filename_match = pattern.search(original_filename or "")
        if filename_match:
            return filename_match.group(0), "filename"
        return "", ""

    def _extract_insurance_type(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> Tuple[str, str]:
        """Extract insurance type with priority for X-marked option rows.

        Canonical layouts keep option labels in column B and an X-mark in
        column D on rows 21/22/23:
          - legal entity: B21='КАСКО', B22='страхование спецтехники'
          - IP (+1 layout): B22='КАСКО', B23='страхование спецтехники'
          - property form: B22/B23='страхование имущества ...'
        """
        marked_value, marked_source, found_marks = self._extract_insurance_type_from_marked_rows(rows)
        if found_marks:
            return (marked_value or "другое"), marked_source

        block_value, block_source, found_block = self._extract_insurance_type_from_block_rows(rows)
        if found_block:
            return (block_value or "другое"), block_source

        # Last-resort fallback for non-canonical forms: inspect only the upper
        # part of the sheet where the insurance-type block usually lives. This
        # avoids false positives from object-table headers like «... имущества».
        top_text = " ".join(cell.normalized for cell in cells if cell.row <= 30)
        source = self._first_source_containing(
            [cell for cell in cells if cell.row <= 30],
            ["имуществ", "спецтехник", "каско"],
        )
        if "имуществ" in top_text:
            return "страхование имущества", source
        if "спецтехник" in top_text:
            return "страхование спецтехники", source
        if "каско" in top_text:
            return "КАСКО", source
        return "другое", source

    def _extract_insurance_type_from_marked_rows(
        self,
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[Optional[str], str, bool]:
        """Return insurance type selected by X-mark in D21/D22/D23."""
        candidates: List[Tuple[str, str]] = []
        mark_sources: List[str] = []

        for row_number in (21, 22, 23):
            row_by_col = {cell.col: cell for cell in rows.get(row_number, [])}
            mark_cell = row_by_col.get(4)
            if not self._is_mark(mark_cell):
                continue
            if mark_cell:
                mark_sources.append(mark_cell.coordinate)

            insurance_type: Optional[str] = None
            for label_col in (2, 3, 5):
                label_cell = row_by_col.get(label_col)
                if not label_cell:
                    continue
                insurance_type = self._map_insurance_type_label(label_cell.value)
                if insurance_type:
                    break
            if insurance_type and mark_cell:
                candidates.append((insurance_type, mark_cell.coordinate))

        if not mark_sources:
            return None, "", False
        if len(candidates) == 1:
            return candidates[0][0], candidates[0][1], True
        if len(candidates) > 1:
            unique_types = {item[0] for item in candidates}
            source = ", ".join(item[1] for item in candidates)
            if len(unique_types) == 1:
                return candidates[0][0], source, True
            return None, source, True
        return None, ", ".join(mark_sources), True

    def _extract_insurance_type_from_block_rows(
        self,
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[Optional[str], str, bool]:
        """Fallback for unmarked forms: infer from rows 21..23 only.

        Preference:
          1) explicit typed value in column D (legacy uploads/tests),
          2) unique option label from B/C/E columns.
        If multiple different options are visible with no mark, return
        unresolved (None) so the preview defaults to manual correction.
        """
        explicit_candidates: List[Tuple[str, str]] = []
        label_candidates: List[Tuple[str, str]] = []

        for row_number in (21, 22, 23):
            row_by_col = {cell.col: cell for cell in rows.get(row_number, [])}
            explicit_cell = row_by_col.get(4)
            if explicit_cell:
                explicit_type = self._map_insurance_type_label(explicit_cell.value)
                if explicit_type:
                    explicit_candidates.append((explicit_type, explicit_cell.coordinate))
            for label_col in (2, 3, 5):
                label_cell = row_by_col.get(label_col)
                if not label_cell:
                    continue
                label_type = self._map_insurance_type_label(label_cell.value)
                if label_type:
                    label_candidates.append((label_type, label_cell.coordinate))

        if explicit_candidates:
            unique_types = {item[0] for item in explicit_candidates}
            if len(unique_types) == 1:
                return explicit_candidates[0][0], explicit_candidates[0][1], True
            return None, ", ".join(item[1] for item in explicit_candidates), True

        if not label_candidates:
            return None, "", False
        unique_label_types = {item[0] for item in label_candidates}
        if len(unique_label_types) == 1:
            return label_candidates[0][0], label_candidates[0][1], True
        return None, "", True

    def _map_insurance_type_label(self, value: Any) -> Optional[str]:
        text = normalize_text(value)
        if not text:
            return None
        if "каско" in text:
            return "КАСКО"
        if "спецтехник" in text:
            return "страхование спецтехники"
        if "имуществ" in text:
            return "страхование имущества"
        return None

    def _extract_insurance_period(self, cells: List[GridCell]) -> Tuple[str, str]:
        full_text = " ".join(cell.normalized for cell in cells)
        if "весь срок лизинга" in full_text or "на весь срок" in full_text:
            return "на весь срок лизинга", self._first_source_containing(cells, ["срок лизинга", "весь срок"])
        if "1 год" in full_text or "один год" in full_text or "12 мес" in full_text:
            return "1 год", self._first_source_containing(cells, ["1 год", "12 мес", "один год"])
        return "", ""

    def _extract_objects(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> List[Dict[str, Any]]:
        # Property applications use a distinct, narrow object table
        # (single description column C, fixed K/L/M/N for year/condition/cost/
        # currency) and a footer block we must not walk into. Detect early
        # and dispatch to the dedicated extractor.
        if self._is_property_form(cells):
            return self._extract_objects_property(cells, rows)

        start_rows = [
            row_number
            for row_number, row_cells in rows.items()
            if self._row_matches(row_cells, ["предмет лизинга", "объект страхования"])
            or (
                self._row_matches(row_cells, ["наименование"])
                and self._row_matches(row_cells, ["год", "стоимость", "vin", "заводской"])
            )
        ]
        objects: List[Dict[str, Any]] = []
        # Map row_text → index in `objects`. A fully identical row does not
        # create a second object: instead it increments source_object_count on
        # the already-created payload and records its coordinates. This keeps
        # the multiplicity of duplicated leasing-export rows instead of
        # silently dropping it (see docs/improvement_plans/
        # parser_v2_identical_object_multiplicity.md).
        index_by_row_text: Dict[str, int] = {}

        for start_row in start_rows[:3]:
            blank_rows = 0
            current_category: Optional[str] = None
            current_category_source = ""
            for row_number in range(start_row, start_row + 35):
                row_cells = rows.get(row_number, [])
                row_text = self._row_text(row_cells)
                row_norm = normalize_text(row_text)
                if not row_text:
                    blank_rows += 1
                    if blank_rows >= 3:
                        break
                    continue
                blank_rows = 0
                if row_number != start_row and self._is_object_stop_row(row_norm):
                    break
                category = self._object_section_category(row_norm)
                if category:
                    current_category = category
                    current_category_source = self._row_source(row_cells)
                    continue
                if self._is_object_header_or_label(row_norm) or self._is_object_template_row(row_norm):
                    continue
                if len(row_text) < 8:
                    continue
                source = self._row_source(row_cells)
                if row_text in index_by_row_text:
                    existing = objects[index_by_row_text[row_text]]
                    existing["source_object_count"] = (existing.get("source_object_count") or 1) + 1
                    existing.setdefault("duplicate_sources", [existing.get("source") or ""])
                    existing["duplicate_sources"].append(source)
                    continue
                index_by_row_text[row_text] = len(objects)
                payload = self._build_object_payload(row_cells, row_text)
                if current_category:
                    payload["vehicle_category"] = current_category
                    payload["vehicle_category_source"] = current_category_source
                    if not payload.get("equipment_type"):
                        payload["equipment_type"] = self._equipment_type_from_category(current_category)
                payload["source_object_count"] = 1
                payload["duplicate_sources"] = [payload.get("source") or source]
                objects.append(payload)

        if not objects:
            value, source = self._extract_labeled_value(cells, rows, label_groups=[("предмет", "лизинга"), ("объект", "страхования")])
            if value and not self._is_object_template_row(normalize_text(value)):
                objects.append(
                    {
                        "description": value,
                        "year": self._year_from_text(value),
                        "source": source,
                        "brand": None,
                        "model": None,
                        "condition": None,
                        "equipment_type": None,
                        "power_or_capacity": None,
                        "acquisition_cost_value": None,
                        "acquisition_cost_currency": None,
                        "vehicle_category": None,
                        "vehicle_category_source": "",
                        "source_object_count": 1,
                        "duplicate_sources": [source] if source else [],
                    }
                )

        return objects[:50]

    def _is_property_form(self, cells: List[GridCell]) -> bool:
        """Detect the property insurance form by the «имуществ»-bearing
        label in B22 (B23 for IP). KASKO/equipment templates leave B22
        empty (the type is selected via an X-mark in column D)."""
        for cell in cells:
            if cell.col != 2:  # column B
                continue
            if cell.row not in (22, 23):
                continue
            if "имуществ" in cell.normalized:
                return True
        return False

    def _extract_objects_property(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> List[Dict[str, Any]]:
        """Property-form object extractor.

        Layout (стабильно по 10 файлам im_request/):
          - header row: B='№ п/п', C='Наименование...', K='Год выпуска',
            M='Стоимость', N='Валюта' (row 41 for юр.лицо, 42 for ИП).
          - object rows directly below: B=<число>, C=<описание>, K=<год>,
            L=<новое/б.у>, M=<цена>, N=<валюта>.
          - the table ends at the first row where B is empty / non-numeric,
            or at a stop-marker row (e.g. «Дополнительные виды страхования»).

        Description is taken **only from column C**, never from row_text,
        so price/year/currency don't leak into the email body.
        """
        header_row: Optional[int] = None
        for row_number in sorted(rows.keys()):
            row_cells = rows[row_number]
            row_by_col = {c.col: c for c in row_cells}
            b_cell = row_by_col.get(2)
            if not b_cell:
                continue
            # normalize_text turns «№ п/п» into «no п/п» (replace + collapse).
            if "no п/п" not in b_cell.normalized:
                continue
            if not self._row_matches(row_cells, ["наименование"]):
                continue
            header_row = row_number
            break

        if header_row is None:
            return []

        objects: List[Dict[str, Any]] = []
        for row_number in range(header_row + 1, header_row + 25):
            row_cells = rows.get(row_number, [])
            if not row_cells:
                break
            row_by_col = {c.col: c for c in row_cells}
            b_cell = row_by_col.get(2)
            c_cell = row_by_col.get(3)

            # Stop on the property-footer block («Дополнительные виды
            # страхования»). _is_object_stop_row already covers its
            # phrases since stage 1.
            row_norm = normalize_text(self._row_text(row_cells))
            if self._is_object_stop_row(row_norm):
                break

            # An object row must have a numeric ordinal in B and a
            # non-empty description in C. Anything else is either a
            # spacer or already past the table.
            if not b_cell or not re.fullmatch(r"\d+", b_cell.value.strip()):
                break
            if not c_cell or not c_cell.value.strip():
                break

            description = c_cell.value.strip()
            year_cell = row_by_col.get(11)  # K
            condition_cell = row_by_col.get(12)  # L
            cost_cell = row_by_col.get(13)  # M
            currency_cell = row_by_col.get(14)  # N

            year = self._year_from_text(year_cell.value) if year_cell else ""
            condition = normalize_condition(condition_cell.value) if condition_cell else None
            cost_value = parse_cost_value(cost_cell.value) if cost_cell else None
            currency = normalize_currency(currency_cell.value) if currency_cell else None

            brand, model = split_brand_model(description, year)
            source = self._row_source(row_cells)

            objects.append(
                {
                    "description": description,
                    "year": year,
                    "source": source,
                    "brand": brand,
                    "model": model,
                    "condition": condition,
                    "equipment_type": None,
                    "power_or_capacity": None,
                    "acquisition_cost_value": str(cost_value) if cost_value is not None else None,
                    "acquisition_cost_currency": currency,
                    "vehicle_category": None,
                    "vehicle_category_source": "",
                    "source_object_count": 1,
                    "duplicate_sources": [source] if source else [],
                }
            )

        return objects[:50]

    def _object_section_category(self, row_norm: str) -> Optional[str]:
        if "транспортные средства категории c" in row_norm or "транспортные средства категории c/e" in row_norm:
            return "C"
        if "транспортные средства категории b" in row_norm:
            return "B"
        if "транспортные средства категории d" in row_norm:
            return "D"
        if "специальная техника" in row_norm:
            return "special_equipment"
        return None

    def _equipment_type_from_category(self, category: Optional[str]) -> Optional[str]:
        return {
            "B": "Категория B",
            "C": "Категория C",
            "D": "Категория D",
            "special_equipment": "Спецтехника",
        }.get(category or "")

    def _extract_casco_ce_from_objects(self, insured_objects: List[Dict[str, Any]]) -> Tuple[bool, str]:
        for obj in insured_objects:
            if obj.get("vehicle_category") == "C":
                return True, obj.get("vehicle_category_source") or obj.get("source", "")
        return False, ""

    def _build_object_payload(self, row_cells: List[GridCell], row_text: str) -> Dict[str, Any]:
        """Turn one object row into a structured payload (stage 3.1).

        Canonical layout (юр.лицо CASCO):
          - columns 3..9 (C..I)   — descriptive text (brand, model, sometimes VIN)
          - column 10 (J)         — manufacturing year
          - column 11 (K)         — condition (новое/б/у)
          - column 12 (L)         — equipment kind OR engine power
          - column 13 (M)         — acquisition cost value
          - column 14 (N)         — acquisition cost currency

        Templates for individual entrepreneurs (and a few regional layouts) shift
        some of these columns by +1. To stay robust we scan columns 10..15 and
        classify each cell by its content (currency lookup, condition lookup,
        numeric value), rather than trusting a single coordinate.
        """
        cell_by_col: Dict[int, GridCell] = {cell.col: cell for cell in row_cells}
        description_cells = [c for c in row_cells if 3 <= c.col <= 9]
        description_text = " ".join(c.value for c in description_cells) if description_cells else row_text

        year = self._year_from_text(row_text)
        brand, model = split_brand_model(description_text, year)

        condition: Optional[str] = None
        currency: Optional[str] = None
        cost_value: Optional[Decimal] = None
        equipment_type: Optional[str] = None
        power_or_capacity: Optional[str] = None

        # Columns 11..15 cover condition/equipment/cost/currency across all
        # observed templates (column 10 is the manufacturing year and is parsed
        # separately above; skipping it avoids classifying e.g. 2025 as a price).
        for col in range(11, 16):
            cell = cell_by_col.get(col)
            if not cell:
                continue
            raw = cell.value
            if not raw:
                continue
            # Skip cells that look like a year — IP templates shift columns,
            # so the year can appear in column 11 instead of 10.
            if re.fullmatch(r"(19|20)\d{2}", raw.strip()):
                continue
            # Try classifications in order of decreasing specificity.
            if condition is None:
                normalized_condition = normalize_condition(raw)
                if normalized_condition:
                    condition = normalized_condition
                    continue
            if currency is None:
                normalized_currency = normalize_currency(raw)
                if normalized_currency:
                    currency = normalized_currency
                    continue
            # Try cost: numeric value of price magnitude (>= 10_000). Engine
            # power (78.05) and small integers stay out of this branch.
            parsed_cost = parse_cost_value(raw)
            if cost_value is None and parsed_cost is not None and parsed_cost >= Decimal("10000"):
                cost_value = parsed_cost
                continue
            # Otherwise classify as equipment kind or engine power.
            kind, power = classify_equipment_or_power(raw)
            if power and power_or_capacity is None:
                power_or_capacity = power
                continue
            if kind and equipment_type is None:
                equipment_type = kind
                continue

        return {
            "description": row_text,
            "year": year,
            "source": self._row_source(row_cells),
            "brand": brand,
            "model": model,
            "condition": condition,
            "equipment_type": equipment_type,
            "power_or_capacity": power_or_capacity,
            "acquisition_cost_value": str(cost_value) if cost_value is not None else None,
            "acquisition_cost_currency": currency,
        }

    def _extract_franchise_type(self, cells: List[GridCell], application_type: str) -> Tuple[str, Dict[str, Any]]:
        cell_map = {(cell.row, cell.col): cell for cell in cells}
        value_rows = [30, 29] if application_type == "individual_entrepreneur" else [29, 30]

        for index, value_row in enumerate(value_rows):
            row_details = self._franchise_row_details(cell_map, value_row)
            if not row_details["selected_columns"]:
                continue
            if row_details["selection_row_looks_like_header"]:
                continue
            if index > 0 and not row_details["label_row_looks_like_franchise_options"]:
                continue
            return row_details["franchise_type"], row_details

        return "none", {
            "value_row": value_rows[0],
            "label_row": value_rows[0] - 1,
            "selected_columns": [],
            "source": "",
            "selection_row_looks_like_header": False,
            "label_row_looks_like_franchise_options": False,
        }

    def _franchise_row_details(self, cell_map: Dict[Tuple[int, int], GridCell], value_row: int) -> Dict[str, Any]:
        label_row = value_row - 1
        selected_columns = []
        for column_letter, col, variant in [
            ("D", 4, "without_franchise"),
            ("E", 5, "percent_franchise"),
            ("F", 6, "absolute_franchise"),
        ]:
            value_cell = cell_map.get((value_row, col))
            if not self._franchise_cell_has_selection(value_cell.value if value_cell else ""):
                continue

            label_cell = cell_map.get((label_row, col))
            selected_columns.append(
                {
                    "column": column_letter,
                    "variant": variant,
                    "value_coordinate": value_cell.coordinate,
                    "value": value_cell.value,
                    "label_coordinate": label_cell.coordinate if label_cell else f"{column_letter}{label_row}",
                    "label": label_cell.value if label_cell else "",
                }
            )

        selected_variants = {item["variant"] for item in selected_columns}
        has_without = "without_franchise" in selected_variants
        has_with = bool({"percent_franchise", "absolute_franchise"} & selected_variants)
        selection_row_looks_like_header = self._franchise_row_looks_like_option_labels(
            [item["value"] for item in selected_columns]
        )
        label_row_looks_like_franchise_options = self._franchise_row_looks_like_option_labels(
            [item["label"] for item in selected_columns]
        )

        if has_without and has_with:
            franchise_type = "both_variants"
        elif has_with:
            franchise_type = "with_franchise"
        else:
            franchise_type = "none"

        source_parts = [
            f"{item['value_coordinate']}={item['value']}"
            + (f" ({item['label_coordinate']}: {item['label']})" if item["label"] else "")
            for item in selected_columns
        ]

        return {
            "value_row": value_row,
            "label_row": label_row,
            "selected_columns": selected_columns,
            "franchise_type": franchise_type,
            "source": "; ".join(source_parts),
            "selection_row_looks_like_header": selection_row_looks_like_header,
            "label_row_looks_like_franchise_options": label_row_looks_like_franchise_options,
        }

    def _franchise_cell_has_selection(self, value: Any) -> bool:
        value = clean_value(value)
        return bool(value) and normalize_text(value) not in {"none", "nan", "null", "false"}

    def _franchise_row_looks_like_option_labels(self, values: Iterable[Any]) -> bool:
        label_like_count = sum(
            1 for value in values
            if self._franchise_text_looks_like_option_label(value)
        )
        return label_like_count >= 2

    def _franchise_text_looks_like_option_label(self, value: Any) -> bool:
        text = normalize_text(value)
        if not text:
            return False
        return any(
            marker in text
            for marker in [
                "без франшиз",
                "с франшиз",
                "франшиз",
                "страхов",
                "абсолют",
                "процент",
                "сумм",
                "руб",
                "%",
            ]
        )

    def _extract_autostart(self, cells: List[GridCell], rows: Dict[int, List[GridCell]]) -> bool:
        # Default to False unless an explicit "да" value is found in the row.
        # An empty value cell next to the "Автозапуск" label means «нет».
        for cell in cells:
            if "автозапуск" not in cell.normalized:
                continue
            tokens: List[str] = []
            for sibling in rows.get(cell.row, []):
                if sibling.col == cell.col:
                    continue
                tokens.extend(t for t in re.split(r"[\s/]+", sibling.normalized) if t)
            # Inline value in the label cell itself, e.g. «Автозапуск: да».
            # Skip when the tail looks like a «(да/нет)» header (both tokens present).
            tail_text = cell.normalized.split("автозапуск", 1)[1]
            tail_tokens = [t for t in re.split(r"[\s/]+", tail_text) if t]
            if not ("да" in tail_tokens and "нет" in tail_tokens):
                tokens.extend(tail_tokens)
            if "да" in tokens:
                return True
        return False

    def _extract_manufacturing_year(self, cells: List[GridCell]) -> Tuple[str, str]:
        value, source = self._extract_labeled_value(cells, self._rows(cells), label_groups=[("год", "выпуск")])
        year = self._year_from_text(value)
        if year:
            return year, source
        for cell in cells:
            year = self._year_from_text(cell.value)
            if year:
                return year, cell.coordinate
        return "", ""

    def _ip_row_offset(self, application_type: str, base_row: int) -> int:
        if application_type == "individual_entrepreneur" and base_row > 8:
            return base_row + 1
        return base_row

    def _extract_asset_status(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
        application_type: str,
    ) -> Tuple[str, str]:
        # 1) Label-based: «статус имущества» / «состояние».
        value, source = self._extract_labeled_value(
            cells,
            rows,
            label_groups=[("статус", "имуществ"), ("состояние",)],
        )
        if value:
            return value, source
        # 2) Fallback by coordinates: column K (11) in object rows 43/45/47/49,
        # shifted by +1 for individual entrepreneur applications. Some templates
        # have an extra leading column, so we also try column L (12) as a backup.
        status_values = {"новое", "новый", "новая", "б/у", "бу", "б-у"}
        parts: List[str] = []
        coords: List[str] = []
        for base_row in (43, 45, 47, 49):
            target_row = self._ip_row_offset(application_type, base_row)
            row_cells = rows.get(target_row, [])
            for col in (11, 12):
                cell = next((c for c in row_cells if c.col == col), None)
                if not cell:
                    continue
                normalized = cell.normalized
                if not normalized:
                    continue
                if normalized in status_values:
                    parts.append(cell.value.strip())
                    coords.append(cell.coordinate)
                    break
                if col == 11 and cell.value.strip():
                    # Legacy V1 behaviour: trust column K even when value is non-canonical.
                    parts.append(cell.value.strip())
                    coords.append(cell.coordinate)
                    break
        if parts:
            return " ".join(parts), ", ".join(coords)
        return "", ""

    def _extract_telematics_complex(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[str, str]:
        # Real applications use two layouts for the telematics block:
        #   inline:    [label] | [«Наименование»] | [value]      (same row)
        #   stacked:   [label] | [«Наименование»]                (row N)
        #              ...     | [value]                          (row N+1 or +2)
        # We anchor on the «телемат» label and scan rightward columns first in
        # the same row, then in the following 1–3 rows, skipping sub-header labels.
        label_cell = next((cell for cell in cells if "телемат" in cell.normalized), None)
        if not label_cell:
            return "", ""
        sub_labels = {"наименование", "название", "модель", "марка", "конфигурация"}

        def is_value(cell: GridCell) -> bool:
            normalized = cell.normalized
            if not normalized or normalized in sub_labels:
                return False
            if self._looks_like_empty_or_label(cell.value):
                return False
            return True

        max_col_offset = 5
        # 1) Same row, to the right of the label.
        for cell in rows.get(label_cell.row, []):
            if cell.col <= label_cell.col or cell.col > label_cell.col + max_col_offset:
                continue
            if is_value(cell):
                return cell.value, cell.coordinate
        # 2) Rows below, columns to the right of the label.
        for row_offset in range(1, 4):
            for cell in rows.get(label_cell.row + row_offset, []):
                if cell.col <= label_cell.col or cell.col > label_cell.col + max_col_offset:
                    continue
                if is_value(cell):
                    return cell.value, cell.coordinate
        return "", ""

    def _extract_transportation_required(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[bool, str]:
        """Detect the additional transportation insurance option.

        Usage-purpose and business-activity cells often mention перевозка
        without requesting the extra transportation risk. Treat only the
        dedicated additional-insurance block as a reliable signal.
        """
        for cell in cells:
            if not self._looks_like_transportation_option(cell):
                continue
            if not self._near_additional_insurance_block(cell, rows):
                continue

            mark_source = self._transportation_mark_source(cell, rows)
            if mark_source:
                return True, mark_source

            detail_source = self._transportation_detail_source(cell, rows)
            if detail_source:
                return True, detail_source

        return False, ""

    def _looks_like_transportation_option(self, cell: GridCell) -> bool:
        text = cell.normalized
        if "перевоз" not in text and "транспортиров" not in text:
            return False
        if self._matches_any_group(
            text,
            [
                ("цель", "использ"),
                ("цели", "использ"),
                ("вид", "деятельност"),
                ("оквэд",),
            ],
        ):
            return False
        return any(
            marker in text
            for marker in [
                "погруз",
                "выгруз",
                "поставщик",
                "лизингополучател",
                "пункт отправления",
                "пункт назначения",
                "срок перевоз",
            ]
        )

    def _near_additional_insurance_block(
        self,
        cell: GridCell,
        rows: Dict[int, List[GridCell]],
    ) -> bool:
        for row_number in range(max(1, cell.row - 3), cell.row + 1):
            row_text = normalize_text(self._row_text(rows.get(row_number, [])))
            if (
                "дополнительн" in row_text
                and "страхован" in row_text
                and ("оборудован" in row_text or "имуществ" in row_text or "вид" in row_text)
            ):
                return True
        return False

    def _transportation_mark_source(
        self,
        option_cell: GridCell,
        rows: Dict[int, List[GridCell]],
    ) -> str:
        for cell in rows.get(option_cell.row, []):
            if cell is option_cell:
                continue
            if abs(cell.col - option_cell.col) > 4:
                continue
            if self._is_mark(cell):
                return cell.coordinate
        return ""

    def _transportation_detail_source(
        self,
        option_cell: GridCell,
        rows: Dict[int, List[GridCell]],
    ) -> str:
        detail_label_groups = [
            ("пункт", "отправлен"),
            ("пункт", "назначен"),
            ("срок", "перевоз"),
        ]
        for row_number in range(option_cell.row + 1, option_cell.row + 7):
            for label_cell in rows.get(row_number, []):
                if not self._matches_any_group(label_cell.normalized, detail_label_groups):
                    continue
                inline = self._inline_value_after_label(label_cell.value)
                if inline:
                    return label_cell.coordinate
                value_cell = self._first_value_right(rows, label_cell)
                if value_cell:
                    return value_cell.coordinate
        return ""

    def _extract_marker_choice(
        self,
        rows: Dict[int, List[GridCell]],
        *,
        left_label_groups: List[Tuple[str, ...]],
        left_value: str,
        right_label_groups: List[Tuple[str, ...]],
        right_value: str,
    ) -> Tuple[Optional[str], str, bool]:
        """Two-label crosshair extraction.

        Find a row that simultaneously holds a left-side label and a
        right-side label, then probe the next two rows (IP +1 shift) for
        an X-mark under either column.

        Returns (value, source_coordinate, found_header):
          - found_header=False signals the caller can fall back to soft text matching;
          - value=None with found_header=True means the layout was found but no
            unambiguous mark — the caller should treat this as "unresolved".
        """
        header_row: Optional[int] = None
        left_col: Optional[int] = None
        right_col: Optional[int] = None
        anchor_coord = ""

        for row_number, row_cells in rows.items():
            left_cell: Optional[GridCell] = None
            right_cell: Optional[GridCell] = None
            for cell in row_cells:
                text = cell.normalized
                if left_cell is None and self._matches_any_group(text, left_label_groups):
                    left_cell = cell
                    continue
                if right_cell is None and self._matches_any_group(text, right_label_groups):
                    right_cell = cell
            if left_cell and right_cell and left_cell.col != right_cell.col:
                header_row = row_number
                left_col = left_cell.col
                right_col = right_cell.col
                anchor_coord = left_cell.coordinate
                break

        if header_row is None:
            return None, "", False

        for offset in (1, 2):
            mark_row = rows.get(header_row + offset, [])
            row_by_col = {c.col: c for c in mark_row}
            left_marked = self._is_mark(row_by_col.get(left_col))
            right_marked = self._is_mark(row_by_col.get(right_col))
            if left_marked and not right_marked:
                cell = row_by_col.get(left_col)
                return left_value, cell.coordinate if cell else anchor_coord, True
            if right_marked and not left_marked:
                cell = row_by_col.get(right_col)
                return right_value, cell.coordinate if cell else anchor_coord, True
            if left_marked and right_marked:
                return None, anchor_coord, True

        return None, anchor_coord, True

    def _extract_insured_party(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[Optional[str], str]:
        """«Страхователь» block: ЛизингоДАТЕЛЬ / ЛизингоПОЛУЧАТЕЛЬ + crosshair."""
        value, source, found_header = self._extract_marker_choice(
            rows,
            left_label_groups=[("лизингодател",)],
            left_value="lessor",
            right_label_groups=[("лизингополучател",)],
            right_value="lessee",
        )
        if found_header:
            return value, source

        # Fallback: file uses free text like «Страхователь: ЛизингоПОЛУЧАТЕЛЬ».
        raw, fallback_source = self._extract_labeled_value(
            cells, rows, label_groups=[("страхователь",)]
        )
        return normalize_insured_party(raw), fallback_source

    def _extract_property_location_right_holder(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[Optional[str], str]:
        """«Правообладатель места расположения» block:
          «собственность лизингополучателя» / «собственность третьего лица» + crosshair.
        """
        value, source, found_header = self._extract_marker_choice(
            rows,
            left_label_groups=[("лизингополучател",), ("собственн", "лизинг")],
            left_value="lessee_owner",
            right_label_groups=[("сторонн",), ("трет",), ("арендодател",)],
            right_value="third_party_owner",
        )
        if found_header:
            return value, source

        raw, fallback_source = self._extract_labeled_value(
            cells, rows,
            label_groups=[("правообладат",), ("собственник", "места")],
        )
        return normalize_property_location_right_holder(raw), fallback_source

    def _is_mark(self, cell: Optional[GridCell]) -> bool:
        if cell is None:
            return False
        return normalize_text(cell.value).strip() in {"х", "x", "+", "v"}

    def _extract_premium_frequency(
        self,
        cells: List[GridCell],
        rows: Dict[int, List[GridCell]],
    ) -> Tuple[Optional[str], str]:
        """Locate the «Порядок уплаты страховой премии» block and read the
        marked frequency variant.

        Layout (юр.лицо CASCO):
          R31: C4='Единовременно' | C5='В рассрочку'
          R32: C5='ежеквартально'
          R33: C5='2 раза в год'    → biannual (within-year installment)
          R34: C5='ежегодно'
          R35: C5='прочее (укажите)' → out of enum
        A mark (e.g. «Х») sits in column 6 next to the selected row, or in
        column 5 next to «Единовременно» / column 4 if the entire row is the
        selection. IP templates shift these rows by +1.
        """
        anchor: Optional[GridCell] = None
        for cell in cells:
            if "порядок" in cell.normalized and "уплат" in cell.normalized:
                anchor = cell
                break
        if anchor is None:
            return None, ""

        # The anchor lives in row N; option labels sit in rows N..N+4.
        for row_offset in range(0, 5):
            row_cells = rows.get(anchor.row + row_offset, [])
            row_by_col = {c.col: c for c in row_cells}
            # Find label cell in cols 4..5 and check whether any cell in cols
            # 4..7 of the same row carries a non-template mark («Х», «X», «+»).
            label_cell = row_by_col.get(5) or row_by_col.get(4)
            if not label_cell:
                continue
            label_value = normalize_premium_frequency_label(label_cell.value)
            if label_value is None:
                continue
            marked = False
            for col in (4, 5, 6, 7):
                cell = row_by_col.get(col)
                if cell is None or cell is label_cell:
                    continue
                normalized = normalize_text(cell.value).strip()
                if normalized in {"х", "x", "+", "v"}:
                    marked = True
                    break
            if marked:
                return label_value, label_cell.coordinate
        return None, ""

    def _build_warnings(self, data: Dict[str, Any], insured_objects: List[Dict[str, str]]) -> List[Dict[str, str]]:
        warnings: List[Dict[str, str]] = []
        required_checks = [
            ("client_name", MISSING_CLIENT, "Клиент не распознан."),
            ("inn", "", "ИНН не распознан."),
            ("dfa_number", MISSING_DFA, "Номер ДФА не распознан."),
            ("branch", "", "Филиал не распознан."),
            ("manager_name", "", "Менеджер не распознан."),
            ("vehicle_info", MISSING_VEHICLE, "Предмет лизинга не распознан."),
        ]
        for field_name, missing_value, message in required_checks:
            if data.get(field_name) == missing_value or not data.get(field_name):
                warnings.append({"level": "manual_required", "field": field_name, "message": message, "source": ""})

        branch_value = data.get("branch")
        if branch_value and branch_value not in AVAILABLE_BRANCHES:
            warnings.append(
                {
                    "level": "manual_required",
                    "field": "branch",
                    "message": f"Филиал «{branch_value}» не входит в список известных — выберите вручную.",
                    "source": "",
                }
            )

        if not data.get("insured_party"):
            warnings.append(
                {
                    "level": "manual_required",
                    "field": "insured_party",
                    "message": "Страхователь не распознан — отметьте «Х» под нужным вариантом или выберите вручную.",
                    "source": "",
                }
            )

        if (
            data.get("insurance_type") == "страхование имущества"
            and not data.get("property_location_right_holder")
        ):
            warnings.append(
                {
                    "level": "manual_required",
                    "field": "property_location_right_holder",
                    "message": "Правообладатель места расположения не распознан — отметьте «Х» под нужным вариантом или выберите вручную.",
                    "source": "",
                }
            )

        if not data.get("insurance_period"):
            warnings.append(
                {
                    "level": "warning",
                    "field": "insurance_period",
                    "message": "Срок страхования не удалось привести к стандартному значению.",
                    "source": "",
                }
            )

        if insured_objects:
            warnings.append(
                {
                    "level": "info",
                    "field": "insured_objects",
                    "message": f"Найдено объектов: {len(insured_objects)}.",
                    "source": "",
                }
            )

        return warnings

    def _build_notes(self, warnings: List[Dict[str, str]]) -> str:
        if not warnings:
            return "Создано через Parser V2. Данные распознаны без критичных предупреждений."
        lines = ["Создано через Parser V2. Проверьте предупреждения разбора:"]
        lines.extend(f"- {warning['message']}" for warning in warnings if warning.get("level") != "info")
        return "\n".join(lines).strip()

    def _calculate_confidence(self, data: Dict[str, Any], insured_objects: List[Dict[str, str]]) -> float:
        checks = [
            bool(data.get("client_name") and data["client_name"] != MISSING_CLIENT),
            bool(data.get("inn")),
            bool(data.get("dfa_number") and data["dfa_number"] != MISSING_DFA),
            data.get("branch") in AVAILABLE_BRANCHES,
            bool(data.get("manager_name")),
            bool(insured_objects),
            data.get("insurance_type") != "другое",
            bool(data.get("insurance_period")),
        ]
        return round(sum(checks) / len(checks), 2)

    def _vehicle_summary(self, insured_objects: List[Dict[str, str]]) -> str:
        if not insured_objects:
            return MISSING_VEHICLE
        descriptions = [obj["description"] for obj in insured_objects if obj.get("description")]
        if len(descriptions) == 1:
            return descriptions[0][:1000]
        return "\n".join(descriptions[:8])[:1000]

    def _detect_application_format(self, data: Dict[str, Any]) -> str:
        if data.get("insurance_type") == "страхование имущества":
            return "property"
        return "casco_equipment"

    def _detect_application_type(self, cells: List[GridCell]) -> str:
        text = " ".join(cell.normalized for cell in cells[:120])
        if "индивидуальн" in text or re.search(r"\bип\b", text):
            return "individual_entrepreneur"
        return "legal_entity"

    def _matches_any_group(self, text: str, groups: List[Tuple[str, ...]]) -> bool:
        return any(all(normalize_text(word) in text for word in group) for group in groups)

    def _row_matches(self, row_cells: List[GridCell], labels: List[str]) -> bool:
        row_norm = normalize_text(self._row_text(row_cells))
        return any(label in row_norm for label in labels)

    def _inline_value_after_label(self, value: str) -> str:
        if ":" not in value:
            return ""
        before, after = value.split(":", 1)
        if len(normalize_text(before)) > 60:
            return ""
        after = clean_value(after)
        return after if len(after) > 1 else ""

    def _first_value_right(self, rows: Dict[int, List[GridCell]], cell: GridCell, max_offset: int = 8) -> Optional[GridCell]:
        same_row = rows.get(cell.row, [])
        for offset in range(1, max_offset + 1):
            candidate = next((item for item in same_row if item.col == cell.col + offset), None)
            if candidate and not self._looks_like_empty_or_label(candidate.value):
                return candidate
        return None

    def _first_value_below(self, rows: Dict[int, List[GridCell]], cell: GridCell, max_rows: int = 3) -> Optional[GridCell]:
        for row_number in range(cell.row + 1, cell.row + max_rows + 1):
            same_column = next((item for item in rows.get(row_number, []) if item.col == cell.col), None)
            if same_column and not self._looks_like_empty_or_label(same_column.value):
                return same_column
        return None

    def _looks_like_empty_or_label(self, value: str) -> bool:
        normalized = normalize_text(value)
        if not normalized:
            return True
        label_words = [
            "страхователь",
            "лизингополучатель",
            "менеджер",
            "инн",
            "филиал",
            "объект",
            "предмет",
            "наименование",
        ]
        if normalized in label_words:
            return True
        # A raw cell that still ends with «:» almost always carries a label,
        # not a value («Юридический адрес:», «Почтовый адрес:» и т.п.).
        raw = clean_value(value)
        if raw.endswith(":") and len(raw) <= 40:
            return True
        # Common form-section headers that should never be returned as values.
        if normalized in {
            "юридический адрес",
            "почтовый адрес",
            "фактический адрес",
            "основной вид деятельности",
            "вид деятельности",
            "параметры страховой сделки",
            "дата рождения",
            "дата подачи",
        }:
            return True
        return False

    def _looks_like_client_name(self, value: str) -> bool:
        normalized = normalize_text(value)
        if len(normalized) < 3:
            return False
        blocked_values = {
            "лизингодатель",
            "лизингополучатель",
            "страхователь",
            "клиент",
            "наименование лизингополучателя",
        }
        if normalized in blocked_values:
            return False
        return "адрес" not in normalized

    def _cell_by_coordinate(self, cells: List[GridCell], coordinate: str) -> Optional[GridCell]:
        if not coordinate:
            return None
        return next((cell for cell in cells if cell.coordinate.upper() == coordinate.upper()), None)

    def _normalize_inn(self, value: str) -> str:
        match = re.search(r"\b\d{10}(\d{2})?\b", clean_value(value))
        return match.group(0) if match else ""

    def _row_text(self, row_cells: List[GridCell]) -> str:
        return " ".join(cell.value for cell in row_cells if cell.value).strip()

    def _row_source(self, row_cells: List[GridCell]) -> str:
        if not row_cells:
            return ""
        return f"{row_cells[0].coordinate}:{row_cells[-1].coordinate}"

    def _is_object_header_or_label(self, row_norm: str) -> bool:
        if any(label in row_norm for label in ["предмет лизинга", "объект страхования"]):
            return True
        header_hits = ["наименование", "год", "стоимость", "vin", "заводской", "серийный"]
        return sum(1 for item in header_hits if item in row_norm) >= 2

    def _is_object_template_row(self, row_norm: str) -> bool:
        return any(marker in row_norm for marker in OBJECT_TEMPLATE_ROW_MARKERS)

    def _is_object_stop_row(self, row_norm: str) -> bool:
        return any(
            marker in row_norm
            for marker in [
                "условия страхования",
                "страховая сумма",
                "франшиз",
                "порядок оплаты",
                "график платеж",
                "дополнительные условия",
                "противоугонные системы и оборудование",
                # Property-form footer block («Дополнительные виды страхования
                # оборудования»): without these markers the parser slurps the
                # transportation/installation rows as fake insured objects.
                "дополнительные виды страхования",
                "перевозка с погрузкой",
                "пункт отправления",
                "пункт назначения",
                "ориентировочный срок перевозки",
                "строительно-монтажн",
                "строительно монтажн",
            ]
        )

    def _year_from_text(self, text: str) -> str:
        match = re.search(r"\b(19\d{2}|20\d{2})\b", clean_value(text))
        return match.group(1) if match else ""

    def _contains_any(self, cells: List[GridCell], markers: List[str]) -> bool:
        full_text = " ".join(cell.normalized for cell in cells)
        return any(normalize_text(marker) in full_text for marker in markers)

    def _first_source_containing(self, cells: List[GridCell], markers: List[str]) -> str:
        normalized_markers = [normalize_text(marker) for marker in markers]
        for cell in cells:
            if any(marker in cell.normalized for marker in normalized_markers):
                return cell.coordinate
        return ""
